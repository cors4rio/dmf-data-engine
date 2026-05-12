import os
import re
import datetime
import xlrd
import openpyxl

def carregar_dp_nao_avancado():
    """
    Lê o arquivo DP NAO.txt mapeando tanto por código quanto por pedaços limpos do nome.
    Garante altíssima resiliência para identificar exceções mesmo com divergência de código.
    """
    arquivo = os.path.join("config", "nao_faz_setor", "DP NAO.txt")
    excecoes_cod = {}
    excecoes_nome = {}
    
    if os.path.exists(arquivo):
        with open(arquivo, 'r', encoding='utf-8', errors='ignore') as f:
            for linha in f:
                linha = linha.strip()
                if not linha or "EMPRESAS QUE" in linha: continue
                
                # Identifica se é consultoria com valor fixo
                é_consultoria = "1:30" in linha
                valor_excecao = 1.5 / 24.0 if é_consultoria else "DP NÃO"
                
                # Limpeza da linha para extrair código e nome
                linha_limpa = linha.replace(';', ' ').replace('\t', ' ')
                partes = linha_limpa.split(' ', 1)
                
                cod_str = None
                nome_str = linha_limpa
                
                if partes and partes[0].isdigit():
                    cod_str = str(int(partes[0]))
                    nome_str = partes[1] if len(partes) > 1 else ""
                else:
                    # Tenta extrair primeiro número se houver
                    m = re.search(r'^(\d+)', linha_limpa)
                    if m:
                        cod_str = str(int(m.group(1)))
                        nome_str = linha_limpa[m.end():].strip()
                        
                if cod_str:
                    excecoes_cod[cod_str] = valor_excecao
                    
                # Limpa o nome para indexação textual
                nome_idx = re.sub(r'\(.*?\)', '', nome_str).strip().upper()
                # Remove sufixos comuns para maximizar match
                for sufixo in [" LTDA", " S.A.", " S/A", " ME", " EPP", " SCP", ".-"]:
                    nome_idx = nome_idx.replace(sufixo, "")
                nome_idx = nome_idx.strip()
                
                if nome_idx and len(nome_idx) > 3:
                    excecoes_nome[nome_idx] = valor_excecao
                    
    return excecoes_cod, excecoes_nome

def normalizar_cnpj(cnpj):
    if cnpj is None: return ""
    return re.sub(r'\D', '', str(cnpj)).zfill(14)

def celula_e_numerica(val):
    """
    Verifica se o valor da célula é numérico/tempo contábil válido.
    Retorna False para strings/textos como 'DP NÃO', 'NAO FAZ CONTABIL'.
    """
    if val is None: return False
    if isinstance(val, (float, int, datetime.time, datetime.timedelta, datetime.datetime)):
        return True
    if isinstance(val, str):
        # Se for string puramente numérica ou no formato de tempo H:M:S
        val_str = val.strip()
        if not val_str: return False
        if re.match(r'^\d+:\d+:\d+$', val_str): return True
        try:
            float(val_str)
            return True
        except ValueError:
            return False
    return False

def revisar_folha_e_totais():
    print("Iniciando revisão completa da Folha (Mês 04.2026) e recálculo seguro da Coluna R...")
    
    excecoes_cod, excecoes_nome = carregar_dp_nao_avancado()
    
    # 1. Carregar dados da Carol
    caminho_carol = os.path.join("ENTRADAS_MANUAIS", "Controle de Empregados (CAROL).xls")
    wb_carol = xlrd.open_workbook(caminho_carol)
    sh_carol = wb_carol.sheet_by_index(0)
    
    clientes_carol = {}
    for r in range(2, sh_carol.nrows):
        row = sh_carol.row_values(r)
        if not row or len(row) < 2 or not row[1]: continue
        try:
            cod_str = str(int(float(str(row[1]).strip())))
            nome = str(row[3]).strip() if len(row) > 3 else ""
            cnpj = normalizar_cnpj(str(row[5]).strip() if len(row) > 5 else "")
            
            func = int(float(row[7] or 0)) if len(row) > 7 and str(row[7]).strip() else 0
            est = int(float(row[9] or 0)) if len(row) > 9 and str(row[9]).strip() else 0
            cont = int(float(row[11] or 0)) if len(row) > 11 and str(row[11]).strip() else 0
            total_ativos = func + est + cont
            
            # Define o valor esperado aplicando a regra de negócio e checando exceções
            esperado = None
            if cod_str in excecoes_cod:
                esperado = excecoes_cod[cod_str]
            else:
                # Checa por nome
                nome_limpo = nome.upper()
                for sufixo in [" LTDA", " S.A.", " S/A", " ME", " EPP", " SCP"]:
                    nome_limpo = nome_limpo.replace(sufixo, "")
                nome_limpo = nome_limpo.strip()
                
                for n_ex, val_ex in excecoes_nome.items():
                    if n_ex in nome_limpo or nome_limpo in n_ex:
                        esperado = val_ex
                        break
                        
            if esperado is None:
                if total_ativos > 0:
                    esperado = ((total_ativos * 0.33) + 1.5) / 24.0
                else:
                    esperado = (5.0 / 60.0) / 24.0 # 5 minutos
                    
            clientes_carol[cod_str] = {
                'nome': nome,
                'cnpj': cnpj,
                'total_ativos': total_ativos,
                'esperado': esperado
            }
        except Exception:
            continue
            
    print(f"Total de clientes extraídos da Carol: {len(clientes_carol)}")
    
    # 2. Carregar Master
    caminho_master = "CONTROLE DE HORAS DMF.xlsm"
    wb_master = openpyxl.load_workbook(caminho_master, data_only=False, keep_vba=True)
    ws_master = wb_master["04.2026"]
    max_r = ws_master.max_row
    
    # Mapear as linhas da Master para garantir que encontramos todas as filiais/linhas de cada cliente
    mapa_linhas_cod = {}
    mapa_linhas_cnpj = {}
    mapa_linhas_nome = {}
    
    for r in range(10, max_r + 1):
        cod = ws_master.cell(r, 8).value
        cnpj = ws_master.cell(r, 10).value
        nome = str(ws_master.cell(r, 11).value or "").strip().upper()
        
        if cod is not None:
            try:
                c_str = str(int(float(str(cod).strip())))
                if c_str not in mapa_linhas_cod: mapa_linhas_cod[c_str] = []
                mapa_linhas_cod[c_str].append(r)
            except ValueError:
                pass
                
        if cnpj is not None:
            cnpj_norm = normalizar_cnpj(cnpj)
            if cnpj_norm:
                if cnpj_norm not in mapa_linhas_cnpj: mapa_linhas_cnpj[cnpj_norm] = []
                mapa_linhas_cnpj[cnpj_norm].append(r)
                
        if nome:
            # Pega os primeiros 15 caracteres para indexação textual
            n_chave = nome[:15]
            if n_chave not in mapa_linhas_nome: mapa_linhas_nome[n_chave] = []
            mapa_linhas_nome[n_chave].append(r)
            
    # 3. Processar Lançamentos de Folha (Coluna Q)
    linhas_atualizadas_dp = 0
    clientes_encontrados = 0
    
    for cod_str, dados in clientes_carol.items():
        linhas = []
        if cod_str in mapa_linhas_cod:
            linhas.extend(mapa_linhas_cod[cod_str])
        elif dados['cnpj'] and dados['cnpj'] in mapa_linhas_cnpj:
            linhas.extend(mapa_linhas_cnpj[dados['cnpj']])
            
        linhas = list(set(linhas))
        
        # Fallback textual extremo
        if not linhas and dados['nome']:
            n_carol = dados['nome'].strip().upper()[:15]
            if n_carol and n_carol in mapa_linhas_nome:
                linhas.extend(mapa_linhas_nome[n_carol])
                
        linhas = list(set(linhas))
        
        if linhas:
            clientes_encontrados += 1
            for r in linhas:
                c = ws_master.cell(row=r, column=17) # Coluna Q
                c.value = dados['esperado']
                if isinstance(dados['esperado'], float):
                    c.number_format = '[h]:mm:ss'
                linhas_atualizadas_dp += 1
                
    print(f"Revisão DP concluída: {clientes_encontrados} clientes devidos encontrados e {linhas_atualizadas_dp} linhas gravadas na Coluna Q.")
    
    # 4. Atualizar Coluna R (Total = Somatório Seguro sem Letras) para TODAS as linhas da Master
    linhas_totais_calculadas = 0
    for r in range(10, max_r + 1):
        # Verifica se a linha tem algum cliente ou dado
        tem_cadastro = any(ws_master.cell(r, col).value is not None for col in [8, 10, 11])
        if not tem_cadastro: continue
        
        val_o = ws_master.cell(r, 15).value # Fiscal
        val_p = ws_master.cell(r, 16).value # Contábil
        val_q = ws_master.cell(r, 17).value # DP
        
        num_o = celula_e_numerica(val_o)
        num_p = celula_e_numerica(val_p)
        num_q = celula_e_numerica(val_q)
        
        # Constrói a fórmula dinamicamente contendo APENAS as colunas numéricas
        colunas_validas = []
        if num_o: colunas_validas.append(f"O{r}")
        if num_p: colunas_validas.append(f"P{r}")
        if num_q: colunas_validas.append(f"Q{r}")
        
        c_total = ws_master.cell(row=r, column=18) # Coluna R
        
        if colunas_validas:
            c_total.value = "=" + "+".join(colunas_validas)
        else:
            # Se nenhuma for numérica (todas contêm letras ou estão vazias)
            c_total.value = ""
            
        c_total.number_format = '[h]:mm:ss'
        linhas_totais_calculadas += 1
        
    print(f"Coluna R recalcular com sucesso em {linhas_totais_calculadas} linhas (ignorando células com letras).")
    
    # 5. Atualizar Subtotais na Linha 7
    for col_idx, letra in [(15, 'O'), (16, 'P'), (17, 'Q'), (18, 'R')]:
        c_sub = ws_master.cell(row=7, column=col_idx)
        c_sub.value = f"=SUBTOTAL(9,{letra}10:{letra}{max_r})"
        c_sub.number_format = '[h]:mm:ss'
        
    print("Linha 7 atualizada com as fórmulas de SUBTOTAL totais.")
    
    # Salvar
    print(f"Salvando planilha final em: {caminho_master}")
    wb_master.save(caminho_master)
    print("SUCESSO: Processo finalizado com perfeição!")

if __name__ == "__main__":
    revisar_folha_e_totais()
