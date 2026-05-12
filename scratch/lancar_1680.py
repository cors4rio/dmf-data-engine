from engine.master_writer import MasterWriter
import openpyxl
import logging

def lancar_1680():
    master_path = r"C:\Users\DMF-AUTOMACAO\Downloads\CONTROLE_DE_HORAS_DMF_thR3zNWl.xlsm"
    cliente_cod = "1680"
    ativos = 24
    
    # Cálculo conforme modulos/dp.py
    horas = (ativos * 0.33) + 1.5
    valor_excel = horas / 24.0
    
    print(f"Lançando Cliente {cliente_cod}: {ativos} ativos -> {horas:.2f} horas ({valor_excel:.6f})")
    
    writer = MasterWriter(master_path)
    if writer.carregar():
        # Forçar a aba 04.2026 pois o writer pega a ativa
        if "04.2026" in writer.wb.sheetnames:
            writer.ws = writer.wb["04.2026"]
            writer._mapear_linhas() # Remapear para a nova aba
            
            print(f"Aba 04.2026 selecionada. Iniciando preenchimento...")
            writer.preencher_dp(cliente_cod, valor_excel)
            writer.recalcular_totais()
            
            if writer.salvar():
                print("SUCESSO: Cliente 1680 lançado com sucesso!")
            else:
                print("ERRO: Falha ao salvar a planilha.")
        else:
            # Se a aba for 042026 sem o ponto
            if "042026" in writer.wb.sheetnames:
                writer.ws = writer.wb["042026"]
                writer._mapear_linhas()
                writer.preencher_dp(cliente_cod, valor_excel)
                writer.recalcular_totais()
                writer.salvar()
                print("SUCESSO: Lançado na aba 042026.")
            else:
                print(f"ERRO: Aba 04.2026 não encontrada. Abas disponíveis: {writer.wb.sheetnames}")

if __name__ == "__main__":
    lancar_1680()
