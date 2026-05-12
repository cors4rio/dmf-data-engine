import os
import xlrd
import openpyxl
import re
import datetime

def carregar_dp_nao():
    arquivo = os.path.join("config", "nao_faz_setor", "DP NAO.txt")
    excecoes = {}
    if os.path.exists(arquivo):
        with open(arquivo, 'r', encoding='utf-8', errors='ignore') as f:
            for linha in f:
                linha = linha.strip()
                if not linha or "Não entra" in linha: continue
                partes = linha.replace(';', ' ').replace('\t', ' ').split(' ', 1)
                try:
                    cod_str = str(int(float(partes[0])))
                    if "1:30" in linha:
                        excecoes[cod_str] = 1.5 / 24.0
                    else:
                        excecoes[cod_str] = "DP NÃO"
                except Exception:
                    continue
    return excecoes

def normalizar_cnpj(cnpj):
    if cnpj is None: return ""
    return re.sub(r'\D', '', str(cnpj)).zfill(14)

def converter_para_fracao_dia(val):
    if val is None: return 0.0
    if isinstance(val, (float, int)): return float(val)
    if isinstance(val, datetime.time):
        segundos = val.hour * 3600 + val.minute * 60 + val.second
        return segundos / 86400.0
    if isinstance(val, datetime.timedelta):
        return val.total_seconds() / 86400.0
    if isinstance(val, str):
        # Tenta parsear "H:M:S"
        m = re.match(r'(\d+):(\d+):(\d+)', val.strip())
        if m:
            h, min_val, s = map(int, m.groups())
            return (h * 3600 + min_val * 60 + s) / 86400.0
    return val

def auditar():
    dp_nao = carregar_dp_nao()
    
    caminho_carol = os.path.join("ENTRADAS_MANUAIS", "Controle de Empregados (CAROL).xls")
    wb_carol = xlrd.open_workbook(caminho_carol)
    sh_carol = wb_carol.sheet_by_index(0)
    
    clientes_carol = {}
    for r in range(2, sh_carol.nrows):
        row = sh_carol.row_values(r)
        if not row or len(row) < 2 or not row[1]: continue
        try:
            cod_str = str(int(float(str(row[1]).strip())))
            nome = str(row[3]).strip() if len(row) > 3 else ""
            cnpj = str(row[5]).strip() if len(row) > 5 else ""
            
            func = int(float(row[7] or 0)) if len(row) > 7 and str(row[7]).strip() else 0
            est = int(float(row[9] or 0)) if len(row) > 9 and str(row[9]).strip() else 0
            cont = int(float(row[11] or 0)) if len(row) > 11 and str(row[11]).strip() else 0
            total_ativos = func + est + cont
            
            clientes_carol[cod_str] = {
                'nome': nome,
                'cnpj': normalizar_cnpj(cnpj),
                'cnpj_raw': str(row[5]).strip() if len(row) > 5 else "",
                'total_ativos': total_ativos
            }
        except Exception:
            continue
            
    caminho_master = "CONTROLE DE HORAS DMF.xlsm"
    wb_master = openpyxl.load_workbook(caminho_master, data_only=False)
    ws_master = wb_master["04.2026"]
    
    mapa_linhas_cod = {}
    mapa_linhas_cnpj = {}
    
    for r in range(10, ws_master.max_row + 1):
        cod = ws_master.cell(r, 8).value
        cnpj = ws_master.cell(r, 10).value
        
        if cod is not None:
            try:
                c_str = str(int(float(str(cod).strip())))
                if c_str not in mapa_linhas_cod: mapa_linhas_cod[c_str] = []
                mapa_linhas_cod[c_str].append(r)
            except ValueError:
                pass
                
        if cnpj is not None:
            cnpj_norm = normalizar_cnpj(cnpj)
            if cnpj_norm:
                if cnpj_norm not in mapa_linhas_cnpj: mapa_linhas_cnpj[cnpj_norm] = []
                mapa_linhas_cnpj[cnpj_norm].append(r)
                
    divergencias = []
    faltantes = []
    encontrados_corretos = 0
    
    MINIMO = (5.0 / 60.0) / 24.0
    
    for cod_str, dados in clientes_carol.items():
        if cod_str in dp_nao:
            esperado = dp_nao[cod_str]
        else:
            ativos = dados['total_ativos']
            if ativos > 0:
                esperado = ((ativos * 0.33) + 1.5) / 24.0
            else:
                esperado = MINIMO
                
        linhas = []
        if cod_str in mapa_linhas_cod:
            linhas.extend(mapa_linhas_cod[cod_str])
        elif dados['cnpj'] and dados['cnpj'] in mapa_linhas_cnpj:
            linhas.extend(mapa_linhas_cnpj[dados['cnpj']])
            
        linhas = list(set(linhas))
        
        if not linhas:
            # Tenta fallback buscando por substring no nome para ter certeza se falta mesmo
            for r in range(10, ws_master.max_row + 1):
                n_master = str(ws_master.cell(r, 11).value or "").upper()
                if dados['nome'] and len(dados['nome']) > 5 and dados['nome'][:10].upper() in n_master:
                    linhas.append(r)
            linhas = list(set(linhas))
            
        if not linhas:
            faltantes.append((cod_str, dados, esperado))
            continue
            
        for r in linhas:
            val_atual = ws_master.cell(r, 17).value # Coluna Q
            
            se_correto = False
            if isinstance(esperado, str):
                if str(val_atual).strip() == esperado:
                    se_correto = True
            else:
                val_float = converter_para_fracao_dia(val_atual)
                if isinstance(val_float, float):
                    if abs(val_float - esperado) < 1e-4:
                        se_correto = True
                        
            if se_correto:
                encontrados_corretos += 1
            else:
                divergencias.append(f"DIVERGÊNCIA Linha {r} | Cód {cod_str} ({dados['nome']}): Atual={val_atual} | Esperado={esperado} ({dados['total_ativos']} ativos)")
                
    print(f"\nResultado da Auditoria com Comparação Precisa:")
    print(f"Corretos: {encontrados_corretos}")
    print(f"Faltantes na Master: {len(faltantes)}")
    for f in faltantes:
        print(f" - FALTANTE: Cód {f[0]} | Nome: {f[1]['nome']} | CNPJ: {f[1]['cnpj_raw']} | Ativos: {f[1]['total_ativos']}")
        
    print(f"\nDivergências de Valor: {len(divergencias)}")
    for d in divergencias:
        print(f" - {d}")

if __name__ == "__main__":
    auditar()
