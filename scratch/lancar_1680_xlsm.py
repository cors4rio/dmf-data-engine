from engine.master_writer import MasterWriter
import openpyxl
import logging
import os
import re

def lancar_1680_xlsm():
    caminho_master = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\CONTROLE DE HORAS DMF.xlsm"
    
    cliente_cod = "1680"
    cnpj_completo = "04.273.258/0001-12"
    cnpj_sem_zero = "4273258000112" # Formato numérico salvo pelo Excel sem zero à esquerda
    ativos = 24
    
    horas = (ativos * 0.33) + 1.5
    valor_excel = horas / 24.0
    
    print(f"Lançando Cliente {cliente_cod}: {ativos} ativos -> {horas:.2f} horas na Master.")
    
    writer = MasterWriter(caminho_master)
    if writer.carregar():
        target_sheet = "04.2026"
        if target_sheet in writer.wb.sheetnames:
            writer.ws = writer.wb[target_sheet]
            writer._mapear_linhas()
            
            # Tenta preencher passando a variação sem zero à esquerda para dar match perfeito no row_map
            linhas = writer.obter_linhas(cliente_cod, cnpj_sem_zero)
            if not linhas:
                # Fallback manual de extrema resiliência buscando por substring na linha
                print("Aviso: Lookup padrão não retornou linhas. Realizando varredura de fallback...")
                for r in range(10, writer.ws.max_row + 1):
                    val_cnpj = str(writer.ws.cell(r, 10).value or "")
                    val_nome = str(writer.ws.cell(r, 11).value or "").upper()
                    if "4273258" in val_cnpj or "IRACI REGINA TEDESCO" in val_nome:
                        linhas.append(r)
                        # Aproveita e preenche o Cód Dominio na Coluna H para fixar o cadastro para o futuro
                        writer.ws.cell(r, 8, value=int(cliente_cod))
                        print(f"Corrigido Cód Dominio na linha {r} para {cliente_cod}.")
            
            if linhas:
                print(f"Linhas encontradas para o cliente 1680: {list(set(linhas))}")
                for r in set(linhas):
                    c = writer.ws.cell(row=r, column=17) # Coluna Q (DP)
                    c.value = valor_excel
                    c.number_format = '[h]:mm:ss'
                    print(f"Preenchido valor {valor_excel:.6f} na célula Q{r}.")
            else:
                print("ERRO CRÍTICO: Cliente 1680 não localizado nem por fallback.")
                return
                
            writer.recalcular_totais()
            if writer.salvar():
                print(f"SUCESSO: Planilha salva e atualizada com perfeição no modelo correto (.xlsm).")
            else:
                print("ERRO ao salvar a planilha.")
        else:
            print(f"ERRO: Aba {target_sheet} não encontrada.")

if __name__ == "__main__":
    lancar_1680_xlsm()
