import openpyxl
import datetime

def time_to_hours(time_str):
    try:
        if not time_str or str(time_str).strip() == '':
            return 0.0
        parts = str(time_str).strip().split(':')
        if len(parts) >= 3:
            h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
            return h + (m / 60.0) + (s / 3600.0)
    except:
        pass
    return 0.0

def analisar_txt():
    FILE_TXT = r"C:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\teste.txt"
    FILE_MASTER = r"C:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm"

    clientes_txt = {}
    total_horas_txt = 0.0
    
    # 1. Load txt
    with open(FILE_TXT, 'r', encoding='utf-8') as f:
        for line in f:
            if not line.strip(): continue
            parts = line.split('\t')
            if len(parts) >= 3:
                cod_raw = parts[0].strip()
                # Some lines have empty code but have name, or layout might vary
                # Actually, looking at the sample, it's code \t name \t \t \t \t time
                # Let's clean empty parts
                clean_parts = [p.strip() for p in parts if p.strip()]
                if len(clean_parts) >= 3:
                    cod_str = clean_parts[0]
                    nome = clean_parts[1]
                    tempo = clean_parts[-1]
                    
                    try:
                        cod_str = str(int(float(cod_str)))
                    except: pass
                    
                    horas = time_to_hours(tempo)
                    if horas > 0 and cod_str:
                        total_horas_txt += horas
                        if cod_str not in clientes_txt:
                            clientes_txt[cod_str] = {'nome': nome, 'horas': 0.0}
                        clientes_txt[cod_str]['horas'] += horas

    print(f"=== TOTAL NA ORIGEM (teste.txt) ===")
    print(f"Total Horas: {total_horas_txt:.2f}")
    print(f"Clientes com tempo > 0 agrupados por código: {len(clientes_txt)}")
    print()

    # 2. Ler Master
    wb_m = openpyxl.load_workbook(FILE_MASTER, data_only=True)
    sh_m = wb_m["12.2025"]
    
    clientes_master = {}
    total_horas_master = 0.0
    
    for row in range(10, sh_m.max_row + 1):
        cod_cell = sh_m.cell(row=row, column=8).value
        n_val = sh_m.cell(row=row, column=14).value
        nome_cell = sh_m.cell(row=row, column=11).value
        
        cod_str = ""
        try:
            if cod_cell is not None:
                cod_str = str(int(float(str(cod_cell).strip().split('.')[0])))
        except: pass
        
        horas_injetadas = 0.0
        if n_val:
            if isinstance(n_val, datetime.time):
                 horas_injetadas = n_val.hour + (n_val.minute / 60.0) + (n_val.second / 3600.0)
            elif isinstance(n_val, datetime.timedelta):
                 horas_injetadas = n_val.total_seconds() / 3600.0
            else:
                 horas_injetadas = float(n_val) * 24.0
            total_horas_master += horas_injetadas
            
        if cod_str:
            clientes_master[cod_str] = {
                'linha': row,
                'nome': str(nome_cell).strip(),
                'horas_n': horas_injetadas
            }

    print(f"=== TOTAL NO MASTER INJETADO (Col N) ===")
    print(f"Total Horas Injetadas: {total_horas_master:.2f}")
    print(f"Clientes contabilizados no master (com codigo): {len(clientes_master)}")
    print()

    # 3. Cruzamento
    faltantes = []
    horas_faltantes = 0.0
    lancados_corretamente = 0
    divergentes = []

    for cod, dados_txt in clientes_txt.items():
        if cod not in clientes_master:
            faltantes.append((cod, dados_txt['nome'], dados_txt['horas']))
            horas_faltantes += dados_txt['horas']
        else:
            h_m = getattr(clientes_master.get(cod), 'horas_n', 0.0) if type(clientes_master.get(cod)) == dict else clientes_master.get(cod).get('horas_n', 0.0)
            diff = abs(dados_txt['horas'] - h_m)
            if diff > 0.05: # Margem de erro de arredondamento de 3 minutos
                divergentes.append((cod, dados_txt['nome'], dados_txt['horas'], h_m, clientes_master[cod]['linha']))
            else:
                lancados_corretamente += 1

    faltantes.sort(key=lambda x: x[2], reverse=True)
    divergentes.sort(key=lambda x: abs(x[2]-x[3]), reverse=True)

    print(f"=== DIAGNÓSTICO FINAL ===")
    print(f"✅ Lançados Corretamente: {lancados_corretamente} clientes (Valor TXT bate com Master)")
    print(f"❌ Estão no TXT mas NÃO estão na Mestra (Faltantes): {len(faltantes)} clientes (Soma: {horas_faltantes:.2f}h perdidas)")
    print(f"⚠️ Estão na Mestra mas com valor DIVERGENTE: {len(divergentes)} clientes")
    
    if divergentes:
        print("\nLista de Divergentes (Valor TXT != Valor Master):")
        for d in divergentes:
            print(f"  Cod: {d[0]:<5} | Nome: {d[1]:<35} | LinhaMestra: {d[4]:<4} | TXT: {d[2]:.2f}h | Master: {d[3]:.2f}h")

    if faltantes:
        print("\nTop 20 Faltantes (Estão no TXT mas NÃO têm linha na planilha Mestra):")
        for f in faltantes[:20]:
            print(f"  Cod: {f[0]:<5} | Nome: {f[1]:<35} | Horas: {f[2]:.2f}h")
            
if __name__ == '__main__':
    analisar_txt()
