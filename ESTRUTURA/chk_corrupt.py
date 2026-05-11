import openpyxl; wb=openpyxl.load_workbook('CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx'); ws=wb['03.2026']; c=0; max_v = 1e6; 
for row in ws.iter_rows(min_row=10, max_row=ws.max_row, min_col=14, max_col=17):
    for cell in row:
        if type(cell.value) in (int, float) and abs(cell.value) > max_v: cell.value = 0; c+=1
wb.save('CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx'); print(f'Fixed {c} corrupt cells!')
