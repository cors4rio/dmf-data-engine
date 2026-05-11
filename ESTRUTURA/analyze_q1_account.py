import openpyxl
import os

FILE_ACCOUTING = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_.xlsx'

def analyze_accounting_q1():
    if not os.path.exists(FILE_ACCOUTING):
        print(f"File not found: {FILE_ACCOUTING}")
        return

    print(f"Analyzing {FILE_ACCOUTING}...")
    
    # Load with data_only=False to see formulas
    wb_form = openpyxl.load_workbook(FILE_ACCOUTING, data_only=False)
    ws_form = wb_form.active
    
    # Load with data_only=True to see values
    wb_val = openpyxl.load_workbook(FILE_ACCOUTING, data_only=True)
    ws_val = wb_val.active
    
    q1_val = ws_val['Q1'].value
    q1_form = ws_form['Q1'].value
    
    print(f"Value of Q1: {q1_val}")
    print(f"Formula of Q1: {q1_form}")
    
    print("\nHeader row (Row 1):")
    headers = [ws_val.cell(1, i).value for i in range(1, ws_val.max_column + 1)]
    print(headers)
    
    print("\nSample Row 2:")
    sample = [ws_val.cell(2, i).value for i in range(1, ws_val.max_column + 1)]
    print(sample)
    
    # Check if Q is a data column or a sum column
    # Let's see how many rows have data in column Q
    rows_with_data = 0
    total_q = 0
    for r in range(2, ws_val.max_row + 1):
        v = ws_val.cell(r, 17).value # 17 is Q
        if v is not None:
            rows_with_data += 1
            if isinstance(v, (int, float)):
                total_q += v
    
    print(f"\nRows with data in Column Q (from Row 2): {rows_with_data}")
    print(f"Manual sum of Column Q (if numeric): {total_q}")

if __name__ == "__main__":
    analyze_accounting_q1()
