import pyodbc
import openpyxl
import xlrd # Para ler o XLS legado da Carol
from datetime import datetime, time
import re
import os
import calendar
import zipfile

# --- MONKEY-PATCH: habilita Zip64 para salvar arquivos xlsm grandes ---
# Necessário porque o openpyxl por padrão não habilita allowZip64=True
_orig_ZipFile_init = zipfile.ZipFile.__init__
def _patched_ZipFile_init(self, *args, **kwargs):
    kwargs['allowZip64'] = True
    _orig_ZipFile_init(self, *args, **kwargs)
zipfile.ZipFile.__init__ = _patched_ZipFile_init

# --- CONFIGURAÇÕES ---
MES_ALVO = 2
ANO_ALVO = 2026
DB_CONFIG = "DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>"
PLANILHA_MASTER = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm"

# Fonte de dados Fallback (Opção 2 - Carol)
PLANILHA_CAROL = rf"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\Controle de Empregados (CAROL)V2.{MES_ALVO:02d}{ANO_ALVO}.xls"

# --- REGRAS DE NEGÓCIO ---
DP_MINIMO_MINUTOS = 5
DP_CONTRIBUINTE_UNICO_TIME = "1:10:00" # Regra específica: 1 Contribuinte apenas = 1h10 (Formatado para Excel)
FATOR_ADICIONAL_FISCAL = 1.70

def clean_cnpj(cnpj):
    if not cnpj: return ""
    return re.sub(r'\D', '', str(cnpj))

def format_seconds_to_hhmmss(seconds):
    if seconds is None: return "0:00:00"
    seconds = int(round(seconds))
    hrs = seconds // 3600
    mins = (seconds % 3600) // 60
    secs = seconds % 60
    return f"{hrs}:{mins:02d}:{secs:02d}"

def seconds_to_excel(total_seconds):
    # O Excel trata 1.0 como 24 horas
    return total_seconds / 86400.0

def get_fiscal_data(start_date, end_date):
    """Extrai produtividade fiscal do Domínio via SQL."""
    print(f"Extraindo dados fiscais ({start_date} a {end_date})...")
    conn = pyodbc.connect(DB_CONFIG)
    cursor = conn.cursor()
    
    query = f"""
    SELECT 
        e.codi_emp,
        e.cgce_emp,
        SUM(DATEDIFF(second, 
            YMD(YEAR(l.data_log), MONTH(l.data_log), DAY(l.data_log)) + l.tini_log, 
            COALESCE(YMD(YEAR(l.dfim_log), MONTH(l.dfim_log), DAY(l.dfim_log)) + l.tfim_log, 
                     YMD(YEAR(l.data_log), MONTH(l.data_log), DAY(l.data_log)) + l.tfim_log)
        )) as total_segundos
    FROM bethadba.geloguser l
    INNER JOIN bethadba.geempre e ON e.codi_emp = l.codi_emp
    WHERE l.sist_log = 5 
      AND l.data_log >= '{start_date}'
      AND l.data_log <= '{end_date}'
      AND l.tfim_log IS NOT NULL
    GROUP BY e.codi_emp, e.cgce_emp
    """
    
    cursor.execute(query)
    data = {}
    for row in cursor.fetchall():
        cnpj = clean_cnpj(row.cgce_emp)
        # Adicional de 70% conforme Spec
        segundos_com_adicional = float(row.total_segundos) * FATOR_ADICIONAL_FISCAL
        data[row.codi_emp] = segundos_com_adicional
        if cnpj:
            data[cnpj] = segundos_com_adicional
    
    conn.close()
    return data

def get_folha_data_from_carol(file_path):
    """Lê ativos da planilha da Carol (Opção 2) e retorna segundos."""
    print(f"Lendo planilha da Carol: {file_path}")
    if not os.path.exists(file_path):
        print(f"AVISO: Planilha da Carol não encontrada.")
        return {}
    
    wb = xlrd.open_workbook(file_path)
    sh = wb.sheet_by_index(0)
    data = {}
    
    # Mapeamento de colunas (conforme inspeção):
    # 1: Código, 7: Funcionários, 9: Estagiários, 11: Contribuintes
    for i in range(1, sh.nrows): # Pula cabeçalho
        row = sh.row_values(i)
        try:
            # Garante que o código é um inteiro
            codi_raw = row[1]
            if not codi_raw: continue
            codi = str(int(float(codi_raw)))
            
            func = float(row[7] or 0)
            estag = float(row[9] or 0)
            contrib = float(row[11] or 0)
            
            total_ativos = func + estag + contrib
            
            # Regra Especial: Apenas 1 Contribuinte (e 0 outros) = 01:10:00 (4200 segundos)
            if contrib == 1 and func == 0 and estag == 0:
                segundos = 4200 # 1h10
            elif total_ativos > 0:
                horas = (total_ativos * 0.33) + 1.5
                segundos = horas * 3600
            else:
                segundos = DP_MINIMO_MINUTOS * 60
                
            data[codi] = segundos
            # CNPJ fallback (Idx 5)
            cnpj = clean_cnpj(row[5])
            if cnpj: data[cnpj] = segundos
            
        except (ValueError, IndexError):
            continue
            
    return data

def load_exceptions():
    """Carrega listas de exceções (DP NÃO, CONSULTORIA, CONTABIL NÃO)."""
    dp_nao = set()
    dp_consultoria = set()
    contabil_nao = set()
    
    # DP NAO
    path_dp = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\nao_faz_setor\DP NAO.txt"
    if os.path.exists(path_dp):
        with open(path_dp, 'r', encoding='latin-1') as f:
            for line in f:
                content = line.strip().upper()
                if not content or content.startswith('#'): continue
                if "(FAZ CONSULTORIA" in content:
                    match = re.search(r'^(\d+)', content)
                    if match: dp_consultoria.add(match.group(1))
                else:
                    match = re.search(r'^(\d+)', content)
                    if match: dp_nao.add(match.group(1))
                    if len(content) > 5: dp_nao.add(content) # Nome como fallback

    # CONTABIL NAO
    path_cont = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\nao_faz_setor\NAO FAZ CONTABIL.txt"
    if os.path.exists(path_cont):
        with open(path_cont, 'r', encoding='latin-1') as f:
            for line in f:
                content = line.strip().upper()
                if not content or content.startswith('#'): continue
                match = re.search(r'^(\d+)', content)
                if match: contabil_nao.add(match.group(1))
                if len(content) > 5: contabil_nao.add(content)
                    
    return dp_nao, dp_consultoria, contabil_nao

def main():
    aba_atual = f"{MES_ALVO:02d}.{ANO_ALVO}"
    # Aba anterior
    mes_ant = MES_ALVO - 1 if MES_ALVO > 1 else 12
    ano_ant = ANO_ALVO if MES_ALVO > 1 else ANO_ALVO - 1
    aba_anterior = f"{mes_ant:02d}.{ano_ant}"
    
    _, last_day = calendar.monthrange(ANO_ALVO, MES_ALVO)
    start_date = f"{ANO_ALVO}-{MES_ALVO:02d}-01"
    end_date = f"{ANO_ALVO}-{MES_ALVO:02d}-{last_day:02d}"

    print(f"--- Processamento {aba_atual} ---")
    
    # 1. Carregar dados de fontes externas
    fisc_data = get_fiscal_data(start_date, end_date)
    folha_data = get_folha_data_from_carol(PLANILHA_CAROL)
    dp_nao, dp_consult, cont_nao = load_exceptions()
    
    # 2. Abrir Planilha Master
    print(f"Abrindo Master: {PLANILHA_MASTER}")
    wb = openpyxl.load_workbook(PLANILHA_MASTER, keep_vba=True)
    if aba_atual not in wb.sheetnames:
        print(f"ERRO: Aba {aba_atual} não encontrada.")
        return
        
    ws = wb[aba_atual]
    ws_ant = wb[aba_anterior] if aba_anterior in wb.sheetnames else None
    
    # 3. Mapear aba anterior para Backfill
    hist_fiscal = {}
    if ws_ant:
        print(f"Mapeando backfill da aba {aba_anterior}...")
        for r in range(10, ws_ant.max_row + 1):
            c = str(ws_ant.cell(row=r, column=8).value or "").split('.')[0]
            j = clean_cnpj(ws_ant.cell(row=r, column=10).value)
            val = ws_ant.cell(row=r, column=15).value # Col O
            if val is not None:
                if c: hist_fiscal[c] = val
                if j: hist_fiscal[j] = val

    # 4. Processar linhas
    # H=Codigo(8), I=Nome(9), J=CNPJ(10), K=Razao(11), N=Mes Ant(14), O=Fiscal(15), P=Contabil(16), Q=Folha(17), R=Total(18)
    processed_count = 0
    for row_idx in range(10, ws.max_row + 1):
        cod_raw = ws.cell(row=row_idx, column=8).value
        if not cod_raw: continue
        cod = str(cod_raw).split('.')[0]
        cnpj = clean_cnpj(ws.cell(row=row_idx, column=10).value)
        nome_fan = str(ws.cell(row=row_idx, column=9).value or "").strip().upper()
        nome_raz = str(ws.cell(row=row_idx, column=11).value or "").strip().upper()
        
        # Step A: Backfill (Col N)
        val_bf = hist_fiscal.get(cod) or hist_fiscal.get(cnpj)
        if val_bf is not None:
            ws.cell(row=row_idx, column=14).value = val_bf
            ws.cell(row=row_idx, column=14).number_format = '[h]:mm:ss'

        # Step B: Fiscal (Col O)
        val_f_sec = fisc_data.get(cod) or fisc_data.get(cnpj)
        if val_f_sec is not None:
            ws.cell(row=row_idx, column=15).value = seconds_to_excel(val_f_sec)
            ws.cell(row=row_idx, column=15).number_format = '[h]:mm:ss'
            
        # Step C: Contábil (Col P) - Exceções
        if cod in cont_nao or nome_fan in cont_nao or nome_raz in cont_nao:
            ws.cell(row=row_idx, column=16).value = "NAO FAZ CONTABIL"
            
        # Step D: Folha/DP (Col Q)
        if cod in dp_consult:
            ws.cell(row=row_idx, column=17).value = seconds_to_excel(5400) # 1:30
            ws.cell(row=row_idx, column=17).number_format = '[h]:mm:ss'
        elif cod in dp_nao or nome_fan in dp_nao or nome_raz in dp_nao:
            ws.cell(row=row_idx, column=17).value = "DP NÃO"
        else:
            val_dp_sec = folha_data.get(cod) or folha_data.get(cnpj)
            if val_dp_sec is not None:
                ws.cell(row=row_idx, column=17).value = seconds_to_excel(val_dp_sec)
                ws.cell(row=row_idx, column=17).number_format = '[h]:mm:ss'

        # Step E: Total (Col R)
        # Fórmula robusta: ignora células com texto (DP NÃO, NAO FAZ CONTABIL)
        # Usa IF(ISNUMBER()) para só somar valores numéricos e evitar #VALOR!
        formula_r = (
            f"=IF(ISNUMBER(O{row_idx}),O{row_idx},0)"
            f"+IF(ISNUMBER(P{row_idx}),P{row_idx},0)"
            f"+IF(ISNUMBER(Q{row_idx}),Q{row_idx},0)"
        )
        ws.cell(row=row_idx, column=18).value = formula_r
        ws.cell(row=row_idx, column=18).number_format = '[h]:mm:ss'
        processed_count += 1

    # --- FÓRMULAS ACUMULADORAS (LINHA 7) ---
    # Soma todos os valores numéricos das colunas O, Q e R (linhas 10 em diante)
    # A função SUM do Excel ignora textos automaticamente
    last_row = ws.max_row
    
    # O7 = Total Fiscal
    ws.cell(row=7, column=15).value = f"=SUM(O10:O{last_row})"
    ws.cell(row=7, column=15).number_format = '[h]:mm:ss'
    
    # Q7 = Total Folha/DP
    ws.cell(row=7, column=17).value = f"=SUM(Q10:Q{last_row})"
    ws.cell(row=7, column=17).number_format = '[h]:mm:ss'
    
    # R7 = Grand Total
    ws.cell(row=7, column=18).value = f"=SUM(R10:R{last_row})"
    ws.cell(row=7, column=18).number_format = '[h]:mm:ss'

    wb.save(PLANILHA_MASTER)
    print(f"Sucesso! {processed_count} linhas processadas e salvas.")

if __name__ == "__main__":
    main()
