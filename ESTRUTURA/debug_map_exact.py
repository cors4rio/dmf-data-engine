import openpyxl
import re

BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = BASE_DIR + r"\CONTROLE DE HORAS DMF - CLIENTE.xlsm"

wb = openpyxl.load_workbook(PLANILHA_MASTER, data_only=True)
ws = wb['03.2026']

def clean_cnpj(cnpj):
    if not cnpj: return ""
    return re.sub(r'\D', '', str(cnpj))

mapa_planilha = {}
mapa_planilha_cnpj = {}

for row in range(10, ws.max_row + 1):
    cod_cell = ws.cell(row=row, column=8).value # Col H
    cnpj_cell = ws.cell(row=row, column=10).value # Col J
    
    if cod_cell and str(cod_cell).strip().isdigit():
        mapa_planilha[int(cod_cell)] = row
    if cnpj_cell:
        mapa_planilha_cnpj[clean_cnpj(cnpj_cell)] = row

print("Did we map 1152 in mapa_planilha?", 1152 in mapa_planilha)
if not (1152 in mapa_planilha):
    for r in range(10, ws.max_row + 1):
        c = ws.cell(row=r, column=8).value
        # If it looks like 1152 ...
        if c == 1152 or str(c).strip() == '1152' or str(c).strip() == '1152.0':
            print(f"Found something like 1152 at row {r}: value={c}, type={type(c)}, isdigit={str(c).strip().isdigit()}")
