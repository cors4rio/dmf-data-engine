import openpyxl
import os

FILE_ACCOUTING = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_.xlsx'

def deep_analyze_account():
    if not os.path.exists(FILE_ACCOUTING):
        print("Arquivo contábil não encontrado.")
        return

    print(f"Analisando {FILE_ACCOUTING}...")
    wb = openpyxl.load_workbook(FILE_ACCOUTING, data_only=False)
    ws = wb['03.2026']
    
    # Inspecionar as primeiras 5 linhas de todas as colunas
    print("\nEstrutura do Cabeçalho (Linhas 1-2):")
    for r in range(1, 3):
        row_vals = []
        for c in range(1, 25): # A até X
            cell = ws.cell(r, c)
            row_vals.append(f"{cell.coordinate}:{cell.value}")
        print(f"Row {r}: {row_vals}")

    # Verificar se existe algum cálculo na coluna Q
    print("\nVerificando fórmulas na coluna Q (17):")
    for r in range(1, 20):
        cell = ws.cell(r, 17)
        if cell.value and (isinstance(cell.value, str) and cell.value.startswith('=')):
            print(f"  [!] Fórmula encontrada em {cell.coordinate}: {cell.value}")
        elif cell.value:
            print(f"  Valor em {cell.coordinate}: {cell.value}")

    # Verificar se há dados lá embaixo
    print("\nVerificando dados no final da planilha:")
    last_row = ws.max_row
    print(f"Max row: {last_row}")
    for r in range(max(1, last_row - 10), last_row + 1):
        v = ws.cell(r, 17).value
        if v:
            print(f"  Linha {r}, Col Q: {v}")

if __name__ == "__main__":
    deep_analyze_account()
