import openpyxl
from datetime import timedelta

# Arquivos
ARQUIVO_ORIGEM = r"C:\Users\DMF-AUTOMACAO\Downloads\CONTROLE_DE_HORAS_DMF_BL4t520G.xlsm"
ABA_ORIGEM = "02.2026"
ARQUIVO_DESTINO = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx"
ABA_DESTINO = "03.2026"

print("Lendo arquivo de origem (Pode demorar por causa do data_only)...")
wb_origem = openpyxl.load_workbook(ARQUIVO_ORIGEM, data_only=True)
ws_origem = wb_origem[ABA_ORIGEM]

# Construir mapa de horas do mês 02
mapa_horas_02 = {}
for row in range(10, ws_origem.max_row + 1):
    cod_cell = ws_origem.cell(row=row, column=8).value
    val_cell = ws_origem.cell(row=row, column=15).value # Coluna O possui as horas Fiscais
    
    if cod_cell is not None:
        try:
            cod = int(float(str(cod_cell).strip()))
            
            # Limpar fantasmas (só adiciona se for um valor de hora razoável)
            valor_salvo = None
            if isinstance(val_cell, timedelta):
                valor_salvo = val_cell.total_seconds() / 86400.0
            elif isinstance(val_cell, (int, float)):
                if abs(val_cell) < 1000: # Proteção contra fórmulas exponenciais (1000 dias de horas contínuas não existem)
                    valor_salvo = float(val_cell)
            
            if valor_salvo is not None and valor_salvo >= 0:
                mapa_horas_02[cod] = valor_salvo
                
        except (ValueError, TypeError):
            continue

wb_origem.close()
print(f"Lidos {len(mapa_horas_02)} registros válidos de horas fiscais do mês anterior.")

print("Abrindo planilha atual...")
wb_destino = openpyxl.load_workbook(ARQUIVO_DESTINO)
ws_destino = wb_destino[ABA_DESTINO]

# Aplicar na Coluna N onde o código bater
alterados = 0
for row_idx in range(10, ws_destino.max_row + 1):
    cod_cell = ws_destino.cell(row=row_idx, column=8).value
    if cod_cell is not None:
        try:
            cod = int(float(str(cod_cell).strip()))
            if cod in mapa_horas_02:
                c_ant = ws_destino.cell(row=row_idx, column=14) # Coluna N
                c_ant.value = mapa_horas_02[cod]
                c_ant.number_format = '[h]:mm:ss'
                alterados += 1
        except (ValueError, TypeError):
            continue

print(f"Salvos {alterados} registros com segurança na Coluna N!")
wb_destino.save(ARQUIVO_DESTINO)
wb_destino.close()
print("Finalizado com Sucesso.")
