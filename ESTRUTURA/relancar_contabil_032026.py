"""
Task 4 (CORRIGIDO) - Relançamento Contábil na Planilha Master (03.2026)
  - Fonte: HORAS CONTABEIS_.xlsx → aba 03.2026 → col R (18) = HORAS VALIDADAS
  - Destino: CONTROLE DE HORAS DMF.xlsx → aba 03.2026 → col P (16) = HORÁRIO CONTÁBIL
  - Validação: Double Match (Código Domínio + CNPJ)
  - REGRA: APENAS coluna P é alterada.
"""
import openpyxl
import re

def clean_cnpj(cnpj_val):
    if cnpj_val is None:
        return None
    return re.sub(r'\D', '', str(cnpj_val).strip())

def main():
    contabil_path = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_.xlsx"
    master_path   = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF.xlsx"
    
    CONTABIL_SHEET = "03.2026"   # Aba correta com coluna HORAS VALIDADAS
    MASTER_SHEET   = "03.2026"
    
    # Colunas da Contábil (fonte)
    COL_COD_CONTABIL  = 1   # A: Código Domínio
    COL_CNPJ_CONTABIL = 3   # C: CNPJ
    COL_NOME_CONTABIL = 4   # D: Nome
    COL_HORAS_VALID   = 18  # R: HORAS VALIDADAS (fórmula - lida com data_only)
    
    # Colunas da Master (destino)
    COL_COD_MASTER    = 8   # H: Código Domínio
    COL_CNPJ_MASTER   = 10  # J: CNPJ
    COL_CONTABIL_DEST = 16  # P: HORÁRIO CONTÁBIL
    START_ROW_MASTER  = 10

    # ============================================================
    # PASSO 1: Carregar dados da planilha contábil (data_only=True para ler fórmulas)
    # ============================================================
    print("=" * 60)
    print("PASSO 1: Carregando HORAS VALIDADAS da contabil (03.2026, col R)...")
    print("=" * 60)
    
    wb_c = openpyxl.load_workbook(contabil_path, data_only=True)
    ws_c = wb_c[CONTABIL_SHEET]
    
    contabil_data = {}
    for row in range(2, ws_c.max_row + 1):
        cod_val = ws_c.cell(row=row, column=COL_COD_CONTABIL).value
        if cod_val is None:
            continue
        try:
            cod = int(float(str(cod_val).strip()))
        except (ValueError, TypeError):
            continue
        
        cnpj = clean_cnpj(ws_c.cell(row=row, column=COL_CNPJ_CONTABIL).value)
        nome = ws_c.cell(row=row, column=COL_NOME_CONTABIL).value
        horas = ws_c.cell(row=row, column=COL_HORAS_VALID).value
        
        contabil_data[cod] = {
            'cnpj': cnpj,
            'nome': nome,
            'horas': horas,
            'row': row
        }
    
    print(f"  Empresas carregadas: {len(contabil_data)}")
    com_horas = sum(1 for v in contabil_data.values() if v['horas'] and v['horas'] != 0)
    sem_horas = sum(1 for v in contabil_data.values() if not v['horas'] or v['horas'] == 0)
    print(f"  Com horas validadas > 0: {com_horas}")
    print(f"  Sem horas (zero/None):   {sem_horas}")
    
    # Amostra para conferência
    print("\n  Amostra (primeiras 5 com horas):")
    count = 0
    for cod, info in contabil_data.items():
        if info['horas'] and info['horas'] != 0:
            print(f"    Cod {cod}: {info['horas']}")
            count += 1
            if count >= 5: break

    # ============================================================
    # PASSO 2: Abrir Master para escrita
    # ============================================================
    print()
    print("=" * 60)
    print("PASSO 2: Abrindo planilha Master...")
    print("=" * 60)
    
    wb_m = openpyxl.load_workbook(master_path)
    ws_m = wb_m[MASTER_SHEET]
    print(f"  Aba alvo: {MASTER_SHEET} | Linhas: {ws_m.max_row}")

    # ============================================================
    # PASSO 3: Double Match + Lançamento
    # ============================================================
    print()
    print("=" * 60)
    print("PASSO 3: Double Match + Lançamento em coluna P...")
    print("=" * 60)
    
    matched = 0
    cod_only = 0
    not_found = 0
    updated = 0
    anomalias = []
    
    for row in range(START_ROW_MASTER, ws_m.max_row + 1):
        cod_master_val = ws_m.cell(row=row, column=COL_COD_MASTER).value
        cnpj_master = clean_cnpj(ws_m.cell(row=row, column=COL_CNPJ_MASTER).value)
        
        if cod_master_val is None and cnpj_master is None:
            continue
        
        cod_master = None
        if cod_master_val is not None:
            try:
                cod_master = int(float(str(cod_master_val).strip()))
            except (ValueError, TypeError):
                pass
        
        match_cod = cod_master in contabil_data if cod_master else False
        
        if match_cod:
            contabil_info = contabil_data[cod_master]
            match_cnpj = False
            if cnpj_master and contabil_info['cnpj']:
                match_cnpj = (cnpj_master == contabil_info['cnpj'])
            
            horas_val = contabil_info['horas']
            
            if match_cnpj:
                matched += 1
            else:
                cod_only += 1
                anomalias.append(
                    f"  L{row}: Cod {cod_master} CNPJ diverge "
                    f"(M={cnpj_master} vs C={contabil_info['cnpj']})"
                )
            
            # Lançar na coluna P (16)
            if horas_val is not None and horas_val != 0:
                ws_m.cell(row=row, column=COL_CONTABIL_DEST).value = horas_val
                updated += 1
            else:
                ws_m.cell(row=row, column=COL_CONTABIL_DEST).value = 0
        else:
            not_found += 1

    # ============================================================
    # PASSO 4: Salvar
    # ============================================================
    print()
    print("=" * 60)
    print("PASSO 4: Salvando...")
    print("=" * 60)
    
    wb_m.save(master_path)
    print(f"  Arquivo salvo: {master_path}")

    # ============================================================
    # RELATÓRIO
    # ============================================================
    print()
    print("=" * 60)
    print("RELATORIO FINAL")
    print("=" * 60)
    print(f"  Double Match (Cod+CNPJ):   {matched}")
    print(f"  Só Código (CNPJ diverge):  {cod_only}")
    print(f"  Não encontrado na contábil:{not_found}")
    print(f"  Valores lançados em P:     {updated}")
    
    if anomalias:
        print(f"\nANOMALIAS ({len(anomalias)}):")
        for a in anomalias[:20]:
            print(a)

if __name__ == "__main__":
    main()
