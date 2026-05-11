import openpyxl

BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = BASE_DIR + r"\CONTROLE DE HORAS DMF - CLIENTE.xlsm"

wb = openpyxl.load_workbook(PLANILHA_MASTER, data_only=False)
ws = wb['03.2026']

mapa_planilha = {}
for row in range(10, ws.max_row + 1):
    cod_cell = ws.cell(row=row, column=8).value
    if cod_cell and str(cod_cell).strip().isdigit():
        mapa_planilha[int(str(cod_cell).strip())] = row
    elif cod_cell:
        # maybe it's float? like 1152.0?
        s = str(cod_cell).strip()
        if s.replace('.0', '').isdigit():
            mapa_planilha[int(s.replace('.0', ''))] = row

print(f"Is 1152 in mapa_planilha? {1152 in mapa_planilha}")
if 1152 in mapa_planilha:
    print(f"Yes, mapped to row {mapa_planilha[1152]}")
else:
    for row in range(10, ws.max_row + 1):
        cod_cell = ws.cell(row=row, column=8).value
        # If it's literally 1152
        if cod_cell == 1152 or cod_cell == "1152":
            print(f"Found 1152 at row {row} literally, but why wasn't it mapped?!")
            print(f"cod_cell={cod_cell}, type={type(cod_cell)}")
            print(f"str(cod_cell).strip().isdigit() -> {str(cod_cell).strip().isdigit()}")
