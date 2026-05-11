import openpyxl

arquivo_final = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm"

try:
    wb = openpyxl.load_workbook(arquivo_final, data_only=False) # Ver fórmulas
    ws = wb["02.2026"]
    
    row = 14
    print(f"Linha {row} (COM FÓRMULAS):")
    for i in range(1, 20):
        col_letter = openpyxl.utils.get_column_letter(i)
        val = ws.cell(row=row, column=i).value
        print(f"  {col_letter}: {val}")

except Exception as e:
    print(f"Erro: {e}")
