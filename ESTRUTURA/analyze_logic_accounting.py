import openpyxl
import os

FILE_ACCOUTING = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_.xlsx'

def analyze_logic():
    print(f"Lendo planilha: {FILE_ACCOUTING}")
    wb = openpyxl.load_workbook(FILE_ACCOUTING, data_only=False)
    
    # Check 03.2026
    ws = wb['03.2026']
    print("\n--- Aba 03.2026 ---")
    print(f"Q1: {ws['Q1'].value}")
    print(f"P2 formula: {ws['P2'].value}")
    print(f"Q2 formula: {ws['Q2'].value}")
    
    # Check Row 1 for other columns
    print("\nRow 1 headers (A-R):")
    row1 = [ws.cell(1, i).value for i in range(1, 19)]
    print(row1)

    # Check EXEMPLO
    if 'EXEMPLO' in wb.sheetnames:
        ws_ex = wb['EXEMPLO']
        print("\n--- Aba EXEMPLO ---")
        print(f"P2 formula: {ws_ex['P2'].value}")
        print(f"Q2 formula: {ws_ex['Q2'].value}")
        print(f"O2 value: {ws_ex['O2'].value}")
    
    # Investigative step: If P2 is missing, what is the logic for faturamento hours?
    # In some DMF sheets, it is: =IF(O2>0, "0:30:00", "00:00:00") or similar.
    
    # Check column O to P relationship in EXEMPLO
    print("\nChecking logic in EXEMPLO rows 2-5...")
    for r in range(2, 6):
        o_val = ws_ex.cell(r, 15).value
        p_val = ws_ex.cell(r, 16).value
        print(f" Row {r}: O={o_val}, P={p_val}")

if __name__ == "__main__":
    analyze_logic()
