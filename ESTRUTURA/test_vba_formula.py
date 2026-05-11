import openpyxl

BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = BASE_DIR + r"\CONTROLE DE HORAS DMF - CLIENTE.xlsm"

wb = openpyxl.load_workbook(PLANILHA_MASTER, data_only=False, keep_vba=True)
ws = wb['03.2026']

cod_cell = ws.cell(row=756, column=8).value
print(f"data_only=False -> cod_cell={cod_cell!r}, type={type(cod_cell)}")

wb_data_only = openpyxl.load_workbook(PLANILHA_MASTER, data_only=True)
ws_data_only = wb_data_only['03.2026']
cod_cell_do = ws_data_only.cell(row=756, column=8).value
print(f"data_only=True -> cod_cell={cod_cell_do!r}, type={type(cod_cell_do)}")
