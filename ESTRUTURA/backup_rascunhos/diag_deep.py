import openpyxl
from datetime import timedelta

# Abrir SEM data_only para ver as fórmulas e os valores reais salvos
wb = openpyxl.load_workbook(
    r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx',
    data_only=False
)
ws = wb['03.2026']

# 1. Verificar o SUBTOTAL em N7
print("=== SUBTOTAL N7 ===")
print(f"N7 formula/value: {ws.cell(7, 14).value!r}")
print(f"N7 number_format: {ws.cell(7, 14).number_format!r}")

# 2. Contar tipos de dados na coluna N
tipos = {}
soma_float = 0
soma_td = 0
count_none = 0
count_str = 0
last_row_with_data = 0

for r in range(10, ws.max_row + 1):
    v = ws.cell(r, 14).value
    t = type(v).__name__
    tipos[t] = tipos.get(t, 0) + 1
    
    if isinstance(v, timedelta):
        soma_td += v.total_seconds() / 86400.0
        last_row_with_data = r
    elif isinstance(v, (int, float)):
        soma_float += float(v)
        last_row_with_data = r
    elif v is None:
        count_none += 1

print(f"\n=== TIPOS DE DADOS NA COLUNA N (linhas 10 a {ws.max_row}) ===")
for t, c in sorted(tipos.items()):
    print(f"  {t}: {c}")

print(f"\nSoma de floats: {soma_float * 24:.2f}h")
print(f"Soma de timedeltas: {soma_td * 24:.2f}h")
print(f"Soma TOTAL: {(soma_float + soma_td) * 24:.2f}h")
print(f"Última linha com dados: {last_row_with_data}")

# 3. Mostrar algumas amostras de timedelta vs float
print("\n=== AMOSTRAS (primeiras 5 de cada tipo) ===")
td_count = 0
fl_count = 0
for r in range(10, ws.max_row + 1):
    v = ws.cell(r, 14).value
    cod = ws.cell(r, 8).value
    if isinstance(v, timedelta) and td_count < 5:
        print(f"  Row {r}, Cod {cod}: timedelta = {v} ({v.total_seconds()/3600:.2f}h)")
        td_count += 1
    elif isinstance(v, (int, float)) and v > 0 and fl_count < 5:
        print(f"  Row {r}, Cod {cod}: float = {v} ({v*24:.2f}h)")
        fl_count += 1

# 4. Verificar dados ABAIXO da linha 588 (limite do SUBTOTAL)
print("\n=== DADOS ABAIXO DE N588 (fora do SUBTOTAL) ===")
soma_fora = 0
count_fora = 0
for r in range(589, ws.max_row + 1):
    v = ws.cell(r, 14).value
    cod = ws.cell(r, 8).value
    if isinstance(v, timedelta):
        soma_fora += v.total_seconds() / 86400.0
        count_fora += 1
    elif isinstance(v, (int, float)) and v > 0:
        soma_fora += float(v)
        count_fora += 1

print(f"  {count_fora} registros com dados fora do range do SUBTOTAL")
print(f"  Soma fora: {soma_fora * 24:.2f}h")

# 5. Comparar com a origem
wb_orig = openpyxl.load_workbook(
    r'C:\Users\DMF-AUTOMACAO\Downloads\CONTROLE_DE_HORAS_DMF_BL4t520G.xlsm', 
    data_only=True
)
ws_orig = wb_orig['02.2026']

# Mapear TODOS os clientes e valores da ORIGEM
mapa_orig = {}
for r in range(10, ws_orig.max_row + 1):
    cod = ws_orig.cell(r, 8).value
    val = ws_orig.cell(r, 15).value
    if cod is not None:
        try:
            cod = int(float(str(cod).strip()))
        except:
            continue
        v = 0
        if isinstance(val, timedelta):
            v = val.total_seconds() / 86400.0
        elif isinstance(val, (int, float)) and abs(val) < 1000:
            v = float(val)
        if cod not in mapa_orig:
            mapa_orig[cod] = 0
        mapa_orig[cod] += v

# Mapear clientes no DESTINO (Col N)
mapa_dest = {}
for r in range(10, ws.max_row + 1):
    cod = ws.cell(r, 8).value
    val = ws.cell(r, 14).value
    if cod is not None:
        try:
            cod = int(float(str(cod).strip()))
        except:
            continue
        v = 0
        if isinstance(val, timedelta):
            v = val.total_seconds() / 86400.0
        elif isinstance(val, (int, float)):
            v = float(val)
        if cod not in mapa_dest:
            mapa_dest[cod] = 0
        mapa_dest[cod] += v

# Clientes que estão na origem mas NÃO no destino
print("\n=== CLIENTES DA ORIGEM SEM CORRESPONDÊNCIA NO DESTINO ===")
perdidos_total = 0
for cod in sorted(mapa_orig.keys()):
    if cod not in mapa_dest and mapa_orig[cod] > 0:
        print(f"  Cod {cod}: {mapa_orig[cod]*24:.2f}h PERDIDAS")
        perdidos_total += mapa_orig[cod]
print(f"  TOTAL PERDIDO (clientes ausentes): {perdidos_total*24:.2f}h")

# Clientes com valores diferentes
print("\n=== CLIENTES COM DIFERENÇA DE VALOR (top 20) ===")
diffs = []
for cod in sorted(mapa_orig.keys()):
    if cod in mapa_dest:
        d = mapa_orig[cod] - mapa_dest[cod]
        if abs(d) > 0.0001:
            diffs.append((cod, mapa_orig[cod]*24, mapa_dest[cod]*24, d*24))
diffs.sort(key=lambda x: abs(x[3]), reverse=True)
for cod, o, d, diff in diffs[:20]:
    print(f"  Cod {cod}: Origem={o:.2f}h, Destino={d:.2f}h, Diferença={diff:.2f}h")
print(f"  Total diferenças: {sum(x[3] for x in diffs):.2f}h")
