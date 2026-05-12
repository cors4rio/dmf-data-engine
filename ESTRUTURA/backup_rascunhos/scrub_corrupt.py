import openpyxl

file_path = 'CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx'
print("Carregando", file_path)
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb['03.2026']

count = 0
# Limpar as colunas N e O de valores absurdos que causam overflow no formato de data do Excel
for row in ws.iter_rows(min_row=10, max_row=ws.max_row, min_col=14, max_col=15):
    for cell in row:
        if type(cell.value) in (int, float):
            if abs(cell.value) > 10000: # Qualquer valor de hora/data absurdo (> 10000 dias)
                cell.value = 0
                cell.number_format = '[h]:mm:ss'
                count += 1
                
print(f"Limpos {count} valores fantasmas radioativos (incompatíveis com data do Excel).")
wb.save(file_path)
print("Salvo!")
