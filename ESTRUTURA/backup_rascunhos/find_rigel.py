import openpyxl

BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = BASE_DIR + r"\CONTROLE DE HORAS DMF - CLIENTE.xlsm"

wb = openpyxl.load_workbook(PLANILHA_MASTER, data_only=True)
ws = wb['03.2026']

for r in range(10, ws.max_row+1):
    nome = ws.cell(row=r, column=4).value
    if nome and "RIGEL" in str(nome).upper():
        cod = ws.cell(row=r, column=8).value
        print(f"Row {r} | Cod: {cod} | Nome: {nome} | Col Q: {ws.cell(row=r, column=17).value}")
