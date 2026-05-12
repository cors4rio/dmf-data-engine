"""
02_CONTABIL/processar.py — Módulo Contábil (HORAS CONTABEIS_.xlsx)
DMF Automação

Preenche as colunas F (QTD Lançamentos) e O (Faturamento)
na planilha HORAS CONTABEIS_.xlsx, e a coluna I (Tem Folha?)
a partir dos dados da planilha Carol ou do Domínio.

Regras de governança:
  - Validação Tripla antes de qualquer escrita (codi_emp + CNPJ + nome)
  - Somente preencher F, O e I
  - Campos A, B, C, D, E, G, H, J, K, L, M, N, P, Q PROIBIDOS de alterar
  - Relatório .md obrigatório com anomalias

Uso:
    python 02_CONTABIL/processar.py --mes 03 --ano 2026
    python 02_CONTABIL/processar.py --mes 03 --ano 2026 --dry-run
"""

import sys
import re
import logging
from datetime import datetime
from pathlib import Path
from difflib import SequenceMatcher

sys.path.insert(0, str(Path(__file__).parent.parent))

from _ENGINE.config import cfg
from _ENGINE.database import executar_query, testar_conexao
from _ENGINE import excel_utils as xl
import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("dmf.contabil")

# ─────────────────────────────────────────────
# COLUNAS DA PLANILHA HORAS CONTABEIS (1-indexed)
# ─────────────────────────────────────────────
C_COD     = 1   # A — Cód Domínio (FIXO — NÃO ALTERAR)
C_GRUPO   = 2   # B — Grupo (FIXO)
C_CNPJ    = 3   # C — CNPJ/CPF (FIXO)
C_EMPRESA = 4   # D — Nome Empresa (FIXO)
C_REGIME  = 5   # E — Regime (PRÉ-PREENCHIDO)
C_LANC    = 6   # F — QTD Lançamentos ← PREENCHER
C_MES     = 7   # G — Mês (PRÉ-PREENCHIDO)
C_HORAS   = 8   # H — Horas (CALCULADO)
C_FOLHA   = 9   # I — Tem Folha? ← PREENCHER
C_HOR_FL  = 10  # J — Horas Folha (CALCULADO)
# K, L, M, N = manuais
C_FATURA  = 15  # O — Total Faturamento ← PREENCHER
# P, Q = calculados

LINHA_INICIO = 2  # Dados começam na linha 2

# ─────────────────────────────────────────────
# SQL — Validação Tripla
# ─────────────────────────────────────────────
SQL_CADASTROS = """
SELECT codi_emp, cgce_emp, nome_emp
FROM bethadba.geempre
ORDER BY codi_emp
"""

SQL_LANCAMENTOS = """
SELECT
    codi_emp                                                  AS Codigo_Cliente,
    SUM(CASE WHEN orig_lan = 1  THEN 1 ELSE 0 END)           AS Lancamentos_Normal,
    SUM(CASE WHEN orig_lan = 39 THEN 1 ELSE 0 END)           AS Lancamentos_Extrato,
    COUNT(*)                                                  AS Total
FROM bethadba.ctlancto
WHERE data_lan >= ?
  AND data_lan <= ?
  AND orig_lan IN (1, 39)
GROUP BY codi_emp
ORDER BY codi_emp
"""

SQL_FATURAMENTO = """
SELECT
    codi_emp            AS Codigo_Cliente,
    SUM(total_contabil) AS Faturamento_Mensal
FROM (
    SELECT codi_emp, SUM(vcon_sai) AS total_contabil
    FROM bethadba.efsaidas
    WHERE dsai_sai >= ? AND dsai_sai <= ?
    GROUP BY codi_emp
    UNION ALL
    SELECT codi_emp, SUM(vcon_ser) AS total_contabil
    FROM bethadba.efservicos
    WHERE dser_ser >= ? AND dser_ser <= ?
    GROUP BY codi_emp
) base
GROUP BY codi_emp
ORDER BY codi_emp
"""

SQL_FOLHA = """
SELECT
    g.codi_emp,
    CASE WHEN COALESCE(f.total, 0) > 0 THEN 'SIM' ELSE 'NAO' END AS Tem_Folha
FROM bethadba.geempre g
LEFT JOIN (
    SELECT
        e.codi_emp,
        SUM(CASE WHEN e.vinculo IN (1,6,11) THEN 1 ELSE 0 END) AS total
    FROM bethadba.foempregados e
    LEFT JOIN bethadba.forescisoes r
        ON r.codi_emp = e.codi_emp
       AND r.i_empregados = e.i_empregados
       AND r.demissao < ?
    WHERE e.admissao <= ?
      AND r.i_empregados IS NULL
    GROUP BY e.codi_emp
) f ON g.codi_emp = f.codi_emp
ORDER BY g.codi_emp
"""


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.upper(), b.upper()).ratio()


def _normalizar_cnpj(v) -> str:
    return re.sub(r"\D", "", str(v or ""))


def _normalizar_nome(v) -> str:
    """Remove CNPJ do prefixo e retorna nome limpo para comparação fuzzy."""
    s = str(v or "")
    # Remove padrões como '12.345.678/0001-99 - ' no início
    s = re.sub(r"^\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\s*[-–]?\s*", "", s)
    return s.strip().upper()


# ─────────────────────────────────────────────
# VALIDAÇÃO TRIPLA
# ─────────────────────────────────────────────

def validar_tripla(
    planilha_rows: list[dict],
    cadastros_bd: list[dict],
) -> tuple[
    dict[int, dict],  # aprovados (2 ou 3 matches): {linha → cadastro_bd}
    list[dict],       # matches parciais (2/3) para relatório
    list[dict],       # rejeitados (0-1/3) para relatório
]:
    """
    Cruza cada linha da planilha com os cadastros do banco.
    Retorna aprovados, parciais e rejeitados.
    """
    # Indexar banco por código, CNPJ e nome
    idx_cod:  dict[str, dict] = {}
    idx_cnpj: dict[str, dict] = {}
    for r in cadastros_bd:
        cod  = xl.normalizar_codigo(r["codi_emp"])
        cnpj = _normalizar_cnpj(r["cgce_emp"])
        if cod:
            idx_cod[cod]  = r
        if cnpj:
            idx_cnpj[cnpj] = r

    aprovados: dict[int, dict] = {}
    parciais:  list[dict] = []
    rejeitados: list[dict] = []

    for row in planilha_rows:
        linha = row["_linha"]

        cod_plan  = xl.normalizar_codigo(row.get("cod"))
        cnpj_plan = _normalizar_cnpj(row.get("cnpj"))
        nome_plan = _normalizar_nome(row.get("nome"))

        match_bd = None
        score = 0
        detalhes = {"cod": False, "cnpj": False, "nome": False}

        # Tentar encontrar candidato via código
        if cod_plan and cod_plan in idx_cod:
            match_bd = idx_cod[cod_plan]
        elif cnpj_plan and cnpj_plan in idx_cnpj:
            match_bd = idx_cnpj[cnpj_plan]
        else:
            # Busca fuzzy por nome como último recurso
            melhor_sim = 0
            for r in cadastros_bd:
                sim = _similarity(nome_plan, _normalizar_nome(r.get("nome_emp", "")))
                if sim > melhor_sim:
                    melhor_sim = sim
                    match_bd = r if sim > 0.85 else None

        if match_bd:
            bd_cod  = xl.normalizar_codigo(match_bd.get("codi_emp"))
            bd_cnpj = _normalizar_cnpj(match_bd.get("cgce_emp"))
            bd_nome = _normalizar_nome(match_bd.get("nome_emp", ""))

            if cod_plan and cod_plan == bd_cod:
                score += 1
                detalhes["cod"] = True
            if cnpj_plan and cnpj_plan == bd_cnpj:
                score += 1
                detalhes["cnpj"] = True
            if _similarity(nome_plan, bd_nome) > 0.85:
                score += 1
                detalhes["nome"] = True

        entrada_log = {
            "linha": linha,
            "cod_planilha": cod_plan,
            "cnpj_planilha": cnpj_plan,
            "nome_planilha": nome_plan,
            "score": score,
            "divergencias": [k for k, v in detalhes.items() if not v],
            "match_bd": match_bd,
        }

        if score >= 2:
            aprovados[linha] = match_bd
            if score == 2:
                parciais.append(entrada_log)
        else:
            rejeitados.append(entrada_log)

    logger.info(
        "Validação Tripla | Aprovados: %d | Parciais: %d | Rejeitados: %d",
        len(aprovados), len(parciais), len(rejeitados)
    )
    return aprovados, parciais, rejeitados


# ─────────────────────────────────────────────
# PIPELINE PRINCIPAL
# ─────────────────────────────────────────────

def processar_contabil(mes: int, ano: int, *, dry_run: bool = False) -> dict:
    """
    Pipeline completo do módulo contábil.
    1. Validação Tripla
    2. Extrai Lançamentos (F), Faturamento (O), Folha (I)
    3. Preenche apenas F, O, I
    """
    if not cfg.contabil:
        raise FileNotFoundError("Planilha HORAS CONTABEIS não encontrada.")

    inicio, fim = cfg.datas_mes(mes, ano)

    # ── 1. Cadastros do banco (validação tripla) ──
    logger.info("Buscando cadastros do banco (validação tripla)...")
    cadastros_bd = executar_query(SQL_CADASTROS)

    # ── 2. Ler planilha contábil ──
    wb = xl.abrir_planilha(cfg.contabil)
    aba = cfg.aba_excel(mes, ano)

    if aba not in wb.sheetnames:
        raise ValueError(f"Aba '{aba}' não encontrada em HORAS CONTABEIS.")

    ws = wb[aba]

    # Ler linhas da planilha
    planilha_rows = []
    for row in ws.iter_rows(min_row=LINHA_INICIO):
        n = row[0].row
        planilha_rows.append({
            "_linha": n,
            "cod":  row[C_COD  - 1].value,
            "cnpj": row[C_CNPJ - 1].value,
            "nome": row[C_EMPRESA - 1].value,
        })

    # ── 3. Validação Tripla ──
    aprovados, parciais, rejeitados = validar_tripla(planilha_rows, cadastros_bd)

    # ── 4. Queries de dados ──
    logger.info("Extraindo lançamentos, faturamento e folha...")
    lancamentos_rows = executar_query(SQL_LANCAMENTOS, params=(inicio, fim))
    faturamento_rows = executar_query(SQL_FATURAMENTO, params=(inicio, fim, inicio, fim))
    folha_rows       = executar_query(SQL_FOLHA, params=(inicio, fim))

    # Indexar por código
    dict_lanc  = {xl.normalizar_codigo(r["Codigo_Cliente"]): int(r["Total"] or 0)
                  for r in lancamentos_rows}
    dict_fat   = {xl.normalizar_codigo(r["Codigo_Cliente"]): float(r["Faturamento_Mensal"] or 0)
                  for r in faturamento_rows}
    dict_folha = {xl.normalizar_codigo(r["codi_emp"]): r["Tem_Folha"]
                  for r in folha_rows}

    # ── 5. Preencher (somente F, O, I) ──
    stats = {"preenchidos": 0, "parciais": len(parciais), "rejeitados": len(rejeitados)}

    for linha, match_bd in aprovados.items():
        cod = xl.normalizar_codigo(match_bd.get("codi_emp")) if match_bd else None
        if not cod:
            continue

        qtd_lanc = dict_lanc.get(cod, 0)
        faturamento = dict_fat.get(cod, 0)
        tem_folha   = dict_folha.get(cod, "NAO")

        if dry_run:
            logger.info("  [DRY-RUN] Linha %d | cod=%s | F=%d | O=%.2f | I=%s",
                        linha, cod, qtd_lanc, faturamento, tem_folha)
        else:
            ws.cell(row=linha, column=C_LANC).value   = qtd_lanc
            ws.cell(row=linha, column=C_FATURA).value = faturamento
            ws.cell(row=linha, column=C_FOLHA).value  = tem_folha
            stats["preenchidos"] += 1

    if not dry_run:
        xl.salvar_planilha(wb, cfg.contabil)

    logger.info(
        "✅ Contábil | Preenchidos: %d | Parciais: %d | Rejeitados: %d",
        stats["preenchidos"], stats["parciais"], stats["rejeitados"]
    )
    return {
        "stats": stats,
        "parciais": parciais,
        "rejeitados": rejeitados,
    }


# ─────────────────────────────────────────────
# RELATÓRIO
# ─────────────────────────────────────────────

def gerar_relatorio(mes: int, ano: int, resultado: dict) -> Path:
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    aba  = cfg.aba_excel(mes, ano)
    nome = cfg.rel_contabil / f"relatorio_contabil_{aba.replace('.', '_')}.md"

    stats    = resultado["stats"]
    parciais = resultado["parciais"]
    rejeit   = resultado["rejeitados"]

    linhas_parciais = ""
    for p in parciais:
        divs = ", ".join(p["divergencias"])
        linhas_parciais += (
            f"| {p['linha']} | {p['cod_planilha']} | {p['cnpj_planilha']} "
            f"| {p['nome_planilha'][:40]} | {divs} |\n"
        )

    linhas_rejeit = ""
    for r in rejeit:
        linhas_rejeit += (
            f"| {r['linha']} | {r['cod_planilha']} | {r['cnpj_planilha']} "
            f"| {r['nome_planilha'][:40]} | {r['score']}/3 |\n"
        )

    conteudo = f"""# Relatório de Execução — Módulo Contábil
**Mês de Apuração:** {aba}
**Executado em:** {ts}

## Resumo

| Métrica | Valor |
|---|---|
| Empresas preenchidas (F, O, I) | {stats['preenchidos']} |
| Matches parciais (2/3) | {stats['parciais']} |
| Empresas rejeitadas (≤1/3) | {stats['rejeitados']} |

## Empresas com Match Parcial (2/3) — ⚠️ Verificar

| Linha | Código | CNPJ | Nome | Campo Divergente |
|---|---|---|---|---|
{linhas_parciais or "*(nenhuma)*"}

## Empresas Rejeitadas (0-1/3) — ❌ Revisão Manual Obrigatória

| Linha | Código | CNPJ | Nome | Score |
|---|---|---|---|---|
{linhas_rejeit or "*(nenhuma)*"}

---
*Gerado automaticamente por 02_CONTABIL/processar.py*
"""
    nome.write_text(conteudo, encoding="utf-8")
    logger.info("📄 Relatório: %s", nome)
    return nome


# ─────────────────────────────────────────────
# ENTRYPOINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    args = cfg.parse_args("DMF — Módulo Contábil")

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    if not testar_conexao():
        logger.error("Conexão ODBC falhou. Abortando.")
        sys.exit(1)

    resultado = processar_contabil(args.mes, args.ano, dry_run=args.dry_run)
    gerar_relatorio(args.mes, args.ano, resultado)
