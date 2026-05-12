import openpyxl

BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = BASE_DIR + r"\CONTROLE DE HORAS DMF - CLIENTE.xlsm"

wb = openpyxl.load_workbook(PLANILHA_MASTER, data_only=True)
ws = wb['03.2026']

for r in range(10, ws.max_row+1):
    cod = ws.cell(row=r, column=8).value
    nome = ws.cell(row=r, column=4).value
    try:
        cod_val = int(float(str(cod).strip()))
        if cod_val in [171, 1480, 274]:
            print(f"Row {r} | Cod: {cod_val} | Nome: {nome} | Col Q: {ws.cell(row=r, column=17).value}")
    except:
        pass
