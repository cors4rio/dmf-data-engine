import openpyxl
import os

file_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_V3_AeqwQXgR.xlsx'

def to_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        try:
            # Tentar converter tratando vírgula como decimal
            return float(val.replace(',', '.').strip())
        except ValueError:
            return 0.0
    return 0.0

try:
    wb_meta = openpyxl.load_workbook(file_path, data_only=False)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    
    sheet_name = wb_meta.sheetnames[0]
    ws_meta = wb_meta[sheet_name]
    ws_data = wb_data[sheet_name]
    
    results = []
    # Analisar até 200 linhas
    max_row = min(ws_meta.max_row, 200)
    for row in range(2, max_row + 1):
        row_info = {
            'row': row,
            'cols': {}
        }
        for col in ['H', 'J', 'L', 'N', 'P', 'Q']:
            val = ws_data[f'{col}{row}'].value
            formula = ws_meta[f'{col}{row}'].value
            row_info['cols'][col] = {
                'val': val,
                'formula': formula,
                'type': type(val).__name__
            }
        results.append(row_info)

    anomalies = []
    for r in results:
        h = to_float(r['cols']['H']['val'])
        j = to_float(r['cols']['J']['val'])
        l = to_float(r['cols']['L']['val'])
        n = to_float(r['cols']['N']['val'])
        p = to_float(r['cols']['P']['val'])
        q = r['cols']['Q']['val']
        
        # Se Q for 0 (ou None) mas a soma dos valores deveria ser > 0
        expected_q = h + j + l + n + p
        actual_q = to_float(q)
        
        if actual_q == 0 and expected_q > 0:
            anomalies.append((r, expected_q))
            
    print(f"Total de linhas analisadas: {len(results)}")
    print(f"Total de anomalias encontradas: {len(anomalies)}")
    
    if anomalies:
        print("\nTOP 3 ANOMALIAS:")
        for r, expected in anomalies[:3]:
            print(f"Linha {r['row']}:")
            for col in ['H', 'J', 'L', 'N', 'P', 'Q']:
                c = r['cols'][col]
                print(f"  {col}: Valor={c['val']} ({c['type']}), Fórmula={c['formula']}")
            print(f"  Soma esperada: {expected}")
    else:
        print("\nNenhuma anomalia encontrada nas primeiras 200 linhas.")
        # Se não houver anomalias, vamos verificar se Q tem fórmula em todas as linhas
        sem_formula = [r['row'] for r in results if not str(r['cols']['Q']['formula']).startswith('=')]
        if sem_formula:
            print(f"Linhas sem fórmula na coluna Q: {sem_formula[:10]}...")

    # Verificar se as "fórmulas" na coluna Q estão apontando para as linhas corretas
    primeira_q = results[0]['cols']['Q']
    print(f"\nExemplo de fórmula em Q2: {primeira_q['formula']}")

except Exception as e:
    import traceback
    print(f"Erro ao processar excel: {e}")
    traceback.print_exc()
