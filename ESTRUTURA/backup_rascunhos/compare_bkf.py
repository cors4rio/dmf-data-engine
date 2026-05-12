import openpyxl; from datetime import timedelta; 
wb1=openpyxl.load_workbook(r'C:\Users\DMF-AUTOMACAO\Downloads\CONTROLE_DE_HORAS_DMF_BL4t520G.xlsm', data_only=True); ws1=wb1['02.2026']; sums_1=0; 
for r in range(10, ws1.max_row+1): 
    v=ws1.cell(r,15).value; 
    if isinstance(v, timedelta): sums_1 += v.total_seconds() / 86400.0 
    elif type(v) in (int, float) and abs(v)<1000: sums_1 += v 
wb2=openpyxl.load_workbook(r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx', data_only=True); ws2=wb2['03.2026']; sums_2=0; 
for r in range(10, ws2.max_row+1): 
    v=ws2.cell(r,14).value;   
    if type(v) in (int, float): sums_2 += v 
    elif isinstance(v, timedelta): sums_2 += v.total_seconds()/86400.0 
print(f'Total OLD (Col O): {sums_1*24:.2f}h'); print(f'Total NEW (Col N): {sums_2*24:.2f}h'); 
