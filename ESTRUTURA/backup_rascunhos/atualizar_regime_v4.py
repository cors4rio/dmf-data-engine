import pyodbc
import openpyxl
import os

def mapear_regime(descricao, rfed=None):
    if not descricao:
        if rfed == 1: return "Lucro Real"
        if rfed in [2, 4]: return "Simples Nacional"
        if rfed == 5: return "Lucro Presumido"
        if rfed == 8: return "Imune / Isenta"
        return "Não Parametrizado"
    
    desc_upper = descricao.upper()
    
    if "SIMPLES" in desc_upper:
        return "Simples Nacional"
    if any(kw in desc_upper for kw in ["PRESUMIDO", "PRESIMIDO", "PRESSUMIDO"]):
        return "Lucro Presumido"
    if "REAL" in desc_upper:
        return "Lucro Real"
    if "PRODUTOR RURAL" in desc_upper:
        return "Produtor Rural"
    if any(kw in desc_upper for kw in ["IMUNE", "ISENTA"]):
        return "Imune / Isenta"
        
    if any(kw in desc_upper for kw in ["INICIAL", "NOVO", "VIGÊNCIA", "PRIMEIRA"]):
        if rfed == 1: return "Lucro Real"
        if rfed in [2, 4]: return "Simples Nacional"
        if rfed == 5: return "Lucro Presumido"
        if rfed == 8: return "Imune / Isenta"

    palavras = descricao.split()
    return " ".join(palavras[:2])

def buscar_regimes_db(data_referencia="2026-01-31"):
    regimes = {}
    conn_str = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        query = f"""
            SELECT p.codi_emp, p.descricao_par, p.rfed_par
            FROM bethadba.efparametro_vigencia p
            WHERE p.vigencia_par = (
                SELECT MAX(vigencia_par)
                FROM bethadba.efparametro_vigencia
                WHERE codi_emp = p.codi_emp
                  AND vigencia_par <= '{data_referencia}'
            )
        """
        cursor.execute(query)
        for row in cursor.fetchall():
            cod_emp = row[0]
            desc = row[1]
            rfed = row[2]
            regimes[cod_emp] = mapear_regime(desc, rfed)
        conn.close()
        return regimes
    except Exception as e:
        print(f"Erro no banco de dados: {e}")
        return {}

def atualizar_aba_012026():
    path = r"C:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS.xlsx"
    
    if not os.path.exists(path):
        print(f"Arquivo não encontrado: {path}")
        return

    print("Carregando regimes de 01/2026 do banco...")
    regimes_db = buscar_regimes_db("2026-01-31")
    print(f"Total de {len(regimes_db)} empresas carregadas.")

    print(f"Abrindo planilha: {path}")
    wb = openpyxl.load_workbook(path)
    
    if '01.2026' not in wb.sheetnames:
        print("Erro: Aba 01.2026 não encontrada.")
        return
        
    ws = wb['01.2026']
    print("Processando aba 01.2026 - Coluna C (REGIME)...")
    
    atualizados = 0
    nao_encontrados = 0
    
    for row in range(2, ws.max_row + 1):
        cod_emp = ws.cell(row=row, column=1).value
        if cod_emp and isinstance(cod_emp, (int, float)):
            cod = int(cod_emp)
            regime = regimes_db.get(cod)
            if regime:
                ws.cell(row=row, column=3).value = regime
                atualizados += 1
            else:
                ws.cell(row=row, column=3).value = "Não Localizado"
                nao_encontrados += 1
                
    try:
        wb.save(path)
        print(f"Sucesso! Aba 01.2026 atualizada.")
        print(f"Empresas atualizadas: {atualizados}")
        print(f"Empresas não localizadas: {nao_encontrados}")
    except Exception as e:
        print(f"Erro ao salvar: {e}")

if __name__ == "__main__":
    atualizar_aba_012026()
