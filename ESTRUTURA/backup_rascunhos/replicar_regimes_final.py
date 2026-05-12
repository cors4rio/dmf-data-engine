import openpyxl
import os

def replicar_regimes_validados():
    path = r"C:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS.xlsx"
    
    if not os.path.exists(path):
        print(f"Erro: Arquivo não encontrado em {path}")
        return

    print(f"Abrindo planilha para réplica: {path}")
    wb = openpyxl.load_workbook(path)
    
    # 1. Carregar o Mapa de Regimes da aba 01.2026 (Referência Validada)
    if '01.2026' not in wb.sheetnames:
        print("Erro: Aba de referência 01.2026 não encontrada.")
        return
        
    ws_ref = wb['01.2026']
    mapa_regimes = {}
    
    print("Mapeando regimes da aba 01.2026...")
    for row in range(2, ws_ref.max_row + 1):
        cod_emp = ws_ref.cell(row=row, column=1).value
        # Na aba 01.2026 a Coluna C (3) já está preenchida corretamente
        regime = ws_ref.cell(row=row, column=3).value
        if cod_emp and regime:
            try:
                mapa_regimes[int(cod_emp)] = regime
            except:
                continue
                
    print(f"Mapa criado com {len(mapa_regimes)} empresas.")

    # 2. Replicar para as demais abas (Exceto 01.2025 e a própria 01.2026)
    abas_processadas = 0
    total_celulas_atualizadas = 0
    
    for sheet_name in wb.sheetnames:
        # Pular aba de referência, aba EXEMPLO e a aba 01.2025 solicitada pelo usuário
        if sheet_name in ['01.2026', '01.2025', 'EXEMPLO']:
            print(f"-> Pulando aba: {sheet_name}")
            continue
            
        print(f"-> Processando aba: {sheet_name}...")
        ws = wb[sheet_name]
        
        # Garantir que a Coluna C seja limpa/atualizada conforme o mapa
        # O usuário disse que o campo REGIME está em C1.
        # Vamos assumir que a Coluna A é COD EMPRESA em todas as abas.
        
        ws.cell(row=1, column=3).value = "REGIME" # Garante o cabeçalho
        
        for row in range(2, ws.max_row + 1):
            cod_val = ws.cell(row=row, column=1).value
            if cod_val:
                try:
                    cod = int(cod_val)
                    regime_para_aplicar = mapa_regimes.get(cod)
                    if regime_para_aplicar:
                        ws.cell(row=row, column=3).value = regime_para_aplicar
                        total_celulas_atualizadas += 1
                    else:
                        ws.cell(row=row, column=3).value = "Não Localizado na Ref"
                except:
                    continue
        
        abas_processadas += 1

    try:
        wb.save(path)
        print(f"\nRéplica concluída!")
        print(f"Abas atualizadas: {abas_processadas}")
        print(f"Total de registros inseridos: {total_celulas_atualizadas}")
    except Exception as e:
        print(f"Erro ao salvar: {e}")

if __name__ == "__main__":
    replicar_regimes_validados()
