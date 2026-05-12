import pyodbc
import openpyxl
import sys

def main():
    conn_str = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
    try:
        print("Conectando ao banco de dados...")
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        
        data_inicio = '2026-04-01'
        data_fim = '2026-04-30'
        
        # 1. QUANTIDADE DE LANÇAMENTOS CONTÁBEIS
        print("Buscando Lancamentos Contabeis...")
        query_lancamentos = f"""
            SELECT 
                codi_emp as Codigo_Cliente,
                SUM(CASE WHEN orig_lan = 1 THEN 1 ELSE 0 END) as Lancamentos_Normal,
                SUM(CASE WHEN orig_lan = 39 THEN 1 ELSE 0 END) as Lancamentos_Extrato_Bancario
            FROM 
                bethadba.ctlancto
            WHERE 
                data_lan >= '{data_inicio}' 
                AND data_lan <= '{data_fim}'
            GROUP BY 
                codi_emp
            HAVING 
                SUM(CASE WHEN orig_lan IN (1, 39) THEN 1 ELSE 0 END) > 0
        """
        cursor.execute(query_lancamentos)
        lancamentos_results = cursor.fetchall()
        dict_lancamentos = {}
        for row in lancamentos_results:
            dict_lancamentos[row.Codigo_Cliente] = (row.Lancamentos_Normal or 0) + (row.Lancamentos_Extrato_Bancario or 0)
            
        # 2. FATURAMENTO
        print("Buscando Faturamento...")
        query_faturamento = f"""
            SELECT 
                codi_emp as Codigo_Cliente, 
                SUM(total_contabil) as Faturamento_Mensal
            FROM (
                SELECT s.codi_emp, SUM(s.vcon_sai) as total_contabil 
                FROM bethadba.efsaidas s
                LEFT JOIN bethadba.efnatureza n ON s.codi_nat = n.codi_nat
                WHERE s.dsai_sai >= '{data_inicio}' AND s.dsai_sai <= '{data_fim}' 
                  AND s.situacao_sai NOT IN (2, 9)
                  AND (n.masc_nat IS NULL OR (
                      n.masc_nat NOT LIKE '5.9%' AND 
                      n.masc_nat NOT LIKE '6.9%' AND 
                      n.masc_nat NOT LIKE '7.9%' AND 
                      n.masc_nat NOT LIKE '%.202%' AND 
                      n.masc_nat NOT LIKE '%.411%'
                  ))
                GROUP BY s.codi_emp
                
                UNION ALL
                
                SELECT codi_emp, SUM(vcon_ser) as total_contabil 
                FROM bethadba.efservicos 
                WHERE dser_ser >= '{data_inicio}' AND dser_ser <= '{data_fim}' 
                  AND situacao_ser NOT IN (2, 9)
                GROUP BY codi_emp
            ) base
            GROUP BY 
                codi_emp
        """
        cursor.execute(query_faturamento)
        faturamento_results = cursor.fetchall()
        dict_faturamento = {}
        for row in faturamento_results:
            dict_faturamento[row.Codigo_Cliente] = float(row.Faturamento_Mensal or 0)

        # 3. TEM FOLHA
        print("Buscando dados de Folha...")
        query_folha = f"""
            SELECT
                e.codi_emp as Codigo_Cliente,
                SUM(CASE WHEN e.vinculo IN (1,6,11) THEN 1 ELSE 0 END) as Qtd_Empregados_Ativos
            FROM bethadba.foempregados e
            LEFT JOIN bethadba.forescisoes r
                ON r.codi_emp      = e.codi_emp
               AND r.i_empregados  = e.i_empregados
               AND r.demissao      < '{data_inicio}'
            WHERE e.admissao <= '{data_fim}'
              AND r.i_empregados IS NULL
            GROUP BY e.codi_emp
        """
        cursor.execute(query_folha)
        folha_results = cursor.fetchall()
        dict_folha = {}
        for row in folha_results:
            dict_folha[row.Codigo_Cliente] = row.Qtd_Empregados_Ativos

        print("Dados extraídos do banco de dados. Atualizando planilha...")
        
        # Atualizar a planilha
        wb_path = 'HORAS CONTABEIS.xlsx'
        sheet_name = '04.2026'
        
        wb = openpyxl.load_workbook(wb_path)
        if sheet_name not in wb.sheetnames:
            print(f"Erro: Aba '{sheet_name}' não encontrada na planilha.")
            return
            
        ws = wb[sheet_name]
        
        COL_COD = 1         # A
        COL_QTD_LANC = 6    # F
        COL_MES_LANC = 7    # G
        COL_TEM_FOLHA = 9   # I
        COL_FATURAMENTO = 15 # O
        
        atualizados_lancamentos = 0
        atualizados_faturamento = 0
        atualizados_folha = 0
        
        for row in range(2, ws.max_row + 1):
            cod_val = ws.cell(row=row, column=COL_COD).value
            if cod_val is None:
                continue
            
            try:
                cod = int(float(str(cod_val).strip()))
            except ValueError:
                continue
            
            # Lançamentos Contábeis
            qtd_lanc = dict_lancamentos.get(cod, 0)
            ws.cell(row=row, column=COL_QTD_LANC).value = qtd_lanc
            if qtd_lanc > 0:
                ws.cell(row=row, column=COL_MES_LANC).value = "04/2026"
            atualizados_lancamentos += 1
            
            # Faturamento
            faturamento = dict_faturamento.get(cod, 0.0)
            ws.cell(row=row, column=COL_FATURAMENTO).value = faturamento
            atualizados_faturamento += 1
            
            # Tem Folha
            qtd_emp = dict_folha.get(cod, 0)
            tem_folha_str = "SIM" if qtd_emp > 0 else "NÃO"
            ws.cell(row=row, column=COL_TEM_FOLHA).value = tem_folha_str
            atualizados_folha += 1

        wb.save(wb_path)
        print(f"Sucesso! Planilha {wb_path} aba {sheet_name} atualizada.")
        print(f"Linhas atualizadas: {atualizados_lancamentos} empresas.")
        
    except Exception as e:
        print(f"Erro Fatal: {e}", file=sys.stderr)

if __name__ == '__main__':
    main()
