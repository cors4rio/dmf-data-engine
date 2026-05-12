import openpyxl
import xlrd
import os
import re

# Configuracoes
MES_ALVO = '03.2026'
BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = os.path.join(BASE_DIR, "CONTROLE DE HORAS DMF - CLIENTE.xlsm")
PLANILHA_CAROL = os.path.join(BASE_DIR, "Controle de Empregados (CAROL)032026.xls")
ARQUIVO_DP_NAO = os.path.join(BASE_DIR, "nao_faz_setor", "DP NAO.txt")
RELATORIO_MD = os.path.join(BASE_DIR, "relatorio_dp_032026.md")

def limpar_cnpj(cnpj):
    if not cnpj: return ""
    return re.sub(r'\D', '', str(cnpj))

def carregar_excecoes_dp():
    dp_nao_codes = set()
    consultoria_codes = set()
    
    if os.path.exists(ARQUIVO_DP_NAO):
        with open(ARQUIVO_DP_NAO, 'r', encoding='utf-8') as f:
            for linha in f:
                linha = linha.strip()
                if not linha or linha.startswith('EMPRESAS'):
                    continue
                
                eh_consultoria = "FAZ CONSULTORIA" in linha.upper()
                
                codigo = None
                
                parts_tab = linha.split('\t')
                parts_semi = linha.split(';')
                
                if len(parts_tab) > 1:
                    cod_str = parts_tab[0].strip()
                    if cod_str.isdigit():
                        codigo = int(cod_str)
                elif len(parts_semi) > 1:
                    cod_str = parts_semi[0].strip()
                    if cod_str.isdigit():
                        codigo = int(cod_str)
                else:
                    primeiro_espaco = linha.find(' ')
                    if primeiro_espaco > 0:
                        cod_str = linha[:primeiro_espaco].strip()
                        if cod_str.isdigit():
                            codigo = int(cod_str)
                
                if eh_consultoria:
                    if codigo: consultoria_codes.add(codigo)
                else:
                    if codigo: dp_nao_codes.add(codigo)
                    
    return dp_nao_codes, consultoria_codes

def get_dados_carol():
    # Lê a planilha de empregados da Carol para obter total_ativos por empresa
    # Coluna 1: Codigo (index 1)
    # Coluna 5: CNPJ (index 5)
    # Colunas de Qtde: 7, 9, 11
    resultados = {}
    cnpjs = {}
    
    if not os.path.exists(PLANILHA_CAROL):
        print(f"Arquivo não encontrado: {PLANILHA_CAROL}")
        return resultados, cnpjs
        
    wb_carol = xlrd.open_workbook(PLANILHA_CAROL)
    sh_carol = wb_carol.sheet_by_index(0)
    for i in range(1, sh_carol.nrows):
        row = sh_carol.row_values(i)
        try:
            cod_raw = row[1]
            if not cod_raw: continue
            cod = int(float(cod_raw))
            
            cnpj = clean_cnpj(row[5] if len(row) > 5 else "")
            
            func = float(row[7] or 0) if len(row) > 7 else 0
            estag = float(row[9] or 0) if len(row) > 9 else 0
            contrib = float(row[11] or 0) if len(row) > 11 else 0
            total_ativos = func + estag + contrib
            
            resultados[cod] = {
                'func': func,
                'estag': estag,
                'contrib': contrib,
                'total': total_ativos
            }
            if cnpj: cnpjs[cod] = str(cnpj)
        except Exception as e:
            pass
            
    return resultados, cnpjs

def clean_cnpj(cnpj):
    if not cnpj: return ""
    return re.sub(r'\D', '', str(cnpj))

def main():
    print("Carregando exceções do DP NÃO...")
    dp_nao_codes, consultoria_codes = carregar_excecoes_dp()
    
    print("Buscando dados da planilha Controle de Empregados (CAROL)...")
    dados_carol, cnpjs_carol = get_dados_carol()
    
    print("Abrindo planilha Master...")
    wb = openpyxl.load_workbook(PLANILHA_MASTER, keep_vba=True)
    if MES_ALVO not in wb.sheetnames:
        print(f"Aba {MES_ALVO} não encontrada.")
        return
    ws = wb[MES_ALVO]
    
    # Mapear linhas da master pelo código (H) e CNPJ (J)
    from collections import defaultdict
    mapa_planilha = defaultdict(list)
    mapa_planilha_cnpj = defaultdict(list)
    
    for row in range(10, ws.max_row + 1):
        cod_cell = ws.cell(row=row, column=8).value # Col H
        cnpj_cell = ws.cell(row=row, column=10).value # Col J
        
        if cod_cell:
            try:
                cod_val = int(float(str(cod_cell).strip()))
                mapa_planilha[cod_val].append(row)
            except ValueError:
                pass
                
        if cnpj_cell:
            mapa_planilha_cnpj[clean_cnpj(cnpj_cell)].append(row)

    não_encontrados = []
    clientes_encontrados = []
    processados = 0
    alterados = 0
    
    print("Aplicando regras e preenchendo coluna Q (Horário Pessoal DP)...")
    MINIMO_5_MIN = (5 / 60) / 24.0
    
    for cod, info in dados_carol.items():
        processados += 1
        total = info['total']
        
        valor_q = None
        if cod in consultoria_codes:
            valor_q = 1.5 / 24.0
        elif cod in dp_nao_codes:
            valor_q = "DP NÃO"
        else:
            if total > 0:
                horas_flt = (total * 0.33) + 1.5
                valor_q = horas_flt / 24.0
            else:
                valor_q = MINIMO_5_MIN
                
        linhas_alvo = []
        if cod in mapa_planilha:
            linhas_alvo = mapa_planilha[cod]
        else:
            cnpj = cnpjs_carol.get(cod, "")
            if cnpj and cnpj in mapa_planilha_cnpj:
                linhas_alvo = mapa_planilha_cnpj[cnpj]

        if linhas_alvo:
            for row_ws in linhas_alvo:
                c = ws.cell(row=row_ws, column=17, value=valor_q) # Coluna Q
                if isinstance(valor_q, float):
                    c.number_format = '[h]:mm:ss'
                alterados += 1
                clientes_encontrados.append((cod, info, valor_q, row_ws))
        else:
            não_encontrados.append((cod, total, valor_q, cnpjs_carol.get(cod, "")))
    
    # Processar DP NAOs que possam nãp estar listados na prancha da Carol
    for cod in (dp_nao_codes | consultoria_codes):
        if cod not in dados_carol:
            valor_q = 1.5 / 24.0 if cod in consultoria_codes else "DP NÃO"
            linhas_alvo = []
            if cod in mapa_planilha:
                linhas_alvo = mapa_planilha[cod]
            
            if linhas_alvo:
                for row_ws in linhas_alvo:
                    c = ws.cell(row=row_ws, column=17, value=valor_q)
                    if isinstance(valor_q, float):
                        c.number_format = '[h]:mm:ss'
                    alterados += 1
                    clientes_encontrados.append((cod, {'func':0, 'estag':0, 'contrib':0, 'total':0}, valor_q, row_ws))
            else:
                não_encontrados.append((cod, 0, valor_q, "N/A"))
    
    # Atualizar fórmula Q7 para refletir as linhas reais
    max_row = ws.max_row
    ws.cell(row=7, column=17, value=f"=SUBTOTAL(9,Q10:Q{max_row})")
    
    print(f"Salvando planilha... ({alterados} linhas alteradas)")
    wb.save(PLANILHA_MASTER)
    wb.close()
    
    print("Gerando relatório...")
    with open(RELATORIO_MD, 'w', encoding='utf-8') as f:
        f.write("# Relatório de Preenchimento DP - 03.2026\n\n")
        f.write(f"- Clientes processados (Carol XLS + Lista de Exceções): {processados + len(dp_nao_codes) + len(consultoria_codes)}\n")
        f.write(f"- Linhas preenchidas na Master com sucesso: {alterados}\n")
        f.write(f"- Clientes NÃO encontrados na planilha Master: {len(não_encontrados)}\n\n")
        
        if não_encontrados:
            f.write("## Clientes não identificados na planilha (Ignorados)\n\n")
            f.write("| Código | CNPJ | Qtd Empregados | Ação Ideal |\n")
            f.write("|---|---|---|---|\n")
            for item in não_encontrados:
                acao = item[2]
                if isinstance(acao, float):
                    acao = f"{acao*24:.2f} h"
                f.write(f"| {item[0]} | {item[3]} | {item[1]} | {acao} |\n")

    ARQUIVO_CALCULOS = os.path.join(BASE_DIR, "calculos_encontrados_dp_032026.md")
    print("Gerando relatório de cálculos...")
    with open(ARQUIVO_CALCULOS, 'w', encoding='utf-8') as f:
        f.write("# Cálculos e Valores Preenchidos - DP (03.2026)\n\n")
        f.write("| Linha (Excel) | Código Domínio | Func. | Estagiário | Contrib. | Total Empregados | Valor Inserido |\n")
        f.write("|---|---|---|---|---|---|---|\n")
        
        # Sort by row number for better readability
        clientes_ordenados = sorted(clientes_encontrados, key=lambda x: x[3])
        for cod, info, val_q, row_excel in clientes_ordenados:
            acao = val_q
            if isinstance(acao, float):
                # converte para exibição amigável sem float point gigante, em formato horas
                horas_flt = acao * 24.0
                horas_int = int(horas_flt)
                minutos_int = int(round((horas_flt - horas_int) * 60))
                acao = f"{horas_int:02d}:{minutos_int:02d}:00"
                
            f.write(f"| {row_excel} | {cod} | {info['func']} | {info['estag']} | {info['contrib']} | {info['total']} | {acao} |\n")

    print(f"Processo finalizado! Relatório salvo: {RELATORIO_MD}")
    print(f"Cálculos detalhados salvos em: {ARQUIVO_CALCULOS}")
    
if __name__ == '__main__':
    main()
