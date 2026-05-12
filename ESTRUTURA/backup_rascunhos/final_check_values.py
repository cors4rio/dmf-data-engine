import openpyxl

def verify_values():
    path = 'c:/Users/DMF-AUTOMACAO/Documents/PROJETOS/N8N automacao/ESTRUTURA/CONTROLE_DE_HORAS_DMF.xlsm'
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sh = wb['02.2026']
    
    # Check these specific rows
    for r in [88, 93, 761, 781]:
        try:
            row = sh[r]
            code = sh.cell(row=r, column=8).value
            name = sh.cell(row=r, column=9).value
            fiscal = sh.cell(row=r, column=15).value
            contabil = sh.cell(row=r, column=16).value
            pessoal = sh.cell(row=r, column=17).value
            print(f"Row {r}: Code={code}, Name={name}, Contabil={contabil}, Fiscal={fiscal}, Pessoal={pessoal}")
        except: pass

if __name__ == '__main__':
    verify_values()
