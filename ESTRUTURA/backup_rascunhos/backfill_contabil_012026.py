"""
Backfill Janeiro/2026 - Planilha HORAS CONTABEIS_.xlsx
Regras:
1. QTD Lançamentos: 01/2026 (orig_lan IN 1, 39)
2. Faturamento: 11/2025 (efsaidas + efservicos)
3. DP: 01/2026 (Controle de Empregados (CAROL) 012026.xls)
"""

import pyodbc
import openpyxl
import xlrd
import os
import re

# Configurações
DB_CONN_STR = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
PLANILHA_CONTABIL = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_.xlsx'
ABA_DESTINO = '01.2026'
ARQUIVO_CAROL = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\Controle de Empregados (CAROL) 012026.xls'

# Datas
INICIO_LANC = '2026-01-01'
FIM_LANC = '2026-01-31'
INICIO_FAT = '2025-11-01'
FIM_FAT = '2025-11-30'

def clean_cnpj(val):
    if val is None: return ""
    return re.sub(r'\D', '', str(val))

def get_data_dominio():
    print("Conectando ao Domínio...")
    conn = pyodbc.connect(DB_CONN_STR)
    cursor = conn.cursor()

    # 1. Lançamentos 01/2026 (Regra 1+39)
    print(f"Buscando lançamentos ({INICIO_LANC} a {FIM_LANC})...")
    query_lanc = f"""
        SELECT codi_emp, COUNT(*) as qtd
        FROM bethadba.ctlancto
        WHERE data_lan >= '{INICIO_LANC}' AND data_lan <= '{FIM_LANC}'
        AND orig_lan IN (1, 39)
        GROUP BY codi_emp
    """
    cursor.execute(query_lanc)
    lancamentos = {str(row[0]): row[1] for row in cursor.fetchall()}

    # 2. Faturamento 11/2025
    print(f"Buscando faturamento ({INICIO_FAT} a {FIM_FAT})...")
    query_fat = f"""
        SELECT codi_emp, SUM(total_contabil) as faturamento
        FROM (
            SELECT codi_emp, SUM(vcon_sai) as total_contabil 
            FROM bethadba.efsaidas 
            WHERE dsai_sai >= '{INICIO_FAT}' AND dsai_sai <= '{FIM_FAT}' 
            GROUP BY codi_emp
            UNION ALL
            SELECT codi_emp, SUM(vcon_ser) as total_contabil 
            FROM bethadba.efservicos 
            WHERE dser_ser >= '{INICIO_FAT}' AND dser_ser <= '{FIM_FAT}' 
            GROUP BY codi_emp
        ) base
        GROUP BY codi_emp
    """
    cursor.execute(query_fat)
    faturamentos = {str(row[0]): float(row[1]) if row[1] else 0.0 for row in cursor.fetchall()}

    conn.close()
    return lancamentos, faturamentos

def get_data_carol():
    print(f"Lendo Dados DP: {ARQUIVO_CAROL}")
    if not os.path.exists(ARQUIVO_CAROL):
        print("AVISO: Arquivo Carol não encontrado.")
        return {}

    wb = xlrd.open_workbook(ARQUIVO_CAROL)
    sh = wb.sheet_by_index(0)
    dp_data = {}
    
    # Coluna 1: Cod, 7: Func, 9: Estag, 11: Contrib
    for i in range(1, sh.nrows):
        row = sh.row_values(i)
        try:
            cod_val = row[1]
            if not cod_val: continue
            cod = str(int(float(cod_val)))
            
            ativos = float(row[7] or 0) + float(row[9] or 0) + float(row[11] or 0)
            dp_data[cod] = "Sim" if ativos > 0 else "Não"
        except:
            continue
    return dp_data

def executar_backfill():
    # Coletar dados
    lancamentos, faturamentos = get_data_dominio()
    dp_data = get_data_carol()

    # Abrir planilha (SEM keep_vba)
    print(f"Abrindo planilha: {PLANILHA_CONTABIL}")
    wb = openpyxl.load_workbook(PLANILHA_CONTABIL)
    ws = wb[ABA_DESTINO]

    # Preencher
    # A=1 (Cod), F=6 (Lanc), I=9 (DP), O=15 (Fat)
    count = 0
    for r in range(2, ws.max_row + 1):
        cod_raw = ws.cell(row=r, column=1).value
        if not cod_raw: continue
        
        try:
            cod = str(int(float(str(cod_raw).strip())))
        except:
            continue

        # Lançamentos (F)
        ws.cell(row=r, column=6).value = lancamentos.get(cod, 0)

        # DP (I)
        ws.cell(row=r, column=9).value = dp_data.get(cod, "Não")

        # Faturamento (O)
        ws.cell(row=r, column=15).value = round(faturamentos.get(cod, 0.0), 2)
        
        count += 1

    print(f"Salvando planilha... ({count} linhas processadas)")
    wb.save(PLANILHA_CONTABIL)
    print("Sucesso!")

if __name__ == "__main__":
    executar_backfill()
