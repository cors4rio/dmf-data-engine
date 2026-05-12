import openpyxl

arquivo_final = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm"

try:
    wb_data = openpyxl.load_workbook(arquivo_final, data_only=True)
    ws = wb_data["02.2026"]
    
    row = 14
    o = ws.cell(row=row, column=15).value # O (HORARIO FISCAL)
    p = ws.cell(row=row, column=16).value # P (HORARIO CONTÁBIL)
    q = ws.cell(row=row, column=17).value # Q (HORARIO PESSOAL)
    r = ws.cell(row=row, column=18).value # R (TOTAL)
    
    print(f"Linha {row}: O={o}, P={p}, Q={q} | R={r}")
    
    # Verificar se R é aproximadamente O+P+Q
    # (Tratando tempos / floats)
    def to_float(v):
        if v is None: return 0.0
        if isinstance(v, (int, float)): return float(v)
        # Se for datetime.time, converter para horas decimais?
        # Ou se for decimal de dia (Excel), manter.
        return 0.0

except Exception as e:
    print(f"Erro: {e}")
