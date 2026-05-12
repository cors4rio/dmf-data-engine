import openpyxl

file_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_V3_AeqwQXgR.xlsx'

try:
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb_data['01.2026'] # Usar a mais recente
    
    print(f"Aba: 01.2026 | Total Linhas: {sheet.max_row}")
    
    count = 0
    for row in range(2, sheet.max_row + 1):
        h = sheet[f'H{row}'].value
        j = sheet[f'J{row}'].value
        l = sheet[f'L{row}'].value
        n = sheet[f'N{row}'].value
        p = sheet[f'P{row}'].value
        q = sheet[f'Q{row}'].value
        
        # Se houver qualquer valor em H,J,L,N,P
        if any([h, j, l, n, p]):
            count += 1
            if count <= 20: # Amostra das primeiras 20 com dados
                print(f"Linha {row}: H={h}, J={j}, L={l}, N={n}, P={p} | Q={q}")
    
    print(f"Total de linhas com dados nas colunas fonte: {count}")

except Exception as e:
    print(f"Erro: {e}")
