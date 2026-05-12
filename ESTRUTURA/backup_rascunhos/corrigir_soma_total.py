import os
import shutil
import openpyxl

def fix_valor_errors():
    xls_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls'
    temp_xlsx = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\temp_fix_valor.xlsx'
    
    print(f"Criando cópia temporária XLSX...")
    shutil.copy(xls_path, temp_xlsx)

    print(f"Abrindo workbook...")
    wb = openpyxl.load_workbook(temp_xlsx)
    sheet_name = '12.2025'
    if sheet_name not in wb.sheetnames:
        print(f"Erro: Aba {sheet_name} não encontrada.")
        return
        
    ws = wb[sheet_name]
    
    col_total = 18 # Coluna R
    
    count = 0
    max_row = ws.max_row
    
    for row_idx in range(10, max_row + 1):
        # Usar a função SOMA (SUM em inglês no openpyxl) que ignora textos
        # O range é de O até Q (colunas 15 a 17)
        formula = f"=SUM(O{row_idx}:Q{row_idx})"
        cell_total = ws.cell(row=row_idx, column=col_total)
        cell_total.value = formula
        cell_total.number_format = '[h]:mm:ss'
        count += 1

    print(f"Fórmulas corrigidas para SUM() em {count} linhas.")

    wb.save(temp_xlsx)
    wb.close()
    
    shutil.copy(temp_xlsx, xls_path)
    os.remove(temp_xlsx)
    print(f"Arquivo {xls_path} corrigido com sucesso!")

if __name__ == "__main__":
    fix_valor_errors()
