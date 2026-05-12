import openpyxl
import datetime

def time_to_hours(val):
    if val is None or str(val).strip() == '':
        return 0.0
    if isinstance(val, datetime.time):
        return val.hour + (val.minute / 60.0) + (val.second / 3600.0)
    if isinstance(val, datetime.timedelta):
        return val.total_seconds() / 3600.0
    
    # Pode vir float se o Excel interpretar a hora como dia decimal
    try:
        f = float(val)
        return f * 24.0
    except:
        pass

    try:
        parts = str(val).split(':')
        if len(parts) >= 3:
            h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
            return h + (m / 60.0) + (s / 3600.0)
        elif len(parts) == 2:
            h, m = int(parts[0]), int(parts[1])
            return h + (m / 60.0)
    except:
        pass
    
    return 0.0

def hours_to_timedelta(h_decimal):
    # Transforma horas decimais num objecto timedelta pro Excel aceitar bem
    if h_decimal < 0: h_decimal = 0.0
    total_seconds = int(h_decimal * 3600)
    return datetime.timedelta(seconds=total_seconds)

def consolidar():
    FILE_ORIGEM = r"C:\Users\DMF-AUTOMACAO\Downloads\CONTROLE_DE_HORAS_DMF_Cg75VG25.xlsm"
    FILE_MESTRA = r"C:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm"
    
    # 1. Ler a base original de O9
    print("1. Lendo Fiscal da Planilha de Origem _Cg75VG25...")
    wb_orig = openpyxl.load_workbook(FILE_ORIGEM, data_only=True)
    sh_orig = wb_orig["12.2025"]
    
    dict_fiscal_orig = {}
    for row in range(10, sh_orig.max_row + 1):
        cod_cell = sh_orig.cell(row=row, column=8).value
        nome_cell = sh_orig.cell(row=row, column=11).value
        # Coluna O é a index 15 (1=A, 15=O)
        val_o = sh_orig.cell(row=row, column=15).value
        
        cod_str = ""
        try:
            if cod_cell is not None:
                cod_str = str(int(float(str(cod_cell).strip().split('.')[0])))
        except: pass
        
        horas_o = time_to_hours(val_o)
        
        if cod_str:
            dict_fiscal_orig[cod_str] = horas_o
            
        if nome_cell:
            dict_fiscal_orig[str(nome_cell).strip().upper()] = horas_o

    # 2. Injetar na Mestra e calcular Totais
    print(f"2. Extraídos dados fiscais de {len(dict_fiscal_orig)//2} empresas. Abrindo Mestra...")
    wb_m = openpyxl.load_workbook(FILE_MESTRA)
    sh_m = wb_m["12.2025"]
    
    linhas_impactadas = 0
    for row in range(10, sh_m.max_row + 1):
        cod_cell = sh_m.cell(row=row, column=8).value
        nome_cell = sh_m.cell(row=row, column=11).value
        
        cod_str = ""
        try:
            if cod_cell is not None:
                cod_str = str(int(float(str(cod_cell).strip().split('.')[0])))
        except: pass
        
        nome_u = str(nome_cell).strip().upper() if nome_cell else ""
        
        # Resgatar valor O original 
        h_fiscal = dict_fiscal_orig.get(cod_str)
        if h_fiscal is None:
            h_fiscal = dict_fiscal_orig.get(nome_u, 0.0)
            
        # Calcular O + 5%
        h_fiscal_novo = h_fiscal * 1.05
        
        # Puxar valores existentes de P e Q na Mestra (para a SOMA do R)
        val_p = sh_m.cell(row=row, column=16).value
        val_q = sh_m.cell(row=row, column=17).value
        h_contabil = time_to_hours(val_p)
        h_pessoal = time_to_hours(val_q)
        
        # Totalizar as 3 colunas (cálculo cru)
        h_total = h_fiscal_novo + h_contabil + h_pessoal
        
        # Somente injetamos se tivermos um cliente na linha (com codigo ou nome valido)
        if cod_str or (nome_u and nome_u != 'NONE'):
            # Injeta O9 -> O(15) -> formatando
            cell_o = sh_m.cell(row=row, column=15)
            cell_o.value = hours_to_timedelta(h_fiscal_novo)
            cell_o.number_format = '[h]:mm:ss'
            
            # Injeta R9 -> R(18) -> formatando
            cell_r = sh_m.cell(row=row, column=18)
            cell_r.value = hours_to_timedelta(h_total)
            cell_r.number_format = '[h]:mm:ss'
            
            linhas_impactadas += 1

    wb_m.save(FILE_MESTRA)
    print(f"Sucesso! {linhas_impactadas} linhas atualizadas na Mestra com O (+5%) e R (Soma).")

if __name__ == '__main__':
    consolidar()
