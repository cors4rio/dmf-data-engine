import os
import shutil
import openpyxl
from openpyxl.utils import get_column_letter

def fix_total_formula():
    xls_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls'
    temp_xlsx = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\temp_fix_total.xlsx'
    
    print(f"Criando cópia temporária XLSX...")
    shutil.copy(xls_path, temp_xlsx)

    print(f"Abrindo workbook...")
    wb = openpyxl.load_workbook(temp_xlsx)
    sheet_name = '12.2025'
    if sheet_name not in wb.sheetnames:
        print(f"Erro: Aba {sheet_name} não encontrada.")
        return
        
    ws = wb[sheet_name]
    col_pessoal = 17 # Coluna Q
    
    count_converted = 0
    # O usuário mencionou Q10:Q740, mas vimos no log que vai até ~963
    # Vou processar até a última linha com dados
    max_row = ws.max_row
    
    for row_idx in range(10, max_row + 1):
        cell = ws.cell(row=row_idx, column=col_pessoal)
        val = cell.value
        
        if isinstance(val, str) and ':' in val:
            try:
                # Converter HH:MM ou HH:MM:SS para fração de dia (Excel Time)
                parts = val.split(':')
                seconds = 0
                if len(parts) == 2: # HH:MM
                    h, m = map(int, parts)
                    seconds = h * 3600 + m * 60
                elif len(parts) == 3: # HH:MM:SS
                    h, m, s = map(int, parts)
                    seconds = h * 3600 + m * 60 + s
                
                # Excel armazena tempo como fração de 24h (1 dia = 1.0)
                excel_time = seconds / 86400.0
                cell.value = excel_time
                cell.number_format = '[h]:mm:ss'
                count_converted += 1
            except Exception as e:
                # Se falhar (ex: "DP NÃO" ou algo que não é tempo), mantém o original
                pass
        elif isinstance(val, (int, float)):
            # Já é número, apenas garantir o formato de hora
            cell.number_format = '[h]:mm:ss'

    print(f"Valores convertidos para formato numérico de tempo: {count_converted}")

    # Corrigir a célula Q7
    # O SUBTOTAL(9, ...) soma os números. 
    # Vou usar a vírgula para fórmulas no openpyxl (o Excel brasileiro converte para ponto e vírgula na exibição)
    q7 = ws['Q7']
    # Atualizar o range para pegar até a última linha (max_row)
    new_formula = f"=SUBTOTAL(9,Q10:Q{max_row})"
    q7.value = new_formula
    q7.number_format = '[h]:mm:ss'
    
    print(f"Fórmula em Q7 atualizada para: {new_formula}")

    wb.save(temp_xlsx)
    wb.close()
    
    shutil.copy(temp_xlsx, xls_path)
    os.remove(temp_xlsx)
    print(f"Arquivo {xls_path} atualizado com sucesso!")

if __name__ == "__main__":
    fix_total_formula()
