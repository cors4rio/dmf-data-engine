import openpyxl
import shutil
import os

def inspect_with_temp_ext(file_path):
    temp_path = file_path + "_temp.xlsx"
    print(f"Copiando {file_path} para {temp_path}")
    try:
        shutil.copy(file_path, temp_path)
        
        # Agora tenta abrir com openpyxl
        wb = openpyxl.load_workbook(temp_path, data_only=True)
        print(f"Abas encontradas: {wb.sheetnames}")
        
        # Inspeciona a aba 'r9' se existir, senão a primeira
        sheet_to_read = "r9" if "r9" in wb.sheetnames else wb.sheetnames[0]
        sh = wb[sheet_to_read]
        print(f"Inspecionando aba: {sh.title}")
        
        for r in range(1, 21):
            row_data = [sh.cell(row=r, column=c).value for c in range(1, 18)] # Colunas A até Q (17)
            print(f"Linha {r}: {row_data}")
            
    except Exception as e:
        print(f"Erro durante a inspeção: {e}")
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)
            print("Arquivo temporário removido.")

if __name__ == "__main__":
    path = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls"
    inspect_with_temp_ext(path)
