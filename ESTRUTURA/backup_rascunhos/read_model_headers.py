import openpyxl

def read_headers():
    file_path = 'C:/Users/DMF-AUTOMACAO/Documents/PROJETOS/RELATORIOS/PLANILHA CONTABIL/Modelo controle de horas contab.xlsx'
    sheet_name = '01.2025'
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Ler a primeira linha
            headers = [cell.value for cell in ws[1] if cell.value is not None]
            print(f"Cabeçalhos encontrados na aba '{sheet_name}':")
            for i, h in enumerate(headers):
                print(f"  Col {i+1}: {h}")
        else:
            print(f"Aba {sheet_name} não existe.")
    except Exception as e:
        print(f"Erro ao ler a planilha: {e}")

if __name__ == '__main__':
    read_headers()
