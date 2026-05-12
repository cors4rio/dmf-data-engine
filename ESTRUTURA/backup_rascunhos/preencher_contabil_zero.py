import os
import shutil
import openpyxl
import re

def fill_zero_contabil():
    txt_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\ZERADOS CONTABIL.txt'
    xls_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls'
    temp_xlsx = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\temp_zeros_update.xlsx'
    
    print(f"Lendo clientes zerados: {txt_path}")
    
    zero_ids = set()
    with open(txt_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            
            # Formato esperado: ID <tab> NOME <tab> 00:00:00
            # Ou ID ; NOME ; 00:00:00
            parts = re.split(r'[\t;]', line)
            if len(parts) >= 3:
                id_match = parts[0].strip()
                time_val = parts[-1].strip()
                
                if time_val == "00:00:00":
                    # Tentar extrair apenas o número do ID
                    match_cod = re.match(r'^(\d+)', id_match)
                    if match_cod:
                        zero_ids.add(match_cod.group(1))

    print(f"IDs identificados para preenchimento com zero: {len(zero_ids)}")

    print(f"Criando cópia temporária XLSX...")
    shutil.copy(xls_path, temp_xlsx)

    print(f"Abrindo workbook...")
    wb = openpyxl.load_workbook(temp_xlsx)
    sheet_name = '12.2025'
    ws = wb[sheet_name]
    
    col_cod = 8
    col_contabil = 16 # Coluna P
    
    count_filled = 0
    
    for row_idx in range(10, ws.max_row + 1):
        cell_cod = ws.cell(row=row_idx, column=col_cod).value
        if cell_cod is not None:
            cod_str = str(cell_cod).strip().split('.')[0]
            
            if cod_str in zero_ids:
                cell_contabil = ws.cell(row=row_idx, column=col_contabil)
                # Regra: Apenas se o campo estiver totalmente vazio
                if cell_contabil.value is None or str(cell_contabil.value).strip() == "":
                    # Inserir zero como valor de tempo
                    cell_contabil.value = 0.0
                    cell_contabil.number_format = '[h]:mm:ss'
                    count_filled += 1

    print(f"Campos preenchidos com '00:00:00' na coluna P: {count_filled}")

    wb.save(temp_xlsx)
    wb.close()
    
    shutil.copy(temp_xlsx, xls_path)
    os.remove(temp_xlsx)
    print(f"Arquivo {xls_path} atualizado com sucesso.")

if __name__ == "__main__":
    fill_zero_contabil()
