import openpyxl

arquivo_final = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm"

try:
    # Load with keep_vba=True for xlsm
    wb = openpyxl.load_workbook(arquivo_final, keep_vba=True, data_only=True)
    
    print(f"Abas: {wb.sheetnames}")
    
    # Vamos olhar a aba que parece ser a principal (provavelmente a última ou a que tem nomes de meses)
    # Procurar por uma aba que tenha "01.2026" ou similar, ou a primeira que pareça de dados.
    sheet_name = wb.sheetnames[0]
    print(f"Analizando aba: {sheet_name}")
    ws = wb[sheet_name]
    
    # Ver cabeçalhos
    row_headers = 8 # O usuário mencionou P9, então talvez o cabeçalho esteja na linha 8?
    headers = [ws.cell(row=row_headers, column=i).value for i in range(1, 20)]
    print(f"Cabeçalhos (Linha {row_headers}): {headers}")
    
    # Amostra de dados da linha 9 em diante
    for row in range(9, 15):
        row_vals = [ws.cell(row=row, column=i).value for i in range(1, 20)]
        print(f"Linha {row}: {row_vals}")

except Exception as e:
    print(f"Erro: {e}")
