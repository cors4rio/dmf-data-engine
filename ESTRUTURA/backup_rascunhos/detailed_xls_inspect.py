import openpyxl
import shutil
import os

def detailed_inspect(file_path, log_path):
    temp_path = file_path + "_detailed.xlsx"
    shutil.copy(file_path, temp_path)
    
    with open(log_path, 'w', encoding='utf-8') as log:
        try:
            wb = openpyxl.load_workbook(temp_path, data_only=True)
            log.write(f"Abas: {wb.sheetnames}\n\n")
            
            for sheet_name in wb.sheetnames:
                sh = wb[sheet_name]
                log.write(f"--- Aba: {sheet_name} (Max Row: {sh.max_row}, Max Col: {sh.max_column}) ---\n")
                
                # Pega as primeiras 50 linhas para ter certeza
                for r in range(1, min(51, sh.max_row + 1)):
                    row_data = [str(sh.cell(row=r, column=c).value) for c in range(1, min(27, sh.max_column + 1))]
                    log.write(f"R{r}: {' | '.join(row_data)}\n")
                log.write("\n")
                
        except Exception as e:
            log.write(f"Erro: {e}\n")
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)

if __name__ == "__main__":
    xls_file = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls"
    log_file = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\xls_structure_log.txt"
    detailed_inspect(xls_file, log_file)
