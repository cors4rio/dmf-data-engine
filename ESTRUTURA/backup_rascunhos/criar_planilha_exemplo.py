import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def criar_planilha_exemplo():
    # Cria uma nova planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "04.2026"  # Exemplo de aba de mês

    # Configuração visual básica para simular a planilha original
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # Adiciona cabeçalhos genéricos nas linhas 1 a 8 para simular o layout da empresa
    ws.merge_cells('A1:R2')
    titulo = ws.cell(row=1, column=1, value="CONTROLE DE PRODUTIVIDADE (EXEMPLO DE ESTRUTURA)")
    titulo.font = Font(size=14, bold=True)
    titulo.alignment = center_align

    # Cabeçalhos da linha 9 (Linha oficial de headers na spec)
    headers = {
        'H': 'Cód. Domínio',
        'I': 'Nome Fantasia',
        'K': 'Razão Social',
        'N': 'Mês Ant. Fiscal',
        'O': 'Horário Fiscal',
        'P': 'Horário Contábil',
        'Q': 'Horário Pessoal (DP)',
        'R': 'Total'
    }

    # Aplica os cabeçalhos
    for col_letter, text in headers.items():
        cell = ws[f"{col_letter}9"]
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        ws.column_dimensions[col_letter].width = 20

    ws.column_dimensions['K'].width = 40  # Razão Social é maior

    # Linha 10 - Dados Fakes de Cliente
    ws['H10'] = 9999
    ws['I10'] = 'EMPRESA FANTASIA EXEMPLO'
    ws['K10'] = 'EMPRESA FANTASIA EXEMPLO LTDA'
    
    # Inserindo dados de tempo (formato numérico para Excel de fração de dia)
    # Por exemplo: 01:30:00 (1.5 horas) = 1.5 / 24 = 0.0625
    ws['N10'] = 0.041666667 # 01:00:00
    ws['O10'] = 0.0625      # 01:30:00
    ws['P10'] = 0.083333333 # 02:00:00
    ws['Q10'] = 0.020833333 # 00:30:00
    
    # Inserindo Fórmula no Total
    ws['R10'] = '=O10+P10+Q10'

    # Formatando as células de tempo para o formato [h]:mm:ss
    for col in ['N', 'O', 'P', 'Q', 'R']:
        ws[f"{col}10"].number_format = '[h]:mm:ss'
        ws[f"{col}10"].alignment = center_align

    # Salva o arquivo no diretório raiz do projeto
    file_path = '../exemplo_controle_de_horas.xlsx'
    wb.save(file_path)
    print(f"Planilha de exemplo criada com sucesso em: {file_path}")

if __name__ == "__main__":
    criar_planilha_exemplo()
