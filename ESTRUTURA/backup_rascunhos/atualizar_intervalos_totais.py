import os
import shutil
import openpyxl

def update_totals_range():
    xls_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls'
    temp_xlsx = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\temp_update_totals.xlsx'
    
    print("Iniciando atualização das fórmulas de total...")
    shutil.copy(xls_path, temp_xlsx)
    wb = openpyxl.load_workbook(temp_xlsx)
    ws = wb['12.2025']
    
    # Definir o limite superior da planilha (conforme observado anteriormente, as linhas vão até 963)
    # Usaremos 963 para garantir que todas as células formatadas sejam incluídas
    limit = 963
    
    # Atualizar O7 (Fiscal)
    ws['O7'].value = f"=SUBTOTAL(9,O10:O{limit})"
    ws['O7'].number_format = '[h]:mm:ss'
    
    # Atualizar P7 (Contábil)
    ws['P7'].value = f"=SUBTOTAL(9,P10:P{limit})"
    ws['P7'].number_format = '[h]:mm:ss'
    
    # Atualizar Q7 (DP)
    # Já estava em 963, mas vamos reforçar para consistência
    ws['Q7'].value = f"=SUBTOTAL(9,Q10:Q{limit})"
    ws['Q7'].number_format = '[h]:mm:ss'
    
    # Adicionar R7 (Total Geral) caso não exista ou esteja incorreto
    # R10:Rlimit são os totais por linha
    ws['R7'].value = f"=SUBTOTAL(9,R10:R{limit})"
    ws['R7'].number_format = '[h]:mm:ss'
    
    print(f"Intervalos atualizados para 10:{limit}")
    
    wb.save(temp_xlsx)
    wb.close()
    shutil.copy(temp_xlsx, xls_path)
    os.remove(temp_xlsx)
    print("Planilha master atualizada com sucesso.")

if __name__ == "__main__":
    update_totals_range()
