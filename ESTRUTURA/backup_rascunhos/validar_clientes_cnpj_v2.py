import os
import shutil
import openpyxl
import re

def validate_by_cnpj_save_log():
    txt_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\VALIDACAO CNPJ.txt'
    xls_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls'
    temp_xlsx = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\temp_validate_cnpj.xlsx'
    log_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\clientes_faltantes_log.txt'
    
    txt_data = [] # List of (cnpj, name, hours)
    with open(txt_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            
            parts = re.split(r'[\t;]', line)
            if len(parts) >= 3:
                cnpj = parts[0].strip().replace('.', '').replace('-', '').replace('/', '')
                name = parts[1].strip()
                hours = parts[-1].strip()
                txt_data.append({'cnpj': cnpj, 'name': name, 'hours': hours})

    shutil.copy(xls_path, temp_xlsx)
    wb = openpyxl.load_workbook(temp_xlsx, data_only=True)
    sheet_name = '12.2025'
    ws = wb[sheet_name]
    
    sheet_cnpjs = set()
    for row_idx in range(10, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=10).value
        if cell_val:
            clean_cnpj = str(cell_val).strip().replace('.', '').replace('-', '').replace('/', '').split('.')[0]
            # Garantir 14 dígitos com zeros à esquerda se for CNPJ
            if len(clean_cnpj) > 0 and len(clean_cnpj) < 14:
                clean_cnpj = clean_cnpj.zfill(14)
            sheet_cnpjs.add(clean_cnpj)

    missing_clients = []
    for client in txt_data:
        clean_txt_cnpj = client['cnpj']
        if len(clean_txt_cnpj) > 0 and len(clean_txt_cnpj) < 14:
             clean_txt_cnpj = clean_txt_cnpj.zfill(14)
             
        if clean_txt_cnpj not in sheet_cnpjs:
            missing_clients.append(client)

    with open(log_path, 'w', encoding='utf-8') as log:
        log.write(f"Total de clientes no TXT: {len(txt_data)}\n")
        log.write(f"Total de CNPJs na planilha: {len(sheet_cnpjs)}\n")
        log.write(f"Total de clientes faltantes: {len(missing_clients)}\n\n")
        log.write("--- LISTA DE FALTANTES ---\n")
        for c in missing_clients:
            log.write(f"CNPJ: {c['cnpj']} | NOME: {c['name']} | HORA: {c['hours']}\n")

    print(f"Log de faltantes gerado em: {log_path}")
    wb.close()
    os.remove(temp_xlsx)

if __name__ == "__main__":
    validate_by_cnpj_save_log()
