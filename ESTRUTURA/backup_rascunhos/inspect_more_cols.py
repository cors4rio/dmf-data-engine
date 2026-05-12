import openpyxl

arquivo_final = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm"

try:
    wb = openpyxl.load_workbook(arquivo_final, data_only=True)
    sn = "02.2026"
    ws = wb[sn]
    
    # Headers até BA
    headers = {}
    for i in range(1, 40):
        col_letter = openpyxl.utils.get_column_letter(i)
        headers[col_letter] = ws.cell(row=8, column=i).value
        sub_header = ws.cell(row=9, column=i).value
        print(f"{col_letter} (L8: {headers[col_letter]} | L9: {sub_header})")

except Exception as e:
    print(f"Erro: {e}")
