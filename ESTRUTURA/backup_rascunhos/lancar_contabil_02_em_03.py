"""
Lançamento de dados contábeis do MÊS 02/2026 na aba 03.2026 da planilha HORAS CONTABEIS_.xlsx

Origem: Banco Domínio (ODBC) - Período Fevereiro 2026
Destino: HORAS CONTABEIS_.xlsx -> Aba 03.2026
Campos: F (QTD Lancamentos Contabeis) e O (Total Faturamento Mês)

Regras aplicadas:
- Lançamentos: orig_lan IN (1, 39) --> (Manual e Conciliação Bancária)
- Faturamento: efsaidas (vcon_sai) + efservicos (vcon_ser)
- Match por código Domínio (Coluna A da planilha)
- Clientes sem dados no BD ficam com 0 (não em branco)
- NÃO usar keep_vba (arquivo é .xlsx)
"""

import pyodbc
import openpyxl
import os
from collections import defaultdict

# Configurações
DB_CONN_STR = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
PLANILHA_CONTABIL = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_.xlsx'
ABA_DESTINO = '03.2026'

# Período de busca: Fevereiro 2026
DATA_INICIO = '2026-02-01'
DATA_FIM = '2026-02-28'

def get_lancamentos(cursor):
    """Busca QTD lançamentos contábeis por empresa (1 = Normal, 39 = Conciliação)."""
    query = f"""
        SELECT codi_emp, COUNT(*) as qtd_lancamentos
        FROM bethadba.ctlancto
        WHERE data_lan >= '{DATA_INICIO}' AND data_lan <= '{DATA_FIM}'
        AND orig_lan IN (1, 39)
        GROUP BY codi_emp
        ORDER BY codi_emp
    """
    cursor.execute(query)
    result = {}
    for row in cursor.fetchall():
        result[str(row[0])] = row[1]
    return result

def get_faturamento(cursor):
    """Busca faturamento mensal por empresa (efsaidas + efservicos)."""
    query = f"""
        SELECT codi_emp, SUM(total_contabil) as faturamento
        FROM (
            SELECT codi_emp, SUM(vcon_sai) as total_contabil 
            FROM bethadba.efsaidas 
            WHERE dsai_sai >= '{DATA_INICIO}' AND dsai_sai <= '{DATA_FIM}' 
            GROUP BY codi_emp
            UNION ALL
            SELECT codi_emp, SUM(vcon_ser) as total_contabil 
            FROM bethadba.efservicos 
            WHERE dser_ser >= '{DATA_INICIO}' AND dser_ser <= '{DATA_FIM}' 
            GROUP BY codi_emp
        ) base
        GROUP BY codi_emp
    """
    cursor.execute(query)
    result = {}
    for row in cursor.fetchall():
        result[str(row[0])] = float(row[1]) if row[1] else 0.0
    return result

def preencher_contabil():
    # 1. Conectar ao Domínio
    print("Conectando ao banco Domínio...")
    try:
        conn = pyodbc.connect(DB_CONN_STR)
        cursor = conn.cursor()
        print("Conexão estabelecida.")
    except Exception as e:
        print(f"ERRO de conexão ODBC: {e}")
        return

    # 2. Buscar dados de Fevereiro 2026
    print(f"\nBuscando lançamentos contábeis ({DATA_INICIO} a {DATA_FIM}) [Regra 1+39]...")
    lancamentos = get_lancamentos(cursor)
    print(f"  {len(lancamentos)} empresas com lançamentos encontradas.")

    print(f"\nBuscando faturamento ({DATA_INICIO} a {DATA_FIM})...")
    faturamento = get_faturamento(cursor)
    print(f"  {len(faturamento)} empresas com faturamento encontradas.")

    conn.close()
    print("Conexão fechada.")

    # 3. Abrir planilha contábil (SEM keep_vba - é .xlsx!)
    print(f"\nAbrindo planilha: {PLANILHA_CONTABIL}")
    wb = openpyxl.load_workbook(PLANILHA_CONTABIL)
    ws = wb[ABA_DESTINO]

    # 4. Iterar todas as linhas e preencher
    count_lanc = 0
    count_fat = 0
    count_zero_lanc = 0
    count_zero_fat = 0

    for row in range(2, ws.max_row + 1):
        cod_raw = ws.cell(row=row, column=1).value  # Coluna A - Cod Dominio
        if cod_raw is None:
            continue
        
        # Tratar código: pode ser int, float ou texto
        try:
            cod = str(int(float(str(cod_raw).strip())))
        except (ValueError, TypeError):
            continue

        # Coluna F (6) - QTD Lançamentos Contábeis
        if cod in lancamentos:
            ws.cell(row=row, column=6).value = lancamentos[cod]
            count_lanc += 1
        else:
            ws.cell(row=row, column=6).value = 0
            count_zero_lanc += 1

        # Coluna O (15) - Total Faturamento Mês
        if cod in faturamento:
            ws.cell(row=row, column=15).value = round(faturamento[cod], 2)
            count_fat += 1
        else:
            ws.cell(row=row, column=15).value = 0
            count_zero_fat += 1

    # 5. Salvar
    print(f"\nSalvando planilha...")
    wb.save(PLANILHA_CONTABIL)

    print(f"\n=== RESUMO DO LANCAMENTO ===")
    print(f"Periodo: {DATA_INICIO} a {DATA_FIM}")
    print(f"Aba destino: {ABA_DESTINO}")
    print(f"Lancamentos preenchidos (1+39): {count_lanc}")
    print(f"Lancamentos zerados: {count_zero_lanc}")
    print(f"Faturamentos preenchidos: {count_fat}")
    print(f"Faturamentos zerados: {count_zero_fat}")
    print("Sucesso!")

if __name__ == "__main__":
    preencher_contabil()
