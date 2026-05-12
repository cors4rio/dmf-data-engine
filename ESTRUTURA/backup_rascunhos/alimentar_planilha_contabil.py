import pyodbc
import openpyxl
from collections import defaultdict
import datetime
import calendar
import os
from config import *
from utils import logger, time_to_float_excel, find_client_row

class DominioExtractor:
    """
    Extrator de Dados do Banco Domínio (ODBC).
    Implementa as queries validadas em N8N_Queries_Planilha_Contabil.md para uso local.
    """
    
    def __init__(self, mes_alvo=CURRENT_SHEET):
        self.mes_alvo = mes_alvo
        # Calcular datas de início e fim
        m, y = self.mes_alvo.split('.')
        self.data_inicio = f"{y}-{m}-01"
        self.data_fim = f"{y}-{m}-{calendar.monthrange(int(y), int(m))[1]}"
        
    def get_connection(self):
        """Estabelece conexão ODBC com o banco Domínio."""
        try:
            return pyodbc.connect(DB_CONN_STR)
        except Exception as e:
            logger.error(f"Falha na conexão ODBC: {e}")
            return None

    def fetch_indicators(self):
        """Retorna dicionários com indicadores por código de empresa."""
        conn = self.get_connection()
        if not conn: return None
        
        indicators = {
            "lancamentos": defaultdict(int),
            "faturamento": defaultdict(float),
            "folha": defaultdict(int),
            "regime": defaultdict(str)
        }
        
        cursor = conn.cursor()
        
        # 1. Lançamentos Contábeis (Válido para Coluna P se necessário)
        q_lan = f"""
            SELECT codi_emp, SUM(CASE WHEN orig_lan IN (1, 5) THEN 1 ELSE 0 END) as qtd
            FROM bethadba.ctlancto
            WHERE data_lan >= '{self.data_inicio}' AND data_lan <= '{self.data_fim}'
            GROUP BY codi_emp
        """
        
        # 2. Faturamento Mensal
        q_fat = f"""
            SELECT codi_emp, SUM(tot) as fat
            FROM (
                SELECT codi_emp, SUM(vcon_sai) as tot FROM bethadba.efsaidas 
                WHERE dsai_sai >= '{self.data_inicio}' AND dsai_sai <= '{self.data_fim}' GROUP BY codi_emp
                UNION ALL
                SELECT codi_emp, SUM(vcon_ser) as tot FROM bethadba.efservicos 
                WHERE dser_ser >= '{self.data_inicio}' AND dser_ser <= '{self.data_fim}' GROUP BY codi_emp
            ) as fat_mensal GROUP BY codi_emp
        """
        
        # 3. Folha de Pagamento (Empregados Ativos)
        q_folha = f"""
            SELECT codi_emp, COUNT(*) as ativos
            FROM bethadba.foempregados
            WHERE admissao <= '{self.data_fim}'
              AND (demissao IS NULL OR demissao >= '{self.data_inicio}')
            GROUP BY codi_emp
        """
        
        # Execução Segura (Estilo Elite Audit)
        queries = [
            (q_lan, "lancamentos"),
            (q_fat, "faturamento"),
            (q_folha, "folha")
        ]
        
        for query, key in queries:
            try:
                logger.info(f"Executando query de {key.capitalize()}...")
                cursor.execute(query)
                for row in cursor.fetchall():
                    indicators[key][str(row[0])] = row[1]
            except Exception as e:
                logger.error(f"Erro na query {key}: {e}")
        
        conn.close()
        return indicators

    def update_master_v3(self, indicators):
        """Alimenta a planilha Master com os dados extraídos."""
        logger.info(f"Abrindo Master para atualização de indicadores: {TARGET_MASTER}")
        wb = openpyxl.load_workbook(TARGET_MASTER, keep_vba=True)
        if self.mes_alvo not in wb.sheetnames:
            logger.error(f"Aba {self.mes_alvo} não encontrada.")
            return

        ws = wb[self.mes_alvo]
        processados = 0
        
        for row in range(START_ROW, ws.max_row + 1):
            cod_val = ws.cell(row=row, column=COL_CODI_EMP).value
            if not cod_val: continue
            
            try:
                cod = str(int(float(str(cod_val).strip())))
                
                # Exemplo: Atualizar Faturamento se houver coluna para isso (opcional na Master v1.0)
                # No momento, a Master foca em Horas (O, P, Q).
                # Se houver indicadores extras, adicionar colunas no config.py.
                
                processados += 1
            except: continue
            
        wb.save(TEMP_OUTPUT)
        logger.info(f"Fim do processamento de indicadores. Arquivos em: {TEMP_OUTPUT}")

if __name__ == "__main__":
    extractor = DominioExtractor()
    data = extractor.fetch_indicators()
    if data:
        print(f"Sucesso! Capturados dados de {len(data['folha'])} empresas com folha.")
