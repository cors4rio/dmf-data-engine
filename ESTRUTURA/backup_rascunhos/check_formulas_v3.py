import openpyxl

def check_formulas():
    xls_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_V3.xlsx'
    
    print(f"Lendo workbook...")
    wb = openpyxl.load_workbook(xls_path, data_only=False)
    sheet_name = wb.sheetnames[-1] # Tentar a ultima aba ou '12.2025'
    if '12.2025' in wb.sheetnames:
        sheet_name = '12.2025'
        
    print(f"Usando aba: {sheet_name}")
    ws = wb[sheet_name]
    max_row = ws.max_row
    
    cells_to_check = ['O7', 'P7', 'Q7', 'R7', 'O9', 'P9', 'Q9', 'R9']
    
    print(f"Aba: {sheet_name} | Total de linhas: {max_row}")
    for cell_ref in cells_to_check:
        try:
            cell = ws[cell_ref]
            val = cell.value
            print(f"[{cell_ref}] -> Fórmula/Valor: {val}")
        except Exception as e:
            pass

if __name__ == "__main__":
    check_formulas()
