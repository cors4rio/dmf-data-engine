import openpyxl

BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = BASE_DIR + r"\CONTROLE DE HORAS DMF - CLIENTE.xlsm"

wb = openpyxl.load_workbook(PLANILHA_MASTER, data_only=True)
ws = wb['03.2026']

for row in range(10, ws.max_row + 1):
    c = ws.cell(row=row, column=8).value
    if c == 1152 or str(c).strip() == '1152' or str(c).strip() == '1152.0':
        print(f"Row {row}: {c}")

