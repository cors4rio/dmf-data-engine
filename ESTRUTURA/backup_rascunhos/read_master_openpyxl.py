import openpyxl

def inspect_master(file_path):
    print(f"Lendo Master com openpyxl: {file_path}")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(f"Abas: {wb.sheetnames}")
        
        # Inspeciona a aba mencionada ou a primeira
        if "12.2025" in wb.sheetnames:
            sh = wb["12.2025"]
        else:
            sh = wb.active
            
        print(f"Aba: {sh.title}, Max Row: {sh.max_row}, Max Col: {sh.max_column}")
        
        # Mostra as primeiras 25 linhas
        for row in range(1, 26):
            vals = [sh.cell(row=row, column=col).value for col in range(1, min(20, sh.max_column + 1))]
            print(f"Linha {row}: {vals}")

    except Exception as e:
        print(f"Erro ao ler com openpyxl: {e}")

if __name__ == "__main__":
    path = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls"
    inspect_master(path)
