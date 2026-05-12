import openpyxl
import xlrd
import os
import re

BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = os.path.join(BASE_DIR, "CONTROLE DE HORAS DMF - CLIENTE.xlsm")
PLANILHA_CAROL = os.path.join(BASE_DIR, "Controle de Empregados (CAROL)032026.xls")
ARQUIVO_DP_NAO = os.path.join(BASE_DIR, "nao_faz_setor", "DP NAO.txt")

def carregar_excecoes_dp():
    dp_nao_codes = set()
    consultoria_codes = set()
    if os.path.exists(ARQUIVO_DP_NAO):
        with open(ARQUIVO_DP_NAO, 'r', encoding='utf-8') as f:
            for linha in f:
                linha = linha.strip()
                if 'FAZ CONSULTORIA' in linha.upper():
                    codigo = 1 # dummy
                else:
                    codigo = 1 # dummy
    return dp_nao_codes, consultoria_codes

dp_nao_codes = set()
consultoria_codes = set()

print("Buscando dados da planilha Controle de Empregados (CAROL)...")
resultados = {}
cnpjs = {}
wb_carol = xlrd.open_workbook(PLANILHA_CAROL)
sh_carol = wb_carol.sheet_by_index(0)
for i in range(1, sh_carol.nrows):
    row = sh_carol.row_values(i)
    try:
        cod_raw = row[1]
        if not cod_raw: continue
        cod = int(float(cod_raw))
        if cod == 1152:
            print("Found 1152 in CAROL.")
            resultados[cod] = 0.0 # hardcoded test
            cnpjs[cod] = "63936347000160"
    except Exception as e:
        pass

print("Abrindo planilha Master...")
wb = openpyxl.load_workbook(PLANILHA_MASTER, keep_vba=True)
ws = wb['03.2026']

mapa_planilha = {}
mapa_planilha_cnpj = {}

for row in range(10, ws.max_row + 1):
    cod_cell = ws.cell(row=row, column=8).value
    cnpj_cell = ws.cell(row=row, column=10).value
    
    if cod_cell and str(cod_cell).strip().isdigit():
        mapa_planilha[int(cod_cell)] = row

print(f"Is 1152 in mapa_planilha? {1152 in mapa_planilha}")
if 1152 in mapa_planilha:
    print(f"Row for 1152 is {mapa_planilha[1152]}")

for cod, total in resultados.items():
    if cod != 1152: continue
    row_ws = None
    if cod in mapa_planilha:
        row_ws = mapa_planilha[cod]
    else:
        print("1152 NOT in mapa_planilha during loop!!")
    
    if row_ws:
        print(f"SUCCESS: 1152 mapped to row {row_ws}")
    else:
        print(f"FAILED: 1152 pushed to não_encontrados")
