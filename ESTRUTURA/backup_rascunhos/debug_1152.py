import openpyxl
import xlrd
import os

BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = os.path.join(BASE_DIR, "CONTROLE DE HORAS DMF - CLIENTE.xlsm")
PLANILHA_CAROL = os.path.join(BASE_DIR, "Controle de Empregados (CAROL)032026.xls")

print("=== LENDO PLANILHA CAROL (CLIENTE 1152) ===")
wb_carol = xlrd.open_workbook(PLANILHA_CAROL)
sh_carol = wb_carol.sheet_by_index(0)
for i in range(1, sh_carol.nrows):
    row = sh_carol.row_values(i)
    try:
        cod_raw = row[1]
        if not cod_raw: continue
        cod = int(float(cod_raw))
        if cod == 1152:
            print(f"Linha {i}: Cod={cod_raw} -> Type: {type(cod_raw)}")
            print(f"  CNPJ: {row[5] if len(row) > 5 else ''}")
            print(f"  Func: {row[7] if len(row) > 7 else ''}")
    except Exception as e:
        pass

print("\n=== LENDO PLANILHA MASTER (CLIENTE 1152) ===")
wb = openpyxl.load_workbook(PLANILHA_MASTER, data_only=True)
ws = wb['03.2026']

encontrados = 0
for row in range(10, ws.max_row + 1):
    cod_cell = ws.cell(row=row, column=8).value
    cnpj_cell = ws.cell(row=row, column=10).value
    
    if cod_cell and str(cod_cell).strip().replace('.0','').isdigit():
        cod = int(float(str(cod_cell).strip()))
        if cod == 1152:
            encontrados += 1
            print(f"Linha {row}: Valor Original={cod_cell!r} -> Type: {type(cod_cell)}")
            print(f"  CNPJ: {cnpj_cell!r}")
            print(f"  Coluna Q atual: {ws.cell(row=row, column=17).value}")

if encontrados == 0:
    print("Nenhuma linha com numero 1152 na coluna 8 da Master.")
