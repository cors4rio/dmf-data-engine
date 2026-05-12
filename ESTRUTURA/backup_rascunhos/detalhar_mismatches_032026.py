import openpyxl
import re

FONTE_CONTABIL = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_.xlsx'
DESTINO_MASTER = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF.xlsx'
ABA = '03.2026'

def clean_cnpj(val):
    if val is None: return ""
    return re.sub(r'\D', '', str(val))

def clean_code(val):
    if val is None: return ""
    try:
        return str(int(float(str(val).strip())))
    except:
        return ""

def main():
    # 1. Carregar Fonte
    print("Mapeando fonte...")
    wb_src = openpyxl.load_workbook(FONTE_CONTABIL, read_only=True, data_only=True)
    ws_src = wb_src[ABA]
    
    src_map = {}
    for r in range(2, ws_src.max_row + 1):
        cod = clean_code(ws_src.cell(row=r, column=1).value)
        cnpj = clean_cnpj(ws_src.cell(row=r, column=3).value)
        if cod:
            src_map[cod] = cnpj
    wb_src.close()

    # 2. Analisar Master
    print("Analisando Master...")
    wb_dst = openpyxl.load_workbook(DESTINO_MASTER, read_only=True)
    ws_dst = wb_dst[ABA]

    print("\n--- LISTA DE EMPRESAS NÃO INTEGRADAS (MOTIVOS) ---\n")
    print(f"{'COD':<6} | {'EMPRESA':<40} | {'MOTIVO':<30}")
    print("-" * 85)

    count_missing = 0
    count_divergent = 0

    for r in range(10, ws_dst.max_row + 1):
        cod = clean_code(ws_dst.cell(row=r, column=8).value)
        cnpj_master = clean_cnpj(ws_dst.cell(row=r, column=10).value)
        nome = ws_dst.cell(row=r, column=11).value or "S/ Nome"

        if not cod: continue

        if cod not in src_map:
            print(f"{cod:<6} | {str(nome)[:40]:<40} | Código não consta na Fonte")
            count_missing += 1
        elif src_map[cod] != cnpj_master:
            print(f"{cod:<6} | {str(nome)[:40]:<40} | CNPJ Divergente (Fonte: {src_map[cod]})")
            count_divergent += 1

    print("\n" + "="*85)
    print(f"Total não localizados na fonte: {count_missing}")
    print(f"Total com divergência de CNPJ: {count_divergent}")
    print(f"Total Geral: {count_missing + count_divergent}")
    
    wb_dst.close()

if __name__ == "__main__":
    main()
