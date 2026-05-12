import os
import sys
import pyodbc
from collections import defaultdict
import openpyxl
from datetime import timedelta

# Configurações
PLANILHA_MASTER = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF.xlsx"
ABA_ALVO = "03.2026"  # Onde vamos lançar
DATA_INICIO = '2026-02-01'
DATA_FIM = '2026-02-28' # Base de dados utilizada
FATOR_ADICIONAL_FISCAL = 1.80
DSN = 'Contabil'
PWD = '<SENHA_NO_ENV>'

def get_fiscal_data_from_db():
    conn_str = f'DSN={DSN};UID=<USER_NO_ENV>;PWD={PWD}'
    
    # Query validada no comparativo com o TXT
    query = f"""
    SELECT 
        l.codi_emp as codigo_cliente,
        SUM(DATEDIFF(second, 
            YMD(YEAR(l.data_log), MONTH(l.data_log), DAY(l.data_log)) + l.tini_log, 
            COALESCE(
                YMD(YEAR(l.dfim_log), MONTH(l.dfim_log), DAY(l.dfim_log)) + l.tfim_log, 
                YMD(YEAR(l.data_log), MONTH(l.data_log), DAY(l.data_log)) + l.tfim_log
            )
        )) as total_segundos
    FROM 
        bethadba.geloguser l
    JOIN 
        bethadba.geempre e ON l.codi_emp = e.codi_emp
    WHERE 
        l.sist_log = 5 -- Módulo Fiscal
        AND l.tfim_log IS NOT NULL
        AND l.data_log BETWEEN '{DATA_INICIO}' AND '{DATA_FIM}'
    GROUP BY 
        l.codi_emp
    """
    
    dict_fiscal = {}
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        print(f"Buscando dados no DB Domínio (Fiscal) - Base {DATA_INICIO} a {DATA_FIM}...")
        cursor.execute(query)
        for row in cursor.fetchall():
            cod, segundos = row
            if segundos:
                dict_fiscal[cod] = float(segundos)
    except Exception as e:
        print(f"Erro ODBC: {e}")
        sys.exit(1)
    finally:
        if 'conn' in locals():
            conn.close()
            
    return dict_fiscal

def segundos_para_hms(total_segundos):
    total_segundos = int(total_segundos)
    h = total_segundos // 3600
    m = (total_segundos % 3600) // 60
    s = total_segundos % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

def main():
    # 1. Obter dados do banco
    dados_fiscal = get_fiscal_data_from_db()
    print(f"Total de empresas com hora fiscal encontrada (Fev): {len(dados_fiscal)}")

    # 2. Abrir Master
    print(f"Abrindo planilha Master ({PLANILHA_MASTER})...")
    wb = openpyxl.load_workbook(PLANILHA_MASTER)

    if ABA_ALVO not in wb.sheetnames:
        print(f"ERRO: Aba {ABA_ALVO} não encontrada!")
        sys.exit(1)
        
    ws = wb[ABA_ALVO]

    # 3. Mapear códigos na planilha (Coluna H = 8)
    mapa_planilha = defaultdict(list)
    for i in range(10, ws.max_row + 1):
        cod_cell = ws.cell(row=i, column=8).value
        # ignorar subcabeçalhos ou linhas mescladas
        if cod_cell is not None:
            try:
                cod = int(float(str(cod_cell).strip()))
                mapa_planilha[cod].append(i)
            except:
                pass

    # 4. Injetar dados
    alterados = 0
    calculos_md = ["# Auditoria Fiscal: Mês 03/2026 (Base 02/2026)\n"]
    calculos_md.append(f"Fator Aplicado: {FATOR_ADICIONAL_FISCAL} (80% extra)\n")
    calculos_md.append("| Cód | Seg. Brutos (Banco) | Seg. Finais (*1.8) | Tempo Master | Status |")
    calculos_md.append("|---|---|---|---|---|")

    for cod, linhas in mapa_planilha.items():
        seg_bruto = dados_fiscal.get(cod, 0)
        seg_final = seg_bruto * FATOR_ADICIONAL_FISCAL
        valor_excel = seg_final / 86400.0 # Formato decimal para data/hora Excel
        hms_str = segundos_para_hms(seg_final)

        for r in linhas:
            cell = ws.cell(row=r, column=15) # Coluna O = 15
            cell.value = valor_excel
            cell.number_format = '[h]:mm:ss'
            alterados += 1
        
        if seg_bruto > 0:
            calculos_md.append(f"| {cod} | {int(seg_bruto)} | {int(seg_final)} | {hms_str} | ✅ Lançado |")

    # 5. Subtotal (O7)
    ws.cell(row=7, column=15, value=f"=SUBTOTAL(9,O10:O{ws.max_row})")

    # 6. Salvar
    print(f"Salvando alterações... ({alterados} células atualizadas)")
    wb.save(PLANILHA_MASTER)
    wb.close()

    # 7. Relatório
    with open("auditoria_fiscal_032026_final.md", "w", encoding="utf-8") as f:
        f.write("\n".join(calculos_md))
    
    print("Processo concluído com sucesso.")
    print("Log gerado: auditoria_fiscal_032026_final.md")

if __name__ == "__main__":
    main()
