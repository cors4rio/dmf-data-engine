import openpyxl
import os

FILE_MASTER = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO_V3.xlsx'
FILE_ACCOUTING = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_.xlsx'

def full_inspection():
    print("--- PESQUISA MASTER SHEET ---")
    if os.path.exists(FILE_MASTER):
        wb = openpyxl.load_workbook(FILE_MASTER, data_only=True)
        ws = wb['03.2026']
        print(f"P1: {ws.cell(1, 16).value}") # P
        print(f"Q1: {ws.cell(1, 17).value}") # Q
        print(f"P9: {ws.cell(9, 16).value}") # Header P
        print(f"Q9: {ws.cell(9, 17).value}") # Header Q
    else:
        print("Master file not found.")

    print("\n--- PESQUISA HORAS CONTABEIS_ MB ---")
    if os.path.exists(FILE_ACCOUTING):
        wb = openpyxl.load_workbook(FILE_ACCOUTING, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            print(f"Aba: {sname}, Q1: {ws.cell(1, 17).value}")
            data_rows = 0
            for r in range(2, min(50, ws.max_row + 1)):
                if ws.cell(r, 17).value is not None:
                    data_rows += 1
            print(f"  Rows 2-50 com dados em Q: {data_rows}")
    else:
        print("Accounting file not found.")

if __name__ == "__main__":
    full_inspection()
