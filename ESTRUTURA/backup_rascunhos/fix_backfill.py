import openpyxl
from datetime import timedelta

FILE = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx'
ABA = '03.2026'

print("Abrindo planilha Master...")
wb = openpyxl.load_workbook(FILE)
ws = wb[ABA]

# =============================================
# PASSO 1: Converter todos os timedeltas em float decimal (fração de dia)
# O Excel interpreta floats como horas corretamente com [h]:mm:ss
# =============================================
convertidos = 0
last_data_row = 10
for r in range(10, ws.max_row + 1):
    c = ws.cell(r, 14)  # Coluna N
    if isinstance(c.value, timedelta):
        c.value = c.value.total_seconds() / 86400.0
        c.number_format = '[h]:mm:ss'
        convertidos += 1
        last_data_row = r
    elif isinstance(c.value, (int, float)) and c.value > 0:
        c.number_format = '[h]:mm:ss'
        last_data_row = r

print(f"Convertidos {convertidos} timedeltas para float decimal.")
print(f"Última linha com dados na coluna N: {last_data_row}")

# =============================================
# PASSO 2: Corrigir o SUBTOTAL para cobrir TODAS as linhas com dados
# =============================================
formula_antiga = ws.cell(7, 14).value
nova_formula = f'=SUBTOTAL(9,N10:N{last_data_row})'
ws.cell(7, 14).value = nova_formula
ws.cell(7, 14).number_format = '[h]:mm:ss'
print(f"SUBTOTAL corrigido: {formula_antiga!r} -> {nova_formula!r}")

# =============================================
# PASSO 3: Verificar a soma total dos dados
# =============================================
soma = 0
count = 0
for r in range(10, last_data_row + 1):
    v = ws.cell(r, 14).value
    if isinstance(v, (int, float)) and v > 0:
        soma += v
        count += 1

total_h = int(soma * 24)
total_m = int((soma * 24 - total_h) * 60)
total_s = int(((soma * 24 - total_h) * 60 - total_m) * 60)
print(f"\nSoma verificada: {count} registros = {total_h}:{total_m:02d}:{total_s:02d}")

# =============================================
# SALVAR
# =============================================
wb.save(FILE)
wb.close()
print("Salvo com sucesso!")
