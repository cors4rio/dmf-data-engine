import openpyxl
import re

FONTE_CONTABIL = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_.xlsx'
DESTINO_MASTER = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF.xlsx'
ABA = '03.2026'

def clean_cnpj(val):
    if val is None: return ""
    return re.sub(r'\D', '', str(val))

def clean_code(val):
    if val is None: return ""
    try:
        return str(int(float(str(val).strip())))
    except:
        return str(val).strip()

def main():
    wb_src = openpyxl.load_workbook(FONTE_CONTABIL, read_only=True)
    ws_src = wb_src[ABA]
    src_map = {}
    for row in ws_src.iter_rows(min_row=2, max_col=3):
        cod = clean_code(row[0].value)
        cnpj = clean_cnpj(row[2].value)
        if cod:
            src_map[cod] = cnpj
    wb_src.close()

    wb_dst = openpyxl.load_workbook(DESTINO_MASTER, read_only=True)
    ws_dst = wb_dst[ABA]
    
    print(f"{'Código':<8} | {'Nome da Empresa':<40} | {'Motivo'}")
    print("-" * 80)
    
    count = 0
    for row in ws_dst.iter_rows(min_row=10, max_col=11):
        cod = clean_code(row[7].value) # Column H (8)
        cnpj_master = clean_cnpj(row[9].value) # Column J (10)
        nome = row[10].value # Column K (11)
        
        if not cod: continue
        
        if cod not in src_map:
            print(f"{cod:<8} | {str(nome)[:40]:<40} | Não consta na planilha Contábil")
            count += 1
        elif src_map[cod] != cnpj_master:
            print(f"{cod:<8} | {str(nome)[:40]:<40} | CNPJ Master ({cnpj_master}) != Fonte ({src_map[cod]})")
            count += 1
            
    print(f"\nTotal: {count} empresas sem integração.")
    wb_dst.close()

if __name__ == "__main__":
    main()
