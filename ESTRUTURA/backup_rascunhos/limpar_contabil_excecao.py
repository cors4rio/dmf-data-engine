import os
import shutil
import openpyxl
import re

def process_contabil_exceptions():
    txt_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\NAO FAZ CONTABIL.txt'
    xls_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls'
    temp_xlsx = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\temp_contabil_update.xlsx'
    
    print(f"Lendo regras de Contabil: {txt_path}")
    
    contabil_nao_codes = set()
    
    with open(txt_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            
            # Extrair código do início da linha
            match_cod = re.match(r'^(\d+)', line)
            if match_cod:
                cod = match_cod.group(1)
                contabil_nao_codes.add(cod)

    print(f"Códigos 'NAO FAZ CONTABIL' identificados: {len(contabil_nao_codes)}")

    print(f"Criando cópia temporária XLSX...")
    shutil.copy(xls_path, temp_xlsx)

    print(f"Abrindo workbook...")
    wb = openpyxl.load_workbook(temp_xlsx)
    sheet_name = '12.2025'
    ws = wb[sheet_name]
    
    col_cod = 8
    col_contabil = 16 # Coluna P
    
    count_updates = 0
    
    for row_idx in range(10, ws.max_row + 1):
        cell_cod = ws.cell(row=row_idx, column=col_cod).value
        if cell_cod is not None:
            cod_str = str(cell_cod).strip().split('.')[0]
            
            if cod_str in contabil_nao_codes:
                ws.cell(row=row_idx, column=col_contabil).value = "NAO FAZ CONTABIL"
                count_updates += 1

    print(f"Atualizações: {count_updates} marcados como 'NAO FAZ CONTABIL' na coluna P.")

    wb.save(temp_xlsx)
    wb.close()
    
    shutil.copy(temp_xlsx, xls_path)
    os.remove(temp_xlsx)
    print(f"Arquivo {xls_path} atualizado com as regras de Contábil.")

if __name__ == "__main__":
    process_contabil_exceptions()
