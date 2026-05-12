import openpyxl

arquivo_atualizado = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF_ATUALIZADO.xlsm"

try:
    wb = openpyxl.load_workbook(arquivo_atualizado, data_only=True)
    ws = wb["02.2026"]
    
    print("Amostra do Arquivo Atualizado (Aba 02.2026):")
    # Tentar algumas linhas interessantes (onde o dry run mostrou dados)
    # Row 10 (347), Row 29 (614), etc.
    rows_to_check = [10, 14, 29]
    for row in rows_to_check:
        cod = ws.cell(row=row, column=8).value
        p = ws.cell(row=row, column=16).value
        r = ws.cell(row=row, column=18).value
        print(f"Linha {row} (Cod {cod}): P={p}, R={r}")

except Exception as e:
    print(f"Erro: {e}")
