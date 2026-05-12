import openpyxl
import os

FILE = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_.xlsx'

def inspect_feb():
    wb = openpyxl.load_workbook(FILE, data_only=True)
    
    # Check 02.2026 structure
    ws02 = wb['02.2026']
    ws03 = wb['03.2026']
    
    print("=== ABA 02.2026 ===")
    print(f"Max row: {ws02.max_row}")
    print(f"Headers: {[ws02.cell(1, c).value for c in range(1, 18)]}")
    
    # Count rows with data in F and O
    count_f = 0
    count_o = 0
    for r in range(2, ws02.max_row + 1):
        f_val = ws02.cell(r, 6).value  # F = QTD Lancamentos
        o_val = ws02.cell(r, 15).value  # O = Faturamento
        if f_val is not None and f_val != 0:
            count_f += 1
        if o_val is not None and o_val != 0:
            count_o += 1
    
    print(f"Linhas com QTD Lancamentos (F) > 0: {count_f}")
    print(f"Linhas com Faturamento (O) > 0: {count_o}")
    
    # Sample first 10 rows with data
    print("\nAmostra (linhas 2-20):")
    for r in range(2, 21):
        cod = ws02.cell(r, 1).value  # A = Cod Dominio
        f_val = ws02.cell(r, 6).value
        o_val = ws02.cell(r, 15).value
        print(f"  Row {r}: Cod={cod}, F(Lanc)={f_val}, O(Fat)={o_val}")
    
    print("\n=== ABA 03.2026 ===")
    print(f"Max row: {ws03.max_row}")
    
    # Check if 03 already has data in F and O
    count_f3 = 0
    count_o3 = 0
    for r in range(2, ws03.max_row + 1):
        f_val = ws03.cell(r, 6).value
        o_val = ws03.cell(r, 15).value
        if f_val is not None and f_val != 0:
            count_f3 += 1
        if o_val is not None and o_val != 0:
            count_o3 += 1
    
    print(f"Linhas com QTD Lancamentos (F) > 0: {count_f3}")
    print(f"Linhas com Faturamento (O) > 0: {count_o3}")
    
    # Check if codes match between tabs
    codes_02 = set()
    codes_03 = set()
    for r in range(2, ws02.max_row + 1):
        v = ws02.cell(r, 1).value
        if v is not None:
            codes_02.add(str(int(float(str(v).strip()))) if isinstance(v, (int, float)) else str(v).strip())
    for r in range(2, ws03.max_row + 1):
        v = ws03.cell(r, 1).value
        if v is not None:
            codes_03.add(str(int(float(str(v).strip()))) if isinstance(v, (int, float)) else str(v).strip())
    
    print(f"\nClientes na aba 02: {len(codes_02)}")
    print(f"Clientes na aba 03: {len(codes_03)}")
    print(f"Clientes em comum: {len(codes_02 & codes_03)}")
    print(f"Só no 02: {len(codes_02 - codes_03)}")
    print(f"Só no 03: {len(codes_03 - codes_02)}")

if __name__ == "__main__":
    inspect_feb()
