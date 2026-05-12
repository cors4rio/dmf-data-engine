import pandas as pd
import openpyxl
import io

def read_master_force_xlsx(file_path):
    print(f"Lendo Master forçando XLSX: {file_path}")
    try:
        # Tenta ler com pandas forçando o motor openpyxl
        df = pd.read_excel(file_path, engine='openpyxl', nrows=25)
        print("Sucesso ao ler primeiras 25 linhas:")
        print(df.to_string())
        
        # Inspecionar abas
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        print(f"Abas: {wb.sheetnames}")
        
    except Exception as e:
        print(f"Erro ao forçar XLSX: {e}")
        
        # Alternativa: copiar para um arquivo temporário com extensão .xlsx
        import shutil
        temp_xlsx = file_path + ".xlsx"
        try:
            shutil.copy(file_path, temp_xlsx)
            wb = load_workbook(temp_xlsx, read_only=True, data_only=True)
            print(f"Abas (via cópia .xlsx): {wb.sheetnames}")
            sh = wb.active # ou "r9" se existir
            print(f"Aba ativa: {sh.title}")
            
            # Remover temp
            import os
            os.remove(temp_xlsx)
        except Exception as e2:
            print(f"Erro ao tentar via cópia: {e2}")

if __name__ == "__main__":
    path = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls"
    read_master_force_xlsx(path)
