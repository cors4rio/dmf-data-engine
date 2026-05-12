import pyodbc
import openpyxl
import os

def mapear_regime(descricao):
    if not descricao:
        return "Não Parametrizado"
    
    desc_upper = descricao.upper()
    
    if "SIMPLES" in desc_upper:
        return "Simples Nacional"
    if any(kw in desc_upper for kw in ["PRESUMIDO", "PRESIMIDO", "PRESSUMIDO"]):
        return "Lucro Presumido"
    if "REAL" in desc_upper:
        return "Lucro Real"
    if "PRODUTOR RURAL" in desc_upper:
        return "Produtor Rural"
    if any(kw in desc_upper for kw in ["IMUNE", "ISENTA"]):
        return "Imune / Isenta"
        
    # Se não bater em nenhum padrão, retorna a descrição original limpa (primeiras 2 palavras)
    palavras = descricao.split()
    return " ".join(palavras[:2])

def atualizar_regimes(mes_alvo="01.2025"):
    # Caminho da planilha
    planilha_path = r"C:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS - REGIMES.xlsx"
    
    if not os.path.exists(planilha_path):
        print(f"Erro: Arquivo não encontrado em {planilha_path}")
        return

    print(f"Iniciando atualização de regimes para {mes_alvo}...")
    
    # 1. Carregar Regimes do Banco
    regimes_db = {}
    conn_str = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        
        # Busca a descrição mais recente de cada empresa até o fim do mês alvo
        # Usando subquery para pegar a última vigência
        query = """
            SELECT p.codi_emp, p.descricao_par, p.rfed_par
            FROM bethadba.efparametro_vigencia p
            WHERE p.vigencia_par = (
                SELECT MAX(vigencia_par)
                FROM bethadba.efparametro_vigencia
                WHERE codi_emp = p.codi_emp
                  AND vigencia_par <= '2025-01-31'
            )
        """
        cursor.execute(query)
        for row in cursor.fetchall():
            cod_emp = row[0]
            desc = row[1]
            rfed = row[2]
            
            # Se a descrição for genérica (ex: "INICIAL", "VIGENCIA INICIAL"), 
            # tenta usar o rfed_par para complementar
            regime_final = mapear_regime(desc)
            if regime_final in ["VIGÊNCIA INICIAL", "INICIAL", "NOVO", "PRIMEIRA VIGENCIA"]:
                if rfed == 1: regime_final = "Lucro Real"
                elif rfed in [2, 4]: regime_final = "Simples Nacional"
                elif rfed == 5: regime_final = "Lucro Presumido"
                elif rfed == 8: regime_final = "Imune / Isenta"
            
            regimes_db[cod_emp] = regime_final
            
        print(f"Dados de {len(regimes_db)} empresas carregados do banco.")
        conn.close()
    except Exception as e:
        print(f"Erro ao acessar banco de dados: {e}")
        return

    # 2. Atualizar Planilha
    try:
        wb = openpyxl.load_workbook(planilha_path)
        if mes_alvo not in wb.sheetnames:
            print(f"Erro: Aba {mes_alvo} não encontrada na planilha.")
            return
            
        ws = wb[mes_alvo]
        print(f"Aba '{mes_alvo}' aberta. Iniciando preenchimento da Coluna D...")
        
        atualizados = 0
        nao_encontrados = 0
        
        # Percorre a partir da linha 2
        for row in range(2, ws.max_row + 1):
            cod_emp = ws.cell(row=row, column=1).value
            
            # Garante que temos um código de empresa válido
            if cod_emp and isinstance(cod_emp, (int, float)):
                cod = int(cod_emp)
                
                regime = regimes_db.get(cod)
                if regime:
                    # Coluna D (4) - REGIME
                    ws.cell(row=row, column=4).value = regime
                    atualizados += 1
                else:
                    ws.cell(row=row, column=4).value = "Não Localizado"
                    nao_encontrados += 1
        
        wb.save(planilha_path)
        print(f"Processo concluído!")
        print(f"Empresas atualizadas: {atualizados}")
        print(f"Empresas não localizadas no banco: {nao_encontrados}")
        print(f"Planilha salva em: {planilha_path}")
        
    except Exception as e:
        print(f"Erro ao manipular Excel: {e}")

if __name__ == "__main__":
    atualizar_regimes("01.2025")
