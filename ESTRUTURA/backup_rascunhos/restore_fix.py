import openpyxl

def restore():
    path = 'c:/Users/DMF-AUTOMACAO/Documents/PROJETOS/N8N automacao/ESTRUTURA/CONTROLE_DE_HORAS_DMF.xlsm'
    print(f"Loading {path}...")
    wb = openpyxl.load_workbook(path, keep_vba=True)
    sh = wb['02.2026']
    
    # 1. Restore rows
    # Insert at 93 (Row 760 becomes 761)
    # Then insert at 781
    print("Inserting rows at 93 and 781...")
    sh.insert_rows(93)
    sh.insert_rows(781)
    
    # Values
    data_1283 = {
        'code': 1283,
        'name': 'BASE ESTRUTURAS LTDA',
        'fiscal': 0.583113425925926,
        'contabil': 0.25,
        'pessoal': 1.125
    }
    data_1195 = {
        'code': 1195,
        'name': 'COMERCIAL VALE DO ARAGUAIA LTDA',
        'fiscal': 0.125,
        'contabil': 0.166666666666667,
        'pessoal': 0.00347222222222222
    }
    
    target_cnpjs = {
        '21425354000158': data_1283,
        '21425354000239': data_1195
    }
    
    print("Updating all occurrences by CNPJ...")
    # Scan up to 1500 rows to find all occurrences
    for r in range(1, 1500):
        try:
            val_j = sh.cell(row=r, column=10).value # Col J (CNPJ)
            if not val_j: continue
            cnpj = str(val_j).strip().replace(".", "").replace("-", "").replace("/", "")
            if cnpj in target_cnpjs:
                d = target_cnpjs[cnpj]
                sh.cell(row=r, column=8, value=d['code'])
                sh.cell(row=r, column=9, value=d['name'])
                sh.cell(row=r, column=15, value=d['fiscal'])
                sh.cell(row=r, column=16, value=d['contabil'])
                sh.cell(row=r, column=17, value=d['pessoal'])
                # Also ensure CNPJ is clean if it was messy
                sh.cell(row=r, column=10, value=cnpj)
                print(f"Updated Row {r} (CNPJ {cnpj})")
        except: break
        
    wb.save(path)
    print(f"Saved changes to {path}.")

if __name__ == '__main__':
    restore()
