import openpyxl

BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = BASE_DIR + r"\CONTROLE DE HORAS DMF - CLIENTE.xlsm"

wb = openpyxl.load_workbook(PLANILHA_MASTER, keep_vba=True)
ws = wb['03.2026']

missing_codes = [1261, 1439, 1649, 780, 1152, 8001]
found_in_sheet = {}

for row in range(10, ws.max_row + 1):
    cod_cell = ws.cell(row=row, column=8).value
    try:
        cod_int = int(float(str(cod_cell).strip()))
        if cod_int in missing_codes:
            found_in_sheet[cod_int] = {
                "row": row,
                "cod_cell_repr": repr(cod_cell),
                "type": type(cod_cell).__name__,
                "isdigit_result": str(cod_cell).strip().isdigit()
            }
    except:
        pass

for k, v in found_in_sheet.items():
    print(f"Code {k}: {v}")
