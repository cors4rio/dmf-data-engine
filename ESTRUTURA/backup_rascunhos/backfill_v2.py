import openpyxl
from datetime import timedelta
from collections import defaultdict

ARQUIVO_ORIGEM = r"C:\Users\DMF-AUTOMACAO\Downloads\CONTROLE_DE_HORAS_DMF_BL4t520G.xlsm"
ABA_ORIGEM = "02.2026"
ARQUIVO_DESTINO = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx"
ABA_DESTINO = "03.2026"

print("Lendo arquivo de origem com data_only=True...")
wb_origem = openpyxl.load_workbook(ARQUIVO_ORIGEM, data_only=True)
ws_origem = wb_origem[ABA_ORIGEM]

# Construir mapa: cod -> lista de valores (preservando duplicatas em ordem)
mapa_horas_02 = defaultdict(list)
for row in range(10, ws_origem.max_row + 1):
    cod_cell = ws_origem.cell(row=row, column=8).value
    val_cell = ws_origem.cell(row=row, column=15).value
    
    if cod_cell is not None:
        try:
            cod = int(float(str(cod_cell).strip()))
        except (ValueError, TypeError):
            continue
        
        valor = 0.0
        if isinstance(val_cell, timedelta):
            valor = val_cell.total_seconds() / 86400.0
        elif isinstance(val_cell, (int, float)):
            if abs(val_cell) < 1000:  # Protetor contra formulas explosivas
                valor = float(val_cell)
        
        if valor >= 0:
            mapa_horas_02[cod].append(valor)

wb_origem.close()
total_registros = sum(len(v) for v in mapa_horas_02.values())
print(f"Lidos {total_registros} registros de {len(mapa_horas_02)} clientes unicos.")

# Abrir destino
print("Abrindo planilha destino...")
wb_destino = openpyxl.load_workbook(ARQUIVO_DESTINO)
ws_destino = wb_destino[ABA_DESTINO]

# Mapear as linhas do destino por codigo (em ordem)
dest_linhas = defaultdict(list)
for row_idx in range(10, ws_destino.max_row + 1):
    cod_cell = ws_destino.cell(row=row_idx, column=8).value
    if cod_cell is not None:
        try:
            cod = int(float(str(cod_cell).strip()))
            dest_linhas[cod].append(row_idx)
        except (ValueError, TypeError):
            continue

# Preencher: casar linha a linha na ORDEM em que aparecem
alterados = 0
problemas = []
last_data_row = 10

for cod, valores_origem in mapa_horas_02.items():
    linhas_destino = dest_linhas.get(cod, [])
    
    if not linhas_destino:
        if any(v > 0 for v in valores_origem):
            problemas.append(f"Cod {cod}: {len(valores_origem)} registros na origem, NAO EXISTE no destino")
        continue
    
    # Casar na ordem: primeiro valor -> primeira linha, segundo -> segunda, etc.
    for i, valor in enumerate(valores_origem):
        if i < len(linhas_destino):
            row = linhas_destino[i]
            c = ws_destino.cell(row=row, column=14)
            c.value = valor
            c.number_format = '[h]:mm:ss'
            alterados += 1
            if row > last_data_row:
                last_data_row = row
        else:
            # Mais valores na origem que linhas no destino
            problemas.append(f"Cod {cod}: valor extra na origem (indice {i}), sem linha correspondente no destino")

# Corrigir SUBTOTAL
ws_destino.cell(7, 14).value = f'=SUBTOTAL(9,N10:N{last_data_row})'
ws_destino.cell(7, 14).number_format = '[h]:mm:ss'

# Verificar soma
soma = 0
for r in range(10, last_data_row + 1):
    v = ws_destino.cell(r, 14).value
    if isinstance(v, (int, float)) and v > 0:
        soma += v

total_h = int(soma * 24)
total_m = int((soma * 24 - total_h) * 60)
total_s = int(((soma * 24 - total_h) * 60 - total_m) * 60)

print(f"Preenchidos {alterados} registros na Coluna N.")
print(f"SUBTOTAL: =SUBTOTAL(9,N10:N{last_data_row})")
print(f"Soma verificada: {total_h}:{total_m:02d}:{total_s:02d}")

if problemas:
    print(f"\nProblemas ({len(problemas)}):")
    for p in problemas:
        print(f"  {p}")

wb_destino.save(ARQUIVO_DESTINO)
wb_destino.close()
print("\nSalvo com sucesso!")
