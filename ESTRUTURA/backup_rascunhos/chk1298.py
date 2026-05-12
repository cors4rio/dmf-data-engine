import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE.xlsm')
ws = wb['03.2026']
print("Format:", ws.cell(94, 17).number_format)
print("Value:", ws.cell(94, 17).value)
