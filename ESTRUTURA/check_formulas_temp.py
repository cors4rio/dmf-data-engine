import openpyxl

def check_formulas():
    xls_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls'
    
    print(f"Lendo workbook...")
    wb = openpyxl.load_workbook(xls_path, data_only=False)
    sheet_name = '12.2025'
    
    if sheet_name not in wb.sheetnames:
        print(f"Erro: Aba {sheet_name} não encontrada.")
        return
        
    ws = wb[sheet_name]
    max_row = ws.max_row
    
    cells_to_check = ['O7', 'P7', 'Q7', 'R7', 'O9', 'P9', 'Q9', 'R9']
    
    print(f"Aba: {sheet_name} | Total de linhas: {max_row}")
    for cell_ref in cells_to_check:
        cell = ws[cell_ref]
        val = cell.value
        print(f"[{cell_ref}] -> Fórmula/Valor: {val}")

if __name__ == "__main__":
    check_formulas()
