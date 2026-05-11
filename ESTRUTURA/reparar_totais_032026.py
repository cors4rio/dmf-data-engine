import openpyxl
from datetime import timedelta

PLANILHA_MASTER = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF.xlsx"
ABA_ALVO = "03.2026"

def total_to_timedelta(val):
    if val is None:
        return timedelta(0)
    if isinstance(val, timedelta):
        return val
    if isinstance(val, (int, float)):
        # No Excel, o valor numerico de data/hora é a fração do dia
        return timedelta(days=val)
    if isinstance(val, str):
        # Provavelmente "NAO FAZ DP" ou similar, retorna 0
        return timedelta(0)
    return timedelta(0)

def main():
    print(f"Abrindo Master para reparação de totais: {PLANILHA_MASTER}")
    wb = openpyxl.load_workbook(PLANILHA_MASTER)
    ws = wb[ABA_ALVO]

    # Identificar última linha
    max_row = ws.max_row
    print(f"Processando de linha 10 até {max_row}...")

    editados = 0
    for i in range(10, max_row + 1):
        # O=15, P=16, Q=17
        val_o = ws.cell(row=i, column=15).value
        val_p = ws.cell(row=i, column=16).value
        val_q = ws.cell(row=i, column=17).value

        # Converter para timedelta para somar
        td_o = total_to_timedelta(val_o)
        td_p = total_to_timedelta(val_p)
        td_q = total_to_timedelta(val_q)

        total_td = td_o + td_p + td_q
        
        # Converter timedelta de volta para valor float do Excel
        # 1 dia = 86400 segundos
        excel_val = total_td.total_seconds() / 86400.0 if total_td.total_seconds() > 0 else 0
        
        cell_r = ws.cell(row=i, column=18)
        cell_r.value = excel_val
        cell_r.number_format = '[h]:mm:ss'
        editados += 1

    # Atualizar R7 com SUBTOTAL
    ws.cell(row=7, column=18, value=f"=SUBTOTAL(9,R10:R{max_row})")
    
    # Conferir se O7, P7, Q7 estão consistentes também
    ws.cell(row=7, column=15, value=f"=SUBTOTAL(9,O10:O{max_row})")
    ws.cell(row=7, column=16, value=f"=SUBTOTAL(9,P10:P{max_row})")
    ws.cell(row=7, column=17, value=f"=SUBTOTAL(9,Q10:Q{max_row})")

    print(f"Salvando alterações... {editados} linhas atualizadas.")
    wb.save(PLANILHA_MASTER)
    wb.close()
    print("Sucesso!")

if __name__ == "__main__":
    main()
