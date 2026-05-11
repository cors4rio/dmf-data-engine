import openpyxl

BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = BASE_DIR + r"\CONTROLE DE HORAS DMF - CLIENTE.xlsm"

wb = openpyxl.load_workbook(PLANILHA_MASTER, data_only=True)
ws = wb['03.2026']

cod = ws.cell(row=94, column=8).value
valor_q = ws.cell(row=94, column=17).value

print(f"Row 94 - Codigo: {cod}")
print(f"Row 94 - Valor Coluna Q: {valor_q} (Type: {type(valor_q)})")
