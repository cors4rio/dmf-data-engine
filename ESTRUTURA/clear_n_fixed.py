import openpyxl

try:
    file_path = 'CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx'
    print("Iniciando limpeza profunda de macros iterativas e lixo na Coluna N...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb['03.2026']

    for row in ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=14, max_col=14):
        for cell in row:
            cell.value = None

    wb.save(file_path)
    print("Coluna N (Retroativo) zerada! Arquivo limpo de formulas mutantes do Excel.")
except Exception as e:
    print(f"Erro: {e}")
