import openpyxl
import processar_horas

def amostra_contabil():
    FILE_CONTABIL = r"C:\Users\DMF-AUTOMACAO\Downloads\HORAS CONTABEIS_V3_PGPeeWVc.xlsx"
    FILE_MASTER = r"C:\Users\DMF-AUTOMACAO\Downloads\CONTROLE DE HORAS DMF - CLIENTE.xlsm"

    wb_c = openpyxl.load_workbook(FILE_CONTABIL, data_only=True)
    sh_c = wb_c['12.2025']
    
    dict_contabil = {}
    print("--- DADOS LIDOS DA PLANILHA CONTABIL (Amostra) ---")
    for row in range(2, 20):
        cod = sh_c.cell(row=row, column=1).value
        nome = sh_c.cell(row=row, column=2).value
        horas_str = sh_c.cell(row=row, column=17).value
        
        # Converte "HH:MM:SS" (ex: "4:30:00") para decimal
        horas_dec = processar_horas.time_to_hours(horas_str)
        if cod:
            try: dict_contabil[int(cod)] = horas_dec
            except: pass
        
        if row < 10:
            print(f"Cód: {cod} | Nome: {nome} | Total Horas Lida: {horas_str} -> Decimal: {horas_dec:.2f}")

    print("\n--- CRUZAMENTO NA PLANILHA MASTER (Amostra) ---")
    wb_m = openpyxl.load_workbook(FILE_MASTER, data_only=True)
    sh_m = wb_m['12.2025']
    
    encontrados = 0
    for row in range(10, sh_m.max_row + 1):
        cod_master = sh_m.cell(row=row, column=8).value
        nome_razao = sh_m.cell(row=row, column=11).value
        
        c_int = -1
        try:
            if cod_master: c_int = int(str(cod_master).strip().split('.')[0])
        except: pass

        if c_int in dict_contabil:
            horas = dict_contabil[c_int]
            if encontrados < 5:
                print(f"Match Master Linha {row}: Cód {c_int} ({nome_razao}) -> Vai receber: {horas:.2f}h na coluna Contábil")
            encontrados += 1
            
    print(f"\nTotal de matches contábeis previstos: {encontrados}")

if __name__ == "__main__":
    amostra_contabil()
