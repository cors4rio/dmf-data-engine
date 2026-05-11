import openpyxl
import os
from config import *
from utils import logger, detect_duplicates, clean_cnpj

def audit_master_integrity(file_path=TEMP_OUTPUT, sheet_name=CURRENT_SHEET):
    """
    Realiza uma auditoria completa de integridade na planilha Master (Estilo Production Audit).
    Verifica: Duplicidade, Fórmulas, Tipos de Dados e Exceções.
    """
    logger.info(f"--- INICIANDO AUDITORIA: {os.path.basename(file_path)} [{sheet_name}] ---")
    
    if not os.path.exists(file_path):
        logger.error(f"Arquivo não encontrado para auditoria: {file_path}")
        return

    # IMPORTANTE: data_only=False para validar as FÓRMULAS na Coluna R
    wb = openpyxl.load_workbook(file_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        logger.error(f"Aba {sheet_name} não encontrada!")
        return
        
    ws = wb[sheet_name]
    
    # 1. VERIFICAÇÃO DE DUPLICIDADE (Métrica de Produção)
    dupes = detect_duplicates(ws)
    if dupes:
        logger.error(f"❌ FALHA GRAVE: {len(dupes)} CNPJs duplicados detectados!")
        for cnpj, lines in dupes.items():
            logger.error(f"   CNPJ {cnpj} nas linhas: {lines}")
    else:
        logger.info("✅ INTEGRIDADE: Nenhum CNPJ duplicado encontrado.")

    # 2. VALIDAÇÃO DE FÓRMULAS (Regra de Ouro #2)
    wrong_formulas = []
    for row in range(START_ROW, ws.max_row + 1):
        cell_r = ws.cell(row=row, column=COL_TOTAL)
        if cell_r.value and isinstance(cell_r.value, str) and cell_r.value.startswith('='):
            expected = f"=O{row}+P{row}+Q{row}"
            if cell_r.value.replace(' ', '') != expected:
                wrong_formulas.append(row)
        elif cell_r.value:
            # Se tiver valor mas não for fórmula
            wrong_formulas.append(row)
            
    if wrong_formulas:
        logger.warning(f"⚠️ AVISO: {len(wrong_formulas)} linhas com fórmulas de total incorretas.")
        if len(wrong_formulas) < 10: logger.warning(f"    Linhas: {wrong_formulas}")
    else:
        logger.info("✅ FÓRMULAS: Coluna R (Total) 100% consistente com as regras.")

    # 3. CONSISTÊNCIA DE EXCEÇÕES
    contabil_nf = 0
    for row in range(START_ROW, ws.max_row + 1):
        p_val = ws.cell(row=row, column=COL_CONTABIL).value
        if p_val == "NAO FAZ CONTABIL":
            contabil_nf += 1
            
    logger.info(f"📊 EXCEÇÕES: {contabil_nf} empresas marcadas como 'NAO FAZ CONTABIL'.")

    # 4. VOLUMETRIA FINAL
    total_linhas = 0
    for row in range(START_ROW, ws.max_row + 1):
        if ws.cell(row=row, column=COL_CODI_EMP).value:
            total_linhas += 1
            
    logger.info(f"📈 TOTAL: {total_linhas} clientes ativos processados na aba {sheet_name}.")
    logger.info("--- FIM DA AUDITORIA ---")

if __name__ == "__main__":
    # Tenta auditar o arquivo temporário gerado pelo run_integration.py
    if os.path.exists(TEMP_OUTPUT):
        audit_master_integrity(TEMP_OUTPUT)
    else:
        logger.info("Arquivo temporário não encontrado. Audite o arquivo Master original ou rode a integração primeiro.")
