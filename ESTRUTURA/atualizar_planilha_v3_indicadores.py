
import pyodbc
import openpyxl
from collections import defaultdict
import datetime
import calendar
import os
import re
import unicodedata

def normalize(v):
    if v is None: return ""
    # Remove accents and normalize
    v = str(v).replace('\n', ' ').strip().upper()
    v = "".join(c for c in unicodedata.normalize('NFD', v) if unicodedata.category(c) != 'Mn')
    return v

def get_month_dates(mes_alvo):
    # mes_alvo format: MM.YYYY (ex: 01.2025)
    try:
        m, y = mes_alvo.split('.')
        data_inicio = f"{y}-{m}-01"
        ultimo_dia = calendar.monthrange(int(y), int(m))[1]
        data_fim = f"{y}-{m}-{ultimo_dia}"
        return data_inicio, data_fim
    except:
        return None, None

def get_db_connection():
    conn_str = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
    try:
        return pyodbc.connect(conn_str)
    except Exception as e:
        print(f"Erro ao conectar ao banco de dados: {e}")
        return None

def atualizar_indicadores_v3():
    file_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_V3.xlsx'
    
    if not os.path.exists(file_path):
        print(f"Erro: Arquivo não encontrado em {file_path}")
        return

    print(f"Iniciando atualização de {os.path.basename(file_path)}...")

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Erro ao abrir Excel: {e}")
        return

    # Regex para identificar abas de meses (ex: 01.2025)
    regex_mes = re.compile(r'^\d{2}\.\d{4}$')
    
    sheets_to_process = [s for s in wb.sheetnames if regex_mes.match(s)]
    print(f"Abas encontradas para processamento: {sheets_to_process}")

    conn = get_db_connection()
    if not conn:
        return
        
    cursor = conn.cursor()

    for sheet_name in sheets_to_process:
        print(f"\n--- Processando Competência: {sheet_name} ---")
        inicio, fim = get_month_dates(sheet_name)
        if not inicio:
            continue
        
        # 1. Buscar Lançamentos (Normal + Extrato)
        print("Extraindo Lançamentos Contábeis...")
        lancamentos_dict = defaultdict(int)
        q_lan = f"""
            SELECT codi_emp, SUM(CASE WHEN orig_lan IN (1, 5) THEN 1 ELSE 0 END) as qtd
            FROM bethadba.ctlancto
            WHERE data_lan >= '{inicio}' AND data_lan <= '{fim}'
            GROUP BY codi_emp
        """
        try:
            cursor.execute(q_lan)
            for row in cursor.fetchall():
                lancamentos_dict[row[0]] = int(row[1]) if row[1] else 0
        except Exception as e:
            print(f"Erro em Lançamentos ({sheet_name}): {e}")

        # 2. Buscar Faturamento Mensal (Saídas + Serviços)
        print("Extraindo Faturamento Mensal...")
        faturamento_dict = defaultdict(float)
        q_fat = f"""
            SELECT codi_emp, SUM(total_contabil) as fat
            FROM (
                SELECT codi_emp, SUM(vcon_sai) as total_contabil 
                FROM bethadba.efsaidas 
                WHERE dsai_sai >= '{inicio}' AND dsai_sai <= '{fim}' 
                GROUP BY codi_emp
                UNION ALL
                SELECT codi_emp, SUM(vcon_ser) as total_contabil 
                FROM bethadba.efservicos 
                WHERE dser_ser >= '{inicio}' AND dser_ser <= '{fim}' 
                GROUP BY codi_emp
            ) base
            GROUP BY codi_emp
        """
        try:
            cursor.execute(q_fat)
            for row in cursor.fetchall():
                faturamento_dict[row[0]] = float(row[1]) if row[1] else 0.0
        except Exception as e:
            print(f"Erro em Faturamento ({sheet_name}): {e}")

        # 3. Atualizar Planilha
        ws = wb[sheet_name]
        raw_headers = [cell.value for cell in ws[1]]
        
        normalized_headers = [normalize(h) for h in raw_headers]
        
        try:
            # Buscar índices de forma mais flexível
            def find_idx(targets):
                for target in targets:
                    norm_target = normalize(target)
                    if norm_target in normalized_headers:
                        return normalized_headers.index(norm_target) + 1
                return None

            idx_cod = find_idx(['COD DOMINIO', 'COD EMPRESA', 'CODIGO', 'COD'])
            idx_lan = find_idx(['QTD LANCAMENTOS CONTABEIS', 'LANCAMENTOS', 'LANCAMENTOS CONTABEIS'])
            idx_fat = find_idx(['TOTAL FATURAMENTO MES', 'FATURAMENTO', 'TOTAL FATURAMENTO'])

            if not idx_cod:
                print(f"Aviso: Coluna de Código não encontrada na aba {sheet_name}. Ignorando.")
                continue
            if not idx_lan or not idx_fat:
                print(f"Aviso: Colunas de indicadores não encontradas na aba {sheet_name}. Ignorando.")
                continue

        except Exception as e:
            print(f"Aviso: Aba {sheet_name} ignorada devido a erro: {e}")
            continue

        rows_updated = 0
        for row_idx in range(2, ws.max_row + 1):
            cod_val = ws.cell(row=row_idx, column=idx_cod).value
            if cod_val and isinstance(cod_val, (int, float)):
                cod = int(cod_val)
                
                # Atualizar Lançamentos
                ws.cell(row=row_idx, column=idx_lan).value = lancamentos_dict.get(cod, 0)
                
                # Atualizar Faturamento
                ws.cell(row=row_idx, column=idx_fat).value = faturamento_dict.get(cod, 0.0)
                
                rows_updated += 1
        
        print(f"Sucesso: {rows_updated} empresas atualizadas na aba {sheet_name}.")

    # Salvar alterações
    try:
        wb.save(file_path)
        print(f"\nPlanilha '{os.path.basename(file_path)}' atualizada com sucesso!")
    except Exception as e:
        print(f"Erro ao salvar arquivo: {e}")
    finally:
        conn.close()

if __name__ == "__main__":
    atualizar_indicadores_v3()
