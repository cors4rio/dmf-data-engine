import openpyxl
from collections import defaultdict
import datetime

source_file = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_V3_AeqwQXgR.xlsx"
target_file = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm"
filter_file = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\NAO FAZ CONTABIL.txt"

def time_to_hours(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val) * 24.0 # Excel stores time as fraction of day
    if isinstance(val, datetime.time):
        return val.hour + val.minute/60.0 + val.second/3600.0
    if isinstance(val, datetime.timedelta):
        return val.total_seconds() / 3600.0
    if isinstance(val, str):
        try:
            parts = val.split(':')
            if len(parts) == 3:
                return int(parts[0]) + int(parts[1])/60.0 + int(parts[2])/3600.0
        except: pass
    return 0.0

def load_filter():
    non_contabil_codes = set()
    try:
        with open(filter_file, 'r', encoding='utf-8') as f:
            for line in f:
                parts = line.split('\t')
                if parts[0].isdigit():
                    non_contabil_codes.add(str(int(parts[0])))
    except:
        pass
    return non_contabil_codes

def run_dry():
    non_contabil = load_filter()
    print(f"Clientes 'Não Faz Contábil': {len(non_contabil)}")
    
    # 1. Somar abas do arquivo fonte
    wb_s = openpyxl.load_workbook(source_file, data_only=True)
    monthly_data = defaultdict(float) # cod -> hours (current month 02.2026)
    accumulated_data = defaultdict(float) # cod -> sum across all 2026? or all tabs?
    
    # Vamos considerar todas as abas que parecem meses do ano atual ou 2025
    # O usuário disse "somar as abas"
    for sn in wb_s.sheetnames:
        if 'MÉDIA' in sn or 'EXEMPLO' in sn: continue
        ws = wb_s[sn]
        print(f"Lendo aba fonte: {sn}")
        for row in range(2, ws.max_row + 1):
            cod = ws.cell(row=row, column=1).value
            hours = time_to_hours(ws.cell(row=row, column=17).value) # Q = 17
            if cod:
                try:
                    cod_str = str(int(float(str(cod).strip())))
                    accumulated_data[cod_str] += hours
                    if sn == "02.2026":
                        monthly_data[cod_str] = hours
                except: pass
                
    # 2. Simular atualização no target
    wb_t = openpyxl.load_workbook(target_file, data_only=True)
    ws_t = wb_t["02.2026"]
    
    print(f"\nProposta de Atualização (Aba 02.2026):")
    count = 0
    for row in range(10, ws_t.max_row + 1):
        cod = ws_t.cell(row=row, column=8).value # B=2... H=8
        if not cod: continue
        try:
            cod_str = str(int(float(str(cod).strip())))
        except: continue
        
        target_p = monthly_data.get(cod_str, 0.0)
        target_r = accumulated_data.get(cod_str, 0.0)
        
        # Filtro
        if cod_str in non_contabil:
            target_p = "NAO FAZ CONTABIL"
            
        if target_p != 0 or target_r != 0:
            count += 1
            if count <= 15:
                print(f"Row {row} (Cod {cod_str}): P -> {target_p}, R -> {target_r}")
                
    print(f"Total de linhas que seriam atualizadas: {count}")

if __name__ == "__main__":
    run_dry()
