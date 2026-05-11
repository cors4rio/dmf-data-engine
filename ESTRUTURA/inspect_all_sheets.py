import openpyxl
import os

file_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_V3_AeqwQXgR.xlsx'

try:
    wb_meta = openpyxl.load_workbook(file_path, data_only=False)
    
    for sheet_name in wb_meta.sheetnames:
        print(f"\n--- Aba: {sheet_name} ---")
        ws = wb_meta[sheet_name]
        
        # Encontrar a coluna Q
        q_col_idx = None
        for cell in ws[1]:
            coord = cell.coordinate
            if coord.startswith('Q'):
                q_col_idx = 'Q'
                break
        
        if not q_col_idx:
            print("Coluna Q não encontrada nesta aba.")
            continue
            
        rows_with_formula = []
        rows_without_formula = []
        rows_with_formula_returning_zero = []
        
        # Carregar a mesma aba com data_only=True para ver os valores
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        ws_data = wb_data[sheet_name]
        
        for row in range(2, ws.max_row + 1):
            formula = ws[f'Q{row}'].value
            val = ws_data[f'Q{row}'].value
            
            if isinstance(formula, str) and formula.startswith('='):
                rows_with_formula.append(row)
                if val == 0:
                    rows_with_formula_returning_zero.append(row)
            else:
                if val is not None:
                    rows_without_formula.append(row)
                    
        print(f"Total de linhas na aba: {ws.max_row}")
        print(f"Linhas com fórmula em Q: {len(rows_with_formula)}")
        print(f"Linhas SEM fórmula em Q (mas com valor): {len(rows_without_formula)}")
        print(f"Linhas com fórmula que resultam em ZERO: {len(rows_with_formula_returning_zero)}")
        
        if rows_with_formula:
            print(f"Exemplo de fórmula na linha {rows_with_formula[0]}: {ws[f'Q{rows_with_formula[0]}'].value}")
            if rows_with_formula_returning_zero:
                row = rows_with_formula_returning_zero[0]
                print(f"Exemplo de anomalia na linha {row}:")
                for col in ['H', 'J', 'L', 'N', 'P', 'Q']:
                    print(f"  {col}: Formula={ws[f'{col}{row}'].value}, Valor={ws_data[f'{col}{row}'].value}, Tipo={type(ws_data[f'{col}{row}'].value).__name__}")

except Exception as e:
    print(f"Erro: {e}")
