import openpyxl
import os
import re
import difflib

BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = os.path.join(BASE_DIR, "CONTROLE DE HORAS DMF.xlsx")
PLANILHA_SIMPLES = os.path.join(BASE_DIR, "Simples Nacional_1778096117893.xlsx")
MES_FONTE = '03.2026'

def clean_cnpj(val):
    """Remove não-dígitos do CNPJ/CPF."""
    if not val:
        return ""
    return re.sub(r'\D', '', str(val).strip())

def normalizar_doc(cnpj_limpo):
    """Normaliza CNPJ para 14 dígitos ou CPF para 11 dígitos com zero-padding."""
    if not cnpj_limpo:
        return ""
    n = len(cnpj_limpo)
    if n <= 11:
        return cnpj_limpo.zfill(11)  # CPF
    else:
        return cnpj_limpo.zfill(14)  # CNPJ

def similar(a, b):
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, str(a).upper().strip(), str(b).upper().strip()).ratio()

def main():
    # 1. Carregar dados da planilha Master (03.2026)
    print("Abrindo planilha Master (CONTROLE DE HORAS DMF)...")
    wb_master = openpyxl.load_workbook(PLANILHA_MASTER, read_only=True, data_only=True)
    
    if MES_FONTE not in wb_master.sheetnames:
        print(f"Aba {MES_FONTE} não encontrada na Master!")
        return
    
    ws_master = wb_master[MES_FONTE]
    
    # Indexar Master por CNPJ normalizado -> {cnpj_norm: {nome, fiscal_valor, row}}
    master_por_cnpj = {}
    for row in range(10, ws_master.max_row + 1):
        cnpj_raw = ws_master.cell(row=row, column=10).value  # Col J
        nome_raw = ws_master.cell(row=row, column=11).value  # Col K
        fiscal_val = ws_master.cell(row=row, column=15).value  # Col O (Horário Fiscal)
        
        cnpj = clean_cnpj(cnpj_raw)
        cnpj_norm = normalizar_doc(cnpj)
        nome = str(nome_raw).strip() if nome_raw else ""
        
        if not cnpj_norm:
            continue
            
        master_por_cnpj[cnpj_norm] = {
            'nome': nome,
            'fiscal': fiscal_val,
            'row': row
        }
    
    wb_master.close()
    print(f"  -> {len(master_por_cnpj)} empresas indexadas da Master (por CNPJ normalizado).")
    
    # 2. Abrir Simples Nacional para escrita
    print("Abrindo planilha Simples Nacional...")
    wb_simples = openpyxl.load_workbook(PLANILHA_SIMPLES)
    ws_simples = wb_simples.active
    
    encontrados = []
    nao_encontrados = []
    rejeitados_nome = []
    alterados = 0
    total_empresas = 0
    
    print("Processando empresas do Simples Nacional...")
    for row in range(2, ws_simples.max_row + 1):
        nome_simples = ws_simples.cell(row=row, column=2).value   # Col B
        cnpj_simples_raw = ws_simples.cell(row=row, column=3).value  # Col C
        
        # Pular linhas completamente vazias
        if not nome_simples and not cnpj_simples_raw:
            continue
        
        total_empresas += 1
        cnpj_simples = clean_cnpj(cnpj_simples_raw)
        cnpj_simples_norm = normalizar_doc(cnpj_simples)
        nome_simples_str = str(nome_simples).strip() if nome_simples else ""
        
        if not cnpj_simples_norm:
            nao_encontrados.append((row, nome_simples_str, "", "CNPJ vazio no Simples"))
            continue
        
        # Buscar pelo CNPJ normalizado na Master
        if cnpj_simples_norm not in master_por_cnpj:
            nao_encontrados.append((row, nome_simples_str, cnpj_simples, "CNPJ não encontrado na Master"))
            continue
        
        dados_master = master_por_cnpj[cnpj_simples_norm]
        
        # Validação dupla: conferir nome (limiar 0.40 para aceitar abreviações)
        sim_ratio = similar(nome_simples_str, dados_master['nome'])
        if sim_ratio < 0.40:
            rejeitados_nome.append((row, nome_simples_str, cnpj_simples, 
                                    dados_master['nome'], f"sim={sim_ratio:.2f}"))
            continue
        
        # Passou na validação dupla - lançar o valor fiscal
        fiscal_val = dados_master['fiscal']
        
        if fiscal_val is not None:
            c = ws_simples.cell(row=row, column=15, value=fiscal_val)  # Col O
            # Preservar formato de horas se for numérico
            if isinstance(fiscal_val, (int, float)):
                c.number_format = '[h]:mm:ss'
            alterados += 1
            encontrados.append((row, nome_simples_str, cnpj_simples, fiscal_val, dados_master['nome']))
        else:
            nao_encontrados.append((row, nome_simples_str, cnpj_simples, "Valor Fiscal vazio na Master (col O)"))
    
    print(f"Salvando planilha Simples Nacional... ({alterados} linhas alteradas)")
    wb_simples.save(PLANILHA_SIMPLES)
    wb_simples.close()
    
    # Relatório
    RELATORIO = os.path.join(BASE_DIR, "relatorio_fiscal_simples_nacional.md")
    with open(RELATORIO, 'w', encoding='utf-8') as f:
        f.write("# Relatório - Lançamento Fiscal no Simples Nacional\n\n")
        f.write(f"- **Fonte:** CONTROLE DE HORAS DMF.xlsx → aba {MES_FONTE}, coluna O (Horário Fiscal)\n")
        f.write(f"- **Destino:** Simples Nacional_1778096117893.xlsx → coluna O (Horário Fiscal)\n\n")
        f.write(f"- Total de empresas no Simples Nacional: **{total_empresas}**\n")
        f.write(f"- Empresas lançadas com sucesso: **{alterados}**\n")
        f.write(f"- Empresas sem lançamento: **{len(nao_encontrados) + len(rejeitados_nome)}**\n")
        f.write(f"  - CNPJ não encontrado na Master: {sum(1 for x in nao_encontrados if 'não encontrado' in x[3])}\n")
        f.write(f"  - CNPJ vazio no Simples: {sum(1 for x in nao_encontrados if 'vazio' in x[3])}\n")
        f.write(f"  - Valor Fiscal vazio na Master: {sum(1 for x in nao_encontrados if 'Fiscal vazio' in x[3])}\n")
        f.write(f"  - Nome divergente: {len(rejeitados_nome)}\n\n")
        
        if encontrados:
            f.write("## Empresas Lançadas\n\n")
            f.write("| Linha | Nome (Simples) | CNPJ | Valor Fiscal |\n")
            f.write("|---|---|---|---|\n")
            for row, nome, cnpj, fiscal, nome_master in encontrados:
                if isinstance(fiscal, (int, float)):
                    h_total = fiscal * 24
                    h = int(h_total)
                    m = int(round((h_total - h) * 60))
                    fiscal_str = f"{h}:{m:02d}:00"
                else:
                    fiscal_str = str(fiscal)
                f.write(f"| {row} | {nome[:40]} | {cnpj} | {fiscal_str} |\n")
        
        if rejeitados_nome:
            f.write("\n## Rejeitados por Nome Divergente\n\n")
            f.write("| Linha | Nome (Simples) | CNPJ | Nome (Master) | Similaridade |\n")
            f.write("|---|---|---|---|---|\n")
            for row, nome, cnpj, nome_m, sim in rejeitados_nome:
                f.write(f"| {row} | {nome[:35]} | {cnpj} | {nome_m[:35]} | {sim} |\n")
        
        if nao_encontrados:
            f.write("\n## Não Encontrados\n\n")
            f.write("| Linha | Nome | CNPJ | Motivo |\n")
            f.write("|---|---|---|---|\n")
            for row, nome, cnpj, motivo in nao_encontrados:
                f.write(f"| {row} | {nome[:40]} | {cnpj} | {motivo} |\n")
    
    print(f"Processo finalizado! Relatório salvo: {RELATORIO}")
    
    # Resumo no console
    print(f"\n=== RESUMO ===")
    print(f"  Total empresas no Simples: {total_empresas}")
    print(f"  Lançados: {alterados}")
    print(f"  Não encontrados: {len(nao_encontrados)}")
    print(f"  Rejeitados (nome): {len(rejeitados_nome)}")

if __name__ == '__main__':
    main()
