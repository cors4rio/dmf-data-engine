import openpyxl
from datetime import timedelta

# Verificar quantos clientes estao na origem com DUPLICATAS
wb_orig = openpyxl.load_workbook(
    r'C:\Users\DMF-AUTOMACAO\Downloads\CONTROLE_DE_HORAS_DMF_BL4t520G.xlsm', data_only=True
)
ws_orig = wb_orig['02.2026']

# Contar ocorrencias de cada codigo na origem
from collections import defaultdict
orig_entries = defaultdict(list)
for r in range(10, ws_orig.max_row + 1):
    cod = ws_orig.cell(r, 8).value
    val = ws_orig.cell(r, 15).value
    if cod is not None:
        try:
            cod = int(float(str(cod).strip()))
        except:
            continue
        v = 0
        if isinstance(val, timedelta):
            v = val.total_seconds() / 86400.0
        elif isinstance(val, (int, float)) and abs(val) < 1000:
            v = float(val)
        orig_entries[cod].append((r, v))

# Contar ocorrencias na Master
wb_dest = openpyxl.load_workbook(
    r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx',
    data_only=False
)
ws_dest = wb_dest['03.2026']

dest_entries = defaultdict(list)
for r in range(10, ws_dest.max_row + 1):
    cod = ws_dest.cell(r, 8).value
    val = ws_dest.cell(r, 14).value
    if cod is not None:
        try:
            cod = int(float(str(cod).strip()))
        except:
            continue
        v = 0
        if isinstance(val, (int, float)):
            v = float(val)
        elif isinstance(val, timedelta):
            v = val.total_seconds() / 86400.0
        dest_entries[cod].append((r, v))

# Clientes duplicados na origem que so tem 1 entrada no destino
print("=== CLIENTES COM DUPLICATAS NA ORIGEM (mais entradas que no destino) ===")
soma_perdida = 0
for cod in sorted(orig_entries.keys()):
    n_orig = len(orig_entries[cod])
    n_dest = len(dest_entries.get(cod, []))
    soma_orig = sum(v for _, v in orig_entries[cod])
    soma_dest = sum(v for _, v in dest_entries.get(cod, []))
    if n_orig > n_dest and soma_orig > 0:
        diff = (soma_orig - soma_dest) * 24
        soma_perdida += diff
        print(f"  Cod {cod}: {n_orig} entradas na origem ({soma_orig*24:.2f}h), {n_dest} no destino ({soma_dest*24:.2f}h). Diff: {diff:.2f}h")

# Clientes na origem que NAO existem no destino
print("\n=== CLIENTES NA ORIGEM QUE NAO EXISTEM NO DESTINO ===")
for cod in sorted(orig_entries.keys()):
    if cod not in dest_entries:
        soma = sum(v for _, v in orig_entries[cod]) * 24
        if soma > 0:
            print(f"  Cod {cod}: {soma:.2f}h PERDIDAS (nao existe na Master)")

print(f"\nTotal horas perdidas por duplicatas: {soma_perdida:.2f}h")

# Somas finais
soma_total_orig = sum(sum(v for _, v in entries) for entries in orig_entries.values()) * 24
soma_total_dest = sum(sum(v for _, v in entries) for entries in dest_entries.values()) * 24
print(f"\nSoma total origem: {soma_total_orig:.2f}h")
print(f"Soma total destino: {soma_total_dest:.2f}h")
print(f"Diferenca: {soma_total_orig - soma_total_dest:.2f}h")
