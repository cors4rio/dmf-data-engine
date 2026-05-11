import openpyxl

def verify():
    path = 'c:/Users/DMF-AUTOMACAO/Documents/PROJETOS/N8N automacao/ESTRUTURA/CONTROLE_DE_HORAS_DMF.xlsm'
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sh = wb['02.2026']
    
    print("CHECKING FOR CODES 1195 AND 1283...")
    found = []
    # Openpyxl index started from 1
    # Note: after deletion, rows shifted, but I'll scan a large range.
    for r in range(1, 1500):
        try:
            cell_h = sh.cell(row=r, column=8) # Col H
            val = cell_h.value
            if val in [1195, 1283]:
                fiscal = sh.cell(row=r, column=15).value
                contabil = sh.cell(row=r, column=16).value
                pessoal = sh.cell(row=r, column=17).value
                name = sh.cell(row=r, column=9).value
                print(f"Row {r}: Code={val}, Name={name}, Fiscal={fiscal}, Contabil={contabil}, Pessoal={pessoal}")
                found.append(val)
        except: break # End of sheet
        
    print(f"\nSummary: Found codes {found}")
    if len(found) == 2 and set(found) == {1195, 1283}:
        print("SUCCESS: Exactly one row for each code found.")
    else:
        print(f"WARNING: Unexpected frequency of codes. Expected 1 each, found {len(found)} total.")

if __name__ == '__main__':
    verify()
