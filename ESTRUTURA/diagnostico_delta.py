import openpyxl
from datetime import timedelta

wb1 = openpyxl.load_workbook(r'C:\Users\DMF-AUTOMACAO\Downloads\CONTROLE_DE_HORAS_DMF_BL4t520G.xlsm', data_only=True)
ws1 = wb1['02.2026']

wb2 = openpyxl.load_workbook(r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx', data_only=True)
ws2 = wb2['03.2026']

def get_h(ws, col_h, col_cod):
    total = 0
    cods = []
    valores_por_cod = {}
    for r in range(10, ws.max_row+1):
        v = ws.cell(r, col_h).value
        cod = ws.cell(r, col_cod).value
        try:
            cod = int(float(str(cod).strip())) if cod is not None else None
        except:
            cod = None

        val = 0
        if isinstance(v, timedelta):
            val = v.total_seconds() / 86400.0
        elif type(v) in (int, float) and abs(v)<1000:
            val = float(v)

        if val > 0:
            total += val
            cods.append((cod, val))
            if cod not in valores_por_cod:
                valores_por_cod[cod] = []
            valores_por_cod[cod].append(val)
    return total, cods, valores_por_cod

tot1, cods1, map1 = get_h(ws1, 15, 8)
tot2, cods2, map2 = get_h(ws2, 14, 8)

print(f"Soma Origem (02.2026 col O): {tot1*24:.2f}h")
print(f"Soma Destino (Master 03.2026 col N): {tot2*24:.2f}h")

horas_orig_lost = {}
soma_perdida = 0
for cod, vals in map1.items():
    if cod not in map2:
        horas_orig_lost[cod] = sum(vals)
        soma_perdida += sum(vals)

print(f"Horas perdidas pq o cliente não existe na Master 03.2026: {soma_perdida*24:.2f}h (em {len(horas_orig_lost)} codigos ausentes)")

# E verificar códigos que existem em ambas mas com soma diferente
delta_soma = 0
for cod, vals in map1.items():
    if cod in map2:
        s1 = sum(vals)
        s2 = sum(map2[cod])
        if abs(s1 - s2) > 0.0001:
            delta_soma += (s1 - s2)

print(f"Horas perdidas em codigos comuns (ex: duplicados com sobrescrita incorreta): {delta_soma*24:.2f}h")
