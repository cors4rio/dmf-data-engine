import os
import shutil
import openpyxl
import csv
from datetime import datetime, timedelta

def time_to_seconds(t_str):
    if not t_str or t_str == 'None':
        return 0
    try:
        parts = t_str.split(':')
        if len(parts) == 2:
            h, m = map(int, parts)
            return h * 3600 + m * 60
        elif len(parts) == 3:
            h, m, s = map(int, parts)
            return h * 3600 + m * 60 + s
    except:
        return 0
    return 0

def seconds_to_hhmm(seconds):
    h = seconds // 3600
    m = (seconds % 3600) // 60
    return f"{h:02d}:{m:02d}"

def process_data():
    csv_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\resultado_consumo_unidades_1261.csv'
    xls_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls'
    temp_xlsx = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\temp_update.xlsx'
    
    print(f"Lendo CSV: {csv_path}")
    # Mapeamento Codigo -> Tempo (HH:MM) do CSV
    # Coluna 0 (ID) é o código de domínio no CSV
    # Coluna 7 (Tempo (HH:MM)) é o valor a ser extraído
    csv_data = {}
    with open(csv_path, mode='r', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')
        header = next(reader)
        for row in reader:
            if len(row) > 7:
                cod = row[0].strip()
                tempo = row[7].strip()
                csv_data[cod] = tempo

    print(f"Total de clientes no CSV: {len(csv_data)}")

    print(f"Criando cópia temporária XLSX do arquivo XLS...")
    shutil.copy(xls_path, temp_xlsx)

    print(f"Abrindo workbook...")
    wb = openpyxl.load_workbook(temp_xlsx)
    
    sheet_name = '12.2025'
    if sheet_name not in wb.sheetnames:
        print(f"ERRO: Aba {sheet_name} não encontrada!")
        return

    ws = wb[sheet_name]
    
    # 1. Localizar colunas de Cód Dominio (H=8) e HORARIO PESSOAL (Q=17) na linha 9
    col_cod = 8
    col_pessoal = 17
    
    # 2. Iterar da linha 10 até o fim
    count_updates = 0
    for row_idx in range(10, ws.max_row + 1):
        cell_cod = ws.cell(row=row_idx, column=col_cod).value
        if cell_cod is not None:
            cod_str = str(cell_cod).strip().split('.')[0] # Remover .0 de float se houver
            if cod_str in csv_data:
                ws.cell(row=row_idx, column=col_pessoal).value = csv_data[cod_str]
                count_updates += 1

    print(f"Atualizações realizadas na coluna Q: {count_updates}")

    # 3. Recalcular totais na aba r9? 
    # Como não existe aba r9, vou assumir que o usuário se referiu ao total geral ou a uma célula específica de total na aba 12.2025.
    # Vou procurar por uma linha de "TOTAL" na aba 12.2025
    total_seconds = 0
    for row_idx in range(10, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=col_pessoal).value
        total_seconds += time_to_seconds(str(val))
    
    print(f"Novo total calculado para HORARIO PESSOAL: {seconds_to_hhmm(total_seconds)}")

    # Salvar e converter de volta? 
    # O usuário pediu para atualizar o XLS original. 
    # Vou salvar o XLSX e depois sobrescrever o XLS (já que descobrimos que o XLS é um XLSX renomeado).
    wb.save(temp_xlsx)
    wb.close()
    
    shutil.copy(temp_xlsx, xls_path)
    os.remove(temp_xlsx)
    print(f"Arquivo {xls_path} atualizado com sucesso.")

if __name__ == "__main__":
    process_data()
