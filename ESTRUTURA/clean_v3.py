import openpyxl

def clean_and_fix_structural_v3():
    # Caminho do arquivo que corrompeu
    file_in = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx'
    file_out = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO_V3.xlsx'

    print(f"Limpando e recuperando estrutura do arquivo...")
    
    # Carregar SEM keep_vba para eliminar qualquer metadado de macro que esteja causando conflito com a extensão .xlsx
    wb = openpyxl.load_workbook(file_in, keep_vba=False)
    
    # Colunas de Tempo: N (14), O (15), P (16), Q (17), R (18)
    time_cols = [14, 15, 16, 17, 18]
    
    for sheet_name in wb.sheetnames:
        print(f"Verificando aba: {sheet_name}")
        ws = wb[sheet_name]
        
        fixed_count = 0
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in time_cols:
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                
                # Se encontrar valores absurdos (> 1000 dias) nas colunas de tempo, zera.
                if isinstance(val, (int, float)) and abs(val) > 1000:
                    print(f"  [{sheet_name}] Corrigindo Célula R{row_idx}C{col_idx}: {val}")
                    cell.value = 0
                    fixed_count += 1
        
        if fixed_count > 0:
            print(f"  Total corrigido na aba {sheet_name}: {fixed_count}")

    print(f"Salvando versão 100% limpa (sem macro parts) em: {file_out}")
    wb.save(file_out)
    print("Processo concluído. Tente abrir a V3.")

if __name__ == "__main__":
    clean_and_fix_structural_v3()
