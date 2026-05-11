import openpyxl
from collections import defaultdict
import datetime
import os
import logging
from config import *
from utils import (
    logger, 
    clean_cnpj, 
    time_to_float_excel, 
    find_client_row, 
    load_filter_text, 
    detect_duplicates
)

class MasterAutomation:
    """
    Motor Principal de Automação DMF (Refatoração Elite).
    Integra dados Fiscal, Contábil e DP na Planilha Master conforme Spec_Planilha_Master.md.
    """
    
    def __init__(self, mes_alvo=CURRENT_SHEET):
        self.mes_alvo = mes_alvo
        self.wb_target = None
        self.ws_target = None
        
        # Carregar Filtros de Exceção Profissionais
        self.filter_contabil = load_filter_text(FILTER_NAO_CONTABIL)
        self.filter_dp = load_filter_text(FILTER_DP_NAO)
        
        logger.info(f"Iniciando MasterAutomation para competência: {self.mes_alvo}")

    def load_accounting_hours(self):
        """Lê a planilha de horas contábeis consolidando mensal e acumulado."""
        if not os.path.exists(SOURCE_CONTABIL):
            logger.error(f"Fonte contábil não encontrada: {SOURCE_CONTABIL}")
            return {}, {}
            
        logger.info(f"Extraindo dados de: {SOURCE_CONTABIL}")
        wb_s = openpyxl.load_workbook(SOURCE_CONTABIL, data_only=True)
        
        monthly_data = defaultdict(float) 
        accumulated_data = defaultdict(float) 
        
        for sn in wb_s.sheetnames:
            # Ignorar abas de cálculo/exemplo conforme padrão do projeto
            if any(x in sn.upper() for x in ['MÉDIA', 'EXEMPLO', 'TOTAL']): continue
            
            ws = wb_s[sn]
            logger.info(f"  Lendo aba: {sn}")
            
            for row in range(2, ws.max_row + 1):
                cod = ws.cell(row=row, column=1).value
                # Coluna Q (index 17) contém o tempo gasto
                hours = time_to_float_excel(ws.cell(row=row, column=17).value)
                
                if cod:
                    try:
                        cod_str = str(int(float(str(cod).strip().split('-')[0])))
                        accumulated_data[cod_str] += hours
                        if sn == self.mes_alvo:
                            monthly_data[cod_str] = hours
                    except: continue
        
        return monthly_data, accumulated_data

    def run(self):
        """Executa o pipeline completo de integração."""
        logger.info("--- INICIANDO INTEGRAÇÃO MASTER ---")
        
        # 1. Coleta de Dados
        monthly_hours, _ = self.load_accounting_hours()
        
        # 2. Acesso à Planilha Master
        if not os.path.exists(TARGET_MASTER):
            logger.error(f"Arquivo Master não encontrado: {TARGET_MASTER}")
            return

        self.wb_target = openpyxl.load_workbook(TARGET_MASTER, keep_vba=True)
        if self.mes_alvo not in self.wb_target.sheetnames:
            logger.error(f"Aba '{self.mes_alvo}' não encontrada no Master. Verifique o config.py.")
            return

        self.ws_target = self.wb_target[self.mes_alvo]
        
        # 3. Auditoria Preventiva
        dupes = detect_duplicates(self.ws_target)
        if dupes:
            logger.warning(f"⚠️ ATENÇÃO: {len(dupes)} CNPJs duplicados detectados na Master!")
            for cnpj, lines in dupes.items():
                logger.warning(f"   - CNPJ {cnpj} aparece nas linhas: {lines}")

        # 4. Preenchimento Sistêmico
        processados = 0
        falhas_match = []

        for row in range(START_ROW, self.ws_target.max_row + 1):
            cod_val = self.ws_target.cell(row=row, column=COL_CODI_EMP).value
            cnpj_val = self.ws_target.cell(row=row, column=COL_CNPJ).value
            
            if not cod_val and not cnpj_val: continue
            
            # Limpeza do código para busca
            cod_str = None
            if cod_val and str(cod_val).strip().isdigit():
                cod_str = str(int(float(str(cod_val).strip())))
            
            # --- COLUNA P: HORÁRIO CONTÁBIL ---
            if cod_str in self.filter_contabil["codes"]:
                self.ws_target.cell(row=row, column=COL_CONTABIL).value = "NAO FAZ CONTABIL"
            else:
                val_p = monthly_hours.get(cod_str, 0.0) if cod_str else 0.0
                if val_p > 0:
                    self.ws_target.cell(row=row, column=COL_CONTABIL).value = val_p
                    self.ws_target.cell(row=row, column=COL_CONTABIL).number_format = '[h]:mm:ss'
                elif cod_str and cod_str not in monthly_hours:
                    falhas_match.append(f"Linha {row}: Cliente {cod_str} sem horas na fonte.")

            # --- COLUNA R: TOTAL (REGRA DE OURO: FÓRMULA) ---
            # Implementamos a fórmula de soma garantindo a flexibilidade do Excel
            self.ws_target.cell(row=row, column=COL_TOTAL).value = f"=O{row}+P{row}+Q{row}"
            self.ws_target.cell(row=row, column=COL_TOTAL).number_format = '[h]:mm:ss'
            
            processados += 1

        # 5. Finalização
        try:
            self.wb_target.save(TEMP_OUTPUT)
            logger.info(f"--- INTEGRAÇÃO CONCLUÍDA ---")
            logger.info(f"Linhas processadas: {processados}")
            logger.info(f"Arquivo salvo em: {TEMP_OUTPUT}")
            
            if falhas_match:
                logger.info(f"Resumo de avisos: {len(falhas_match)} clientes sem dados de horas.")
                
        except Exception as e:
            logger.critical(f"ERRO FATAL AO SALVAR: {e}")

if __name__ == "__main__":
    app = MasterAutomation()
    app.run()
