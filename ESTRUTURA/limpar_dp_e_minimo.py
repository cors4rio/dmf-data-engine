import os
import shutil
import openpyxl
import re

def process_dp_exceptions():
    txt_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\DP NAO.txt'
    xls_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls'
    temp_xlsx = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\temp_dp_update.xlsx'
    
    print(f"Lendo regras de DP: {txt_path}")
    
    dp_nao_codes = set()
    consultoria_codes = {} # Cod -> Valor (ex: '1:30')
    
    with open(txt_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or "EMPRESAS QUE NÃO FAZ DP" in line:
                continue
            
            # Tentar extrair código da linha
            # Formatos comuns: "1107 GLOBAL...", "988	LE BRUT...", "853;AGROMIX..."
            match_cod = re.match(r'^(\d+)', line)
            if match_cod:
                cod = match_cod.group(1)
                if "1:30" in line or "consultoria" in line.lower():
                    consultoria_codes[cod] = "1:30"
                else:
                    dp_nao_codes.add(cod)
            elif "Não entra - sistema próprio" in line:
                # Aqui pode haver código depois ou apenas o nome. 
                # Se não houver código fácil, vamos ignorar nomes por agora ou tentar o match do nome depois se necessário.
                pass

    print(f"Códigos DP NÃO identificados: {len(dp_nao_codes)}")
    print(f"Códigos Consultoria (1:30) identificados: {len(consultoria_codes)}")

    print(f"Criando cópia temporária XLSX do arquivo XLS...")
    shutil.copy(xls_path, temp_xlsx)

    print(f"Abrindo workbook...")
    wb = openpyxl.load_workbook(temp_xlsx)
    sheet_name = '12.2025'
    ws = wb[sheet_name]
    
    col_cod = 8
    col_pessoal = 17 # Coluna Q
    
    count_dp_nao = 0
    count_consultoria = 0
    
    for row_idx in range(10, ws.max_row + 1):
        cell_cod = ws.cell(row=row_idx, column=col_cod).value
        if cell_cod is not None:
            cod_str = str(cell_cod).strip().split('.')[0]
            
            if cod_str in consultoria_codes:
                ws.cell(row=row_idx, column=col_pessoal).value = consultoria_codes[cod_str]
                count_consultoria += 1
            elif cod_str in dp_nao_codes:
                ws.cell(row=row_idx, column=col_pessoal).value = "DP NÃO"
                count_dp_nao += 1
            
            # Adicional: Se a coluna D (PESSOAL) da aba 12.2025 diz "DP NÃO", garante que a Q também seja "DP NÃO"
            # Isso ajuda nos casos onde o nome da empresa bate mas o código não foi pego no TXT
            col_nome_dp = 4 # Coluna D
            cell_nome_dp = ws.cell(row=row_idx, column=col_nome_dp).value
            if cell_nome_dp == "DP NÃO" and ws.cell(row=row_idx, column=col_pessoal).value not in ["DP NÃO", "1:30"]:
                 ws.cell(row=row_idx, column=col_pessoal).value = "DP NÃO"
                 count_dp_nao += 1

    print(f"Atualizações: {count_dp_nao} marcados como 'DP NÃO', {count_consultoria} marcados como '1:30'.")

    wb.save(temp_xlsx)
    wb.close()
    
    shutil.copy(temp_xlsx, xls_path)
    os.remove(temp_xlsx)
    print(f"Arquivo {xls_path} atualizado com as regras de DP.")

if __name__ == "__main__":
    process_dp_exceptions()
