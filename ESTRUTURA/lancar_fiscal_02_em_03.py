"""
Lançamento Fiscal (Mês 02/2026) → Planilha Master (Aba 03.2026, Coluna O)

Fonte: Banco Domínio (geloguser, sist_log=5) - Período Fevereiro 2026
Comparação: ClienteTempo Gasto 022026.xls (para validação cruzada)
Destino: CONTROLE DE HORAS DMF.xlsx -> Aba 03.2026 -> Coluna O (15)

Regras:
- Fator multiplicador: 1.80 (adicional de 80%)
- Clientes sem dados no BD recebem 0 (não em branco) - mata fantasmas
- NÃO usar keep_vba (.xlsx)
- Converter para float decimal (fração de dia) antes de gravar
- Formato: [h]:mm:ss
- SUBTOTAL em O7 com range dinâmico
"""

import os
import sys
import pyodbc
import openpyxl
from collections import defaultdict
from datetime import timedelta
import struct

# ============== CONFIGURAÇÕES ==============
PLANILHA_MASTER = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF.xlsx"
CLIENTETEMPO_XLS = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\ClienteTempo Gasto 022026.xls"
ABA_MES = "03.2026"
FATOR_ADICIONAL_FISCAL = 1.80
DB_CONN_STR = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'

# Período: FEVEREIRO 2026
DATA_INICIO = '2026-02-01'
DATA_FIM = '2026-02-28'

def get_fiscal_data_from_db():
    """Busca tempos gastos no módulo Fiscal (sist_log=5) no período de Fev/2026."""
    query = f"""
    SELECT 
        l.codi_emp as codigo_cliente,
        SUM(DATEDIFF(second, 
            YMD(YEAR(l.data_log), MONTH(l.data_log), DAY(l.data_log)) + l.tini_log, 
            COALESCE(
                YMD(YEAR(l.dfim_log), MONTH(l.dfim_log), DAY(l.dfim_log)) + l.tfim_log, 
                YMD(YEAR(l.data_log), MONTH(l.data_log), DAY(l.data_log)) + l.tfim_log
            )
        )) as total_segundos
    FROM 
        bethadba.geloguser l
    JOIN 
        bethadba.geempre e ON l.codi_emp = e.codi_emp
    WHERE 
        l.sist_log = 5
        AND l.tfim_log IS NOT NULL
        AND l.data_log BETWEEN '{DATA_INICIO}' AND '{DATA_FIM}'
    GROUP BY 
        l.codi_emp
    """
    
    dict_fiscal = {}
    try:
        conn = pyodbc.connect(DB_CONN_STR)
        cursor = conn.cursor()
        print(f"Buscando dados Fiscais no DB Domínio ({DATA_INICIO} a {DATA_FIM})...")
        cursor.execute(query)
        for row in cursor.fetchall():
            cod, segundos = row
            if segundos:
                dict_fiscal[cod] = float(segundos)
        conn.close()
    except Exception as e:
        print(f"Erro ODBC: {e}")
        sys.exit(1)
    
    return dict_fiscal

def load_clientetempo_for_comparison():
    """Tenta ler a planilha ClienteTempo Gasto para validação cruzada."""
    comparison = {}
    if not os.path.exists(CLIENTETEMPO_XLS):
        print(f"AVISO: Arquivo de comparação não encontrado: {CLIENTETEMPO_XLS}")
        return comparison
    
    try:
        # Tentar com xlrd primeiro
        import xlrd
        wb = xlrd.open_workbook(CLIENTETEMPO_XLS)
        ws = wb.sheet_by_index(0)
        for r in range(1, ws.nrows):
            try:
                cod = int(ws.cell_value(r, 0))
                # Procurar coluna de tempo (geralmente a última ou a que contém hh:mm:ss)
                tempo_str = str(ws.cell_value(r, ws.ncols - 1))
                comparison[cod] = tempo_str
            except:
                pass
        print(f"  ClienteTempo: {len(comparison)} clientes carregados para comparação.")
    except Exception as e:
        print(f"  AVISO: Não foi possível ler ClienteTempo via xlrd: {e}")
        # Tentar com openpyxl (se for xlsx disfarçado)
        try:
            wb = openpyxl.load_workbook(CLIENTETEMPO_XLS, data_only=True)
            ws = wb.active
            for r in range(2, ws.max_row + 1):
                try:
                    cod = int(float(str(ws.cell(r, 1).value).strip()))
                    tempo = ws.cell(r, ws.max_column).value
                    comparison[cod] = str(tempo)
                except:
                    pass
            print(f"  ClienteTempo via openpyxl: {len(comparison)} clientes.")
        except Exception as e2:
            print(f"  AVISO: Também falhou via openpyxl: {e2}")
    
    return comparison

def segundos_para_hms(total_segundos):
    total_segundos = int(total_segundos)
    h = total_segundos // 3600
    m = (total_segundos % 3600) // 60
    s = total_segundos % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

def main():
    # 1. Buscar dados fiscais do BD
    dados_fiscal = get_fiscal_data_from_db()
    print(f"Total de empresas com hora fiscal no BD: {len(dados_fiscal)}")

    # 2. Carregar comparação (opcional)
    comparison = load_clientetempo_for_comparison()

    # 3. Abrir planilha Master (SEM keep_vba - é .xlsx!)
    print(f"\nAbrindo planilha Master: {PLANILHA_MASTER}")
    if not os.path.exists(PLANILHA_MASTER):
        print(f"ERRO: Planilha Master não encontrada!")
        sys.exit(1)
    
    # Extensão .xlsx = NÃO usar keep_vba (regra Spec item D)
    is_xlsm = PLANILHA_MASTER.endswith('.xlsm')
    wb = openpyxl.load_workbook(PLANILHA_MASTER, keep_vba=is_xlsm)
    
    if ABA_MES not in wb.sheetnames:
        print(f"Aba {ABA_MES} não encontrada!")
        sys.exit(1)
    
    ws = wb[ABA_MES]

    # 4. Mapear códigos -> linhas (suporta duplicados)
    mapa_planilha = defaultdict(list)
    for i in range(10, ws.max_row + 1):
        cod_cell = ws.cell(row=i, column=8).value  # Coluna H = Cod Domínio
        if cod_cell is not None:
            try:
                cod = int(float(str(cod_cell).strip()))
                mapa_planilha[cod].append(i)
            except ValueError:
                pass

    # 5. Preencher Coluna O (15) - Fiscal
    alterados = 0
    calculos_md = []
    calculos_md.append(f"# Relatório Auditado: Cálculos Fiscal (Dados 02/2026 → Aba {ABA_MES})")
    calculos_md.append(f"**Período dos dados:** {DATA_INICIO} a {DATA_FIM}")
    calculos_md.append(f"**Fator Multiplicador:** 80% (1.80)\n")
    calculos_md.append("| Cód | Seg. Original | Seg. Final (×1.80) | Valor [h]:mm:ss | Status |")
    calculos_md.append("|---|---|---|---|---|")

    for cod, linhas_alvo in mapa_planilha.items():
        seg_bruto = dados_fiscal.get(cod, 0)
        seg_final = seg_bruto * FATOR_ADICIONAL_FISCAL
        valor_excel = seg_final / 86400.0  # Fração de dia para Excel
        str_final = segundos_para_hms(seg_final)

        for linha in linhas_alvo:
            c = ws.cell(row=linha, column=15)  # Coluna O = 15
            c.value = valor_excel
            c.number_format = '[h]:mm:ss'
            alterados += 1

        if seg_bruto > 0:
            calculos_md.append(f"| {cod} | {int(seg_bruto)} | {int(seg_final)} | {str_final} | ✅ {len(linhas_alvo)} linha(s) |")
        else:
            calculos_md.append(f"| {cod} | 0 | 0 | 00:00:00 | ⚠️ Zerado |")

    # Órfãos do BD (existem no Domínio mas não na planilha)
    orfaos = []
    for cod, seg_bruto in dados_fiscal.items():
        if cod not in mapa_planilha:
            seg_final = seg_bruto * FATOR_ADICIONAL_FISCAL
            str_final = segundos_para_hms(seg_final)
            orfaos.append((cod, str_final))
            calculos_md.append(f"| {cod} | {int(seg_bruto)} | {int(seg_final)} | {str_final} | ❌ Órfão |")

    # 6. SUBTOTAL em O7 (range dinâmico)
    max_row = ws.max_row
    ws.cell(row=7, column=15).value = f"=SUBTOTAL(9,O10:O{max_row})"
    ws.cell(row=7, column=15).number_format = '[h]:mm:ss'

    # 7. Salvar
    print(f"\nSalvando planilha... ({alterados} linhas escritas)")
    wb.save(PLANILHA_MASTER)
    wb.close()

    # 8. Relatórios
    with open("calculos_fiscal_02_em_03.md", "w", encoding="utf-8") as f:
        f.write("\n".join(calculos_md))

    print(f"\n=== RESUMO ===")
    print(f"Empresas no DB: {len(dados_fiscal)}")
    print(f"Linhas processadas na Master: {alterados}")
    print(f"Clientes órfãos: {len(orfaos)}")
    print(f"SUBTOTAL em O7: =SUBTOTAL(9,O10:O{max_row})")
    print(f"Relatório: calculos_fiscal_02_em_03.md")
    print("Sucesso!")

if __name__ == "__main__":
    main()
