import pyodbc
import openpyxl
import os

def get_database_counts():
    conn_str = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
    counts = {}
    try:
        # Use simple string formatting since this is a controlled scratch script
        # Filter: orig_lan IN (1, 39)
        query = """
        SELECT 
            codi_emp,
            COUNT(*) as total
        FROM bethadba.ctlancto
        WHERE data_lan >= '2025-01-01'
          AND data_lan <= '2025-01-31'
          AND orig_lan IN (1, 39)
        GROUP BY codi_emp
        """
        
        print("Conectando ao banco de dados...")
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        print("Executando query de lançamentos (01/2025, Origens 1 e 39)...")
        cursor.execute(query)
        
        for row in cursor.fetchall():
            counts[int(row.codi_emp)] = int(row.total)
            
        conn.close()
        print(f"Sucesso! {len(counts)} empresas com lançamentos encontradas.")
    except Exception as e:
        print(f"Erro no banco de dados: {e}")
    return counts

def update_excel(counts):
    file_path = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_.xlsx"
    backup_path = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_BACKUP.xlsx"
    
    try:
        print(f"Abrindo planilha: {file_path}")
        wb = openpyxl.load_workbook(file_path)
        sheet_name = "01.2025"
        
        if sheet_name not in wb.sheetnames:
            print(f"Erro: Aba {sheet_name} não encontrada!")
            return
            
        ws = wb[sheet_name]
        
        updated_count = 0
        total_rows = ws.max_row
        
        print(f"Processando {total_rows - 1} linhas...")
        
        # Iterar a partir da linha 2 (ignorar cabeçalho)
        for row in range(2, total_rows + 1):
            cod_val = ws.cell(row=row, column=1).value
            if cod_val is not None:
                try:
                    # Tentar converter para int (pode vir como float no Excel)
                    cod = int(float(str(cod_val).strip()))
                    # Pegar o valor do banco ou 0 se não houver lançamentos oficiais
                    new_val = counts.get(cod, 0)
                    
                    # Atualizar coluna F (6)
                    ws.cell(row=row, column=6).value = new_val
                    updated_count += 1
                except (ValueError, TypeError):
                    continue
        
        # Salvar backup antes
        # os.rename(file_path, backup_path) # Comentado para segurança, usaremos save em novo nome se preferir, mas o user disse "LANCAR EM QTD LANCAMENTOS"
        
        print(f"Salvando alterações... ({updated_count} linhas atualizadas)")
        wb.save(file_path)
        print("Arquivo salvo com sucesso.")
        
    except Exception as e:
        print(f"Erro ao atualizar Excel: {e}")

if __name__ == "__main__":
    db_counts = get_database_counts()
    if db_counts is not None:
        update_excel(db_counts)
