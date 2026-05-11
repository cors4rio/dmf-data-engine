import openpyxl

def deep_scrub_and_recover():
    file_in = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx'
    file_out = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO_V2.xlsx'

    print(f"Iniciando limpeza profunda e recuperação do arquivo...")
    
    # Abrir sem keep_vba para garantir que o XLSX seja limpo de partes binárias de macro
    wb = openpyxl.load_workbook(file_in, keep_vba=False)
    
    for sheet_name in wb.sheetnames:
        print(f"Escanendo aba: {sheet_name}")
        ws = wb[sheet_name]
        
        count_fixed = 0
        for row in ws.iter_rows():
            for cell in row:
                # Se o valor for numérico e absurdo para uma data/hora excel (> 1.000.000 dias)
                if isinstance(cell.value, (int, float)):
                    if abs(cell.value) > 1000000:
                        print(f"  Fixing R{cell.row}C{cell.column}: {cell.value}")
                        cell.value = 0
                        count_fixed += 1
        
        if count_fixed > 0:
            print(f"  Aba {sheet_name}: {count_fixed} células limpas.")

    print(f"Salvando versão limpa como {file_out}...")
    wb.save(file_out)
    print("Sucesso! Tente abrir o arquivo _V2.")

if __name__ == "__main__":
    deep_scrub_and_recover()
