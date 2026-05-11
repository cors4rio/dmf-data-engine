import re
import datetime
import logging
import os
import openpyxl
from config import *

# --- CONFIGURAÇÃO DE LOGGING ---
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(BASE_DIR, "automacao_dmf.log"), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("DMF_AUTOMACAO")

# --- UTILITÁRIOS DE DADOS ---

def clean_cnpj(cnpj):
    """Remove pontos, barras e traços do CNPJ."""
    if not cnpj: return ""
    return re.sub(r'\D', '', str(cnpj))

def time_to_float_excel(val):
    """Converte valores de tempo para o formato decimal do Excel (1.0 = 24h)."""
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    if isinstance(val, datetime.time):
        return (val.hour + val.minute/60.0 + val.second/3600.0) / 24.0
    if isinstance(val, datetime.timedelta):
        return val.total_seconds() / (3600.0 * 24.0)
    if isinstance(val, str):
        try:
            parts = val.split(':')
            if len(parts) == 3:
                h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
                return (h + m/60.0 + s/3600.0) / 24.0
        except: pass
    return 0.0

# --- UTILITÁRIOS DE EXCEL (PLANEJAMENTO ELITE) ---

def find_client_row(sheet, codi_emp, cnpj):
    """
    Implementa o ALGORITMO DE LOOKUP DUPLO da Spec_Planilha_Master.
    1. Tenta pelo Código Domínio (Coluna H)
    2. Fallback pelo CNPJ (Coluna J)
    """
    cnpj_limpo = clean_cnpj(cnpj)
    
    # 1ª Tentativa: Código Domínio
    for row_idx in range(START_ROW, sheet.max_row + 1):
        cod_cell = sheet.cell(row=row_idx, column=COL_CODI_EMP).value
        # Ignorar textos especiais no campo H durante lookup por código
        if cod_cell and str(cod_cell).strip().isdigit():
            if str(int(float(str(cod_cell).split('-')[0]))).strip() == str(codi_emp):
                return row_idx
    
    # 2ª Tentativa: CNPJ
    if cnpj_limpo:
        for row_idx in range(START_ROW, sheet.max_row + 1):
            cnpj_cell = sheet.cell(row=row_idx, column=COL_CNPJ).value
            if cnpj_cell and clean_cnpj(cnpj_cell) == cnpj_limpo:
                return row_idx

    return None

def load_filter_text(file_path):
    """Lê arquivos de filtro (DP NAO, NAO FAZ CONTABIL) e retorna set de códigos/nomes."""
    data = {"codes": set(), "names": set()}
    if not os.path.exists(file_path):
        logger.warning(f"Arquivo de filtro não encontrado: {file_path}")
        return data
        
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line: continue
                
                # Tratar formatos: 603\tNOME ou 603;NOME ou Apenas NOME
                parts = re.split(r'[\t;]', line)
                code_part = parts[0].strip()
                
                if code_part.isdigit():
                    data["codes"].add(code_part)
                elif "Não entra" in code_part or "Não esta" in code_part:
                    # Casos especiais: pegar o nome que vem depois do tab/ponto vírgula
                    if len(parts) > 1:
                        data["names"].add(parts[1].strip().upper())
                else:
                    data["names"].add(code_part.upper())
    except Exception as e:
        logger.error(f"Erro ao ler filtro {file_path}: {e}")
        
    return data

def detect_duplicates(sheet):
    """Identifica CNPJs em duplicidade na aba mensal."""
    from collections import defaultdict
    mapa = defaultdict(list)
    for row_idx in range(START_ROW, sheet.max_row + 1):
        cnpj_cell = sheet.cell(row=row_idx, column=COL_CNPJ).value
        if cnpj_cell:
            cnpj_limpo = clean_cnpj(cnpj_cell)
            if cnpj_limpo:
                mapa[cnpj_limpo].append(row_idx)
    return {cnpj: linhas for cnpj, linhas in mapa.items() if len(linhas) > 1}
