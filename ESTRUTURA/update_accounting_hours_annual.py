import pyodbc
import openpyxl
import os
import calendar

def get_database_counts(month, year):
    conn_str = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
    counts = {}
    
    # Datas do mês
    last_day = calendar.monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{last_day}"
    
    try:
        query = f"""
        SELECT 
            codi_emp,
            COUNT(*) as total
        FROM bethadba.ctlancto
        WHERE data_lan >= '{start_date}'
          AND data_lan <= '{end_date}'
          AND orig_lan IN (1, 39)
        GROUP BY codi_emp
        """
        
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        cursor.execute(query)
        
        for row in cursor.fetchall():
            counts[int(row.codi_emp)] = int(row.total)
            
        conn.close()
        return counts
    except Exception as e:
        print(f"Erro no banco de dados para {month}/{year}: {e}")
        return None

def update_all_months():
    file_path = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_.xlsx"
    
    try:
        print(f"Abrindo planilha: {file_path}")
        wb = openpyxl.load_workbook(file_path)
        
        for month in range(2, 13):
            year = 2025
            sheet_name = f"{month:02d}.{year}"
            
            if sheet_name not in wb.sheetnames:
                print(f"Aba {sheet_name} não encontrada, pulando...")
                continue
                
            print(f"\n--- Processando competência: {sheet_name} ---")
            
            counts = get_database_counts(month, year)
            if counts is None:
                continue
                
            print(f"Empresas com lançamentos (1, 39) no banco: {len(counts)}")
            
            ws = wb[sheet_name]
            updated_count = 0
            total_rows = ws.max_row
            
            for row in range(2, total_rows + 1):
                cod_val = ws.cell(row=row, column=1).value
                if cod_val is not None:
                    try:
                        cod = int(float(str(cod_val).strip()))
                        new_val = counts.get(cod, 0)
                        
                        # Atualizar coluna F (6)
                        ws.cell(row=row, column=6).value = new_val
                        updated_count += 1
                    except (ValueError, TypeError):
                        continue
            
            print(f"Linhas atualizadas na aba {sheet_name}: {updated_count}")
        
        print(f"\nSalvando alterações finais no arquivo...")
        wb.save(file_path)
        print("Arquivo salvo com sucesso para todos os meses (02 a 12/2025).")
        
    except Exception as e:
        print(f"Erro ao atualizar Excel: {e}")

if __name__ == "__main__":
    update_all_months()
