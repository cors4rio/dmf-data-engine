import openpyxl
from datetime import timedelta
import os

PLANILHA_MASTER = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx"
ABA_MES = "03.2026"

def val_to_seconds(val):
    if val is None:
        return 0
    if isinstance(val, timedelta):
        return val.total_seconds()
    if isinstance(val, (int, float)):
        return val * 86400.0  # Excel numeric value to seconds
    return 0

def processar_calculo_total():
    if not os.path.exists(PLANILHA_MASTER):
        print(f"Erro: Planilha não encontrada em {PLANILHA_MASTER}")
        return

    print(f"Abrindo planilha Master: {PLANILHA_MASTER}...")
    # keep_vba=True para preservar macros se houver (mesmo sendo .xlsx agora, carregar com segurança)
    wb = openpyxl.load_workbook(PLANILHA_MASTER, keep_vba=True)
    ws = wb[ABA_MES]

    max_row = ws.max_row
    count_calculados = 0
    count_pula = 0

    print(f"Iniciando cálculo da Coluna R (Total) da linha 10 até {max_row}...")

    for row in range(10, max_row + 1):
        # 1. Verificar Nome para regra de exclusão
        nome = ws.cell(row=row, column=9).value # Coluna I
        if nome and "NAO FAZ DP" in str(nome).upper():
            ws.cell(row=row, column=18).value = 0 # Zera o total se não faz DP? Ou deixa em branco?
            # O usuário disse: "não devem ser calculados o nomes para não dar erro"
            # Vamos deixar 0 para não quebrar fórmulas de soma externas
            ws.cell(row=row, column=18).number_format = '[h]:mm:ss'
            count_pula += 1
            continue

        # 2. Capturar valores de Fiscal(O), Contabil(P) e Pessoal(Q)
        v_fiscal = ws.cell(row=row, column=15).value # O
        v_contabil = ws.cell(row=row, column=16).value # P
        v_pessoal = ws.cell(row=row, column=17).value # Q

        # 3. Somar em segundos para precisão
        s_fiscal = val_to_seconds(v_fiscal)
        s_contabil = val_to_seconds(v_contabil)
        s_pessoal = val_to_seconds(v_pessoal)

        total_segundos = s_fiscal + s_contabil + s_pessoal
        
        # 4. Gravar em R (18) como valor numérico (fração de dia)
        if total_segundos > 0:
            ws.cell(row=row, column=18).value = total_segundos / 86400.0
        else:
            ws.cell(row=row, column=18).value = 0
            
        ws.cell(row=row, column=18).number_format = '[h]:mm:ss'
        count_calculados += 1

    # 5. Inserir fórmula de SUBTOTAL em R7
    ws.cell(row=7, column=18).value = f"=SUBTOTAL(9,R10:R{max_row})"
    ws.cell(row=7, column=18).number_format = '[h]:mm:ss'

    print(f"Salvando planilha...")
    wb.save(PLANILHA_MASTER)
    print(f"Sucesso! {count_calculados} linhas calculadas. {count_pula} linhas puladas (NAO FAZ DP).")
    print(f"Subtotal inserido em R7: =SUBTOTAL(9,R10:R{max_row})")

if __name__ == "__main__":
    processar_calculo_total()
