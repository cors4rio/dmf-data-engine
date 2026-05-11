import os
import shutil
import openpyxl

def update_total_with_dp():
    xls_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls'
    temp_xlsx = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\temp_update_total.xlsx'
    
    print(f"Criando cópia temporária XLSX...")
    shutil.copy(xls_path, temp_xlsx)

    print(f"Abrindo workbook...")
    wb = openpyxl.load_workbook(temp_xlsx)
    sheet_name = '12.2025'
    if sheet_name not in wb.sheetnames:
        print(f"Erro: Aba {sheet_name} não encontrada.")
        return
        
    ws = wb[sheet_name]
    
    # Colunas: O (15), P (16), Q (17), R (18)
    # No openpyxl, as colunas são 1-indexed: O=15, P=16, Q=17, R=18
    # Verificando os índices baseados no log anterior:
    # 14: HORARIO FISCAL (O)
    # 15: HORARIO CONTÁBIL (P)
    # 16: HORARIO PESSOAL (Q)
    # 17: TOTAL (R)
    
    col_fiscal = 15
    col_contabil = 16
    col_dp = 17
    col_total = 18
    
    count = 0
    max_row = ws.max_row
    
    for row_idx in range(10, max_row + 1):
        # Inserir a fórmula na coluna R
        # Fórmula: =SOMA(O10:Q10) ou =O10+P10+Q10
        # Usaremos + para ser direto
        formula = f"=O{row_idx}+P{row_idx}+Q{row_idx}"
        cell_total = ws.cell(row=row_idx, column=col_total)
        cell_total.value = formula
        cell_total.number_format = '[h]:mm:ss'
        
        # Garantir que as colunas de origem também tenham o formato correto se forem números
        for col_idx in [col_fiscal, col_contabil, col_dp]:
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '[h]:mm:ss'
        
        count += 1

    print(f"Fórmulas de TOTAL atualizadas em {count} linhas.")

    wb.save(temp_xlsx)
    wb.close()
    
    shutil.copy(temp_xlsx, xls_path)
    os.remove(temp_xlsx)
    print(f"Arquivo {xls_path} atualizado com sucesso!")

if __name__ == "__main__":
    update_total_with_dp()
