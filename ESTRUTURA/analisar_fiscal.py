import xlrd
import openpyxl
import datetime

def time_to_hours(time_str):
    try:
        if not time_str or str(time_str).strip() == '':
            return 0.0
        if isinstance(time_str, datetime.time):
            return time_str.hour + (time_str.minute / 60.0) + (time_str.second / 3600.0)
        if isinstance(time_str, datetime.timedelta):
            return time_str.total_seconds() / 3600.0
        parts = str(time_str).split(':')
        if len(parts) >= 3:
            h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
            return h + (m / 60.0) + (s / 3600.0)
    except:
        pass
    return 0.0

def func_analisar_fiscal():
    FILE_FISCAL = r"C:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\ClienteTempo Gasto.xls"
    FILE_MASTER = r"C:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm"

    # 1. Totalizar arquivo de origem (Fiscal)
    clientes_fiscal = {}
    total_horas_fiscal_origem = 0.0
    
    wb_f = xlrd.open_workbook(FILE_FISCAL)
    sh_f = wb_f.sheet_by_index(0)
    
    for i in range(1, sh_f.nrows):
        r = sh_f.row_values(i)
        cod_raw = r[0]   # Col A
        nome_raw = r[4]  # Col E
        tempo_raw = r[17] if len(r) > 17 else ''
        
        horas = time_to_hours(tempo_raw)
        if horas > 0:
            total_horas_fiscal_origem += horas
            cod_str = ""
            if cod_raw:
                try: cod_str = str(int(float(str(cod_raw).strip())))
                except: cod_str = str(cod_raw).strip()
            
            nome = str(nome_raw).strip().upper() if nome_raw else ""
            
            if cod_str or nome:
                chave = cod_str if cod_str else nome
                clientes_fiscal[chave] = {
                    'cod': cod_str,
                    'nome': nome,
                    'horas': horas
                }

    print(f"=== TOTAL NA ORIGEM (ClienteTempo Gasto.xls) ===")
    print(f"Total Horas: {total_horas_fiscal_origem:.2f}")
    print(f"Clientes com tempo > 0: {len(clientes_fiscal)}")
    print()

    # 2. Totalizar arquivo master gerado (Coluna N)
    clientes_master = {}
    total_horas_master_inj = 0.0
    
    wb_m = openpyxl.load_workbook(FILE_MASTER, data_only=True)
    sh_m = wb_m["12.2025"]
    
    for row in range(10, sh_m.max_row + 1):
        cod_cell = sh_m.cell(row=row, column=8).value
        nome_cell = sh_m.cell(row=row, column=11).value
        n_val = sh_m.cell(row=row, column=14).value
        
        if n_val:
            # The value is a fraction of a day, multiply by 24 to get hours
            # wait, if it's already a float or datetime.time? openpyxl returning datetime.time for [h]:mm:ss formats might occur.
            if isinstance(n_val, datetime.time):
                 horas = n_val.hour + (n_val.minute / 60.0) + (n_val.second / 3600.0)
            elif isinstance(n_val, datetime.timedelta):
                 horas = n_val.total_seconds() / 3600.0
            else:
                 horas = float(n_val) * 24.0 # It was injected as horas / 24.0
            
            total_horas_master_inj += horas
            
            cod_str = ""
            try:
                if cod_cell is not None:
                    cod_str = str(int(float(str(cod_cell).strip().split('.')[0])))
            except: pass
            
            nome = str(nome_cell).strip().upper() if nome_cell else ""
            chave = cod_str if cod_str else nome
            
            clientes_master[chave] = {
                'cod': cod_str,
                'nome': nome,
                'horas': horas,
                'linha': row
            }

    print(f"=== TOTAL NO MASTER INJETADO (CONTROLE_DE_HORAS_DMF.xlsm - Coluna N) ===")
    print(f"Total Horas Injetadas: {total_horas_master_inj:.2f}")
    print(f"Clientes preenchidos: {len(clientes_master)}")
    print()
    
    print(f"=== DIFERENÇA ===")
    diferenca = total_horas_fiscal_origem - total_horas_master_inj
    print(f"Diferença de horas: {diferenca:.2f} \n")
    
    # 3. Identificar clientes faltantes (Estavam na Origem, mas NÃO estão no Master)
    faltantes = []
    horas_faltantes = 0.0
    for chave, c_fisc in clientes_fiscal.items():
        if chave not in clientes_master:
            # Check by name in master keys if key was code, just in case
            found_by_nome = False
            for m_key, c_mast in clientes_master.items():
                if c_fisc['nome'] and c_fisc['nome'] in c_mast['nome']:
                    found_by_nome = True
                    break
            
            if not found_by_nome:
                faltantes.append(c_fisc)
                horas_faltantes += c_fisc['horas']
                
    # Sort descending by hours
    faltantes.sort(key=lambda x: x['horas'], reverse=True)
    
    print(f"Encontrados {len(faltantes)} clientes que possuem horas na origem mas NÃO foram lançados no Master.")
    print(f"A soma de horas desses clientes faltantes é: {horas_faltantes:.2f}")
    print("\nTop 20 Clientes Faltantes com mais horas:")
    for f in faltantes[:20]:
        print(f"  Cod: {f['cod']:<6} | Nome: {f['nome']:<40} | Horas: {f['horas']:.2f}")

if __name__ == '__main__':
    func_analisar_fiscal()
