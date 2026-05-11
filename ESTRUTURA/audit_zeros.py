import openpyxl

file_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_V3_AeqwQXgR.xlsx'

try:
    wb_meta = openpyxl.load_workbook(file_path, data_only=False)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    
    print(f"Abas encontradas: {wb_meta.sheetnames}")
    
    for sheet_name in wb_meta.sheetnames:
        ws_meta = wb_meta[sheet_name]
        ws_data = wb_data[sheet_name]
        
        # Tentar localizar a coluna Q (ou buscar por cabeçalho "TOTAL" no caso de Q1 ser cabeçalho)
        # O usuário mencionou "CAMPO Q1", talvez Q seja o nome ou a posição.
        # Na aba 01.2025, Q2 tinha a fórmula.
        
        q1_val = ws_data['Q1'].value
        print(f"\nAba: {sheet_name} | Q1: {q1_val}")
        
        zeros_with_source = []
        for row in range(2, min(ws_meta.max_row + 1, 1000)):
            q_val = ws_data[f'Q{row}'].value
            if q_val == 0 or q_val is None:
                # Verificar se fontes são > 0
                h = ws_data[f'H{row}'].value or 0
                j = ws_data[f'J{row}'].value or 0
                l = ws_data[f'L{row}'].value or 0
                n = ws_data[f'N{row}'].value or 0
                p = ws_data[f'P{row}'].value or 0
                
                # Tratar se for string
                try:
                    h = float(str(h).replace(',', '.')) if isinstance(h, str) else h
                    j = float(str(j).replace(',', '.')) if isinstance(j, str) else j
                    l = float(str(l).replace(',', '.')) if isinstance(l, str) else l
                    n = float(str(n).replace(',', '.')) if isinstance(n, str) else n
                    p = float(str(p).replace(',', '.')) if isinstance(p, str) else p
                except:
                    pass
                
                if isinstance(h, (int, float)) and h > 0 or \
                   isinstance(j, (int, float)) and j > 0 or \
                   isinstance(l, (int, float)) and l > 0 or \
                   isinstance(n, (int, float)) and n > 0 or \
                   isinstance(p, (int, float)) and p > 0:
                    zeros_with_source.append(row)
        
        print(f"Linhas com Q=0 mas fontes > 0: {len(zeros_with_source)}")
        if zeros_with_source:
            print(f"Exemplos de linhas: {zeros_with_source[:10]}")
            row = zeros_with_source[0]
            print(f"Dados da linha {row}:")
            for col in ['H', 'J', 'L', 'N', 'P', 'Q']:
                print(f"  {col}: Formula={ws_meta[f'{col}{row}'].value}, Valor={ws_data[f'{col}{row}'].value}")

except Exception as e:
    print(f"Erro: {e}")
