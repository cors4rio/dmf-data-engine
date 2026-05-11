import openpyxl

arquivo_final = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm"

try:
    wb = openpyxl.load_workbook(arquivo_final, data_only=True)
    sn = "02.2026"
    if sn not in wb.sheetnames:
        sn = wb.sheetnames[-3] # Tentar uma das últimas se 02.2026 não existir
        
    print(f"Usando aba: {sn}")
    ws = wb[sn]
    
    # Ver cabeçalhos até a coluna Z na linha 8
    headers = {}
    for i in range(1, 27):
        col_letter = openpyxl.utils.get_column_letter(i)
        headers[col_letter] = ws.cell(row=8, column=i).value
        
    print(f"Cabeçalhos (Linha 8): {headers}")
    
    # Ver dados da linha 9 (P9, R9)
    row = 9
    print(f"Dados da Linha {row}:")
    for i in range(1, 27):
        col_letter = openpyxl.utils.get_column_letter(i)
        val = ws.cell(row=row, column=i).value
        print(f"  {col_letter}: {val}")

except Exception as e:
    print(f"Erro: {e}")
