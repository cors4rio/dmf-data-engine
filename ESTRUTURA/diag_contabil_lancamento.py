"""
Diagnóstico de precisão do lançamento contábil.
Verifica:
1. Empresa 1283 - detalhamento do faturamento
2. Contagem de lançamentos com diferentes filtros de orig_lan
3. Comparação com a planilha 02.2026 existente
"""
import pyodbc
import openpyxl

DB_CONN_STR = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
PLANILHA = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_.xlsx'

DATA_INICIO = '2026-02-01'
DATA_FIM = '2026-02-28'

def diagnosticar():
    conn = pyodbc.connect(DB_CONN_STR)
    cursor = conn.cursor()

    # ======= DIAGNÓSTICO 1: Empresa 1283 - Faturamento Detalhado =======
    print("=" * 60)
    print("DIAGNÓSTICO 1: FATURAMENTO EMPRESA 1283")
    print("=" * 60)

    # Saídas (efsaidas)
    cursor.execute(f"""
        SELECT SUM(vcon_sai) as total_saidas
        FROM bethadba.efsaidas
        WHERE codi_emp = 1283
        AND dsai_sai >= '{DATA_INICIO}' AND dsai_sai <= '{DATA_FIM}'
    """)
    row = cursor.fetchone()
    total_saidas = float(row[0]) if row and row[0] else 0
    print(f"  efsaidas (vcon_sai): R$ {total_saidas:,.2f}")

    # Serviços (efservicos)
    cursor.execute(f"""
        SELECT SUM(vcon_ser) as total_servicos
        FROM bethadba.efservicos
        WHERE codi_emp = 1283
        AND dser_ser >= '{DATA_INICIO}' AND dser_ser <= '{DATA_FIM}'
    """)
    row = cursor.fetchone()
    total_servicos = float(row[0]) if row and row[0] else 0
    print(f"  efservicos (vcon_ser): R$ {total_servicos:,.2f}")
    print(f"  TOTAL (saidas + servicos): R$ {total_saidas + total_servicos:,.2f}")
    print(f"  Valor esperado pelo usuario: R$ 1,179,957.53")
    print(f"  Diferença: R$ {(total_saidas + total_servicos) - 1179957.53:,.2f}")

    # ======= DIAGNÓSTICO 2: Lançamentos com diferentes filtros =======
    print("\n" + "=" * 60)
    print("DIAGNÓSTICO 2: CONTAGEM DE LANÇAMENTOS POR ORIG_LAN")
    print("=" * 60)

    # Quais valores de orig_lan existem?
    cursor.execute(f"""
        SELECT orig_lan, COUNT(*) as qtd
        FROM bethadba.ctlancto
        WHERE data_lan >= '{DATA_INICIO}' AND data_lan <= '{DATA_FIM}'
        GROUP BY orig_lan
        ORDER BY orig_lan
    """)
    print("  Distribuição de orig_lan no período:")
    for row in cursor.fetchall():
        print(f"    orig_lan={row[0]}: {row[1]} lançamentos")

    # Total sem filtro
    cursor.execute(f"""
        SELECT COUNT(DISTINCT codi_emp) as empresas,
               COUNT(*) as total_lancamentos
        FROM bethadba.ctlancto
        WHERE data_lan >= '{DATA_INICIO}' AND data_lan <= '{DATA_FIM}'
    """)
    row = cursor.fetchone()
    print(f"\n  SEM filtro orig_lan: {row[0]} empresas, {row[1]} lançamentos total")

    # Apenas orig_lan = 1
    cursor.execute(f"""
        SELECT COUNT(DISTINCT codi_emp) as empresas,
               COUNT(*) as total_lancamentos
        FROM bethadba.ctlancto
        WHERE data_lan >= '{DATA_INICIO}' AND data_lan <= '{DATA_FIM}'
        AND orig_lan = 1
    """)
    row = cursor.fetchone()
    print(f"  COM filtro orig_lan=1: {row[0]} empresas, {row[1]} lançamentos total")

    # Sem filtro de orig_lan (todos)
    cursor.execute(f"""
        SELECT codi_emp, COUNT(*) as qtd
        FROM bethadba.ctlancto
        WHERE data_lan >= '{DATA_INICIO}' AND data_lan <= '{DATA_FIM}'
        GROUP BY codi_emp
        ORDER BY codi_emp
    """)
    all_launches = {str(r[0]): r[1] for r in cursor.fetchall()}
    print(f"\n  Empresas COM lançamentos (qualquer orig_lan): {len(all_launches)}")

    # ======= DIAGNÓSTICO 3: Comparar com aba 02.2026 =======
    print("\n" + "=" * 60)
    print("DIAGNÓSTICO 3: COMPARAÇÃO COM ABA 02.2026")
    print("=" * 60)

    wb = openpyxl.load_workbook(PLANILHA, data_only=True)
    ws02 = wb['02.2026']

    # Empresa 1283 na aba 02
    for r in range(2, ws02.max_row + 1):
        cod = ws02.cell(r, 1).value
        if cod is None:
            continue
        try:
            cod_str = str(int(float(str(cod).strip())))
        except (ValueError, TypeError):
            continue
        if cod_str == '1283':
            f_val = ws02.cell(r, 6).value
            o_val = ws02.cell(r, 15).value
            print(f"  Aba 02.2026 - Cod 1283 (Row {r}):")
            print(f"    F (Lancamentos): {f_val}")
            print(f"    O (Faturamento): {o_val}")
            break

    # Quantas empresas no 02 têm lancamentos > 0?
    count_f02 = 0
    for r in range(2, ws02.max_row + 1):
        try:
            v = ws02.cell(r, 6).value
            if v and v > 0:
                count_f02 += 1
        except:
            pass
    print(f"\n  Total com lancamentos > 0 na aba 02.2026: {count_f02}")

    # ======= DIAGNÓSTICO 4: Empresa 1283 - faturamento SÓ por saídas =======
    print("\n" + "=" * 60)
    print("DIAGNÓSTICO 4: EMP 1283 - DETALHAMENTO SAÍDAS")
    print("=" * 60)
    cursor.execute(f"""
        SELECT dsai_sai, SUM(vcon_sai) as total
        FROM bethadba.efsaidas
        WHERE codi_emp = 1283
        AND dsai_sai >= '{DATA_INICIO}' AND dsai_sai <= '{DATA_FIM}'
        GROUP BY dsai_sai
        ORDER BY dsai_sai
    """)
    total_check = 0
    for row in cursor.fetchall():
        val = float(row[1])
        total_check += val
        print(f"  Data: {row[0]}, Valor: R$ {val:,.2f}")
    print(f"  TOTAL saidas 1283: R$ {total_check:,.2f}")

    conn.close()

if __name__ == "__main__":
    diagnosticar()
