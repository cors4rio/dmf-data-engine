import openpyxl; wb=openpyxl.load_workbook('CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx'); ws=wb['03.2026']; 
for row in ws.iter_rows(min_row=10, max_row=ws.max_row, min_col=14, max_col=14): 
    for cell in row: cell.value = None
wb.save('CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx'); print('Column N cleared!')
