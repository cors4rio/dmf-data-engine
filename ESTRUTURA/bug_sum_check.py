import openpyxl

file_path = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm"
wb = openpyxl.load_workbook(file_path, data_only=False)
sh = wb['02.2026']

print(f"--- Fórmulas e Tipos (Aba {sh.title}) ---")
print(f"Q7 Formula: {sh['Q7'].value}")

# Verificar algumas linhas da coluna Q e R
for r in range(10, 20):
    val_q = sh.cell(row=r, column=17).value
    type_q = type(val_q).__name__
    val_o = sh.cell(row=r, column=15).value
    type_o = type(val_o).__name__
    val_r = sh.cell(row=r, column=18).value
    
    print(f"Linha {r}:")
    print(f"  Col O (Fiscal): {val_o} ({type_o})")
    print(f"  Col Q (DP):     {val_q} ({type_q})")
    print(f"  Col R (Total):  {val_r}")
