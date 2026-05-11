import pyodbc
import openpyxl
import os
from datetime import datetime

def mapear_regime(descricao, rfed=None):
    if not descricao:
        if rfed == 1: return "Lucro Real"
        if rfed in [2, 4]: return "Simples Nacional"
        if rfed == 5: return "Lucro Presumido"
        if rfed == 8: return "Imune / Isenta"
        return "Não Parametrizado"
    
    desc_upper = descricao.upper()
    
    if "SIMPLES" in desc_upper:
        return "Simples Nacional"
    if any(kw in desc_upper for kw in ["PRESUMIDO", "PRESIMIDO", "PRESSUMIDO"]):
        return "Lucro Presumido"
    if "REAL" in desc_upper:
        return "Lucro Real"
    if "PRODUTOR RURAL" in desc_upper:
        return "Produtor Rural"
    if any(kw in desc_upper for kw in ["IMUNE", "ISENTA"]):
        return "Imune / Isenta"
        
    if any(kw in desc_upper for kw in ["INICIAL", "NOVO", "VIGÊNCIA", "PRIMEIRA"]):
        if rfed == 1: return "Lucro Real"
        if rfed in [2, 4]: return "Simples Nacional"
        if rfed == 5: return "Lucro Presumido"
        if rfed == 8: return "Imune / Isenta"

    palavras = descricao.split()
    return " ".join(palavras[:2])

def get_regimes_database(data_referencia):
    regimes = {}
    conn_str = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        query = f"""
            SELECT p.codi_emp, p.descricao_par, p.rfed_par
            FROM bethadba.efparametro_vigencia p
            WHERE p.vigencia_par = (
                SELECT MAX(vigencia_par)
                FROM bethadba.efparametro_vigencia
                WHERE codi_emp = p.codi_emp
                  AND vigencia_par <= '{data_referencia}'
            )
        """
        cursor.execute(query)
        for row in cursor.fetchall():
            cod_emp = row[0]
            desc = row[1]
            rfed = row[2]
            regimes[cod_emp] = mapear_regime(desc, rfed)
        conn.close()
        return regimes
    except Exception as e:
        print(f"Erro ao acessar banco de dados para {data_referencia}: {e}")
        return {}

def atualizar_planilha_regimes():
    planilha_path = r"C:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS - REGIMES.xlsx"
    
    if not os.path.exists(planilha_path):
        print(f"Erro: Arquivo não encontrado em {planilha_path}")
        return

    print("Carregando planilha...")
    wb = openpyxl.load_workbook(planilha_path)
    
    # 1. Carregar cache de 01.2026 (Referência principal)
    # IMPORTANTE: Vamos usar Janeiro/2026 como base absoluta para todos os meses <= 2026-01
    print("Buscando regimes de referência em 01/2026...")
    cache_2026 = get_regimes_database("2026-01-31")
    print(f"Cache de 01/2026 carregado com {len(cache_2026)} empresas.")

    cache_dinamico = {}

    for sheet_name in wb.sheetnames:
        if "." not in sheet_name or len(sheet_name) != 7:
            continue
            
        print(f"\nProcessando aba: {sheet_name}")
        ws = wb[sheet_name]
        
        # FASE 1: LIMPEZA TOTAL DA COLUNA D (REGIME)
        print(f"-> Limpando Coluna D em {sheet_name}...")
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=4).value = None

        # FASE 2: PREENCHIMENTO
        try:
            mes, ano = map(int, sheet_name.split("."))
            data_aba = datetime(ano, mes, 1)
            data_ref_2026 = datetime(2026, 1, 1)
            
            if data_aba <= data_ref_2026:
                print(f"-> Aplicando réplica de 01/2026 para {sheet_name}")
                regimes_atuais = cache_2026
            else:
                import calendar
                ultimo_dia = calendar.monthrange(ano, mes)[1]
                str_ref = f"{ano}-{mes:02d}-{ultimo_dia}"
                print(f"-> Buscando regime real para {sheet_name} (Ref: {str_ref})")
                if str_ref not in cache_dinamico:
                    cache_dinamico[str_ref] = get_regimes_database(str_ref)
                regimes_atuais = cache_dinamico[str_ref]

            atualizados = 0
            for row in range(2, ws.max_row + 1):
                cod_emp = ws.cell(row=row, column=1).value
                if cod_emp and isinstance(cod_emp, (int, float)):
                    cod = int(cod_emp)
                    regime = regimes_atuais.get(cod)
                    if regime:
                        ws.cell(row=row, column=4).value = regime
                        atualizados += 1
                    else:
                        ws.cell(row=row, column=4).value = "Não Localizado"
            
            print(f"-> Aba {sheet_name} finalizada: {atualizados} regimes preenchidos.")
            
        except ValueError:
            print(f"-> Pulando aba {sheet_name}")
            continue

    print("\nSalvando alterações...")
    try:
        wb.save(planilha_path)
        print("Sucesso! Planilha limpa e atualizada com lógica de réplica.")
    except Exception as e:
        print(f"Erro ao salvar planilha: {e}")

if __name__ == "__main__":
    atualizar_planilha_regimes()
