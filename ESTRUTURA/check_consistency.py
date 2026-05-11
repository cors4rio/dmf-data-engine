import openpyxl
import re

file_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_V3_AeqwQXgR.xlsx'

def check_formula_consistency(sheet_name, ws):
    print(f"\n--- Aba: {sheet_name} ---")
    inconsistent_rows = []
    
    for row in range(2, ws.max_row + 1):
        formula = ws[f'Q{row}'].value
        if not isinstance(formula, str) or not formula.startswith('='):
            continue
            
        # Regex para extrair números de linha da fórmula (H2, J2, L2, N2, P2)
        row_refs = re.findall(r'[A-Z]+(\d+)', formula)
        
        # Verificar se todas as referências apontam para a linha atual
        is_consistent = all(int(r_ref) == row for r_ref in row_refs)
        
        if not is_consistent:
            inconsistent_rows.append((row, formula))
            
    print(f"Total de linhas com fórmulas: {len(inconsistent_rows) + (ws.max_row - 1 - len(inconsistent_rows))}")
    print(f"Linhas com fórmulas inconsistentes: {len(inconsistent_rows)}")
    
    if inconsistent_rows:
        print("Exemplos de inconsistência:")
        for row, formula in inconsistent_rows[:5]:
            print(f"  Linha {row}: Formula={formula} (Deveria ser relacionado à linha {row})")

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    for sn in wb.sheetnames:
        if 'MÉDIA' in sn or 'EXEMPLO' in sn: continue
        check_formula_consistency(sn, wb[sn])
except Exception as e:
    print(f"Erro: {e}")
