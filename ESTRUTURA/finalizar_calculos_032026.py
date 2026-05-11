import openpyxl

PLANILHA_MASTER = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF.xlsx"
ABA_ALVO = "03.2026"

def main():
    print(f"Recalculando Totais na aba {ABA_ALVO}...")
    wb = openpyxl.load_workbook(PLANILHA_MASTER)
    ws = wb[ABA_ALVO]

    recalculados = 0
    # O=15 (Fiscal), P=16 (Contabil), Q=17 (DP), R=18 (Total)
    for i in range(10, ws.max_row + 1):
        cod_cell = ws.cell(row=i, column=8).value
        
        # Ignorar linhas sem cdigo (mesma lgica do script fiscal)
        if cod_cell is None:
            continue
            
        try:
            # Garante que conseguimos ler o cdigo
            cod_str = str(cod_cell).strip()
            if not cod_str or cod_str.lower() in ['none', 'null']:
                continue
                
            # Ler valores das colunas O, P e Q
            # Se for string (ex: "NAO FAZ DP"), tratamos como 0 no somatrio
            def get_val(col):
                val = ws.cell(row=i, column=col).value
                if val is None: return 0.0
                if isinstance(val, (int, float)): return float(val)
                # Se for string, tentamos converter ou retornamos 0
                try:
                    return float(val)
                except:
                    return 0.0

            ov = get_val(15)
            pv = get_val(16)
            qv = get_val(17)
            
            total = ov + pv + qv
            
            # Escrever no Total (R)
            cell_r = ws.cell(row=i, column=18)
            cell_r.value = total
            cell_r.number_format = '[h]:mm:ss'
            recalculados += 1
            
        except Exception as e:
            # Se falhar em uma linha, continua para as outras
            pass

    # Atualizar Subtotal do Total (R7)
    ws.cell(row=7, column=18, value=f"=SUBTOTAL(9,R10:R{ws.max_row})")
    
    print(f"Sucesso: {recalculados} linhas do TOTAL (R) atualizadas.")
    wb.save(PLANILHA_MASTER)
    wb.close()

if __name__ == "__main__":
    main()
