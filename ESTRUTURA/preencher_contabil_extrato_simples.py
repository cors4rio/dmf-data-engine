import openpyxl
import os
import re
import difflib

BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = os.path.join(BASE_DIR, "CONTROLE DE HORAS DMF.xlsx")
PLANILHA_EXTRATO = os.path.join(BASE_DIR, "Empresas do simples somente com extrato_1778096183676.xlsx")
MES_FONTE = '03.2026'

def clean_cnpj(val):
    if not val:
        return ""
    return re.sub(r'\D', '', str(val).strip())

def normalizar_doc(cnpj_limpo):
    if not cnpj_limpo:
        return ""
    n = len(cnpj_limpo)
    if n <= 11:
        return cnpj_limpo.zfill(11)
    else:
        return cnpj_limpo.zfill(14)

def similar(a, b):
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, str(a).upper().strip(), str(b).upper().strip()).ratio()

def main():
    # 1. Carregar dados da planilha Master (03.2026)
    print("Abrindo planilha Master (CONTROLE DE HORAS DMF)...")
    wb_master = openpyxl.load_workbook(PLANILHA_MASTER, read_only=True, data_only=True)
    
    if MES_FONTE not in wb_master.sheetnames:
        print(f"Aba {MES_FONTE} não encontrada na Master!")
        return
    
    ws_master = wb_master[MES_FONTE]
    
    # Indexar Master por código -> {cod: {cnpj_norm, nome, contabil_valor, row}}
    master_por_cod = {}
    master_por_cnpj = {}
    for row in range(10, ws_master.max_row + 1):
        cod_raw = ws_master.cell(row=row, column=8).value   # Col H (Cód Domínio)
        cnpj_raw = ws_master.cell(row=row, column=10).value  # Col J (CNPJ)
        nome_raw = ws_master.cell(row=row, column=11).value  # Col K (Nome)
        contabil_val = ws_master.cell(row=row, column=16).value  # Col P (Horário Contábil)
        
        cnpj = clean_cnpj(cnpj_raw)
        cnpj_norm = normalizar_doc(cnpj)
        nome = str(nome_raw).strip() if nome_raw else ""
        
        cod = None
        if cod_raw is not None:
            try:
                cod = int(float(str(cod_raw).strip()))
            except (ValueError, TypeError):
                pass
        
        dados = {
            'cod': cod,
            'cnpj': cnpj,
            'cnpj_norm': cnpj_norm,
            'nome': nome,
            'contabil': contabil_val,
            'row': row
        }
        
        if cod is not None:
            master_por_cod[cod] = dados
        if cnpj_norm:
            master_por_cnpj[cnpj_norm] = dados
    
    wb_master.close()
    print(f"  -> {len(master_por_cod)} empresas indexadas por código, {len(master_por_cnpj)} por CNPJ.")
    
    # 2. Abrir planilha Extrato para escrita
    print("Abrindo planilha Extrato Simples...")
    wb_extrato = openpyxl.load_workbook(PLANILHA_EXTRATO)
    ws_extrato = wb_extrato.active
    
    encontrados = []
    nao_encontrados = []
    rejeitados = []
    alterados = 0
    total_empresas = 0
    
    print("Processando empresas do Extrato Simples com Validação Tripla...")
    for row in range(2, ws_extrato.max_row + 1):
        cnpj_extrato_raw = ws_extrato.cell(row=row, column=1).value   # Col A (CNPJ)
        cod_extrato_raw = ws_extrato.cell(row=row, column=2).value    # Col B (Código)
        nome_extrato = ws_extrato.cell(row=row, column=5).value       # Col E (Nome)
        
        # Pular linhas vazias
        if not cnpj_extrato_raw and not cod_extrato_raw and not nome_extrato:
            continue
        
        total_empresas += 1
        cnpj_extrato = clean_cnpj(cnpj_extrato_raw)
        cnpj_extrato_norm = normalizar_doc(cnpj_extrato)
        nome_extrato_str = str(nome_extrato).strip() if nome_extrato else ""
        
        cod_extrato = None
        if cod_extrato_raw is not None:
            try:
                cod_extrato = int(float(str(cod_extrato_raw).strip()))
            except (ValueError, TypeError):
                pass
        
        # VALIDAÇÃO TRIPLA: buscar por código primeiro, depois CNPJ
        dados_master = None
        busca_metodo = ""
        
        if cod_extrato is not None and cod_extrato in master_por_cod:
            dados_master = master_por_cod[cod_extrato]
            busca_metodo = "código"
        elif cnpj_extrato_norm and cnpj_extrato_norm in master_por_cnpj:
            dados_master = master_por_cnpj[cnpj_extrato_norm]
            busca_metodo = "CNPJ"
        
        if dados_master is None:
            nao_encontrados.append((row, cod_extrato, nome_extrato_str, cnpj_extrato, 
                                    "Não encontrado na Master (nem por código nem por CNPJ)"))
            continue
        
        # Validar CNPJ
        motivos_rejeicao = []
        if cnpj_extrato_norm and dados_master['cnpj_norm'] and cnpj_extrato_norm != dados_master['cnpj_norm']:
            motivos_rejeicao.append(f"CNPJ difere: Extrato={cnpj_extrato} | Master={dados_master['cnpj']}")
        
        # Validar Código
        if cod_extrato is not None and dados_master['cod'] is not None and cod_extrato != dados_master['cod']:
            motivos_rejeicao.append(f"Código difere: Extrato={cod_extrato} | Master={dados_master['cod']}")
        
        # Validar Nome
        sim_ratio = similar(nome_extrato_str, dados_master['nome'])
        if sim_ratio < 0.40:
            motivos_rejeicao.append(f"Nome difere (sim={sim_ratio:.2f}): Extrato={nome_extrato_str[:30]} | Master={dados_master['nome'][:30]}")
        
        if motivos_rejeicao:
            rejeitados.append((row, cod_extrato, nome_extrato_str, cnpj_extrato, 
                              "; ".join(motivos_rejeicao)))
            continue
        
        # Passou na validação tripla - lançar o valor contábil
        contabil_val = dados_master['contabil']
        
        if contabil_val is not None:
            c = ws_extrato.cell(row=row, column=6, value=contabil_val)  # Col F
            if isinstance(contabil_val, (int, float)):
                c.number_format = '[h]:mm:ss'
            alterados += 1
            encontrados.append((row, cod_extrato, nome_extrato_str, cnpj_extrato, 
                               contabil_val, dados_master['nome']))
        else:
            nao_encontrados.append((row, cod_extrato, nome_extrato_str, cnpj_extrato,
                                   "Valor Contábil vazio na Master (col P)"))
    
    print(f"Salvando planilha Extrato... ({alterados} linhas alteradas)")
    wb_extrato.save(PLANILHA_EXTRATO)
    wb_extrato.close()
    
    # Relatório
    RELATORIO = os.path.join(BASE_DIR, "relatorio_contabil_extrato_simples.md")
    with open(RELATORIO, 'w', encoding='utf-8') as f:
        f.write("# Relatório - Lançamento Contábil no Extrato Simples\n\n")
        f.write(f"- **Fonte:** CONTROLE DE HORAS DMF.xlsx → aba {MES_FONTE}, coluna P (Horário Contábil)\n")
        f.write(f"- **Destino:** Extrato Simples → coluna F (Horas Contábeis)\n")
        f.write(f"- **Validação:** Tripla (Código + CNPJ + Nome)\n\n")
        f.write(f"- Total de empresas no Extrato: **{total_empresas}**\n")
        f.write(f"- Empresas lançadas com sucesso: **{alterados}**\n")
        f.write(f"- Empresas sem lançamento: **{len(nao_encontrados) + len(rejeitados)}**\n")
        f.write(f"  - Não encontradas na Master: {len(nao_encontrados)}\n")
        f.write(f"  - Rejeitadas na validação tripla: {len(rejeitados)}\n\n")
        
        if encontrados:
            f.write("## Empresas Lançadas\n\n")
            f.write("| Linha | Código | Nome | CNPJ | Valor Contábil |\n")
            f.write("|---|---|---|---|---|\n")
            for row, cod, nome, cnpj, contabil, nome_master in encontrados:
                if isinstance(contabil, (int, float)):
                    h_total = contabil * 24
                    h = int(h_total)
                    m = int(round((h_total - h) * 60))
                    contabil_str = f"{h}:{m:02d}:00"
                else:
                    contabil_str = str(contabil)
                f.write(f"| {row} | {cod} | {nome[:35]} | {cnpj} | {contabil_str} |\n")
        
        if rejeitados:
            f.write("\n## Rejeitados na Validação Tripla\n\n")
            f.write("| Linha | Código | Nome | CNPJ | Motivo |\n")
            f.write("|---|---|---|---|---|\n")
            for row, cod, nome, cnpj, motivo in rejeitados:
                f.write(f"| {row} | {cod} | {nome[:30]} | {cnpj} | {motivo} |\n")
        
        if nao_encontrados:
            f.write("\n## Não Encontrados\n\n")
            f.write("| Linha | Código | Nome | CNPJ | Motivo |\n")
            f.write("|---|---|---|---|---|\n")
            for row, cod, nome, cnpj, motivo in nao_encontrados:
                f.write(f"| {row} | {cod} | {nome[:35]} | {cnpj} | {motivo} |\n")
    
    print(f"Processo finalizado! Relatório salvo: {RELATORIO}")
    print(f"\n=== RESUMO ===")
    print(f"  Total empresas no Extrato: {total_empresas}")
    print(f"  Lançados: {alterados}")
    print(f"  Não encontrados: {len(nao_encontrados)}")
    print(f"  Rejeitados (validação tripla): {len(rejeitados)}")

if __name__ == '__main__':
    main()
