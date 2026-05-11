import os
import sys
import pyodbc
from collections import defaultdict
import openpyxl

PLANILHA_MASTER = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx"
ABA_MES = "03.2026"
ABA_MES_ANTERIOR = "02.2026"
FATOR_ADICIONAL_FISCAL = 1.80

def get_fiscal_data_from_db():
    conn_str = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
    
    query = """
    SELECT 
        l.codi_emp as codigo_cliente,
        SUM(DATEDIFF(second, 
            YMD(YEAR(l.data_log), MONTH(l.data_log), DAY(l.data_log)) + l.tini_log, 
            COALESCE(
                YMD(YEAR(l.dfim_log), MONTH(l.dfim_log), DAY(l.dfim_log)) + l.tfim_log, 
                YMD(YEAR(l.data_log), MONTH(l.data_log), DAY(l.data_log)) + l.tfim_log
            )
        )) as total_segundos
    FROM 
        bethadba.geloguser l
    JOIN 
        bethadba.geempre e ON l.codi_emp = e.codi_emp
    WHERE 
        l.sist_log = 5 -- Módulo Fiscal
        AND l.tfim_log IS NOT NULL -- Sessões finalizadas
        AND l.data_log BETWEEN '2026-03-01' AND '2026-03-31'
    GROUP BY 
        l.codi_emp
    """
    
    dict_fiscal = {}
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        print("Buscando dados no DB Domínio (Fiscal)...")
        cursor.execute(query)
        for row in cursor.fetchall():
            cod, segundos = row
            if segundos:
                dict_fiscal[cod] = float(segundos)
    except Exception as e:
        print(f"Erro ODBC: {e}")
        sys.exit(1)
    finally:
        if 'conn' in locals():
            conn.close()
            
    return dict_fiscal

def segundos_para_hms(total_segundos):
    total_segundos = int(total_segundos)
    h = total_segundos // 3600
    m = (total_segundos % 3600) // 60
    s = total_segundos % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

def main():
    dados_fiscal = get_fiscal_data_from_db()
    print(f"Total de empresas com hora fiscal encontrada: {len(dados_fiscal)}")

    print(f"Abrindo planilha Master ({PLANILHA_MASTER})...")
    wb = openpyxl.load_workbook(PLANILHA_MASTER, keep_vba=True)

    if ABA_MES not in wb.sheetnames:
        print(f"Aba {ABA_MES} não encontrada!")
        sys.exit(1)
        
    ws = wb[ABA_MES]

    # Mapear linhas para os códigos do mês atual
    mapa_planilha = defaultdict(list)
    for i in range(10, ws.max_row + 1):
        cod_cell = ws.cell(row=i, column=8).value
        # ignorar linhas vazias ou sem código
        if cod_cell is not None:
            try:
                cod = int(float(str(cod_cell).strip()))
                mapa_planilha[cod].append(i)
            except ValueError:
                pass

    alterados = 0
    nao_encontrados = []
    calculos_md = []

    calculos_md.append("# Relatório Auditado: Cálculos Fiscal (03/2026)")
    calculos_md.append(f"**Data:** {sys.argv[0]} - Executado em processamento batch\n")
    calculos_md.append("Fator Multiplicador: 80% (1.80)\n")
    calculos_md.append("| Cód | Segundos Originais | Segundos Finais (80%) | Valor [h]:mm:ss | Status Master |")
    calculos_md.append("|---|---|---|---|---|")

    # Injetar os tempos na Master sobre todos os cadastrados, limpando fantasmas
    for cod, linhas_alvo in mapa_planilha.items():
        # Captura o valor se houver ou zera completamente
        seg_bruto = dados_fiscal.get(cod, 0)
        seg_final = seg_bruto * FATOR_ADICIONAL_FISCAL
        valor_q = seg_final / 86400.0 # Transforma em decimal para o Excel
        str_final = segundos_para_hms(seg_final)

        for linha in linhas_alvo:
            c = ws.cell(row=linha, column=15) # Coluna O = 15
            c.value = valor_q
            c.number_format = '[h]:mm:ss'
            alterados += 1
            
        if seg_bruto > 0:
            calculos_md.append(f"| {cod} | {int(seg_bruto)} | {int(seg_final)} | {str_final} | ✅ {len(linhas_alvo)} linha(s) |")
        else:
            calculos_md.append(f"| {cod} | 0 | 0 | 00:00:00 | ⚠️ Não Trabalhou (Zerado) |")

    # Identifica o que retornou do BD e não existe na base (Órfãos do Domínio)
    for cod, seg_bruto in dados_fiscal.items():
        if cod not in mapa_planilha:
            seg_final = seg_bruto * FATOR_ADICIONAL_FISCAL
            str_final = segundos_para_hms(seg_final)
            nao_encontrados.append((cod, str_final))
            calculos_md.append(f"| {cod} | {int(seg_bruto)} | {int(seg_final)} | {str_final} | ❌ Orfão não existe na Master |")
            
    # Autocalcular a Subtotalização (Coluna O = 15)
    max_row = ws.max_row
    ws.cell(row=7, column=15, value=f"=SUBTOTAL(9,O10:O{max_row})")

    print(f"Salvando planilha... ({alterados} linhas escritas na Master)")
    wb.save(PLANILHA_MASTER)
    wb.close()

    # MD 1: Erros e ausentes.
    relatorio_md = f"# Relatório de Preenchimento Fiscal ({ABA_MES})\n\n"
    relatorio_md += f"- **Empresas extraídas do DB:** {len(dados_fiscal)}\n"
    relatorio_md += f"- **Linhas alteradas na Master:** {alterados}\n"
    relatorio_md += f"- **Clientes não encontrados na Master:** {len(nao_encontrados)}\n\n"

    if nao_encontrados:
        relatorio_md += "## Clientes Ignorados (Não encontrados na Planilha)\n\n"
        relatorio_md += "| Código | Horas a preencher |\n|---|---|\n"
        for item in nao_encontrados:
            relatorio_md += f"| {item[0]} | {item[1]} |\n"

    with open("relatorio_fiscal_032026.md", "w", encoding="utf-8") as f:
        f.write(relatorio_md)

    with open("calculos_encontrados_fiscal_032026.md", "w", encoding="utf-8") as f:
        f.write("\n".join(calculos_md))

    print(f"Relatório salvo: relatorio_fiscal_032026.md")
    print(f"Log gravado: calculos_encontrados_fiscal_032026.md")


if __name__ == "__main__":
    main()
