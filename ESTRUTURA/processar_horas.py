import pyodbc
import openpyxl
import xlrd
import os
import datetime

def time_to_hours(time_str):
    try:
        if not time_str or str(time_str).strip() == '':
            return 0.0
        if isinstance(time_str, datetime.time):
            return time_str.hour + (time_str.minute / 60.0) + (time_str.second / 3600.0)
        if isinstance(time_str, datetime.timedelta):
            return time_str.total_seconds() / 3600.0
        parts = str(time_str).split(':')
        if len(parts) >= 3:
            h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
            return h + (m / 60.0) + (s / 3600.0)
    except:
        pass
    return 0.0

def clean_name(name):
    return str(name).strip().upper()

def extrair_dados_folha(mes='12', ano='2025'):
    print("1. Lendo novo relatorio do Excel (Folha)...")
    FILE_FOLHA = r"C:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\Controle de Empregados 12(CAROL).xls"
    
    dict_folha = {}
    try:
        wb = xlrd.open_workbook(FILE_FOLHA)
        sh = wb.sheet_by_index(0)
        
        # As colunas que contêm quantidades a somar baseadas no cabeçalho (linha 0)
        colunas_alvo = [7, 9, 11, 13, 15, 17, 19, 21, 23]
        
        # Pula as duas primeiras linhas (0=Header1, 1=Header2)
        for row in range(2, sh.nrows):
            linha = sh.row_values(row)
            
            cod_raw = linha[0]
            nome_raw = linha[1]
            if not cod_raw or not nome_raw:
                continue
                
            try:
                cod_str = str(int(float(cod_raw)))
            except:
                continue
            
            soma_ativos = 0.0
            for col_idx in colunas_alvo:
                if col_idx < len(linha):
                    valor = linha[col_idx]
                    try:
                        soma_ativos += float(valor)
                    except:
                        pass
            
            if soma_ativos > 0:
                horas = (soma_ativos * 0.33) + 1.5
            else:
                horas = 0.0
                
            dict_folha[cod_str] = horas
            nome_limpo = str(nome_raw).strip().upper()
            dict_folha[nome_limpo] = horas
            
        print(f"Sucesso. Extraiu folha de {len(dict_folha)//2} empresas com codigo/nome.")
        return dict_folha
        
    except Exception as e:
        print("Erro lendo folha do excel:", e)
        return {}

def process():
    print("Iniciando processamento...")
    DATA_INICIO = "2025-12-01"
    DATA_FIM = "2025-12-31"

    FILE_FISCAL = r"C:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\ClienteTempo Gasto.xls"
    FILE_CONTABIL = r"C:\Users\DMF-AUTOMACAO\Downloads\HORAS CONTABEIS_V3_PGPeeWVc.xlsx"
    FILE_MASTER = r"C:\Users\DMF-AUTOMACAO\Downloads\CONTROLE DE HORAS DMF - CLIENTE.xlsm"

    # --- 1. FOLHA ---
    dict_folha = extrair_dados_folha()

    # --- 2. FISCAL ---
    dict_fiscal = {}
    print("2. Lendo", FILE_FISCAL)
    try:
        wb_f = xlrd.open_workbook(FILE_FISCAL)
        sh_f = wb_f.sheet_by_index(0)
        for i in range(1, sh_f.nrows):
            r = sh_f.row_values(i)
            cod_raw = r[0]   # Col A = Código
            nome_raw = r[4]  # Col E = Cliente
            tempo_raw = r[17] if len(r) > 17 else ''  # Col R = Tempo Gasto

            horas = time_to_hours(tempo_raw)
            if horas <= 0:
                continue

            # Indexa pelo código (string limpa sem .0)
            if cod_raw:
                try:
                    cod_str = str(int(float(str(cod_raw).strip())))
                    if cod_str not in dict_fiscal:
                        dict_fiscal[cod_str] = 0.0
                    dict_fiscal[cod_str] += horas
                except:
                    pass

            # Indexa também pelo nome para fallback
            if nome_raw and str(nome_raw).strip() not in ('', 'Cliente'):
                nome_limpo = clean_name(nome_raw)
                if nome_limpo not in dict_fiscal:
                    dict_fiscal[nome_limpo] = 0.0
                dict_fiscal[nome_limpo] += horas

    except Exception as e:
        print("Erro fiscal:", e)
    print(f"  Fiscal carregado: {len(dict_fiscal)} entradas. '347' in dict: {'347' in dict_fiscal}, val: {dict_fiscal.get('347')}")

    # --- 3. CONTABIL ---
    dict_contabil = {}
    print("3. Lendo", FILE_CONTABIL)
    try:
        wb_c = openpyxl.load_workbook(FILE_CONTABIL, data_only=True)
        sheet_names = [s for s in wb_c.sheetnames if len(s.split('.')) == 2 and s.split('.')[1].isdigit()]
        
        medias_sum = {}
        medias_count = {}
        
        media_col = 17 # TOTAL HORAS MÊS
        cod_col = 1    # CÓD DOMINIO
        nome_col = 2   # GRUPO
        
        for sname in sheet_names:
            sh_c = wb_c[sname]
            for row in range(2, sh_c.max_row + 1):
                val_codigo = sh_c.cell(row=row, column=cod_col).value
                val_nome = str(sh_c.cell(row=row, column=nome_col).value).strip().upper() if sh_c.cell(row=row, column=nome_col).value else ""
                val_media = sh_c.cell(row=row, column=media_col).value

                try: 
                    media_float = float(time_to_hours(val_media)) if isinstance(val_media, (str, datetime.time, datetime.timedelta)) else float(val_media) if val_media else 0.0
                except: 
                    media_float = 0.0
                
                if val_codigo:
                    try: 
                        cod_str = str(int(float(str(val_codigo).strip())))
                        if cod_str not in medias_sum:
                            medias_sum[cod_str] = 0.0
                            medias_count[cod_str] = 0

                        if val_media is not None:
                            medias_sum[cod_str] += media_float
                            medias_count[cod_str] += 1
                    except: pass
                
                # Ignorando media pelo str NOME para ser mais confiável o CODIGO, caso hajam duplos
                pass 
                
        for cod_str, total_val in medias_sum.items():
            if medias_count.get(cod_str, 0) > 0:
                dict_contabil[cod_str] = total_val / medias_count[cod_str]
                    
    except Exception as e:
        print("Erro contabil:", e)

    # --- 4. PREENCHER MASTER ---
    print("4. Atualizando planilha Master...")
    wb_m = openpyxl.load_workbook(FILE_MASTER, keep_vba=True)
    if "12.2025" not in wb_m.sheetnames:
        print("Aba 12.2025 nao encontrada!")
        return
        
    sh_m = wb_m["12.2025"]
    atualizados = 0
    # Buscar nome na coluna D (4) a partir da linha 10
    # Preencher:
    # N9 / col 14: TEMPO DOMÍNIO FISCAL (Fiscal bruto)
    # O9 / col 15: HORÁRIO FISCAL (Fiscal bruto + 65%)
    # P9 / col 16: HORÁRIO CONTÁBIL (Contábil media)
    # Q9 / col 17: HORÁRIO PESSOAL (Folha horas)

    for row in range(10, sh_m.max_row + 1):
        cod_cell = sh_m.cell(row=row, column=8).value
        nome_razao = sh_m.cell(row=row, column=11).value
        nome_fantasia = sh_m.cell(row=row, column=9).value
        
        n_limpo = clean_name(nome_razao) if nome_razao else ""
        nf_limpo = clean_name(nome_fantasia) if nome_fantasia else ""
        
        c_str = ""
        try:
            if cod_cell is not None:
                c_str = str(int(float(str(cod_cell).strip().split('.')[0])))
        except: pass
        
        found = False
        if n_limpo or nf_limpo or c_str:
            # Fiscal (match por código primeiro, depois por nome)
            v_fisc = dict_fiscal.get(c_str) or dict_fiscal.get(n_limpo) or dict_fiscal.get(nf_limpo)
            if v_fisc is not None:
                sh_m.cell(row=row, column=14, value=v_fisc / 24.0)
                sh_m.cell(row=row, column=15, value=(v_fisc * 1.65) / 24.0)
                found = True
                
            # Contabil (Matches purely by cod usually, or name)
            v_cont = dict_contabil.get(c_str) or dict_contabil.get(n_limpo) or dict_contabil.get(nf_limpo)
            if v_cont is not None:
                sh_m.cell(row=row, column=16, value=v_cont / 24.0)
                found = True
                
            # Folha (Matches by cod or name)
            v_folha = dict_folha.get(c_str) or dict_folha.get(n_limpo) or dict_folha.get(nf_limpo)
            if v_folha is not None:
                sh_m.cell(row=row, column=17, value=v_folha / 24.0)
                found = True
                
        if found:
            atualizados += 1
        elif row < 25 and (n_limpo or nf_limpo or c_str):
            print(f"NO MATCH ROW {row}: c_str='{c_str}', n_limpo='{n_limpo}', nf_limpo='{nf_limpo}'")
            print(f"  -> In Folha? c_str:{c_str in dict_folha} n_limpo:{n_limpo in dict_folha} nf_limpo:{nf_limpo in dict_folha}")
            print(f"  -> In Fiscal? n_limpo:{n_limpo in dict_fiscal} nf_limpo:{nf_limpo in dict_fiscal}")
            print(f"  -> In Contabil? c_str:{c_str in dict_contabil} n_limpo:{n_limpo in dict_contabil} nf_limpo:{nf_limpo in dict_contabil}")

    # Salvando
    out_file = r"C:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm"
    wb_m.save(out_file)
    print(f"Sucesso! {atualizados} linhas atualizadas. Salvo em: {out_file}")

if __name__ == "__main__":
    process()
