import openpyxl

def execute():
    path = 'c:/Users/DMF-AUTOMACAO/Documents/PROJETOS/N8N automacao/ESTRUTURA/CONTROLE_DE_HORAS_DMF.xlsm'
    print(f"Loading {path}...")
    wb = openpyxl.load_workbook(path, keep_vba=True)
    sh = wb['02.2026']
    
    # 1. Update Row 88 (1283 - Base Estruturas)
    # Col H=8, O=15, P=16, Q=17
    sh.cell(row=88, column=8, value=1283)
    sh.cell(row=88, column=15, value=0.583113425925926) # 13:59:41
    sh.cell(row=88, column=16, value=0.25)             # 06:00:00
    sh.cell(row=88, column=17, value=1.125)            # 27:00:00
    print("Updated Row 88.")

    # 2. Update Row 761 (1195 - Comercial Vale)
    sh.cell(row=761, column=8, value=1195)
    sh.cell(row=761, column=15, value=0.125)           # 03:00:00
    sh.cell(row=761, column=16, value=0.166666666666667) # 04:00:00
    sh.cell(row=761, column=17, value=0.00347222222222222) # 00:05:00
    print("Updated Row 761.")

    # 3. Delete duplicates
    # We must delete from bottom to top to avoid index shifts
    print("Deleting duplicates (93 and 781)...")
    sh.delete_rows(781)
    sh.delete_rows(93)
    
    wb.save(path)
    print(f"Saved changes to {path}.")

if __name__ == '__main__':
    execute()
