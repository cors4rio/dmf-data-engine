import os
import shutil
import openpyxl
import re

def validate_duplicates():
    txt_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\VALIDACAO CNPJ.txt'
    xls_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls'
    temp_xlsx = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\temp_val_dup.xlsx'
    log_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\faltantes_duplicados_log.txt'
    
    print(f"Lendo dados de validação: {txt_path}")
    
    txt_entries = []
    with open(txt_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            
            parts = re.split(r'[\t;]', line)
            if len(parts) >= 3:
                cnpj = parts[0].strip().replace('.', '').replace('-', '').replace('/', '').split('.')[0]
                if cnpj and len(cnpj) < 14:
                    cnpj = cnpj.zfill(14)
                name = parts[1].strip()
                hours = parts[-1].strip()
                txt_entries.append({'cnpj': cnpj, 'name': name, 'hours': hours})

    print(f"Clientes no TXT: {len(txt_entries)}")

    print(f"Lendo planilha: {xls_path}")
    shutil.copy(xls_path, temp_xlsx)
    wb = openpyxl.load_workbook(temp_xlsx, data_only=True)
    ws = wb['12.2025']
    
    # Contar ocorrências de cada CNPJ na planilha
    xls_cnpj_counts = {}
    for row_idx in range(10, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=10).value
        if cell_val:
            clean_cnpj = str(cell_val).strip().replace('.', '').replace('-', '').replace('/', '').split('.')[0]
            if clean_cnpj and len(clean_cnpj) < 14:
                clean_cnpj = clean_cnpj.zfill(14)
            
            xls_cnpj_counts[clean_cnpj] = xls_cnpj_counts.get(clean_cnpj, 0) + 1

    print(f"CNPJs únicos na planilha: {len(xls_cnpj_counts)}")

    missing_clients = []
    
    # Lógica de Matching:
    # Para cada entrada do TXT, verificamos se há um "slot" disponível na contagem da planilha para aquele CNPJ.
    for client in txt_entries:
        cnpj = client['cnpj']
        if cnpj in xls_cnpj_counts and xls_cnpj_counts[cnpj] > 0:
            # Encontrou uma unidade na planilha para esse CNPJ, "consome" o slot
            xls_cnpj_counts[cnpj] -= 1
        else:
            # Não há mais unidades na planilha para esse CNPJ (ou não existe)
            missing_clients.append(client)

    # Calcular Total de Horas Faltantes
    total_seconds = 0
    for c in missing_clients:
        h_str = c['hours']
        try:
            parts = h_str.split(':')
            if len(parts) >= 2:
                h = int(parts[0])
                m = int(parts[1])
                s = int(parts[2]) if len(parts) > 2 else 0
                total_seconds += h * 3600 + m * 60 + s
        except:
            pass
            
    total_h = total_seconds // 3600
    total_m = (total_seconds % 3600) // 60
    total_s = total_seconds % 60

    print(f"Faltantes identificados: {len(missing_clients)}")

    with open(log_path, 'w', encoding='utf-8') as log:
        log.write(f"--- RELATÓRIO DE FALTANTES (CONSIDERANDO DUPLICIDADE DE CNPJ) ---\n")
        log.write(f"Total de registros no Contábil (TXT): {len(txt_entries)}\n")
        log.write(f"Total de Unidades Faltantes: {len(missing_clients)}\n")
        log.write(f"Total de Horas Faltantes: {total_h:02d}:{total_m:02d}:{total_s:02d}\n\n")
        log.write("LISTA DE UNIDADES FALTANTES:\n")
        log.write("-" * 80 + "\n")
        for c in missing_clients:
            log.write(f"CNPJ: {c['cnpj']:14} | HORA: {c['hours']:8} | NOME: {c['name']}\n")

    print(f"Log detalhado gerado em: {log_path}")
    wb.close()
    os.remove(temp_xlsx)

if __name__ == "__main__":
    validate_duplicates()
