import pyodbc
import openpyxl
import xlrd
import re
import difflib
import sys
import datetime

def clean_cnpj(val):
    if not val:
        return ""
    return re.sub(r'\D', '', str(val).strip())

def similar(a, b):
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, str(a).upper(), str(b).upper()).ratio()

def main():
    conn_str = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
    data_inicio = '2026-03-01'
    data_fim = '2026-03-31'
    
    # Dicionarios
    db_cadastros = {}
    db_lancamentos = {}
    db_faturamento = {}
    carol_folha = {}
    
    print("1. Extraindo cadastros do banco de dados (geempre)...")
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        
        # 1.1 Cadastros
        cursor.execute("SELECT codi_emp, cgce_emp, nome_emp FROM bethadba.geempre")
        for row in cursor.fetchall():
            db_cadastros[row.codi_emp] = {
                'cnpj': clean_cnpj(row.cgce_emp),
                'nome': row.nome_emp
            }
            
        # 1.2 Lancamentos (1 e 39)
        print("2. Extraindo quantidade de lançamentos...")
        q_lanc = f"""
            SELECT 
                codi_emp,
                COUNT(*) as qtd_lancamentos
            FROM bethadba.ctlancto
            WHERE data_lan >= '{data_inicio}' AND data_lan <= '{data_fim}'
            AND orig_lan IN (1, 39)
            GROUP BY codi_emp
        """
        cursor.execute(q_lanc)
        for row in cursor.fetchall():
            db_lancamentos[row.codi_emp] = int(row.qtd_lancamentos)
            
        # 1.3 Faturamento
        print("3. Extraindo faturamento...")
        q_fat = f"""
            SELECT 
                codi_emp, 
                SUM(total_contabil) as faturamento
            FROM (
                SELECT codi_emp, SUM(vcon_sai) as total_contabil 
                FROM bethadba.efsaidas 
                WHERE dsai_sai >= '{data_inicio}' AND dsai_sai <= '{data_fim}' 
                GROUP BY codi_emp
                UNION ALL
                SELECT codi_emp, SUM(vcon_ser) as total_contabil 
                FROM bethadba.efservicos 
                WHERE dser_ser >= '{data_inicio}' AND dser_ser <= '{data_fim}' 
                GROUP BY codi_emp
            ) base
            GROUP BY codi_emp
        """
        cursor.execute(q_fat)
        for row in cursor.fetchall():
            db_faturamento[row.codi_emp] = float(row.faturamento)
            
        conn.close()
    except Exception as e:
        print(f"Erro ODBC: {e}")
        return
        
    print("4. Lendo Planilha da Carol (Controle de Empregados)...")
    try:
        wb_carol = xlrd.open_workbook('Controle de Empregados (CAROL)032026.xls')
        ws_carol = wb_carol.sheet_by_index(0)
        
        for rowx in range(1, ws_carol.nrows): # skip header maybe, check row 0
            row = ws_carol.row_values(rowx)
            cod_val = row[1]
            if not cod_val:
                continue
            try:
                cod = int(float(cod_val))
            except ValueError:
                continue
                
            # Col 7 = func, 9 = estag, 11 = contrib
            def get_int(val):
                try:
                    return int(float(val)) if val else 0
                except:
                    return 0
                    
            func = get_int(row[7])
            estag = get_int(row[9])
            contrib = get_int(row[11])
            total_ativos = func + estag + contrib
            
            carol_folha[cod] = "SIM" if total_ativos > 0 else "NAO"
            
    except Exception as e:
        print(f"Erro ao ler planilha da Carol: {e}")
        return

    print("5. Atualizando HORAS CONTABEIS_.xlsx (Aba 04.2026)...")
    wb_path = 'HORAS CONTABEIS_.xlsx'
    sheet_name = '04.2026'
    
    try:
        wb = openpyxl.load_workbook(wb_path)
        if sheet_name not in wb.sheetnames:
            print(f"Aba {sheet_name} não encontrada!")
            return
            
        ws = wb[sheet_name]
        
        # Estatísticas
        sucessos = 0
        matches_parciais = []
        rejeitadas = []
        
        for row in range(2, ws.max_row + 1):
            cod_val = ws.cell(row=row, column=1).value
            cnpj_val = ws.cell(row=row, column=3).value
            nome_val = ws.cell(row=row, column=4).value
            
            if cod_val is None:
                continue
                
            try:
                cod = int(float(str(cod_val).strip()))
            except ValueError:
                continue
                
            cnpj_planilha = clean_cnpj(cnpj_val)
            nome_planilha = str(nome_val or "").strip()
            
            # Validação Tripla
            db_info = db_cadastros.get(cod)
            if not db_info:
                rejeitadas.append(f"Linha {row}: Cód {cod} não encontrado no banco Domínio.")
                continue
                
            cnpj_db = db_info['cnpj']
            nome_db = db_info['nome']
            
            match_cod = True # já usamos para buscar
            match_cnpj = (cnpj_planilha == cnpj_db) if cnpj_planilha else False
            match_nome = similar(nome_planilha, nome_db) > 0.8
            
            matches = 1 + int(match_cnpj) + int(match_nome)
            
            if matches >= 2:
                if matches == 2:
                    falha = "CNPJ" if not match_cnpj else "NOME"
                    matches_parciais.append(f"Linha {row}: Cód {cod} | Falha em {falha}. Planilha: {cnpj_planilha} / {nome_planilha} | DB: {cnpj_db} / {nome_db}")
                
                # PREENCHER DADOS
                # F (6): Qtd Lançamentos
                qtd = db_lancamentos.get(cod, 0)
                ws.cell(row=row, column=6).value = qtd
                
                # I (9): Tem Folha
                tem_folha = carol_folha.get(cod, "NAO")
                ws.cell(row=row, column=9).value = tem_folha
                
                # O (15): Faturamento
                fat = db_faturamento.get(cod, 0.0)
                ws.cell(row=row, column=15).value = fat
                
                sucessos += 1
                
            else:
                rejeitadas.append(f"Linha {row}: Cód {cod} falhou validação. Planilha: {cnpj_planilha} / {nome_planilha} | DB: {cnpj_db} / {nome_db}")
                
        wb.save(wb_path)
        print(f"Salvo {wb_path}. {sucessos} empresas atualizadas.")
        
        # Gerar Relatorio .md
        md_content = f"# Auditoria de Preenchimento Contábil - Mês 04.2026\n"
        md_content += f"> **Executado em:** {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        md_content += f"## 1. Resumo\n"
        md_content += f"- **Empresas Atualizadas (Match 3/3 ou 2/3):** {sucessos}\n"
        md_content += f"- **Matches Parciais (2/3):** {len(matches_parciais)}\n"
        md_content += f"- **Rejeitadas (0-1/3 ou Cód ñ encontrado):** {len(rejeitadas)}\n\n"
        
        if matches_parciais:
            md_content += f"## 2. Empresas com Match Parcial (Alerta)\n"
            for m in matches_parciais:
                md_content += f"- {m}\n"
            md_content += "\n"
            
        if rejeitadas:
            md_content += f"## 3. Empresas Rejeitadas (Necessitam Revisão)\n"
            for r in rejeitadas:
                md_content += f"- {r}\n"
                
        with open('auditoria_contabil_042026.md', 'w', encoding='utf-8') as f:
            f.write(md_content)
            
        print("Relatório auditoria_contabil_042026.md gerado com sucesso.")

    except Exception as e:
        print(f"Erro ao salvar planilha: {e}")

if __name__ == '__main__':
    main()
