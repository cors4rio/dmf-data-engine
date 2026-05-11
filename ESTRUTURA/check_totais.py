import openpyxl

file_path = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm"
wb = openpyxl.load_workbook(file_path, data_only=False)
sh = wb['02.2026']

print("--- Fórmulas/Valores nas células de totais (Linha 7) ---")
print(f"O7 (Fiscal Total):  {sh['O7'].value}")
print(f"Q7 (DP Total):      {sh['Q7'].value}")
print(f"R7 (Grand Total):   {sh['R7'].value}")

# Conta quantas linhas têm dados em O e Q
count_o = sum(1 for r in range(10, sh.max_row + 1) if sh.cell(r, 15).value is not None)
count_q = sum(1 for r in range(10, sh.max_row + 1) if sh.cell(r, 17).value is not None)
print(f"\nLinhas com valor em O (Fiscal): {count_o}")
print(f"Linhas com valor em Q (DP):     {count_q}")
print(f"Max row: {sh.max_row}")
