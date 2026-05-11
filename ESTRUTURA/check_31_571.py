import openpyxl
from datetime import timedelta

wb_orig = openpyxl.load_workbook(
    r'C:\Users\DMF-AUTOMACAO\Downloads\CONTROLE_DE_HORAS_DMF_BL4t520G.xlsm', data_only=True
)
ws_orig = wb_orig['02.2026']

wb_dest = openpyxl.load_workbook(
    r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF - CLIENTE RECUPERADO.xlsx'
)
ws_dest = wb_dest['03.2026']

# Verificar cod 31 e 571 em ambas as planilhas
for target_cod in [31, 571]:
    print(f"\n=== CODIGO {target_cod} ===")
    print("  ORIGEM (02.2026, Col O):")
    for r in range(10, ws_orig.max_row + 1):
        cod = ws_orig.cell(r, 8).value
        if cod is not None:
            try:
                c = int(float(str(cod).strip()))
            except:
                continue
            if c == target_cod:
                val = ws_orig.cell(r, 15).value
                nome = ws_orig.cell(r, 9).value
                if isinstance(val, timedelta):
                    h = val.total_seconds() / 3600
                    print(f"    Row {r}: nome={nome}, valor={val} ({h:.2f}h)")
                elif isinstance(val, (int, float)):
                    print(f"    Row {r}: nome={nome}, valor={val} ({val*24:.2f}h)")
                else:
                    print(f"    Row {r}: nome={nome}, valor={val!r}")

    print("  DESTINO (03.2026, Col N):")
    for r in range(10, ws_dest.max_row + 1):
        cod = ws_dest.cell(r, 8).value
        if cod is not None:
            try:
                c = int(float(str(cod).strip()))
            except:
                continue
            if c == target_cod:
                val = ws_dest.cell(r, 14).value
                nome = ws_dest.cell(r, 9).value
                if isinstance(val, timedelta):
                    h = val.total_seconds() / 3600
                    print(f"    Row {r}: nome={nome}, valor={val} ({h:.2f}h)")
                elif isinstance(val, (int, float)):
                    print(f"    Row {r}: nome={nome}, valor={val} ({val*24:.2f}h)")
                else:
                    print(f"    Row {r}: nome={nome}, valor={val!r}")
