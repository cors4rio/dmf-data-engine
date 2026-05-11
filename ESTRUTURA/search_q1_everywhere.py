import os
import openpyxl

folder = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA'
files = os.listdir(folder)

print("Files in directory:")
for f in files:
    if "V3" in f or "CONTABIL" in f.upper() or "HORAS" in f.upper():
        print(f" - {f}")

target_files = [f for f in files if "HORAS CONTABEIS_.xlsx" in f or "V3" in f or "CONTROLE DE HORAS DMF - CLIENTE" in f]

for f in target_files:
    full_path = os.path.join(folder, f)
    print(f"\n--- Checking {f} ---")
    try:
        wb = openpyxl.load_workbook(full_path, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            q1 = ws['Q1'].value if 'Q1' in ws else "N/A"
            print(f" Sheet: {sname}, Q1: {q1}")
            if isinstance(q1, (int, float)) or (isinstance(q1, str) and "=" in q1):
                print(f"   [!] Potential target found in {f} sheet {sname}")
    except Exception as e:
        print(f" Error opening {f}: {e}")
