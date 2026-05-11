"""
check_setup.py — Verificação de Setup e Dry-Run de Conexão
DMF Automação

Valida o ambiente antes de qualquer execução:
  ✅ Dependências Python instaladas
  ✅ Arquivo .env com senha ODBC
  ✅ Conexão ODBC ao Domínio
  ✅ Planilha Master encontrada e aba do mês existe
  ✅ Planilha HORAS CONTABEIS encontrada
  ✅ Planilha Carol encontrada (se disponível)
  ✅ Arquivos de exceção (DP NAO.txt, NAO FAZ CONTABIL.txt)

Uso:
    python check_setup.py --mes 03 --ano 2026
"""

import sys
import logging
from pathlib import Path
from datetime import date

sys.path.insert(0, str(Path(__file__).parent))

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger("dmf.check")

OK  = "  ✅"
NOK = "  ❌"
WRN = "  ⚠️ "

erros = 0
avisos = 0


def ok(msg: str)  -> None: logger.info("%s %s", OK,  msg)
def err(msg: str) -> None:
    global erros; erros += 1
    logger.error("%s %s", NOK, msg)
def warn(msg: str) -> None:
    global avisos; avisos += 1
    logger.warning("%s %s", WRN, msg)


def check_deps() -> None:
    logger.info("\n── Dependências ──")
    deps = ["openpyxl", "pyodbc", "dotenv"]
    for dep in deps:
        mod = dep.replace("dotenv", "dotenv")
        try:
            __import__(mod if dep != "dotenv" else "dotenv")
            ok(dep)
        except ImportError:
            if dep == "pyodbc":
                try:
                    import pypyodbc  # type: ignore
                    ok("pypyodbc (substituto do pyodbc)")
                except ImportError:
                    err(f"{dep} não instalado — pip install pyodbc")
            else:
                err(f"{dep} não instalado — pip install {dep}")

    try:
        import xlrd
        ok("xlrd (fallback Carol)")
    except ImportError:
        warn("xlrd não instalado. Fallback Carol indisponível — pip install xlrd==1.2.0")


def check_env() -> None:
    logger.info("\n── Arquivo .env ──")
    env_path = Path(__file__).parent / ".env"
    if not env_path.exists():
        err(".env não encontrado. Crie com ODBC_PASS=suasenha")
        return
    ok(".env encontrado")
    content = env_path.read_text()
    if "ODBC_PASS" in content:
        ok("ODBC_PASS definido")
    else:
        err("ODBC_PASS não encontrado no .env")


def check_odbc() -> None:
    logger.info("\n── Conexão ODBC ──")
    try:
        from _ENGINE.database import testar_conexao
        if testar_conexao():
            ok("Conexão ODBC bem-sucedida")
        else:
            err("Conexão ODBC falhou")
    except Exception as e:
        err(f"Erro ao testar conexão: {e}")


def check_planilhas(mes: int, ano: int) -> None:
    logger.info("\n── Planilhas ──")
    from _ENGINE.config import cfg

    # Master
    if cfg.master:
        ok(f"Master encontrada: {cfg.master.name}")
        try:
            import openpyxl
            wb = openpyxl.load_workbook(cfg.master, read_only=True, data_only=True,
                                        keep_vba=cfg.master.suffix == ".xlsm")
            aba = cfg.aba_excel(mes, ano)
            if aba in wb.sheetnames:
                ok(f"Aba '{aba}' existe na Master")
            else:
                err(f"Aba '{aba}' NÃO encontrada na Master. Abas disponíveis: {wb.sheetnames[:5]}")
            wb.close()
        except Exception as e:
            err(f"Erro ao abrir Master: {e}")
    else:
        err("Planilha Master NÃO encontrada (CONTROLE_DE_HORAS_DMF.xlsm)")

    # Contábil
    if cfg.contabil:
        ok(f"HORAS CONTABEIS encontrada: {cfg.contabil.name}")
        try:
            import openpyxl
            wb = openpyxl.load_workbook(cfg.contabil, read_only=True, data_only=True)
            aba = cfg.aba_excel(mes, ano)
            if aba in wb.sheetnames:
                ok(f"Aba '{aba}' existe em HORAS CONTABEIS")
            else:
                warn(f"Aba '{aba}' não encontrada em HORAS CONTABEIS")
            wb.close()
        except Exception as e:
            err(f"Erro ao abrir HORAS CONTABEIS: {e}")
    else:
        warn("Planilha HORAS CONTABEIS não encontrada")

    # Carol
    carol = cfg.carol_path(mes, ano)
    if carol:
        ok(f"Planilha Carol: {carol.name}")
    else:
        warn(f"Planilha Carol para {mes:02d}/{ano} não encontrada (fallback indisponível)")


def check_excecoes() -> None:
    logger.info("\n── Arquivos de Exceção ──")
    from _ENGINE.config import cfg

    if cfg.dp_nao and Path(cfg.dp_nao).exists():
        linhas = len(Path(cfg.dp_nao).read_text(encoding="utf-8", errors="ignore").splitlines())
        ok(f"DP NAO.txt: {linhas} linhas")
    else:
        warn(f"DP NAO.txt não encontrado: {cfg.dp_nao}")

    if cfg.nao_contabil and Path(cfg.nao_contabil).exists():
        linhas = len(Path(cfg.nao_contabil).read_text(encoding="utf-8", errors="ignore").splitlines())
        ok(f"NAO FAZ CONTABIL.txt: {linhas} linhas")
    else:
        warn(f"NAO FAZ CONTABIL.txt não encontrado: {cfg.nao_contabil}")


def main() -> None:
    hoje = date.today()
    import argparse
    parser = argparse.ArgumentParser(description="DMF — Verificação de Setup")
    parser.add_argument("-m", "--mes", type=int, default=hoje.month)
    parser.add_argument("-y", "--ano", type=int, default=hoje.year)
    parser.add_argument("--skip-odbc", action="store_true",
                        help="Pula teste de conexão ODBC.")
    args = parser.parse_args()

    logger.info("=" * 50)
    logger.info("DMF Automação — Check Setup (%02d/%d)", args.mes, args.ano)
    logger.info("=" * 50)

    check_deps()
    check_env()
    if not args.skip_odbc:
        check_odbc()
    check_planilhas(args.mes, args.ano)
    check_excecoes()

    logger.info("\n" + "=" * 50)
    if erros == 0 and avisos == 0:
        logger.info("✅ Tudo OK! Pronto para executar automacao.py.")
    elif erros == 0:
        logger.info("⚠️  Setup OK com %d aviso(s). Verifique antes de executar.", avisos)
    else:
        logger.info("❌ %d erro(s) encontrado(s). Corrija antes de executar.", erros)
    logger.info("=" * 50)

    sys.exit(0 if erros == 0 else 1)


if __name__ == "__main__":
    main()
