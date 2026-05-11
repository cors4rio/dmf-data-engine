"""
excel_utils.py — Manipulação Segura de Planilhas (openpyxl)
DMF Automação | _ENGINE

Centraliza TODAS as regras de ouro de escrita no Excel para evitar
corrupção, ghosts, timedelta vs float, SUBTOTAL truncado, etc.

Regras de ouro implementadas aqui:
  [A] keep_vba=True somente para .xlsm
  [B] data_only=True na leitura para pegar valores calculados
  [C] Formato [h]:mm:ss obrigatório em células de tempo
  [D] Conversão timedelta → float antes de gravar
  [E] Código → int(float(str(x))) para evitar '1152.0'
  [F] Zeros explícitos em clientes sem dados (matar ghosts)
  [G] SUBTOTAL dinâmico usando ws.max_row
  [H] Lookup duplo: Código Domínio (col H) → CNPJ (col J)
  [I] dict com lista de linhas (não sobrescrever clientes duplicados)
"""

import re
import logging
from collections import defaultdict
from datetime import timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

logger = logging.getLogger("dmf.excel")

# ─────────────────────────────────────────────
# CONSTANTES DE COLUNAS (Master) — 1-indexed
# ─────────────────────────────────────────────
COL_FLAG_EXCECAO  = 4   # D  — ex: "DP NÃO"
COL_COD_DOMINIO   = 8   # H  — codi_emp
COL_NOME_FANTASIA = 9   # I  — nome fantasia
COL_CNPJ          = 10  # J  — CNPJ formatado
COL_RAZAO_SOCIAL  = 11  # K  — razão social
COL_MES_ANT_FISC  = 14  # N  — mês anterior fiscal (backfill)
COL_FISCAL        = 15  # O  — horário fiscal
COL_CONTABIL      = 16  # P  — horário contábil
COL_DP            = 17  # Q  — horário pessoal (DP)
COL_TOTAL         = 18  # R  — total O+P+Q

LINHA_INICIO_DADOS  = 10  # Dados começam na linha 10
LINHA_SUBTOTAL      = 7   # Subtotais ficam na linha 7

VALORES_ESPECIAIS_H = {
    "Não entra - sistema próprio",
    "Não esta na Dominio",
    "Não entra - Consultoria",
}

# ─────────────────────────────────────────────
# ABERTURA / FECHAMENTO SEGURO
# ─────────────────────────────────────────────

def abrir_planilha(caminho: str | Path, somente_leitura: bool = False) -> Workbook:
    """
    Abre uma planilha Excel com os parâmetros corretos.
    [A] keep_vba condicionado à extensão (.xlsm vs .xlsx)
    [B] data_only=True para ler valores calculados (não fórmulas)
    """
    caminho = Path(caminho)
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")

    is_xlsm = caminho.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(
        caminho,
        keep_vba=is_xlsm,        # [A]
        data_only=True,           # [B]
        read_only=somente_leitura,
    )
    logger.debug("Aberta planilha '%s' (xlsm=%s).", caminho.name, is_xlsm)
    return wb


def salvar_planilha(wb: Workbook, caminho: str | Path) -> None:
    """Salva a planilha preservando a extensão correta."""
    caminho = Path(caminho)
    wb.save(caminho)
    logger.info("💾 Planilha salva: %s", caminho)


# ─────────────────────────────────────────────
# HELPERS DE TIPO
# ─────────────────────────────────────────────

def normalizar_cnpj(raw: Any) -> str:
    """Remove pontuação e retorna CNPJ como string de dígitos."""
    return re.sub(r"\D", "", str(raw or ""))


def normalizar_codigo(raw: Any) -> str | None:
    """
    Converte qualquer representação de código para string inteira.
    [E] Trata '1152.0', 1152, 1152.0 → '1152'
    Retorna None se não for um código numérico válido.
    """
    if raw is None:
        return None
    raw_str = str(raw).strip()
    if raw_str in VALORES_ESPECIAIS_H:
        return None
    try:
        return str(int(float(raw_str)))
    except (ValueError, TypeError):
        return None


def tempo_para_frac_dia(segundos: float | int) -> float:
    """Converte segundos para fração de dia (formato interno do Excel)."""
    return segundos / 86400.0


def frac_dia_para_segundos(frac: float) -> float:
    """Converte fração de dia para segundos."""
    return frac * 86400.0


def garantir_float_tempo(valor: Any) -> float | None:
    """
    [D] Converte timedelta, float, int ou None para float (fração de dia).
    Retorna None se o valor for texto (ex: 'DP NÃO').
    """
    if valor is None:
        return None
    if isinstance(valor, timedelta):
        return valor.total_seconds() / 86400.0
    if isinstance(valor, (int, float)):
        return float(valor)
    return None  # string como 'DP NÃO' — não converter


# ─────────────────────────────────────────────
# ESCRITA SEGURA DE CÉLULAS
# ─────────────────────────────────────────────

def escrever_tempo(ws: Worksheet, linha: int, col: int, valor: Any) -> None:
    """
    [C] [D] Escreve valor de tempo com formato [h]:mm:ss.
    Aceita float (fração dia), timedelta ou segundos (int/float > 1).
    """
    cell = ws.cell(row=linha, column=col)
    if isinstance(valor, str):
        # Exceções textuais como 'DP NÃO', 'NAO FAZ CONTABIL'
        cell.value = valor
        cell.number_format = "@"
        return
    if isinstance(valor, timedelta):
        valor = valor.total_seconds() / 86400.0
    if isinstance(valor, (int, float)):
        # Se vier como segundos brutos (> 10), converter para frac dia
        if valor > 10:
            valor = valor / 86400.0
    cell.value = valor
    cell.number_format = "[h]:mm:ss"


def escrever_zero(ws: Worksheet, linha: int, col: int) -> None:
    """
    [F] Escreve zero explícito para matar valores residuais (ghosts).
    """
    cell = ws.cell(row=linha, column=col)
    cell.value = 0
    cell.number_format = "[h]:mm:ss"


def escrever_formula_total(ws: Worksheet, linha: int) -> None:
    """Escreve =O{linha}+P{linha}+Q{linha} na coluna R."""
    o = get_column_letter(COL_FISCAL)
    p = get_column_letter(COL_CONTABIL)
    q = get_column_letter(COL_DP)
    r_cell = ws.cell(row=linha, column=COL_TOTAL)
    r_cell.value = f"={o}{linha}+{p}{linha}+{q}{linha}"
    r_cell.number_format = "[h]:mm:ss"


def atualizar_subtotal(ws: Worksheet, col: int) -> None:
    """
    [G] Recalcula e injeta SUBTOTAL dinâmico na linha 7 para a coluna dada.
    Usa ws.max_row para cobrir todas as linhas atuais.
    """
    col_letter = get_column_letter(col)
    formula = (
        f"=SUBTOTAL(9,{col_letter}{LINHA_INICIO_DADOS}"
        f":{col_letter}{ws.max_row})"
    )
    cell = ws.cell(row=LINHA_SUBTOTAL, column=col)
    cell.value = formula
    cell.number_format = "[h]:mm:ss"
    logger.debug("Subtotal coluna %s: %s", col_letter, formula)


# ─────────────────────────────────────────────
# LOOKUP E MAPEAMENTO
# ─────────────────────────────────────────────

def mapear_linhas_master(ws: Worksheet) -> tuple[
    dict[str, list[int]],   # por código domínio
    dict[str, list[int]],   # por CNPJ
]:
    """
    [H] [I] Varre a planilha e retorna dois índices:
      - por_codigo: {codi_emp_str → [lista de linhas]}
      - por_cnpj:   {cnpj_digits → [lista de linhas]}

    O uso de listas permite detectar e tratar clientes duplicados.
    """
    por_codigo: dict[str, list[int]] = defaultdict(list)
    por_cnpj:   dict[str, list[int]] = defaultdict(list)

    for row in ws.iter_rows(min_row=LINHA_INICIO_DADOS):
        n_linha = row[0].row

        cod_raw  = row[COL_COD_DOMINIO - 1].value
        cnpj_raw = row[COL_CNPJ - 1].value

        cod = normalizar_codigo(cod_raw)
        if cod:
            por_codigo[cod].append(n_linha)

        cnpj = normalizar_cnpj(cnpj_raw)
        if len(cnpj) in (11, 14):  # CPF ou CNPJ válido
            por_cnpj[cnpj].append(n_linha)

    return dict(por_codigo), dict(por_cnpj)


def encontrar_linhas(
    codi_emp: int | str | None,
    cnpj: str | None,
    por_codigo: dict,
    por_cnpj: dict,
) -> list[int]:
    """
    [H] Lookup duplo: tenta por código primeiro, depois por CNPJ.
    Retorna lista de linhas (pode ser mais de uma para duplicados).
    """
    if codi_emp is not None:
        cod = normalizar_codigo(codi_emp)
        if cod and cod in por_codigo:
            return por_codigo[cod]

    if cnpj:
        cnpj_limpo = normalizar_cnpj(cnpj)
        if cnpj_limpo in por_cnpj:
            return por_cnpj[cnpj_limpo]

    return []


def detectar_cnpj_duplicados(por_cnpj: dict[str, list[int]]) -> dict[str, list[int]]:
    """Retorna apenas os CNPJs que aparecem em mais de uma linha."""
    return {cnpj: linhas for cnpj, linhas in por_cnpj.items() if len(linhas) > 1}


# ─────────────────────────────────────────────
# LEITURA DE DADOS DA PLANILHA
# ─────────────────────────────────────────────

def ler_coluna(ws: Worksheet, col: int, min_row: int = LINHA_INICIO_DADOS) -> dict[int, Any]:
    """Retorna {linha: valor} para a coluna especificada."""
    return {
        row[0].row: row[col - 1].value
        for row in ws.iter_rows(min_row=min_row)
        if row[0].row is not None
    }
