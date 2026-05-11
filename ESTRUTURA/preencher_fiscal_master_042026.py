import openpyxl
import os
import re

# Configurações
MES_ALVO = '04.2026'
BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = os.path.join(BASE_DIR, "CONTROLE DE HORAS DMF.xlsx")
ARQUIVO_FISCAL = os.path.join(BASE_DIR, "horas fiscal 042026.txt")
RELATORIO_MD = os.path.join(BASE_DIR, "relatorio_fiscal_042026.md")

FATOR_ADICIONAL_FISCAL = 1.80

def parse_hms(tempo_str):
    """Converte HH:MM:SS para total de segundos."""
    tempo_str = tempo_str.strip()
    partes = tempo_str.split(':')
    if len(partes) == 3:
        h, m, s = int(partes[0]), int(partes[1]), int(partes[2])
        return h * 3600 + m * 60 + s
    elif len(partes) == 2:
        h, m = int(partes[0]), int(partes[1])
        return h * 3600 + m * 60
    return 0

def segundos_para_hms(total_segundos):
    h = total_segundos // 3600
    m = (total_segundos % 3600) // 60
    s = total_segundos % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

def main():
    # 1. Ler o arquivo TXT com os tempos fiscais
    print("Lendo arquivo de horas fiscais...")
    fiscal_por_cod = {}  # {cod: total_segundos_brutos}
    
    with open(ARQUIVO_FISCAL, 'r', encoding='utf-8') as f:
        for linha in f:
            linha = linha.strip()
            if not linha:
                continue
            
            # Separar por tabs
            partes = linha.split('\t')
            # Remover partes vazias
            partes = [p.strip() for p in partes if p.strip()]
            
            if len(partes) < 2:
                continue
            
            # Primeira parte = código, última = tempo
            cod_raw = partes[0]
            tempo_raw = partes[-1]
            
            # Pular header
            if cod_raw == 'Código' or cod_raw == '' or tempo_raw == 'Tempo Gasto':
                continue
            
            try:
                cod = int(float(cod_raw))
            except (ValueError, TypeError):
                continue
            
            segundos = parse_hms(tempo_raw)
            
            # Agrupar (mesmo código pode ter múltiplas linhas)
            if cod not in fiscal_por_cod:
                fiscal_por_cod[cod] = 0
            fiscal_por_cod[cod] += segundos
    
    print(f"  -> {len(fiscal_por_cod)} empresas únicas lidas do TXT.")
    
    # 2. Aplicar adicional de 80%
    fiscal_final = {}
    for cod, seg_bruto in fiscal_por_cod.items():
        seg_final = int(seg_bruto * FATOR_ADICIONAL_FISCAL)
        fiscal_final[cod] = {
            'bruto': seg_bruto,
            'final': seg_final,
            'excel': seg_final / 86400.0  # Fração de dia para o Excel
        }
    
    # 3. Abrir planilha Master
    print("Abrindo planilha Master...")
    wb = openpyxl.load_workbook(PLANILHA_MASTER)
    
    if MES_ALVO not in wb.sheetnames:
        print(f"Aba {MES_ALVO} não encontrada!")
        return
    ws = wb[MES_ALVO]
    
    encontrados = []
    nao_encontrados_master = []
    zerados = []
    alterados = 0
    total_linhas = 0
    
    print("Preenchendo coluna O (Horário Fiscal)...")
    
    # REGRA DE OURO (Troubleshooting B): iterar TODAS as linhas da Master,
    # não apenas as que têm dados no TXT, para matar fantasmas
    for row in range(10, ws.max_row + 1):
        cod_val = ws.cell(row=row, column=8).value  # Col H (Código Domínio)
        
        if cod_val is None:
            continue
        
        try:
            cod = int(float(str(cod_val).strip()))
        except (ValueError, TypeError):
            continue
        
        total_linhas += 1
        
        if cod in fiscal_final:
            dados = fiscal_final[cod]
            c = ws.cell(row=row, column=15, value=dados['excel'])  # Col O
            c.number_format = '[h]:mm:ss'
            alterados += 1
            encontrados.append((row, cod, dados['bruto'], dados['final']))
        else:
            # Matar fantasma: gravar 0 para empresas sem dados no período
            c = ws.cell(row=row, column=15, value=0)
            c.number_format = '[h]:mm:ss'
            zerados.append((row, cod))
    
    # Empresas no TXT que não estão na Master
    codigos_master = set()
    for row in range(10, ws.max_row + 1):
        cod_val = ws.cell(row=row, column=8).value
        if cod_val:
            try:
                codigos_master.add(int(float(str(cod_val).strip())))
            except:
                pass
    
    for cod in fiscal_final:
        if cod not in codigos_master:
            nao_encontrados_master.append((cod, fiscal_final[cod]['bruto']))
    
    print(f"Salvando planilha... ({alterados} linhas com dados, {len(zerados)} zeradas)")
    wb.save(PLANILHA_MASTER)
    wb.close()
    
    # 4. Relatório
    print("Gerando relatório...")
    with open(RELATORIO_MD, 'w', encoding='utf-8') as f:
        f.write("# Relatório - Lançamento Fiscal na Master (04.2026)\n\n")
        f.write(f"- **Fonte:** horas fiscal 042026.txt\n")
        f.write(f"- **Destino:** CONTROLE DE HORAS DMF.xlsx → aba {MES_ALVO}, coluna O\n")
        f.write(f"- **Fator adicional:** {FATOR_ADICIONAL_FISCAL}x ({int((FATOR_ADICIONAL_FISCAL-1)*100)}% sobre o bruto)\n\n")
        f.write(f"- Total empresas no TXT: **{len(fiscal_por_cod)}**\n")
        f.write(f"- Total linhas processadas na Master: **{total_linhas}**\n")
        f.write(f"- Linhas preenchidas com dados: **{alterados}**\n")
        f.write(f"- Linhas zeradas (sem dados no TXT): **{len(zerados)}**\n")
        f.write(f"- Empresas no TXT sem match na Master: **{len(nao_encontrados_master)}**\n\n")
        
        if encontrados:
            f.write("## Empresas Lançadas\n\n")
            f.write("| Linha | Código | Tempo Bruto | Tempo Final (×1.80) |\n")
            f.write("|---|---|---|---|\n")
            for row, cod, bruto, final in sorted(encontrados, key=lambda x: x[0]):
                f.write(f"| {row} | {cod} | {segundos_para_hms(bruto)} | {segundos_para_hms(final)} |\n")
        
        if nao_encontrados_master:
            f.write("\n## Empresas no TXT sem match na Master\n\n")
            f.write("| Código | Tempo Bruto |\n")
            f.write("|---|---|\n")
            for cod, bruto in sorted(nao_encontrados_master):
                f.write(f"| {cod} | {segundos_para_hms(bruto)} |\n")
    
    print(f"Processo finalizado! Relatório: {RELATORIO_MD}")
    print(f"\n=== RESUMO ===")
    print(f"  Empresas no TXT: {len(fiscal_por_cod)}")
    print(f"  Linhas preenchidas: {alterados}")
    print(f"  Linhas zeradas: {len(zerados)}")
    print(f"  Sem match na Master: {len(nao_encontrados_master)}")

if __name__ == '__main__':
    main()
