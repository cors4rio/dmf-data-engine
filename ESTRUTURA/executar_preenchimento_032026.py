import pyodbc
import openpyxl
import xlrd
import os
import re
from datetime import datetime

# CONFIGS
MES_ALVO = '03.2026'
DATA_INICIO = '2026-03-01'
DATA_FIM = '2026-03-31'
BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_ALVO = os.path.join(BASE_DIR, "HORAS CONTABEIS_.xlsx")
PLANILHA_CAROL = os.path.join(BASE_DIR, "Controle de Empregados (CAROL)032026.xls")
MD_REPORT = os.path.join(BASE_DIR, "qtdlancamento032026.md")
DB_DSN = "DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>"

def clean_cnpj(cnpj):
    if not cnpj: return ""
    return re.sub(r'\D', '', str(cnpj))

def clean_str(s):
    return str(s).strip().upper() if s else ""

def generate_report(report_matches, report_partial, report_rejected, report_updates):
    with open(MD_REPORT, "w", encoding="utf-8") as f:
        f.write("# Relatório de Preenchimento Contábil - 03.2026\n\n")
        
        f.write("## 🛠 Metodologia e Regras de Negócio\n")
        f.write("- **Validação Tripla:** Cada cliente foi validado cruzando Código Domínio (A), CNPJ (C) e Nome da Empresa (D).\n")
        f.write("  - ✅ **Aprovado:** Match de 3/3 campos.\n")
        f.write("  - ⚠️ **Parcial:** Match de 2/3 campos (com alerta no relatório).\n")
        f.write("  - ❌ **Rejeitado:** Match < 2 campos (não preenchido).\n")
        f.write("- **Filtro de Lançamentos (Coluna F):** Extração exclusiva de `orig_lan = 1` (Lançamento Normal). \n")
        f.write("  - *Nota: A origem 5 foi removida para garantir fidelidade aos relatórios internos da Domínio, evitando inflação de dados.*\n")
        f.write("- **Faturamento (Coluna O):** Soma de saídas e serviços do banco de dados.\n")
        f.write("- **Dados de Folha (Coluna I):** Cruzamento realizado com a planilha da Carol (`Controle de Empregados (CAROL)032026.xls`).\n\n")

        f.write(f"## REJEITADAS / ÓRFÃS ({len(report_rejected)})\n")
        for l in report_rejected: f.write(l + "\n")
        
        f.write(f"\n## APROVADAS COM ALERTA (PARCIAIS) ({len(report_partial)})\n")
        for l in report_partial: f.write(l + "\n")
        
        f.write(f"\n## DADOS ATUALIZADOS ({len(report_updates)})\n")
        for l in report_updates: f.write(l + "\n")
    print(f"Report salvo em: {MD_REPORT}")

def main():
    print("Iniciando processo...")
    try:
        conn = pyodbc.connect(DB_DSN)
        cursor = conn.cursor()
    except Exception as e:
        print(f"Erro ao conectar ao DB (teste rodar manual se for erro de arquitetura): {e}")
        return

    # 1. Validação Tripla (geempre)
    print("Buscando cadastro de empresas...")
    cursor.execute("SELECT codi_emp, cgce_emp, nome_emp FROM bethadba.geempre")
    db_emp = {}
    for r in cursor.fetchall():
        cod = str(r[0])
        db_emp[cod] = {
            'cnpj': clean_cnpj(r[1]),
            'nome': clean_str(r[2])
        }

    # 2. Lançamentos Contábeis
    print("Buscando lançamentos...")
    cursor.execute(f"SELECT codi_emp, COUNT(*) as qtd "
                   f"FROM bethadba.ctlancto "
                   f"WHERE data_lan >= '{DATA_INICIO}' AND data_lan <= '{DATA_FIM}' "
                   f"AND orig_lan = 1 "
                   f"GROUP BY codi_emp")
    db_lanc = {str(r[0]): r[1] for r in cursor.fetchall()}

    # 3. Faturamento
    print("Buscando faturamento...")
    cursor.execute(f"SELECT codi_emp, SUM(total) as fat FROM ("
                   f"SELECT codi_emp, SUM(vcon_sai) as total FROM bethadba.efsaidas WHERE dsai_sai >= '{DATA_INICIO}' AND dsai_sai <= '{DATA_FIM}' GROUP BY codi_emp UNION ALL "
                   f"SELECT codi_emp, SUM(vcon_ser) as total FROM bethadba.efservicos WHERE dser_ser >= '{DATA_INICIO}' AND dser_ser <= '{DATA_FIM}' GROUP BY codi_emp"
                   f") base GROUP BY codi_emp")
    db_fat = {str(r[0]): float(r[1]) for r in cursor.fetchall()}
    conn.close()

    # 4. Dados Folha (Carol)
    print("Buscando dados da Folha (Carol)...")
    folha_data = {}
    try:
        wb_carol = xlrd.open_workbook(PLANILHA_CAROL)
        sh_carol = wb_carol.sheet_by_index(0)
        for i in range(1, sh_carol.nrows):
            row = sh_carol.row_values(i)
            try:
                cod_raw = row[1]
                if not cod_raw: continue
                cod = str(int(float(cod_raw)))
                func = float(row[7] or 0)
                estag = float(row[9] or 0)
                contrib = float(row[11] or 0)
                total_ativos = func + estag + contrib
                folha_data[cod] = "SIM" if total_ativos > 0 else "NAO"
            except: pass
    except Exception as e:
        print(f"Erro lendo Carol ({PLANILHA_CAROL}): {e}")

    # 5. Processamento Planilha Alvo
    print("Atualizando planilha alvo...")
    wb = openpyxl.load_workbook(PLANILHA_ALVO, data_only=False)
    
    if MES_ALVO not in wb.sheetnames:
        print(f"Aba {MES_ALVO} não encontrada.")
        return
        
    ws = wb[MES_ALVO]

    # Listas p/ Report
    report_matches = []
    report_partial = []
    report_rejected = []
    report_updates = []

    for row in range(2, ws.max_row + 1):
        col_a_cod = ws.cell(row=row, column=1).value
        col_c_cnpj = ws.cell(row=row, column=3).value
        col_d_nome = ws.cell(row=row, column=4).value
        
        if not col_a_cod: continue
        try:
            plan_cod = str(int(float(str(col_a_cod).strip())))
        except ValueError:
            continue
        plan_cnpj = clean_cnpj(col_c_cnpj)
        plan_nome = clean_str(col_d_nome)
        
        # Tratamento fuzzy nome (remover CNPJ inicial no nome)
        plan_nome_clean = plan_nome
        m = re.match(r'^[\d\.\-\/\s]+(.*)', plan_nome)
        if m:
            plan_nome_clean = m.group(1).strip()
        
        db_info = db_emp.get(plan_cod)
        if not db_info:
            report_rejected.append(f"- Órfã: Cód {plan_cod} - {plan_nome} não existe na Base de Dados Domínio.")
            continue
        
        score = 1 # cod já bateu pra chegar aqui
        issues = []
        
        if plan_cnpj and db_info['cnpj']:
            if plan_cnpj == db_info['cnpj'] or plan_cnpj.zfill(14) == db_info['cnpj'].zfill(14):
                score += 1
            else:
                issues.append(f"CNPJ divergiu Planilha({plan_cnpj}) vs DB({db_info['cnpj']})")
        
        db_nome_clean = db_info['nome']
        if plan_nome_clean and db_nome_clean:
            if plan_nome_clean in db_nome_clean or db_nome_clean in plan_nome_clean:
                score += 1
            else:
                issues.append(f"Nome divergiu Planilha({plan_nome_clean}) vs DB({db_nome_clean})")
                
        # Regra aprovação
        if score >= 2:
            if score == 2:
                report_partial.append(f"- Cód {plan_cod}: {', '.join(issues)}")
            
            # Preencher F: Lançamentos (col 6)
            val_lanc = db_lanc.get(plan_cod)
            if val_lanc is not None:
                ws.cell(row=row, column=6).value = val_lanc
            
            # Preencher O: Faturamento (col 15)
            val_fat = db_fat.get(plan_cod)
            if val_fat is not None:
                ws.cell(row=row, column=15).value = val_fat
                
            # Preencher I: Folha (col 9)
            val_folha = folha_data.get(plan_cod)
            if val_folha is not None:
                ws.cell(row=row, column=9).value = val_folha
                
            report_updates.append(f"- Cód {plan_cod} ({plan_nome}): Lanc={val_lanc or 0}, Fat={val_fat or 0}, Folha={val_folha or 'N/A'}")
        else:
            report_rejected.append(f"- REJEITADA Cód {plan_cod}: Múltiplas divergências: {', '.join(issues)}")

    print("Salvando planilha...")
    wb.save(PLANILHA_ALVO)
    generate_report(report_matches, report_partial, report_rejected, report_updates)
    print("Processo concluído.")

if __name__ == "__main__":
    main()
