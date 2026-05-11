import pyodbc
import openpyxl
import os

def mapear_rfed(rfed_code):
    """
    Mapeamento técnico baseado no campo rfed_par do Domínio:
    1 - Lucro Real
    2 - Simples Nacional (ME)
    4 - Simples Nacional (EPP)
    5 - Lucro Presumido
    7 - Lucro Arbitrado
    8 - Imune / Isenta
    """
    mapping = {
        1: "Lucro Real",
        2: "Simples Nacional",
        4: "Simples Nacional",
        5: "Lucro Presumido",
        7: "Lucro Arbitrado",
        8: "Imune / Isenta"
    }
    return mapping.get(rfed_code, "Não Parametrizado")

def buscar_enquadramentos_db(data_referencia="2026-01-31"):
    enquadramentos = {}
    conn_str = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
    try:
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        # Query focada no campo rfed_par (Tabela Fiscal -> Parâmetros -> Federal -> Enquadramento)
        query = f"""
            SELECT p.codi_emp, p.rfed_par
            FROM bethadba.efparametro_vigencia p
            WHERE p.vigencia_par = (
                SELECT MAX(vigencia_par)
                FROM bethadba.efparametro_vigencia
                WHERE codi_emp = p.codi_emp
                  AND vigencia_par <= '{data_referencia}'
            )
        """
        cursor.execute(query)
        for row in cursor.fetchall():
            cod_emp = row[0]
            rfed = row[1]
            enquadramentos[cod_emp] = mapear_rfed(rfed)
        conn.close()
        return enquadramentos
    except Exception as e:
        print(f"Erro no banco de dados: {e}")
        return {}

def atualizar_enquadramento_012026():
    path = r"C:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS.xlsx"
    
    if not os.path.exists(path):
        print(f"Arquivo não encontrado: {path}")
        return

    print("Carregando enquadramentos federais de 01/2026...")
    res_db = buscar_enquadramentos_db("2026-01-31")
    print(f"Dados de {len(res_db)} empresas carregados do Domínio.")

    print(f"Abrindo planilha: {path}")
    wb = openpyxl.load_workbook(path)
    
    sheet_name = '01.2026'
    if sheet_name not in wb.sheetnames:
        print(f"Erro: Aba {sheet_name} não encontrada.")
        return
        
    ws = wb[sheet_name]
    print(f"Atualizando 'Enquadramento' na Coluna C da aba {sheet_name}...")
    
    atualizados = 0
    nao_localizados = 0
    
    for row in range(2, ws.max_row + 1):
        cod_emp = ws.cell(row=row, column=1).value # Coluna A: Cód Empresa
        if cod_emp and isinstance(cod_emp, (int, float)):
            cod = int(cod_emp)
            regime_real = res_db.get(cod)
            
            # Atualiza a Coluna C (Campo C1 é REGIME)
            if regime_real:
                ws.cell(row=row, column=3).value = regime_real
                atualizados += 1
            else:
                ws.cell(row=row, column=3).value = "Não Localizado"
                nao_localizados += 1
                
    try:
        wb.save(path)
        print(f"\nFinalizado com sucesso!")
        print(f"Empresas atualizadas (Regime Real): {atualizados}")
        print(f"Empresas ignoradas/não encontradas: {nao_localizados}")
    except Exception as e:
        print(f"Erro ao salvar arquivo: {e}")

if __name__ == "__main__":
    atualizar_enquadramento_012026()
