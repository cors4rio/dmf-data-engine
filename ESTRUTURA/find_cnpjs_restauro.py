import openpyxl

def find_by_cnpj():
    path = 'c:/Users/DMF-AUTOMACAO/Documents/PROJETOS/N8N automacao/ESTRUTURA/CONTROLE_DE_HORAS_DMF.xlsm'
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sh = wb['02.2026']
    
    target_cnpjs = ['21425354000158', '21425354000239']
    print("SEARCHING BY CNPJ...")
    for r in range(1, 1500):
        try:
            # CNPJ is in Col J (index 10)
            cnpj = str(sh.cell(row=r, column=10).value or "").strip().replace(".", "").replace("-", "").replace("/", "")
            if cnpj in target_cnpjs:
                code = sh.cell(row=r, column=8).value
                name = sh.cell(row=r, column=9).value
                print(f"Row {r}: Code={code}, CNPJ={cnpj}, Name={name}")
        except: break

if __name__ == '__main__':
    find_by_cnpj()
