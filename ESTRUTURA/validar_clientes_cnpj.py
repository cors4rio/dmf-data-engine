import os
import shutil
import openpyxl
import re

def validate_by_cnpj():
    txt_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\VALIDACAO CNPJ.txt'
    xls_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls'
    temp_xlsx = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\temp_validate_cnpj.xlsx'
    
    print(f"Lendo dados de validação: {txt_path}")
    
    txt_data = [] # List of (cnpj, name, hours)
    with open(txt_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            
            # Formato esperado: CNPJ <tab> NOME <tab> HH:MM:SS
            parts = re.split(r'[\t;]', line)
            if len(parts) >= 3:
                cnpj = parts[0].strip().replace('.', '').replace('-', '').replace('/', '')
                name = parts[1].strip()
                hours = parts[-1].strip()
                txt_data.append({'cnpj': cnpj, 'name': name, 'hours': hours})

    print(f"Clientes no TXT para validar: {len(txt_data)}")

    print(f"Criando cópia temporária XLSX...")
    shutil.copy(xls_path, temp_xlsx)

    print(f"Abrindo workbook...")
    wb = openpyxl.load_workbook(temp_xlsx, data_only=True)
    sheet_name = '12.2025'
    ws = wb[sheet_name]
    
    # CNPJ está na Coluna J (10)
    # Pegar todos os CNPJs da planilha para comparação rápida
    sheet_cnpjs = set()
    for row_idx in range(10, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=10).value
        if cell_val:
            # Limpar CNPJ da planilha
            clean_cnpj = str(cell_val).strip().replace('.', '').replace('-', '').replace('/', '').split('.')[0]
            sheet_cnpjs.add(clean_cnpj)

    print(f"CNPJs únicos encontrados na planilha: {len(sheet_cnpjs)}")

    missing_clients = []
    for client in txt_data:
        if client['cnpj'] not in sheet_cnpjs:
            missing_clients.append(client)

    print(f"\n--- CLIENTES FALTANTES NA PLANILHA ---")
    if not missing_clients:
        print("Nenhum cliente faltando.")
    else:
        for c in missing_clients:
            print(f"CNPJ: {c['cnpj']} | NOME: {c['name']} | HORA: {c['hours']}")
    
    print(f"\nTotal de clientes faltantes: {len(missing_clients)}")

    wb.close()
    os.remove(temp_xlsx)

if __name__ == "__main__":
    validate_by_cnpj()
