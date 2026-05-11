import os
import shutil
import openpyxl
import re

def add_missing_to_xls():
    log_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\faltantes_duplicados_log.txt'
    xls_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xls'
    temp_xlsx = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\temp_add_faltantes.xlsx'
    
    # 1. Ler o relatório de faltantes
    faltantes = []
    print(f"Lendo faltantes do log: {log_path}")
    with open(log_path, 'r', encoding='utf-8') as f:
        reading_list = False
        for line in f:
            if 'LISTA DE UNIDADES FALTANTES:' in line:
                reading_list = True
                continue
            if reading_list and line.startswith('CNPJ:'):
                # Extração via RegEx: CNPJ: <val> | HORA: <val> | NOME: <val>
                match = re.match(r'CNPJ:\s*(.*?)\s*\|\s*HORA:\s*(.*?)\s*\|\s*NOME:\s*(.*)', line)
                if match:
                    faltantes.append({
                        'cnpj': match.group(1).strip().replace('"', ''),
                        'hora': match.group(2).strip(),
                        'nome': match.group(3).strip()
                    })

    print(f"Total de faltantes para adicionar: {len(faltantes)}")

    # 2. Abrir a planilha
    print("Preparando planilha...")
    shutil.copy(xls_path, temp_xlsx)
    wb = openpyxl.load_workbook(temp_xlsx)
    ws = wb['12.2025']
    
    # 3. Encontrar a última linha preenchida (nas colunas de interesse)
    start_row = 10
    for r in range(start_row, ws.max_row + 1):
        if ws.cell(row=r, column=2).value or ws.cell(row=r, column=10).value:
            start_row = r + 1
            
    print(f"Iniciando inserção na linha: {start_row}")

    # 4. Inserir os dados
    # Mapeamentos solicitados: 
    # J (10): CNPJ
    # K (11): NOME
    # P (16): HORA CONTABIL
    # R (18): TOTAL (Precisa da fórmula SUM)
    
    count = 0
    current_row = start_row
    for item in faltantes:
        # CNPJ
        ws.cell(row=current_row, column=10).value = item['cnpj']
        # NOME
        ws.cell(row=current_row, column=11).value = item['nome']
        
        # HORA (Converter para número do Excel)
        try:
            h_str = item['hora']
            parts = h_str.split(':')
            seconds = 0
            if len(parts) >= 2:
                h = int(parts[0])
                m = int(parts[1])
                s = int(parts[2]) if len(parts) > 2 else 0
                seconds = h * 3600 + m * 60 + s
            
            # Excel Time (fração de dia)
            excel_val = seconds / 86400.0
            cell_p = ws.cell(row=current_row, column=16)
            cell_p.value = excel_val
            cell_p.number_format = '[h]:mm:ss'
        except Exception as e:
            print(f"Erro ao converter hora '{item['hora']}': {e}")
            ws.cell(row=current_row, column=16).value = item['hora']

        # FÓRMULA TOTAL (Coluna R - 18)
        # Manter a consistência com o que fizemos antes: SUM(O:Q)
        ws.cell(row=current_row, column=18).value = f"=SUM(O{current_row}:Q{current_row})"
        ws.cell(row=current_row, column=18).number_format = '[h]:mm:ss'
        
        count += 1
        current_row += 1

    print(f"Sucesso: {count} unidades adicionadas.")
    
    # 5. Salvar e substituir
    wb.save(temp_xlsx)
    wb.close()
    shutil.copy(temp_xlsx, xls_path)
    os.remove(temp_xlsx)
    print("Arquivo master atualizado.")

if __name__ == "__main__":
    add_missing_to_xls()
