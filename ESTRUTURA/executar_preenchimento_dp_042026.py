import pyodbc
import openpyxl
import xlrd
import os
import re
import difflib
import sys

# Configuracoes
MES_ALVO = '04.2026'
BASE_DIR = r"c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA"
PLANILHA_MASTER = os.path.join(BASE_DIR, "CONTROLE DE HORAS DMF.xlsx")
PLANILHA_CAROL = os.path.join(BASE_DIR, "Controle de Empregados 042226 (CAROL).xls")
ARQUIVO_DP_NAO = os.path.join(BASE_DIR, "nao_faz_setor", "DP NAO.txt")
RELATORIO_MD = os.path.join(BASE_DIR, "relatorio_dp_042026.md")
ARQUIVO_CALCULOS = os.path.join(BASE_DIR, "calculos_encontrados_dp_042026.md")

def clean_cnpj(val):
    if not val:
        return ""
    return re.sub(r'\D', '', str(val).strip())

def similar(a, b):
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, str(a).upper(), str(b).upper()).ratio()

def carregar_excecoes_dp():
    dp_nao_codes = set()
    consultoria_codes = set()
    
    if os.path.exists(ARQUIVO_DP_NAO):
        with open(ARQUIVO_DP_NAO, 'r', encoding='utf-8') as f:
            for linha in f:
                linha = linha.strip()
                if not linha or linha.startswith('EMPRESAS'):
                    continue
                
                eh_consultoria = "FAZ CONSULTORIA" in linha.upper()
                
                codigo = None
                
                parts_tab = linha.split('\t')
                parts_semi = linha.split(';')
                
                if len(parts_tab) > 1:
                    cod_str = parts_tab[0].strip()
                    if cod_str.isdigit():
                        codigo = int(cod_str)
                elif len(parts_semi) > 1:
                    cod_str = parts_semi[0].strip()
                    if cod_str.isdigit():
                        codigo = int(cod_str)
                else:
                    primeiro_espaco = linha.find(' ')
                    if primeiro_espaco > 0:
                        cod_str = linha[:primeiro_espaco].strip()
                        if cod_str.isdigit():
                            codigo = int(cod_str)
                
                if eh_consultoria:
                    if codigo: consultoria_codes.add(codigo)
                else:
                    if codigo: dp_nao_codes.add(codigo)
                    
    return dp_nao_codes, consultoria_codes

def get_dados_carol():
    """Lê a planilha de empregados da Carol (fonte primária para contagem)."""
    # Layout: Col1=Código, Col3=Nome, Col5=CNPJ, Col7=Funcionários, Col9=Estagiários, Col11=Contribuintes
    resultados = {}
    cnpjs = {}
    
    if not os.path.exists(PLANILHA_CAROL):
        print(f"  ERRO: Arquivo não encontrado: {PLANILHA_CAROL}")
        return resultados, cnpjs
        
    wb_carol = xlrd.open_workbook(PLANILHA_CAROL)
    sh_carol = wb_carol.sheet_by_index(0)
    
    for i in range(1, sh_carol.nrows):  # Pula header (linha 0)
        row = sh_carol.row_values(i)
        try:
            cod_raw = row[1]
            if not cod_raw:
                continue
            cod = int(float(str(cod_raw)))
            
            cnpj = clean_cnpj(row[5] if len(row) > 5 else "")
            
            func = float(row[7] or 0) if len(row) > 7 and row[7] != '' else 0
            estag = float(row[9] or 0) if len(row) > 9 and row[9] != '' else 0
            contrib = float(row[11] or 0) if len(row) > 11 and row[11] != '' else 0
            total_ativos = func + estag + contrib
            
            resultados[cod] = {
                'func': int(func),
                'estag': int(estag),
                'contrib': int(contrib),
                'total': int(total_ativos)
            }
            if cnpj:
                cnpjs[cod] = str(cnpj)
        except Exception:
            pass
            
    return resultados, cnpjs

def main():
    print("Carregando exceções do DP NÃO...")
    dp_nao_codes, consultoria_codes = carregar_excecoes_dp()
    
    # 1. Ler dados da planilha Carol (fonte de contagem de empregados)
    print("Buscando dados da planilha Controle de Empregados (CAROL)...")
    dados_carol, cnpjs_carol = get_dados_carol()
    print(f"  -> {len(dados_carol)} empresas carregadas da Carol.")
    
    # 2. Buscar cadastros do banco (para validação tripla)
    db_cadastros = {}
    conn_str = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
    try:
        print("Conectando ao banco de dados (apenas cadastros)...")
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        
        cursor.execute("SELECT codi_emp, cgce_emp, nome_emp FROM bethadba.geempre")
        for row in cursor.fetchall():
            db_cadastros[row.codi_emp] = {
                'cnpj': clean_cnpj(row.cgce_emp),
                'nome': str(row.nome_emp).strip() if row.nome_emp else ""
            }
        conn.close()
        print(f"  -> {len(db_cadastros)} cadastros carregados do banco.")
    except Exception as e:
        print(f"  AVISO: Erro ao consultar banco: {e}")
        print("  Continuando sem validação tripla (apenas código)...")

    # 3. Abrir planilha Master
    print("Abrindo planilha Master...")
    try:
        wb = openpyxl.load_workbook(PLANILHA_MASTER)
    except Exception as e:
        print(f"Erro ao carregar planilha {PLANILHA_MASTER}: {e}")
        return
        
    if MES_ALVO not in wb.sheetnames:
        print(f"Aba {MES_ALVO} não encontrada.")
        return
    ws = wb[MES_ALVO]
    
    rejeitadas = []
    clientes_encontrados = []
    processados = 0
    alterados = 0
    
    print("Aplicando regras com Validação Tripla e preenchendo coluna Q (Horário Pessoal DP)...")
    MINIMO_5_MIN = (5 / 60) / 24.0
    
    for row in range(10, ws.max_row + 1):
        cod_val = ws.cell(row=row, column=8).value  # Col H
        nome_val = ws.cell(row=row, column=11).value  # Col K (EMPRESA CARTÃO CNPJ)
        cnpj_val = ws.cell(row=row, column=10).value  # Col J
        
        # Ignorar linhas vazias
        if cod_val is None and cnpj_val is None and nome_val is None:
            continue
            
        if isinstance(cod_val, str) and cod_val.startswith('Não'):
            continue
            
        if cod_val is None:
            continue
            
        try:
            cod = int(float(str(cod_val).strip()))
        except ValueError:
            continue
            
        processados += 1
        cnpj_planilha = clean_cnpj(cnpj_val)
        nome_planilha = str(nome_val or "").strip()
        
        # TRIPLA VALIDAÇÃO (contra o banco geempre)
        valido = True
        motivo_rejeicao = ""
        
        if db_cadastros:  # Só valida se temos dados do banco
            if cod not in db_cadastros:
                valido = False
                motivo_rejeicao = "Código não existe no banco"
            else:
                db_cnpj = db_cadastros[cod]['cnpj']
                db_nome = db_cadastros[cod]['nome']
                
                if cnpj_planilha and db_cnpj and cnpj_planilha != db_cnpj:
                    valido = False
                    motivo_rejeicao = f"CNPJ difere. Planilha: {cnpj_planilha} | DB: {db_cnpj}"
                elif nome_planilha:
                    sim_ratio = similar(nome_planilha, db_nome)
                    if sim_ratio < 0.6:
                        valido = False
                        motivo_rejeicao = f"Nome difere (sim={sim_ratio:.2f}). Planilha: {nome_planilha} | DB: {db_nome}"

        if not valido:
            rejeitadas.append((cod, nome_planilha, cnpj_planilha, row, motivo_rejeicao))
            continue
            
        # Buscar dados na Carol pelo código
        total = 0
        info = {'func': 0, 'estag': 0, 'contrib': 0, 'total': 0}
        if cod in dados_carol:
            info = dados_carol[cod]
            total = info['total']
        
        # Calcula valor para Q
        valor_q = None
        if cod in consultoria_codes:
            valor_q = 1.5 / 24.0
        elif cod in dp_nao_codes:
            valor_q = "DP NÃO"
        else:
            if total > 0:
                horas_flt = (total * 0.33) + 1.5
                valor_q = horas_flt / 24.0
            else:
                valor_q = MINIMO_5_MIN
                
        # Escrever na Coluna Q (17) — APENAS coluna Q
        c = ws.cell(row=row, column=17, value=valor_q)
        if isinstance(valor_q, float):
            c.number_format = '[h]:mm:ss'
            
        alterados += 1
        clientes_encontrados.append((cod, info, valor_q, row))
        
    print(f"Salvando planilha... ({alterados} linhas alteradas)")
    wb.save(PLANILHA_MASTER)
    wb.close()
    
    # Relatórios
    print("Gerando relatório...")
    with open(RELATORIO_MD, 'w', encoding='utf-8') as f:
        f.write("# Relatório de Preenchimento DP - 04.2026\n\n")
        f.write(f"- **Fonte de dados:** Planilha Carol (Controle de Empregados)\n")
        f.write(f"- Linhas válidas processadas na Master: {processados}\n")
        f.write(f"- Linhas preenchidas na Master com sucesso: {alterados}\n")
        f.write(f"- Clientes rejeitados por Validação Tripla: {len(rejeitadas)}\n\n")
        
        if rejeitadas:
            f.write("## Clientes Rejeitados na Validação (Ignorados)\n\n")
            f.write("| Linha | Código | Nome Planilha | CNPJ | Motivo |\n")
            f.write("|---|---|---|---|---|\n")
            for item in rejeitadas:
                f.write(f"| {item[3]} | {item[0]} | {item[1]} | {item[2]} | {item[4]} |\n")
                
    print("Gerando relatório de cálculos...")
    with open(ARQUIVO_CALCULOS, 'w', encoding='utf-8') as f:
        f.write("# Cálculos e Valores Preenchidos - DP (04.2026)\n\n")
        f.write("- **Fonte:** Planilha Carol (Controle de Empregados 042226)\n\n")
        f.write("| Linha | Código | Func | Estag | Contrib | Total | Valor Inserido |\n")
        f.write("|---|---|---|---|---|---|---|\n")
        
        clientes_ordenados = sorted(clientes_encontrados, key=lambda x: x[3])
        for cod, info, val_q, row_excel in clientes_ordenados:
            acao = val_q
            if isinstance(acao, float):
                horas_flt = acao * 24.0
                horas_int = int(horas_flt)
                minutos_int = int(round((horas_flt - horas_int) * 60))
                acao = f"{horas_int:02d}:{minutos_int:02d}:00"
                
            f.write(f"| {row_excel} | {cod} | {info['func']} | {info['estag']} | {info['contrib']} | {info['total']} | {acao} |\n")

    print(f"Processo finalizado! Relatório salvo: {RELATORIO_MD}")
    print(f"Cálculos detalhados salvos em: {ARQUIVO_CALCULOS}")
    
if __name__ == '__main__':
    main()
