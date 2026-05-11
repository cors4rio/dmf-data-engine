import pyodbc
import sys
from datetime import datetime
from openpyxl import Workbook

def extract_annual_entries_report():
    conn_str = 'DSN=Contabil;UID=<USER_NO_ENV>;PWD=<SENHA_NO_ENV>'
    try:
        print("Conectando ao banco de dados...")
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        
        # Cria a pasta de trabalho (Workbook)
        wb = Workbook()
        
        # Remove a aba default criada
        default_sheet = wb.active
        
        # Iterar de mes 1 a 11
        for mes in range(1, 12):
            mes_str = f"{mes:02d}"
            ano = "2025"
            
            # Qual o ultimo dia do mes para a query?
            # 2025 nao é bissexto
            dias_no_mes = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30}
            ultimo_dia = dias_no_mes[mes]
            
            data_ini = f"{ano}-{mes_str}-01"
            data_fim = f"{ano}-{mes_str}-{ultimo_dia}"
            
            nome_aba = f"{mes_str}-{ano}"
            print(f"Extraindo dados para: {nome_aba} ({data_ini} a {data_fim})")
            
            query = f"""
                SELECT 
                    l.codi_emp as Codigo_Cliente,
                    SUM(CASE WHEN l.orig_lan = 1 THEN 1 ELSE 0 END) as Lancamentos_Normal,
                    SUM(CASE WHEN l.orig_lan = 39 THEN 1 ELSE 0 END) as Lancamentos_Extrato_Bancario,
                    COUNT(*) as Total_Lancamentos_Gerais
                FROM 
                    bethadba.ctlancto l
                WHERE 
                    l.data_lan >= '{data_ini}' 
                    AND l.data_lan <= '{data_fim}'
                GROUP BY 
                    l.codi_emp
                HAVING 
                    SUM(CASE WHEN l.orig_lan IN (1, 39) THEN 1 ELSE 0 END) > 0
                ORDER BY 
                    l.codi_emp
            """
            
            cursor.execute(query)
            results = cursor.fetchall()
            
            # Cria uma aba com o nome do mês
            ws = wb.create_sheet(title=nome_aba)
            
            # Escreve o cabecalho
            ws.append(['Codigo_Cliente', 'Lancamentos_Normal', 'Lancamentos_Extrato_Bancario', 'Total_Lancamentos_Gerais'])
            
            if not results:
                print(f"  -> Nenhum dado encontrado para {nome_aba}.")
            else:
                print(f"  -> {len(results)} empresas processadas.")
                for row in results:
                    ws.append([row[0], row[1], row[2], row[3]])
                    
        # Remove a planilha de criacao (Sheet) caso ela estiver la, mas já lidamos com default_sheet
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb["Sheet"])
            
        output_file = 'Relatorio_Lancamentos_Contabeis_01_a_11_2025.xlsx'
        wb.save(output_file)
        print(f"\nRelatório Excel completo gerado: {output_file}")
        
    except Exception as e:
        print(f"Erro Fatal: {e}", file=sys.stderr)

if __name__ == '__main__':
    extract_annual_entries_report()
