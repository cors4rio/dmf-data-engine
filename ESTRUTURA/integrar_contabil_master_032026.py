import openpyxl
import re

# Configurações
FONTE_CONTABIL = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_.xlsx'
DESTINO_MASTER = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE DE HORAS DMF.xlsx'
ABA = '03.2026'

def clean_cnpj(val):
    if val is None: return ""
    return re.sub(r'\D', '', str(val))

def clean_code(val):
    if val is None: return ""
    try:
        return str(int(float(str(val).strip())))
    except:
        return str(val).strip()

def main():
    # 1. Carregar Dados da Fonte (Contábil) - data_only=True para pegar o valor das formulas
    print(f"Lendo fonte: {FONTE_CONTABIL}")
    wb_src = openpyxl.load_workbook(FONTE_CONTABIL, data_only=True)
    ws_src = wb_src[ABA]
    
    mapping = {}
    # A=1 (Cod), C=3 (CNPJ), R=18 (Horas Validadas)
    # Dados começam na linha 2
    for r in range(2, ws_src.max_row + 1):
        cod = clean_code(ws_src.cell(row=r, column=1).value)
        cnpj = clean_cnpj(ws_src.cell(row=r, column=3).value)
        valor = ws_src.cell(row=r, column=18).value
        
        if cod and cnpj:
            mapping[(cod, cnpj)] = valor

    wb_src.close()
    print(f"  {len(mapping)} empresas mapeadas na fonte.")

    # 2. Abrir Destino (Master)
    print(f"Abrindo Master: {DESTINO_MASTER}")
    wb_dst = openpyxl.load_workbook(DESTINO_MASTER)
    ws_dst = wb_dst[ABA]

    # 3. Integrar - H=8 (Cod), J=10 (CNPJ), P=16 (Horario Contabil)
    # Dados começam na linha 10
    match_count = 0
    mismatch_count = 0
    
    for r in range(10, ws_dst.max_row + 1):
        cod = clean_code(ws_dst.cell(row=r, column=8).value)
        cnpj = clean_cnpj(ws_dst.cell(row=r, column=10).value)
        
        if not cod: continue
        
        key = (cod, cnpj)
        if key in mapping:
            valor = mapping[key]
            cell_p = ws_dst.cell(row=r, column=16)
            cell_p.value = valor
            cell_p.number_format = '[h]:mm:ss'
            match_count += 1
        else:
            mismatch_count += 1

    # 4. Salvar
    print(f"Salvando Master... ({match_count} matches, {mismatch_count} não encontrados)")
    wb_dst.save(DESTINO_MASTER)
    print("Sucesso!")

if __name__ == "__main__":
    main()
