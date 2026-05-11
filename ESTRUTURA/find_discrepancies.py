import openpyxl
from datetime import time, timedelta, datetime

file_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_V3_AeqwQXgR.xlsx'

def to_timedelta(val):
    if val is None:
        return timedelta(0)
    if isinstance(val, time):
        return timedelta(hours=val.hour, minutes=val.minute, seconds=val.second)
    if isinstance(val, (int, float)):
        # Excel armazena tempo como fração do dia (1.0 = 24h)
        return timedelta(days=val)
    if isinstance(val, bool):
        return timedelta(0)
    if isinstance(val, str):
        try:
            # Tentar parsear "HH:MM:SS" ou "HH:MM"
            parts = list(map(int, val.split(':')))
            if len(parts) == 3:
                return timedelta(hours=parts[0], minutes=parts[1], seconds=parts[2])
            if len(parts) == 2:
                return timedelta(hours=parts[0], minutes=parts[1])
        except:
            pass
    return timedelta(0)

try:
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    
    for sn in wb_data.sheetnames:
        if 'MÉDIA' in sn or 'EXEMPLO' in sn: continue
        ws = wb_data[sn]
        
        discrepancies = []
        for row in range(2, ws.max_row + 1):
            h = to_timedelta(ws[f'H{row}'].value)
            j = to_timedelta(ws[f'J{row}'].value)
            l = to_timedelta(ws[f'L{row}'].value)
            n = to_timedelta(ws[f'N{row}'].value)
            p = to_timedelta(ws[f'P{row}'].value)
            q = to_timedelta(ws[f'Q{row}'].value)
            
            expected = h + j + l + n + p
            # Tolerar pequena diferença de arredondamento
            if abs((expected - q).total_seconds()) > 1:
                discrepancies.append((row, expected, q))
                
        print(f"Aba: {sn} | Desvios encontrados: {len(discrepancies)}")
        if discrepancies:
            print("Exemplos (Linha | Esperado | Cache):")
            for r, exp, cache in discrepancies[:5]:
                print(f"  Linha {r}: {exp} != {cache}")

except Exception as e:
    print(f"Erro: {e}")
