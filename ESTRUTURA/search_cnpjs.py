import openpyxl
import os
import re

def clean_cnpj(c):
    return re.sub(r'\D', '', str(c))

cnpjs_to_find = ['21425354000158', '21425354000239']
files_to_check = [
    r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_V3_AeqwQXgR.xlsx',
    r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm'
]

results = []

for file_path in files_to_check:
    if not os.path.exists(file_path):
        results.append(f"File not found: {file_path}")
        continue
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                row_str = " ".join([str(cell) for cell in row if cell is not None])
                for cnpj in cnpjs_to_find:
                    if cnpj in clean_cnpj(row_str):
                        results.append(f"Found {cnpj} in {file_path} [{sheet_name}] line {row_idx}: {row}")
    except Exception as e:
        results.append(f"Error reading {file_path}: {e}")

with open('c:/Users/DMF-AUTOMACAO/Documents/PROJETOS/N8N automacao/ESTRUTURA/search_results.txt', 'w') as f:
    f.write("\n".join(results))

print("Search completed. Results in search_results.txt")
