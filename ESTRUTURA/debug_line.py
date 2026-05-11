import openpyxl

file_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_V3_AeqwQXgR.xlsx'

try:
    wb_meta = openpyxl.load_workbook(file_path, data_only=False)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    
    sheet_name = '01.2026'
    ws_meta = wb_meta[sheet_name]
    ws_data = wb_data[sheet_name]
    
    # Imprimir cabeçalhos até a coluna Q
    headers = {}
    for i in range(1, 19): # A até R
        col_letter = openpyxl.utils.get_column_letter(i)
        headers[col_letter] = ws_data.cell(row=1, column=i).value
        
    print(f"Cabeçalhos: {headers}")
    
    # Detalhar linha 2
    row = 2
    cols = ['H', 'J', 'L', 'N', 'P', 'Q']
    print(f"\nDetalhes da Linha {row}:")
    for col in cols:
        m_val = ws_meta[f'{col}{row}'].value
        d_val = ws_data[f'{col}{row}'].value
        print(f"  {col} ({headers[col]}): Formula={m_val}, Valor={repr(d_val)}, Tipo={type(d_val).__name__}")
        
    # Detalhar uma linha onde Q é 0 (ou parece estar errado)
    # Procurar por uma linha onde H, J, L, N ou P tem valor, mas Q é 0:00:00
    for r in range(2, 50):
        q_val = ws_data[f'Q{r}'].value
        if q_val is None or (isinstance(q_val, (int, float)) and q_val == 0) or (hasattr(q_val, 'hour') and q_val.hour == 0 and q_val.minute == 0):
             # Verificar se fontes tem algo
             sources = [ws_data[f'{c}{r}'].value for c in ['H', 'J', 'L', 'N', 'P']]
             if any(sources):
                 print(f"\nANOMALIA ENCONTRADA NA LINHA {r}:")
                 for col in cols:
                     print(f"  {col} ({headers[col]}): Formula={ws_meta[f'{col}{row}'].value}, Valor={repr(ws_data[f'{col}{r}'].value)}")
                 break

except Exception as e:
    print(f"Erro: {e}")
