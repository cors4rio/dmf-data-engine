import openpyxl
import re

def clean_cnpj(c):
    return re.sub(r'\D', '', str(c))

cnpjs = ['21425354000158', '21425354000239']
# Updated codes mentioned by user: 21425354000158 -> 1283, 21425354000239 -> 1195

source_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\HORAS CONTABEIS_V3_AeqwQXgR.xlsx'
master_path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF.xlsm'

print("--- SOURCE: HORAS CONTABEIS ---")
wb_s = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
sh_s = wb_s.active
for row in sh_s.iter_rows(values_only=True):
    row_str = "".join([str(c) for c in row if c])
    for cnpj in cnpjs:
        if cnpj in row_str:
            print(f"CNPJ {cnpj}: {row}")

print("\n--- MASTER: CONTROLE_DE_HORAS_DMF (Aba 02.2026) ---")
wb_m = openpyxl.load_workbook(master_path, data_only=True, read_only=True)
if '02.2026' in wb_m.sheetnames:
    sh_m = wb_m['02.2026']
    # Column H is index 7, Column J is index 9, Column P is index 15, Column Q is index 16
    for row in sh_m.iter_rows(min_row=10, values_only=True):
        code = row[7]
        cnpj_val = clean_cnpj(row[9])
        if cnpj_val in cnpjs or str(code) in ['1283', '1195']:
            print(f"Row {row[0]}: Code={code}, CNPJ={row[9]}, Contabil(P)={row[15]}, Pessoal(Q)={row[16]}, Fiscal(O)={row[14]}")
else:
    print("Sheet 02.2026 not found")
