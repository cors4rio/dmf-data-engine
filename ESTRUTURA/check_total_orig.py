import openpyxl
from datetime import timedelta

wb = openpyxl.load_workbook(
    r'C:\Users\DMF-AUTOMACAO\Downloads\CONTROLE_DE_HORAS_DMF_BL4t520G.xlsm', data_only=True
)
ws = wb['02.2026']

# Onde esta o total 2247:55:39 na planilha original?
# Verificar linha 7 (onde costuma ficar o SUBTOTAL) e linha 9
for r in [7, 8, 9]:
    for c in [14, 15, 16, 17]:
        v = ws.cell(r, c).value
        print(f"  Row {r}, Col {c}: {v!r}")

# Calcular a soma bruta de TODA a coluna O
soma = 0
count = 0
for r in range(10, ws.max_row + 1):
    v = ws.cell(r, 15).value
    if isinstance(v, timedelta):
        soma += v.total_seconds() / 86400.0
        count += 1
    elif isinstance(v, (int, float)) and abs(v) < 1000:
        soma += float(v)
        count += 1

total_h = int(soma * 24)
total_m = int((soma * 24 - total_h) * 60)
total_s = int(((soma * 24 - total_h) * 60 - total_m) * 60)
print(f"\nSoma pura de toda col O: {count} registros = {total_h}:{total_m:02d}:{total_s:02d}")
print(f"Em horas decimais: {soma*24:.4f}")

# Verificar a formula do SUBTOTAL na origem
wb2 = openpyxl.load_workbook(
    r'C:\Users\DMF-AUTOMACAO\Downloads\CONTROLE_DE_HORAS_DMF_BL4t520G.xlsm', data_only=False
)
ws2 = wb2['02.2026']
print(f"\nFormula em O7 origem: {ws2.cell(7, 15).value!r}")
print(f"Max row: {ws2.max_row}")
