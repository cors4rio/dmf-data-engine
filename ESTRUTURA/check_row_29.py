import openpyxl
from datetime import time, timedelta

path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA\CONTROLE_DE_HORAS_DMF_ATUALIZADO.xlsm'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['02.2026']
r = 29
p = ws.cell(row=r, column=16).value
r_val = ws.cell(row=r, column=18).value

def format_val(v):
    if v is None: return "None"
    return str(v)

print(f"Row 29 (614): P={format_val(p)}, R={format_val(r_val)}")
