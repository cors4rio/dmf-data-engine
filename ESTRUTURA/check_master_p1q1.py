import os
import openpyxl

def find_master_and_check():
    path = r'c:\Users\DMF-AUTOMACAO\Documents\PROJETOS\N8N automacao\ESTRUTURA'
    files = os.listdir(path)
    print("Files found:")
    for f in files:
        if "RECUPERADO" in f or "MASTER" in f.upper():
            print(f" - {f}")
            full_path = os.path.join(path, f)
            try:
                wb = openpyxl.load_workbook(full_path, data_only=False)
                if '03.2026' in wb.sheetnames:
                    ws = wb['03.2026']
                    print(f"   [03.2026] P1: {ws['P1'].value}, Q1: {ws['Q1'].value}")
                    print(f"   [03.2026] P7: {ws['P7'].value}, Q7: {ws['Q7'].value}")
                    print(f"   [03.2026] P9: {ws['P9'].value}, Q9: {ws['Q9'].value}")
                else:
                    print(f"   Sheet '03.2026' not found in {f}")
            except Exception as e:
                print(f"   Error reading {f}: {e}")

if __name__ == "__main__":
    find_master_and_check()
