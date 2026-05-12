import os
import openpyxl
import xlrd

print("Testando Carol...")
caminho = "ENTRADAS_MANUAIS/Controle de Empregados (CAROL).xls"
try:
    wb = xlrd.open_workbook(caminho)
    sheet = wb.sheet_by_index(0)
    for rowx in range(0, 10):
        try:
            print(f"Row {rowx}: {sheet.row_values(rowx)[:15]}")
        except:
            pass
except Exception as e:
    print("Erro:", e)

print("\nTestando Contabil...")
caminho2 = "ENTRADAS_MANUAIS/HORAS CONTABEIS.xlsx"
try:
    wb2 = openpyxl.load_workbook(caminho2, data_only=True)
    sheet2 = wb2.active
    for row in range(1, 10):
        try:
            print(f"Row {row}: Col A={sheet2.cell(row=row, column=1).value}, Col R={sheet2.cell(row=row, column=18).value}")
        except:
            pass
except Exception as e:
    print("Erro:", e)
