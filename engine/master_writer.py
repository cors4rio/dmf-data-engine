import openpyxl
import logging
import os
import re

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class MasterWriter:
    """
    Componente central responsável por ler e escrever de forma segura na Planilha Master.
    Aplica validação dupla (Codi_emp e CNPJ) e formatação correta de horas.
    """
    def __init__(self, caminho_master):
        self.caminho = caminho_master
        self.wb = None
        self.ws = None
        self.row_map = {}
        
    def carregar(self):
        if not os.path.exists(self.caminho):
            logging.error(f"[MASTER] Planilha Master não encontrada: {self.caminho}")
            return False
            
        logging.info(f"[MASTER] Carregando Planilha Master na memória: {self.caminho}")
        try:
            self.wb = openpyxl.load_workbook(self.caminho, keep_vba=True)
            # Pega a aba ativa (normalmente a do mês atual se salva assim)
            self.ws = self.wb.active
            self._mapear_linhas()
            return True
        except Exception as e:
            logging.error(f"[MASTER] Erro crítico ao abrir Planilha Master: {e}")
            return False
            
    def _mapear_linhas(self):
        """Mapeia os códigos e CNPJs para as linhas correspondentes."""
        self.row_map = {}
        # Na Master, os dados reais começam na linha 10
        for row in range(10, self.ws.max_row + 1):
            cod = self.ws.cell(row=row, column=8).value  # Coluna H
            cnpj = self.ws.cell(row=row, column=10).value # Coluna J
            
            # Mapeamento Primário (Código)
            if cod is not None:
                try:
                    cod_str = str(int(float(str(cod).strip())))
                    if cod_str not in self.row_map:
                        self.row_map[cod_str] = []
                    self.row_map[cod_str].append(row)
                except ValueError:
                    pass
                    
            # Mapeamento Secundário (CNPJ)
            if cnpj is not None:
                cnpj_limpo = re.sub(r'\D', '', str(cnpj))
                if cnpj_limpo:
                    if cnpj_limpo not in self.row_map:
                        self.row_map[cnpj_limpo] = []
                    # Evitar adicionar a mesma linha 2 vezes na mesma chave
                    if row not in self.row_map[cnpj_limpo]:
                        self.row_map[cnpj_limpo].append(row)
                        
        logging.info(f"[MASTER] Mapeamento concluído. Indexadas linhas seguras.")

    def obter_linhas(self, cod_str=None, cnpj_str=None):
        linhas = []
        if cod_str and cod_str in self.row_map:
            linhas.extend(self.row_map[cod_str])
        elif cnpj_str: # Fallback seguro
            cnpj_limpo = re.sub(r'\D', '', str(cnpj_str))
            if cnpj_limpo in self.row_map:
                linhas.extend(self.row_map[cnpj_limpo])
        return list(set(linhas))

    def preencher_fiscal(self, cod_str, valor_horas, cnpj_str=None):
        """Escreve na Coluna O"""
        self._escrever(cod_str, cnpj_str, 15, valor_horas)

    def preencher_contabil(self, cod_str, valor_horas, cnpj_str=None):
        """Escreve na Coluna P"""
        self._escrever(cod_str, cnpj_str, 16, valor_horas)
        
    def preencher_dp(self, cod_str, valor_horas, cnpj_str=None):
        """Escreve na Coluna Q"""
        self._escrever(cod_str, cnpj_str, 17, valor_horas)

    def _escrever(self, cod_str, cnpj_str, col_idx, valor):
        linhas = self.obter_linhas(cod_str, cnpj_str)
        for r in linhas:
            c = self.ws.cell(row=r, column=col_idx)
            c.value = valor
            if isinstance(valor, (float, int)):
                c.number_format = '[h]:mm:ss'
                
    def recalcular_totais(self):
        """Atualiza dinamicamente a Coluna R (Total = O+P+Q) e o SUBTOTAL (linha 7)."""
        max_r = self.ws.max_row
        for r in range(10, max_r + 1):
            cod = self.ws.cell(row=r, column=8).value
            if cod is not None:
                c = self.ws.cell(row=r, column=18)
                c.value = f"=O{r}+P{r}+Q{r}"
                c.number_format = '[h]:mm:ss'
                
        # Atualizar subtotais
        for col_idx, letra in [(15, 'O'), (16, 'P'), (17, 'Q'), (18, 'R')]:
            c = self.ws.cell(row=7, column=col_idx)
            c.value = f"=SUBTOTAL(9,{letra}10:{letra}{max_r})"
            c.number_format = '[h]:mm:ss'
            
        logging.info("[MASTER] Fórmulas de total recalculadas com sucesso.")

    def salvar(self):
        """Salva com bloqueio de arquivo."""
        logging.info("[MASTER] Salvando Planilha Master...")
        try:
            self.wb.save(self.caminho)
            logging.info(f"[MASTER] Salvo com sucesso em: {self.caminho}")
            return True
        except PermissionError:
            logging.error("[MASTER] ERRO CRÍTICO: Planilha aberta no Excel pelo usuário. Feche-a e rode novamente.")
            return False
        except Exception as e:
            logging.error(f"[MASTER] Erro ao salvar: {e}")
            return False
