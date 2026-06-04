"""
dmf_engine/api.py
Api slim — bridge PyWebView ↔ ModuleRegistry.

Zero lógica de negócio aqui. Cada método público fica disponível no JS via:
    window.pywebview.api.nome_do_metodo(args)

Recebe dependências pelo construtor (injeção), sem globais de main.py.
"""
import os
import re
import json
import logging
from datetime import datetime, timedelta

log = logging.getLogger("Api")


class Api:
    """
    Bridge slim: despacha chamadas JS para o ModuleRegistry ou lida
    com auth/config/utilidades que não são módulos de negócio.

    Parâmetros de construtor:
      registry    — ModuleRegistry
      config      — ConfigManager
      auth        — módulo dmf_engine.auth
      bus         — EventBus
      base_dir    — diretório base (onde fica config.json, logs)
      project_root — raiz do projeto (onde fica a master .xlsm)
      window_fn   — callable que retorna a janela PyWebView atual
    """

    def __init__(self, registry, config, auth, bus, base_dir: str, project_root: str, window_fn):
        self._registry = registry
        self._config = config
        self._auth = auth
        self._bus = bus
        self._base_dir = base_dir
        self._project_root = project_root
        self._win = window_fn    # callable → janela PyWebView
        self._sessao = None

        try:
            auth.inicializar_supervisores_padrao(base_dir)
        except Exception as e:
            log.error(f"[BOOT] Falha ao inicializar supervisores: {e}")

        try:
            cfg = self._config.load()
            from engine.database import db
            db.configurar(
                dsn=cfg.get("db_dsn"),
                uid=cfg.get("db_uid"),
                pwd=cfg.get("db_pwd") or db.pwd,
                timeout=cfg.get("db_timeout", 5),
            )
        except Exception as e:
            log.error(f"[BOOT] Falha ao aplicar credenciais do config: {e}")

    # ── Helpers internos ─────────────────────────────────────────────────────

    def _autorizado(self, modulo: str) -> bool:
        if not self._sessao:
            return False
        return self._auth.pode_executar(self._sessao.get("papel"), modulo)

    def _sessao_snapshot(self) -> dict:
        return dict(self._sessao) if self._sessao else {}

    def _master_path(self, opcoes: dict = None) -> str:
        cfg = self._config.load()
        return (
            (opcoes or {}).get("master_path")
            or cfg.get("master_path")
            or os.path.join(self._project_root, "CONTROLE DE HORAS DMF.xlsm")
        )

    # ── Autenticação ─────────────────────────────────────────────────────────

    def verificar_estado_login(self):
        return {"sessao": self._sessao}

    def login(self, nome, senha):
        r = self._auth.autenticar(self._base_dir, nome, senha)
        if r.get("ok") and not r.get("precisa_trocar_senha"):
            self._sessao = {
                "nome": r["nome"], "label": r.get("label"),
                "papel": r.get("papel"), "maquina": r.get("maquina"),
            }
            log.info(f"[AUTH] Sessão iniciada: {r['nome']} ({r.get('papel')})")
        return r

    def trocar_senha_primeiro_acesso(self, nome, senha_atual, senha_nova):
        r = self._auth.trocar_senha_primeiro_acesso(self._base_dir, nome, senha_atual, senha_nova)
        if r.get("ok"):
            self._sessao = {
                "nome": r["nome"], "label": r.get("label"),
                "papel": r.get("papel"), "maquina": r.get("maquina"),
            }
        return r

    def logout(self):
        anterior = self._sessao.get("nome") if self._sessao else None
        self._sessao = None
        log.info(f"[AUTH] Sessão encerrada: {anterior}")
        return {"ok": True}

    def trocar_senha(self, senha_atual, senha_nova):
        if not self._sessao:
            return {"ok": False, "erro": "Sessão não autenticada."}
        return self._auth.trocar_senha(self._base_dir, self._sessao["nome"], senha_atual, senha_nova)

    def admin_listar_usuarios(self):
        if not self._sessao or self._sessao.get("papel") != "admin":
            return {"ok": False, "erro": "Apenas o admin pode listar usuários."}
        return {"ok": True, "usuarios": self._auth.listar_supervisores(self._base_dir)}

    def admin_liberar_maquina(self, nome):
        if not self._sessao or self._sessao.get("papel") != "admin":
            return {"ok": False, "erro": "Apenas o admin pode liberar máquinas."}
        return self._auth.liberar_maquina(self._base_dir, nome)

    def admin_resetar_senha(self, nome):
        if not self._sessao or self._sessao.get("papel") != "admin":
            return {"ok": False, "erro": "Apenas o admin pode resetar senhas."}
        return self._auth.resetar_senha(self._base_dir, nome)

    def abrir_guia_uso(self):
        import webbrowser
        url = ("https://any.coop/AAKr8LpD6zoPuFhRs2EXVjQQEVdcaqM2JW8XPhRpS9JjL4FV/"
               "guia-de-uso-controle-de-horas-dmf")
        try:
            webbrowser.open(url, new=2)
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "erro": str(e)}

    # ── Configuração ─────────────────────────────────────────────────────────

    def carregar_parametros(self):
        return self._config.load()

    def salvar_parametros(self, params: dict):
        ok = self._config.save(params)
        return {"ok": ok}

    def testar_conexao(self):
        from engine.database import db
        cfg = self._config.load()
        db.configurar(
            dsn=cfg.get("db_dsn"),
            uid=cfg.get("db_uid"),
            pwd=cfg.get("db_pwd") or db.pwd,
            timeout=cfg.get("db_timeout", 5),
        )
        ok = db.connect()
        resultado = {
            "ok": ok,
            "msg": "Conexão com Domínio estabelecida com sucesso." if ok else "Falha na conexão ODBC.",
            "erro": db.ultimo_erro if not ok else None,
        }
        self._config.save({
            "db_ultimo_teste_ok": ok,
            "db_ultimo_teste_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "db_ultimo_erro": db.ultimo_erro if not ok else None,
        })
        return resultado

    def verificar_primeira_execucao(self):
        """Retorna True se config.json não existe (primeira execução)."""
        import os
        return {"primeira_execucao": not os.path.exists(self._config._path)}

    def diagnosticar_odbc(self):
        import sys, platform
        cfg = self._config.load()
        dsn_conf = cfg.get("db_dsn", "")
        resultado = {
            "python_version": sys.version.split()[0],
            "python_bits": platform.architecture()[0],
            "drivers": [],
            "dsns_user": {},
            "dsn_configurado": dsn_conf,
            "dsn_existe": False,
        }
        try:
            import pyodbc
            resultado["drivers"] = sorted(pyodbc.drivers())
            resultado["dsns_user"] = dict(pyodbc.dataSources())
            resultado["dsn_existe"] = dsn_conf in resultado["dsns_user"]
        except Exception as e:
            resultado["erro_pyodbc"] = str(e)
        return resultado

    # ── Catálogo + dispatcher genérico ───────────────────────────────────────

    def get_catalog(self) -> dict:
        """Catálogo filtrado pelo papel da sessão atual."""
        papel = (self._sessao or {}).get("papel", "")
        return self._registry.catalog_for_role(papel)

    def executar_modulo(self, module_id: str, opcoes: dict = None) -> dict:
        """Dispatcher genérico — único ponto de entrada para execução de módulos."""
        opcoes = opcoes or {}
        papel = (self._sessao or {}).get("papel", "")
        cat = self._registry.catalog()
        mod_meta = next(
            (m for mods in cat.values() for m in mods if m["id"] == module_id),
            None,
        )
        if mod_meta is None:
            return {"ok": False, "erro": f"Módulo '{module_id}' não encontrado."}
        if papel not in mod_meta["papeis"] and papel != "admin":
            return {"ok": False, "erro": "Sem permissão para este módulo."}
        opcoes["_sessao"] = self._sessao_snapshot()

        # Módulos "sync" (ex: launchers que abrem subprocesso) executam de forma
        # direta e o resultado real volta ao JS. Sem isso, o retorno seria
        # engolido pelo ThreadRunner (que sempre devolve {"ok": True}) e um erro
        # do launcher ficaria invisível — clica e "nada acontece".
        if mod_meta.get("execucao") == "sync":
            mod = self._registry.get(module_id)
            try:
                return mod.execute(opcoes)
            except Exception as e:
                log.error(f"[{module_id}] execução sync falhou: {e}")
                return {"ok": False, "erro": str(e)}

        return self._registry.execute(module_id, opcoes)

    def get_module_status(self, module_id: str) -> dict:
        return self._registry.get_status(module_id)

    # ── Seleção de arquivos ───────────────────────────────────────────────────

    def selecionar_arquivo(self, tipos=None):
        import webview
        if tipos is None:
            tipos = ["Planilha Excel (*.xlsx;*.xlsm)", "Todos os arquivos (*.*)"]
        win = self._win()
        result = win.create_file_dialog(webview.OPEN_DIALOG, allow_multiple=False, file_types=tipos)
        if result and result[0]:
            caminho = result[0]
            self._config.save({"master_path": caminho})
            return caminho
        return None

    def abrir_seletor_pasta(self):
        import webview
        r = self._win().create_file_dialog(webview.FOLDER_DIALOG, directory=os.path.expanduser("~"))
        return {"caminho": r[0] if r else None}

    def abrir_arquivo(self, caminho: str):
        import subprocess
        try:
            os.startfile(caminho)
            return {"ok": True}
        except AttributeError:
            subprocess.Popen(["xdg-open", caminho])
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "erro": str(e)}

    def validar_arquivos(self, caminhos: dict):
        return {k: os.path.exists(v) if v else False for k, v in caminhos.items()}

    # ── Estado por módulo ─────────────────────────────────────────────────────

    def obter_estado_fiscal(self, data_inicio):
        from engine import estado_compartilhado as estado_sh
        master_path = self._master_path()
        comp = (data_inicio or "")[:7]
        sh = estado_sh.obter(master_path, "fiscal", comp)
        lancado = sh.get("lancado") or {}
        return {
            "competencia": comp,
            "lancado": bool(lancado),
            "lancado_em": lancado.get("em"),
            "lancado_por": lancado.get("por"),
        }

    def obter_estado_dp(self, data_inicio):
        from engine import estado_compartilhado as estado_sh
        cfg = self._config.load()
        comp = (data_inicio or "")[:7]
        master_path = self._master_path()
        sh = estado_sh.obter(master_path, "dp", comp)
        carol = sh.get("carol_importada") or {}
        lancado = sh.get("lancado") or {}
        return {
            "competencia": comp,
            "importado": bool(carol) or bool(cfg.get(f"dp_carol_importado_{comp}")),
            "importado_em": carol.get("em") or cfg.get(f"dp_carol_importado_em_{comp}"),
            "importado_por": carol.get("por"),
            "caminho": cfg.get(f"dp_carol_path_{comp}"),
            "total": carol.get("total") or cfg.get(f"dp_carol_total_{comp}"),
            "com_ativos": carol.get("com_ativos") or cfg.get(f"dp_carol_com_ativos_{comp}"),
            "lancado": bool(lancado) or bool(cfg.get(f"dp_lancado_{comp}")),
            "lancado_em": lancado.get("em") or cfg.get(f"dp_lancado_em_{comp}"),
            "lancado_por": lancado.get("por"),
        }

    def obter_estado_contabil(self, data_inicio):
        from engine import estado_compartilhado as estado_sh
        cfg = self._config.load()
        comp = (data_inicio or "")[:7]
        master_path = self._master_path()
        sh = estado_sh.obter(master_path, "contabil", comp)
        return {
            "competencia": comp,
            "processado": bool(sh.get("processado")) or bool(cfg.get(f"contabil_processado_{comp}")),
            "processado_em": (sh.get("processado") or {}).get("em") or cfg.get(f"contabil_processado_em_{comp}"),
            "processado_por": (sh.get("processado") or {}).get("por"),
            "lancado": bool(sh.get("lancado")) or bool(cfg.get(f"contabil_lancado_{comp}")),
            "lancado_em": (sh.get("lancado") or {}).get("em") or cfg.get(f"contabil_lancado_em_{comp}"),
            "lancado_por": (sh.get("lancado") or {}).get("por"),
            "caminho": cfg.get("contabil_origem_path"),
        }

    def obter_resumo_competencia(self, data_inicio):
        comp = (data_inicio or "")[:7]
        try:
            ano_f, mes_f = int(comp[:4]), int(comp[5:7]) - 1
            if mes_f == 0:
                mes_f, ano_f = 12, ano_f - 1
            comp_fiscal = f"{ano_f}-{mes_f:02d}"
        except Exception:
            comp_fiscal = comp
        return {
            "competencia": comp,
            "competencia_fiscal": comp_fiscal,
            "fiscal": self.obter_estado_fiscal(f"{comp_fiscal}-01"),
            "contabil": self.obter_estado_contabil(data_inicio),
            "dp": self.obter_estado_dp(data_inicio),
        }

    def obter_lock_master(self):
        from engine.lock_master import verificar_lock
        master_path = self._master_path()
        if not os.path.exists(master_path):
            return {"lock": None}
        return {"lock": verificar_lock(master_path)}

    # ── Dashboard ─────────────────────────────────────────────────────────────

    def obter_metricas_dashboard(self, competencia=None):
        d = self.obter_dashboard_completo(competencia)
        return {
            "ok": d.get("ok", False),
            "clientes": d.get("clientes", 0),
            "matches": d.get("matches", 0),
            "pendencias": d.get("pendencias", 0),
            "cnpj_dup": d.get("cnpj_dup", 0),
            "master_path": d.get("master_path"),
            "msg": d.get("msg"),
        }

    def obter_dashboard_completo(self, competencia=None):
        from engine.onedrive_helper import esta_online_only, forcar_download
        master_path = self._master_path()
        vazio = {
            "ok": False, "clientes": 0, "matches": 0, "pendencias": 0, "cnpj_dup": 0,
            "horas_por_setor": {"fiscal": 0, "contabil": 0, "dp": 0},
            "top_clientes": [], "distribuicao": [],
            "preenchimento_setor": {"fiscal": 0, "contabil": 0, "dp": 0},
            "aba_usada": None, "master_path": master_path,
        }
        if not master_path or not os.path.exists(master_path):
            vazio["msg"] = "Planilha master não encontrada"
            return vazio

        if esta_online_only(master_path):
            forcar_download(master_path, timeout=20)

        aba_alvo = None
        if competencia and len(competencia) >= 7:
            aba_alvo = f"{competencia[5:7]}.{competencia[:4]}"

        try:
            import openpyxl
            wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
            if aba_alvo and aba_alvo in wb.sheetnames:
                ws = wb[aba_alvo]
                aba_usada = aba_alvo
            else:
                ws = wb.active
                aba_usada = ws.title

            clientes = matches = pendencias = 0
            cnpj_count: dict = {}
            codigo_count: dict = {}
            soma_fiscal = soma_contabil = soma_dp = 0.0
            pres_fiscal = pres_contabil = pres_dp = 0
            buckets = [0, 0, 0, 0, 0, 0]
            clientes_buf = []
            pendentes_buf = []

            def _num(v):
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    return float(v)
                if isinstance(v, timedelta):
                    return v.total_seconds() / 86400.0
                return None

            for linha_idx, row in enumerate(ws.iter_rows(min_row=10, values_only=True), start=10):
                if not row or len(row) < 18:
                    continue
                cod = row[7]
                nome = row[10] if row[10] not in (None, "") else row[8]
                cnpj = row[9]
                if cod is None and cnpj is None and nome is None:
                    continue
                if cod is None:
                    continue
                clientes += 1
                campos = sum(1 for v in (cod, nome, cnpj) if v not in (None, ""))
                if campos >= 2:
                    matches += 1
                else:
                    pendencias += 1
                    pendentes_buf.append({
                        "linha": linha_idx,
                        "codigo": str(cod) if cod is not None else "",
                        "nome": str(nome or ""),
                        "cnpj": str(cnpj or ""),
                        "campos_preenchidos": campos,
                    })
                if cnpj:
                    cnpj_limpo = re.sub(r"\D", "", str(cnpj))
                    if cnpj_limpo:
                        cnpj_count.setdefault(cnpj_limpo, []).append(
                            {"linha": linha_idx, "codigo": str(cod), "nome": str(nome or ""), "cnpj": cnpj_limpo}
                        )
                try:
                    codigo_norm = str(int(float(str(cod).strip())))
                    codigo_count.setdefault(codigo_norm, []).append(
                        {"linha": linha_idx, "codigo": codigo_norm, "nome": str(nome or ""), "cnpj": str(cnpj or "")}
                    )
                except (ValueError, TypeError):
                    pass

                v_f = _num(row[14]) or 0.0
                v_c = _num(row[15]) or 0.0
                v_d = _num(row[16]) or 0.0
                soma_fiscal += v_f; soma_contabil += v_c; soma_dp += v_d
                if v_f > 0: pres_fiscal += 1
                if v_c > 0: pres_contabil += 1
                if v_d > 0: pres_dp += 1

                total_r = _num(row[17])
                if total_r is None:
                    total_r = v_f + v_c + v_d
                total_h = total_r * 24.0

                if total_h <= 0: buckets[0] += 1
                elif total_h <= 2: buckets[1] += 1
                elif total_h <= 5: buckets[2] += 1
                elif total_h <= 10: buckets[3] += 1
                elif total_h <= 20: buckets[4] += 1
                else: buckets[5] += 1

                try:
                    codigo_str = str(int(float(str(cod).strip())))
                except (ValueError, TypeError):
                    continue
                clientes_buf.append((total_r, str(nome or "")[:42], v_f, v_c, v_d, codigo_str))

            wb.close()
            clientes_buf.sort(key=lambda t: t[0], reverse=True)
            top_clientes = [
                {"codigo": c[5], "nome": c[1] or c[5],
                 "horas_total": round(c[0] * 24, 2),
                 "horas_fiscal": round(c[2] * 24, 2),
                 "horas_contabil": round(c[3] * 24, 2),
                 "horas_dp": round(c[4] * 24, 2)}
                for c in clientes_buf[:10]
            ]
            cnpj_dup_lista = [{"cnpj": k, "ocorrencias": v} for k, v in cnpj_count.items() if len(v) > 1]
            codigo_dup_lista = [{"codigo": k, "ocorrencias": v} for k, v in codigo_count.items() if len(v) > 1]

            return {
                "ok": True, "aba_usada": aba_usada,
                "competencia_solicitada": competencia,
                "clientes": clientes, "matches": matches,
                "pendencias": pendencias, "cnpj_dup": len(cnpj_dup_lista),
                "codigo_dup": len(codigo_dup_lista),
                "horas_por_setor": {
                    "fiscal": round(soma_fiscal * 24, 2),
                    "contabil": round(soma_contabil * 24, 2),
                    "dp": round(soma_dp * 24, 2),
                },
                "preenchimento_setor": {"fiscal": pres_fiscal, "contabil": pres_contabil, "dp": pres_dp},
                "top_clientes": top_clientes,
                "distribuicao": [
                    {"faixa": "0h", "qtd": buckets[0]},
                    {"faixa": "0–2h", "qtd": buckets[1]},
                    {"faixa": "2–5h", "qtd": buckets[2]},
                    {"faixa": "5–10h", "qtd": buckets[3]},
                    {"faixa": "10–20h", "qtd": buckets[4]},
                    {"faixa": "20h+", "qtd": buckets[5]},
                ],
                "drill": {
                    "pendencias": pendentes_buf[:200],
                    "cnpj_dup": cnpj_dup_lista[:100],
                    "codigo_dup": codigo_dup_lista[:100],
                },
                "master_path": master_path,
            }
        except Exception as e:
            log.error(f"[DASHBOARD] {e}")
            vazio["msg"] = str(e)
            return vazio

    def obter_variacoes_mes_a_mes(self, competencia=None, top_n=8):
        master_path = self._master_path()
        vazio = {"ok": False, "top_altas": [], "top_quedas": [], "aba_atual": None, "aba_anterior": None}
        if not master_path or not os.path.exists(master_path):
            vazio["msg"] = "Master não encontrada"
            return vazio
        if not competencia or len(competencia) < 7:
            vazio["msg"] = "Competência inválida"
            return vazio

        ano_a, mes_a = int(competencia[:4]), int(competencia[5:7])
        mes_p = mes_a - 1
        ano_p = ano_a
        if mes_p == 0:
            mes_p, ano_p = 12, ano_a - 1
        aba_atual = f"{mes_a:02d}.{ano_a}"
        aba_anterior = f"{mes_p:02d}.{ano_p}"

        try:
            import openpyxl
            wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
            if aba_atual not in wb.sheetnames or aba_anterior not in wb.sheetnames:
                wb.close()
                vazio["msg"] = f"Aba ausente ({aba_atual} ou {aba_anterior})"
                vazio["aba_atual"] = aba_atual if aba_atual in wb.sheetnames else None
                vazio["aba_anterior"] = aba_anterior if aba_anterior in wb.sheetnames else None
                return vazio

            def _ler_aba(sheet):
                out = {}
                for row in sheet.iter_rows(min_row=10, values_only=True):
                    if not row or len(row) < 18:
                        continue
                    cod = row[7]
                    if cod is None:
                        continue
                    try:
                        cod_str = str(int(float(str(cod).strip())))
                    except (ValueError, TypeError):
                        continue
                    nome = row[10] if row[10] not in (None, "") else row[8]
                    total = row[17]
                    if isinstance(total, timedelta):
                        total_dia = total.total_seconds() / 86400.0
                    elif isinstance(total, (int, float)) and not isinstance(total, bool):
                        total_dia = float(total)
                    else:
                        partes = []
                        for idx in (14, 15, 16):
                            v = row[idx]
                            if isinstance(v, timedelta):
                                partes.append(v.total_seconds() / 86400.0)
                            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                                partes.append(float(v))
                        total_dia = sum(partes) if partes else 0.0
                    out[cod_str] = (str(nome or ""), total_dia)
                return out

            atual = _ler_aba(wb[aba_atual])
            anterior = _ler_aba(wb[aba_anterior])
            wb.close()

            comparacoes = []
            for cod, (nome, t_atual) in atual.items():
                t_ant = anterior.get(cod, (nome, 0.0))[1]
                delta_dia = t_atual - t_ant
                comparacoes.append({
                    "codigo": cod, "nome": nome[:40],
                    "horas_atual": round(t_atual * 24, 2),
                    "horas_anterior": round(t_ant * 24, 2),
                    "delta_horas": round(delta_dia * 24, 2),
                    "delta_pct": round((delta_dia / t_ant * 100) if t_ant > 0.001 else 0, 1),
                })

            altas = sorted([c for c in comparacoes if c["delta_horas"] > 0.1],
                           key=lambda c: c["delta_horas"], reverse=True)[:top_n]
            quedas = sorted([c for c in comparacoes if c["delta_horas"] < -0.1],
                            key=lambda c: c["delta_horas"])[:top_n]
            return {"ok": True, "aba_atual": aba_atual, "aba_anterior": aba_anterior,
                    "top_altas": altas, "top_quedas": quedas}
        except Exception as e:
            log.error(f"[VARIACOES] {e}")
            vazio["msg"] = str(e)
            return vazio

    def listar_competencias_master(self):
        from engine.onedrive_helper import esta_online_only, forcar_download
        master_path = self._master_path()
        if not master_path or not os.path.exists(master_path):
            return {"ok": False, "msg": "Master não encontrada", "competencias": []}
        if esta_online_only(master_path):
            forcar_download(master_path, timeout=20)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
            padrao = re.compile(r"^(\d{2})\.(\d{4})$")
            result = []
            for aba in wb.sheetnames:
                m = padrao.match(aba)
                if m:
                    mes, ano = m.group(1), m.group(2)
                    result.append({"aba": aba, "value": f"{ano}-{mes}", "label": f"{mes}/{ano}"})
            wb.close()
            result.sort(key=lambda x: x["value"], reverse=True)
            return {"ok": True, "competencias": result}
        except Exception as e:
            log.error(f"[COMPETENCIAS] {e}")
            return {"ok": False, "msg": str(e), "competencias": []}

    # ── Diagnóstico Contábil ──────────────────────────────────────────────────

    def obter_status_contabil(self):
        from engine.onedrive_helper import esta_online_only, forcar_download
        cfg = self._config.load()
        candidatos = [
            cfg.get("contabil_origem_path"),
            os.path.join(self._project_root, "ENTRADAS_MANUAIS", "HORAS CONTABEIS.xlsx"),
        ]
        caminho = next((p for p in candidatos if p and os.path.exists(p)), candidatos[0])
        info = {
            "caminho": caminho, "configurado": cfg.get("contabil_origem_path"),
            "existe": bool(caminho and os.path.exists(caminho)),
            "tamanho_kb": None, "modificado_em": None,
            "linhas": None, "abas": [], "online_only": False,
        }
        if info["existe"]:
            try:
                if esta_online_only(caminho):
                    info["online_only"] = True
                    forcar_download(caminho, timeout=20)
                st = os.stat(caminho)
                info["tamanho_kb"] = round(st.st_size / 1024, 1)
                info["modificado_em"] = datetime.fromtimestamp(st.st_mtime).strftime("%d/%m/%Y %H:%M")
                import openpyxl
                wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
                info["abas"] = wb.sheetnames
                info["linhas"] = wb.active.max_row
                wb.close()
            except Exception as e:
                log.error(f"[CONTABIL_STATUS] {e}")
                info["erro"] = str(e)
        return info

    # ── Relatórios ────────────────────────────────────────────────────────────

    def listar_relatorios_md(self):
        padroes = ("relatorio_contabil_", "log_execucao", "relatorio_dp_", "relatorio_fiscal_")
        encontrados = []
        try:
            for nome in os.listdir(self._project_root):
                if not nome.lower().endswith(".md"):
                    continue
                if not any(nome.startswith(p) for p in padroes):
                    continue
                caminho = os.path.join(self._project_root, nome)
                try:
                    st = os.stat(caminho)
                    encontrados.append({
                        "nome": nome, "caminho": caminho,
                        "tamanho_kb": round(st.st_size / 1024, 1),
                        "modificado_em": datetime.fromtimestamp(st.st_mtime).strftime("%d/%m/%Y %H:%M"),
                    })
                except Exception:
                    pass
        except Exception as e:
            log.error(f"[RELATORIOS] {e}")
        encontrados.sort(key=lambda d: d["modificado_em"], reverse=True)
        return {"relatorios": encontrados}

    def ler_relatorio_md(self, nome):
        if not nome or ".." in nome or "/" in nome or "\\" in nome or not nome.lower().endswith(".md"):
            return {"ok": False, "erro": "Nome inválido."}
        caminho = os.path.join(self._project_root, nome)
        if not os.path.exists(caminho):
            return {"ok": False, "erro": "Arquivo não encontrado."}
        try:
            with open(caminho, encoding="utf-8") as f:
                return {"ok": True, "conteudo": f.read(), "nome": nome}
        except Exception as e:
            return {"ok": False, "erro": str(e)}

    # ── Log ───────────────────────────────────────────────────────────────────

    def ler_log(self):
        log_path = os.path.join(self._base_dir, "dmf_engine.log")
        if os.path.exists(log_path):
            with open(log_path, encoding="utf-8") as f:
                return {"conteudo": "".join(f.readlines()[-100:])}
        return {"conteudo": "Nenhuma execução registrada."}

    def ler_log_erros(self, limite=40):
        log_path = os.path.join(self._base_dir, "dmf_engine_errors.log")
        if not os.path.exists(log_path):
            return {"eventos": [], "conteudo": ""}
        try:
            with open(log_path, encoding="utf-8") as f:
                linhas = f.readlines()
        except Exception as e:
            return {"eventos": [], "erro": str(e)}
        padrao = re.compile(r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ \[(WARNING|ERROR|CRITICAL)\]")
        eventos = []
        atual = None
        for linha in linhas:
            m = padrao.match(linha)
            if m:
                if atual is not None:
                    eventos.append(atual)
                atual = {"timestamp": m.group(1), "nivel": m.group(2), "linha": linha.rstrip("\n")}
            elif atual is not None:
                atual["linha"] += "\n" + linha.rstrip("\n")
        if atual is not None:
            eventos.append(atual)
        return {"eventos": eventos[-limite:], "total_arquivo": len(eventos)}

    def limpar_log_erros(self):
        log_path = os.path.join(self._base_dir, "dmf_engine_errors.log")
        try:
            open(log_path, "w", encoding="utf-8").close()
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "erro": str(e)}

    def obter_ultima_execucao(self):
        log_path = os.path.join(self._base_dir, "dmf_engine.log")
        if not os.path.exists(log_path):
            return {"eventos": []}
        try:
            with open(log_path, encoding="utf-8") as f:
                linhas = f.readlines()[-400:]
            padrao = re.compile(
                r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}).+?\[(FISCAL|DP|CONTABIL|CONTÁBIL|MASTER|CICLO|METRICAS)\]\s*(.+)"
            )
            eventos = []
            for ln in linhas:
                m = padrao.search(ln)
                if not m:
                    continue
                ts, modulo, msg = m.groups()
                eventos.append({
                    "hora": ts[11:16], "data": ts[:10],
                    "modulo": modulo.replace("Á", "A").title(),
                    "msg": msg.strip(),
                })
            return {"eventos": eventos[-6:]}
        except Exception as e:
            log.error(f"[ULTIMA_EXEC] {e}")
            return {"eventos": [], "erro": str(e)}

    def obter_estado_modulos(self):
        estado_path = os.path.join(self._base_dir, "ultimo_ciclo.json")
        if not os.path.exists(estado_path):
            return {"modulos": {}, "timestamp": None, "competencia": None}
        try:
            with open(estado_path, encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            log.error(f"[ESTADO] {e}")
            return {"modulos": {}, "timestamp": None, "competencia": None}

    # ── Buscar XML (Padrão A) ─────────────────────────────────────────────────

    def _bx_svc(self):
        """Retorna o BuscarXMLService via adaptador do registry."""
        mod = self._registry.get("buscar_xml")
        if mod is None:
            raise RuntimeError("Módulo buscar_xml não registrado.")
        return mod._get_service()

    def bx_listar_lojas(self) -> list:
        try:
            return self._bx_svc().listar_lojas()
        except Exception as e:
            log.error(f"[BX] bx_listar_lojas: {e}")
            return []

    def bx_listar_clientes_tokai(self) -> list:
        try:
            return self._bx_svc().listar_clientes_tokai()
        except Exception as e:
            log.error(f"[BX] bx_listar_clientes_tokai: {e}")
            return []

    def bx_executar_nfe(self, lojas: list, periodo: dict) -> dict:
        try:
            return self._bx_svc().executar_nfe(lojas, periodo)
        except Exception as e:
            log.error(f"[BX] bx_executar_nfe: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_executar_nfce(self, lojas: list, periodo: dict) -> dict:
        try:
            return self._bx_svc().executar_nfce(lojas, periodo)
        except Exception as e:
            log.error(f"[BX] bx_executar_nfce: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_executar_sped(self, lojas: list, periodo: dict) -> dict:
        try:
            return self._bx_svc().executar_sped(lojas, periodo)
        except Exception as e:
            log.error(f"[BX] bx_executar_sped: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_executar_tokai(self, clientes: list = None, periodo: dict = None) -> dict:
        try:
            return self._bx_svc().executar_tokai(clientes=clientes, periodo=periodo)
        except Exception as e:
            log.error(f"[BX] bx_executar_tokai: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_compactar_tokai(self, clientes: list = None, periodo: dict = None) -> dict:
        try:
            return self._bx_svc().compactar_tokai(clientes=clientes, periodo=periodo)
        except Exception as e:
            log.error(f"[BX] bx_compactar_tokai: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_previa_deletar_competencia(self, grupo: str, competencia: str) -> dict:
        try:
            return self._bx_svc().previa_deletar_competencia(grupo, competencia)
        except Exception as e:
            log.error(f"[BX] bx_previa_deletar_competencia: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_deletar_competencia(self, grupo: str, competencia: str,
                               confirmacao: str = "", caminhos: list = None) -> dict:
        try:
            return self._bx_svc().deletar_competencia(grupo, competencia, confirmacao, caminhos)
        except Exception as e:
            log.error(f"[BX] bx_deletar_competencia: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_cancelar(self, modulo: str) -> dict:
        try:
            return self._bx_svc().cancelar(modulo)
        except Exception as e:
            log.error(f"[BX] bx_cancelar: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_get_status(self) -> dict:
        try:
            return self._bx_svc().get_status()
        except Exception as e:
            log.error(f"[BX] bx_get_status: {e}")
            return {"executando": [], "daemons": {}}

    def bx_get_status_daemons(self) -> dict:
        try:
            return self._bx_svc().get_status_daemons()
        except Exception as e:
            log.error(f"[BX] bx_get_status_daemons: {e}")
            return {}

    def bx_start_daemon(self, nome: str) -> dict:
        try:
            return self._bx_svc().start_daemon(nome)
        except Exception as e:
            log.error(f"[BX] bx_start_daemon: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_stop_daemon(self, nome: str) -> dict:
        try:
            return self._bx_svc().stop_daemon(nome)
        except Exception as e:
            log.error(f"[BX] bx_stop_daemon: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_carregar_config(self) -> dict:
        try:
            return self._bx_svc().carregar_config()
        except Exception as e:
            log.error(f"[BX] bx_carregar_config: {e}")
            return {}

    def bx_salvar_config(self, dados: dict) -> dict:
        try:
            return self._bx_svc().salvar_config(dados)
        except Exception as e:
            log.error(f"[BX] bx_salvar_config: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_testar_conexao_erp(self) -> dict:
        try:
            return self._bx_svc().testar_conexao_erp()
        except Exception as e:
            log.error(f"[BX] bx_testar_conexao_erp: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_verificar_disco_z(self) -> dict:
        try:
            return self._bx_svc().verificar_disco_z()
        except Exception as e:
            log.error(f"[BX] bx_verificar_disco_z: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_mapear_disco_z(self) -> dict:
        try:
            return self._bx_svc().mapear_disco_z()
        except Exception as e:
            log.error(f"[BX] bx_mapear_disco_z: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_get_logs(self, modulo: str = "", limite: int = 200) -> list:
        try:
            return self._bx_svc().get_logs(modulo, limite)
        except Exception as e:
            log.error(f"[BX] bx_get_logs: {e}")
            return []

    def bx_limpar_logs(self) -> dict:
        try:
            return self._bx_svc().limpar_logs()
        except Exception as e:
            log.error(f"[BX] bx_limpar_logs: {e}")
            return {"ok": False, "erro": str(e)}

    def bx_get_historico(self) -> list:
        try:
            return self._bx_svc().get_historico()
        except Exception as e:
            log.error(f"[BX] bx_get_historico: {e}")
            return []

    # ── Sem Movimento NFS-e Salvador ─────────────────────────────────────────

    def _sm_svc(self):
        mod = self._registry.get("sem_movimento_nfse")
        if mod is None:
            raise RuntimeError("Módulo sem_movimento_nfse não registrado.")
        return mod._get_service()

    def sm_abrir_template(self, tipo: str) -> dict:
        """Abre o template TXT ou XLSX de exemplo no aplicativo padrão do SO."""
        import subprocess
        nome = f"template_sem_movimento.{'txt' if tipo == 'txt' else 'xlsx'}"
        caminho = os.path.join(self._base_dir, "ui", nome)
        if not os.path.exists(caminho):
            return {"ok": False, "erro": f"Template não encontrado: {caminho}"}
        try:
            os.startfile(caminho)
            return {"ok": True}
        except Exception as e:
            try:
                subprocess.Popen(["explorer", caminho])
                return {"ok": True}
            except Exception:
                return {"ok": False, "erro": str(e)}

    def sm_selecionar_planilha(self) -> dict:
        """Abre diálogo para selecionar TXT ou xlsx de empresas SM."""
        import webview
        tipos = ("Planilha TXT Excel (*.txt;*.xlsx;*.xlsm)", "Todos os arquivos (*.*)")
        result = self._win().create_file_dialog(
            webview.OPEN_DIALOG, allow_multiple=False, file_types=tipos
        )
        if result and result[0]:
            return {"ok": True, "caminho": result[0]}
        return {"ok": False, "caminho": None}

    def sm_carregar_planilha(self, caminho: str) -> dict:
        """Parse TXT/xlsx com CNPJ+senha. Retorna preview com senha mascarada."""
        try:
            self._sm_svc()  # garante que o path do serviço já foi injetado
            from sm_planilha import carregar  # noqa: E402
            resultado = carregar(caminho)
            # Mascara senhas antes de enviar ao JS
            preview = [
                {"cnpj": e["cnpj"], "senha": "****", "linha": e["linha"]}
                for e in resultado["empresas"]
            ]
            return {
                "ok":       True,
                "empresas": preview,
                "total":    len(preview),
                "invalidas": resultado["invalidas"],
            }
        except Exception as e:
            log.error(f"[SM] sm_carregar_planilha: {e}")
            return {"ok": False, "erro": str(e)}

    def sm_executar(self, empresas_com_senha: list, mes: int, ano: int,
                    pasta_destino: str) -> dict:
        """
        Inicia o lote. empresas_com_senha vem do JS com senha real
        (o frontend guarda a lista original carregada em memória).
        """
        try:
            if not os.path.isdir(pasta_destino):
                return {"ok": False, "erro": f"Pasta não encontrada: {pasta_destino}"}
            return self._sm_svc().executar(empresas_com_senha, mes, ano, pasta_destino)
        except Exception as e:
            log.error(f"[SM] sm_executar: {e}")
            return {"ok": False, "erro": str(e)}

    def sm_executar_por_caminho(self, caminho: str, mes: int, ano: int,
                               pasta_destino: str) -> dict:
        """Recarrega a planilha com senhas reais e inicia o lote."""
        try:
            if not os.path.isdir(pasta_destino):
                return {"ok": False, "erro": f"Pasta não encontrada: {pasta_destino}"}
            self._sm_svc()  # garante path injetado
            from sm_planilha import carregar  # noqa: E402
            resultado = carregar(caminho)
            if not resultado["empresas"]:
                return {"ok": False, "erro": "Nenhuma empresa válida na planilha."}
            return self._sm_svc().executar(resultado["empresas"], mes, ano, pasta_destino)
        except Exception as e:
            log.error(f"[SM] sm_executar_por_caminho: {e}")
            return {"ok": False, "erro": str(e)}

    def sm_cancelar(self) -> dict:
        try:
            return self._sm_svc().cancelar()
        except Exception as e:
            log.error(f"[SM] sm_cancelar: {e}")
            return {"ok": False, "erro": str(e)}

    def sm_get_status(self) -> dict:
        try:
            return {"ok": True, **self._sm_svc().get_status()}
        except Exception as e:
            return {"ok": False, "erro": str(e)}

    def sm_carregar_config(self) -> dict:
        try:
            cfg = self._config.load()
            return {
                "ok":                       True,
                "sm_anticaptcha_api_key":   cfg.get("sm_anticaptcha_api_key", ""),
                "sm_headless":              cfg.get("sm_headless", True),
                "sm_captcha_timeout_s":     cfg.get("sm_captcha_timeout_s", 60),
                "sm_pausa_entre_empresas_s": cfg.get("sm_pausa_entre_empresas_s", 2),
            }
        except Exception as e:
            return {"ok": False, "erro": str(e)}

    def sm_salvar_config(self, dados: dict) -> dict:
        try:
            cfg = self._config.load()
            for chave in ("sm_anticaptcha_api_key", "sm_headless",
                          "sm_captcha_timeout_s", "sm_pausa_entre_empresas_s"):
                if chave in dados:
                    cfg[chave] = dados[chave]
            self._config.save(cfg)
            return {"ok": True}
        except Exception as e:
            log.error(f"[SM] sm_salvar_config: {e}")
            return {"ok": False, "erro": str(e)}

