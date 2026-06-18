"""
dmf_engine/modules/m_relatorio_pagamento.py
Módulo inline (Padrão 0) — Relatório de Pagamento (extrato de conta corrente Domínio).

Organiza o extrato bruto de conta corrente (layout do Grupo Tokai / ERP Domínio) em
duas abas de lançamentos individuais (Entradas / Saídas) + Resumo com os totais e o
cabeçalho do cliente. A lógica de negócio é autocontida neste arquivo (Python puro,
sem dependência de services/) — roda no mesmo processo da Central via ThreadRunner.

Regras de negócio (validadas contra a planilha real e a aba "Planilha1" manual):
- Cada lançamento ocupa 2 linhas no raw: a linha de dados (H=valor parcial, I=débito,
  J=crédito) e a linha seguinte "===> Atv. Financeira: ...".
- Vários lançamentos consecutivos com H preenchido e I/J vazios formam um BLOCO, que
  fecha na linha cujo I ou J vem preenchido com a soma de todos os H do bloco.
- Valor de cada item = sinal_do_bloco × H (preserva estorno: H negativo continua
  negativo). Categoria (Entrada/Saída) vem de onde o bloco fechou: I (débito)→Saída,
  J (crédito)→Entrada. Estorno negativo abate a própria categoria.
- Histórico = Atv.Financeira (sem o prefixo) + " - " + col E + " - " + col F.
- A aba do extrato é sempre a primeira (índice 0); o nome varia por cliente/banco.
"""
from dmf_engine.modules.base import BaseModule, ModuleMeta

PREFIXO_ATV = "===> Atv. Financeira:"
AZUL = "FF0066CC"      # entradas
VERMELHO = "FFCC0000"  # saídas


class RelatorioPagamentoModule(BaseModule):

    @property
    def meta(self) -> ModuleMeta:
        return ModuleMeta(
            id="relatorio_pagamento",
            nome="Relatório de Pagamento",
            desc="Organiza o extrato de conta corrente do Domínio em Entradas/Saídas + Resumo.",
            setor="CONTÁBIL",
            icon="ti-cash-banknote",
            color="#1E7C5A",
            papeis=["admin", "contabil"],
        )

    def execute(self, opcoes: dict) -> dict:
        import os
        import re
        import logging
        import traceback
        from pathlib import Path

        import openpyxl
        from openpyxl.styles import Font

        log = logging.getLogger("RelatorioPagamento")

        # ── helpers locais ────────────────────────────────────────────────────
        def _num(v):
            return isinstance(v, (int, float))

        def _str(v):
            return str(v).strip() if v is not None else ""

        def _extrair_cabecalho(ws):
            """Lê empresa/titular, banco/conta e período das ~15 primeiras linhas.
            O título (TITULAR/BANCO) fica na coluna A; o período pode estar em A ou B."""
            cab = {"empresa": "", "banco": "", "periodo": ""}
            for r in range(1, min(ws.max_row, 16) + 1):
                for col in (1, 2):
                    txt = _str(ws.cell(row=r, column=col).value)
                    if not txt:
                        continue
                    up = txt.upper()
                    if not cab["empresa"] and up.startswith("TITULAR:"):
                        cab["empresa"] = txt.split(":", 1)[1].strip()
                    elif not cab["banco"] and up.startswith("BANCO/CONTA CORRENTE:"):
                        cab["banco"] = txt.split(":", 1)[1].strip()
                    elif not cab["periodo"] and up.startswith("PERÍODO"):
                        # remove o prefixo "Período de " — o rótulo já diz "Período:"
                        cab["periodo"] = re.sub(r"^per[íi]odo\s+de\s+", "", txt.strip(), flags=re.IGNORECASE)
            return cab

        def _higienizar(campo):
            """Remove caracteres proibidos pelo Windows de um campo, preservando
            hífens e pontos válidos (ex.: 'LOTTI.MARINA.BA', '01-12-2025')."""
            campo = re.sub(r'[<>:"/\\|?*]', " ", campo)   # proibidos → espaço
            campo = re.sub(r"\s+", " ", campo).strip()     # colapsa só espaços
            return campo

        def _montar_nome_arquivo(cab, base_fallback):
            """Monta 'Cliente - Banco - DD-MM-AAAA a DD-MM-AAAA' (formato enxuto):
            só o nome do cliente (sem código), nome do banco (sem códigos) e as datas.
            Se faltar dado essencial, cai no nome original da planilha."""
            # Cliente: "1 - LOTTI.MARINA.BA" → "LOTTI.MARINA.BA"
            empresa = cab.get("empresa", "")
            cliente = empresa.split(" - ", 1)[1].strip() if " - " in empresa else empresa.strip()

            # Banco: "7 - SANTANDER - 130098713" → "SANTANDER" (parte sem código)
            partes_banco = [p.strip() for p in cab.get("banco", "").split(" - ") if p.strip()]
            banco = next((p for p in partes_banco if not p.isdigit()), "")

            # Período: "01/12/2025 até 31/12/2025" → "01-12-2025 a 31-12-2025"
            periodo = cab.get("periodo", "")
            periodo = re.sub(r"\s*at[ée]\s*", " a ", periodo, flags=re.IGNORECASE)
            periodo = periodo.replace("/", "-").strip()

            partes = [_higienizar(p) for p in (cliente, banco, periodo) if p]
            if not partes:
                return f"{base_fallback}_organizado"
            return " - ".join(partes)

        def _processar(ws):
            """Percorre a aba e retorna lista de {data_pag, historico, valor, tipo}."""
            max_row = ws.max_row
            lancamentos = []
            bloco = []  # itens pendentes: {data_pag, atv, col_e, col_f, h}

            r = 1
            while r <= max_row:
                a = ws.cell(row=r, column=1).value
                if isinstance(a, str) and a.strip().startswith(PREFIXO_ATV):
                    r += 1
                    continue

                h = ws.cell(row=r, column=8).value
                i = ws.cell(row=r, column=9).value
                j = ws.cell(row=r, column=10).value

                if not _num(h):
                    # cabeçalho/rodapé/linha em branco — pula sem descartar lançamento
                    r += 1
                    continue

                c = ws.cell(row=r, column=3).value
                e = ws.cell(row=r, column=5).value
                f = ws.cell(row=r, column=6).value

                atv_raw = ""
                if r + 1 <= max_row:
                    prox = ws.cell(row=r + 1, column=1).value
                    if isinstance(prox, str) and prox.strip().startswith(PREFIXO_ATV):
                        atv_raw = prox.strip()
                atv = re.sub(r"^===>\s*Atv\.\s*Financeira:\s*", "", atv_raw).strip()

                data_pag = c if c is not None else (bloco[0]["data_pag"] if bloco else None)

                bloco.append({
                    "data_pag": data_pag,
                    "atv": atv,
                    "col_e": _str(e),
                    "col_f": _str(f),
                    "h": h,
                })

                if _num(i) or _num(j):
                    sinal = -1 if _num(i) else 1
                    tipo = "S" if _num(i) else "E"  # débito→Saída, crédito→Entrada
                    for item in bloco:
                        partes = [p for p in (item["atv"], item["col_e"], item["col_f"]) if p]
                        lancamentos.append({
                            "data_pag": item["data_pag"],
                            "historico": " - ".join(partes),
                            # sinal nativo de H preservado (estorno continua negativo)
                            "valor": round(sinal * item["h"], 2),
                            "tipo": tipo,
                        })
                    bloco = []

                r += 2 if atv_raw else 1

            if bloco:
                log.warning("Bloco aberto com %d item(ns) nunca fechou (sem I/J).", len(bloco))
            return lancamentos

        def _montar_aba(ws, lancs, cor, negativar):
            """A=Data, B=Valor, C=Histórico. Valor colorido; se negativar, força <0."""
            ws.append(["Data", "Valor", "Histórico"])
            for cel in ws[1]:
                cel.font = Font(bold=True)
            ws.freeze_panes = "A2"

            fonte_valor = Font(color=cor)
            for lanc in lancs:
                v = -abs(lanc["valor"]) if negativar else lanc["valor"]
                ws.append([lanc["data_pag"], v, lanc["historico"]])
                rr = ws.max_row
                ws.cell(row=rr, column=1).number_format = "dd/mm/yyyy"
                cv = ws.cell(row=rr, column=2)
                cv.number_format = "#,##0.00"
                cv.font = fonte_valor

            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 16
            ws.column_dimensions["C"].width = 90

        def _gerar_saida(lancamentos, cab, caminho_saida):
            wb = openpyxl.Workbook()

            entradas = [l for l in lancamentos if l["tipo"] == "E"]
            saidas = [l for l in lancamentos if l["tipo"] == "S"]

            ws_ent = wb.active
            ws_ent.title = "Entradas"
            _montar_aba(ws_ent, entradas, AZUL, negativar=False)

            ws_sai = wb.create_sheet("Saídas")
            _montar_aba(ws_sai, saidas, VERMELHO, negativar=True)

            total_entradas = round(sum(l["valor"] for l in entradas), 2)
            total_saidas = round(sum(-abs(l["valor"]) for l in saidas), 2)
            saldo = round(total_entradas + total_saidas, 2)

            ws2 = wb.create_sheet("Resumo")
            # Cabeçalho do cliente no topo
            negrito = Font(bold=True)
            cinza = Font(color="FF857060")
            for rotulo, valor in (
                ("Cliente:", cab.get("empresa", "")),
                ("Banco/Conta:", cab.get("banco", "")),
                ("Período:", cab.get("periodo", "")),
            ):
                ws2.append([rotulo, valor])
                ws2.cell(row=ws2.max_row, column=1).font = cinza
            ws2.append([])

            inicio_tot = ws2.max_row + 1
            ws2.append(["Indicador", "Valor"])
            for cel in ws2[inicio_tot]:
                cel.font = negrito
            for rotulo, valor, cor in (
                ("Total de Entradas", total_entradas, AZUL),
                ("Total de Saídas", total_saidas, VERMELHO),
                ("Saldo do Período", saldo, None),
            ):
                ws2.append([rotulo, valor])
                cel = ws2.cell(row=ws2.max_row, column=2)
                cel.number_format = "#,##0.00"
                if cor:
                    cel.font = Font(color=cor)
            ws2.column_dimensions["A"].width = 22
            ws2.column_dimensions["B"].width = 40

            wb.save(caminho_saida)
            return total_entradas, total_saidas, saldo, len(entradas), len(saidas)

        # ── fluxo principal ───────────────────────────────────────────────────
        try:
            arquivo = opcoes.get("arquivo")
            output_dir = opcoes.get("output_dir") or str(Path.home() / "Downloads")

            if not arquivo:
                return {"ok": False, "erro": "Nenhuma planilha selecionada."}
            if not os.path.exists(arquivo):
                return {"ok": False, "erro": "A planilha selecionada não foi encontrada."}

            self.progress(10, "Abrindo planilha...")
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = wb.worksheets[0]

            self.progress(30, "Lendo dados do cliente...")
            cab = _extrair_cabecalho(ws)

            self.progress(45, "Organizando lançamentos...")
            lancamentos = _processar(ws)
            if not lancamentos:
                return {"ok": True, "aviso": "Nenhum lançamento encontrado na planilha.",
                        "entradas": 0, "saidas": 0}

            self.progress(75, f"Gerando planilha ({len(lancamentos)} lançamentos)...")
            base = os.path.splitext(os.path.basename(arquivo))[0]
            nome_saida = _montar_nome_arquivo(cab, base)
            destino = os.path.join(output_dir, f"{nome_saida}.xlsx")
            tot_ent, tot_sai, saldo, n_ent, n_sai = _gerar_saida(lancamentos, cab, destino)

            self.progress(100, "Concluído.")
            return {
                "ok": True,
                "arquivo": destino,
                "entradas": n_ent,
                "saidas": n_sai,
                "total_entradas": tot_ent,
                "total_saidas": tot_sai,
                "saldo": saldo,
                "cliente": cab,
            }

        except PermissionError:
            return {"ok": False, "erro": "A planilha de saída está aberta. Feche-a e tente novamente."}
        except Exception as e:
            log.error(traceback.format_exc())
            return {"ok": False, "erro": str(e)}
