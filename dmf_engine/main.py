"""
DMF Engine — Ponto de entrada PyWebView
Abre janela nativa com o HTML já existente e expõe a bridge Python ↔ JS.
"""
import webview
import json
import os
import re
import sys
import platform
import threading
import logging
from logging.handlers import RotatingFileHandler
import traceback
from datetime import datetime, timedelta
from decimal import Decimal

# Fix para o ícone na barra de tarefas do Windows
if platform.system() == "Windows":
    import ctypes
    try:
        myappid = "dmf.engine.app.1"
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
    except Exception:
        pass

# Resolve diretórios:
#   BASE_DIR      → escrita (logs, config.json, supervisores.json) — ao lado do .exe em frozen
#   RESOURCES_DIR → leitura de assets empacotados (HTML, .ico) — dentro de _internal em frozen
#   PROJECT_ROOT  → raiz para importar engine/ e modulos/
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
    RESOURCES_DIR = os.path.join(sys._MEIPASS, "dmf_engine")
    PROJECT_ROOT = BASE_DIR
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    RESOURCES_DIR = BASE_DIR
    PROJECT_ROOT = os.path.dirname(BASE_DIR)
sys.path.insert(0, PROJECT_ROOT)

from engine.database import db
from modulos.fiscal import extrair_e_preencher_fiscal
from modulos.dp import extrair_e_preencher_dp
from modulos.contabil_preenchedor import preencher_horas_contabeis
from modulos.contabil_integrador import injetar_contabil_na_master
from modulos.excecoes import caminho_excecao, ler_codigos_setor, metadados_setor, PASTA as PASTA_EXCECOES
from engine.master_writer import MasterWriter
from engine.onedrive_helper import esta_online_only, forcar_download
from engine.lock_master import adquirir_lock, liberar_lock, verificar_lock
from engine import estado_compartilhado as estado_sh
from dmf_engine import auth as auth_mod

# B09: RotatingFileHandler em vez de FileHandler — impede crescimento ilimitado.
# 10MB por arquivo × 5 backups = 50MB máximo por handler. Suficiente para meses
# de operação dos 5 supervisores, e ler_log() não fica lento com arquivo gigante.
_log_geral = RotatingFileHandler(
    os.path.join(BASE_DIR, "dmf_engine.log"),
    maxBytes=10 * 1024 * 1024, backupCount=5, encoding="utf-8",
)
_log_geral.setLevel(logging.INFO)
_log_geral.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))

# Handler dedicado a WARNING/ERROR/CRITICAL para facilitar depuração.
_log_erros = RotatingFileHandler(
    os.path.join(BASE_DIR, "dmf_engine_errors.log"),
    maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8",
)
_log_erros.setLevel(logging.WARNING)
_log_erros.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(name)s · %(message)s"))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[_log_geral, _log_erros, logging.StreamHandler()],
)

# Captura exceções não-tratadas (incluindo das threads de execução).
def _logar_excecao_nao_tratada(exc_type, exc_value, exc_tb):
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_tb)
        return
    logging.critical("Exceção não tratada", exc_info=(exc_type, exc_value, exc_tb))

sys.excepthook = _logar_excecao_nao_tratada
try:
    threading.excepthook = lambda args: logging.critical(
        f"Exceção não tratada na thread {args.thread.name}",
        exc_info=(args.exc_type, args.exc_value, args.exc_traceback),
    )
except AttributeError:
    pass

# Guard de arquitetura: o driver ODBC do Sybase é 32-bit. Rodar em 64-bit causa IM014.
PYTHON_BITS = platform.architecture()[0]
if PYTHON_BITS != "32bit":
    logging.warning(
        "[BOOT] Python rodando em %s. O driver ODBC do Domínio (SQL Anywhere) é "
        "32-bit — conexões falharão com IM014. Use 'run.bat' ou rode "
        "'py -3-32 dmf_engine\\main.py'.",
        PYTHON_BITS,
    )

CONFIG_FILE = os.path.join(BASE_DIR, "config.json")

# Referência global à janela (preenchida após create_window)
window = None


# B08: serializador JSON robusto contra tipos não-padrão.
# Usar SEMPRE para resultados que vão para window.evaluate_js (que recebem o JSON
# embutido na string JS executada). datetime/timedelta/Decimal/bytes/set quebravam
# json.dumps puro com TypeError, e o except Exception engolia o erro → callback JS
# nunca era invocada → UI travava em spinner indefinidamente.
def _json_default(o):
    """Converte tipos não-serializáveis em representações JSON-safe."""
    if isinstance(o, datetime):
        return o.isoformat()
    if isinstance(o, timedelta):
        return str(o)  # ex: "0:30:00"
    if isinstance(o, Decimal):
        return float(o)
    if isinstance(o, bytes):
        return o.decode("utf-8", errors="replace")
    if isinstance(o, set):
        return list(o)
    # Último recurso: força string. Garante que NUNCA propaga TypeError.
    return str(o)


def _json_safe(obj, **kwargs):
    """Wrapper sobre json.dumps com default tolerante. Mesma assinatura: passa kwargs."""
    return json.dumps(obj, default=_json_default, ensure_ascii=False, **kwargs)


class Api:
    """
    Cada método público aqui fica disponível no JS como:
        window.pywebview.api.nome_do_metodo(args)
    """

    def __init__(self):
        # Sessão de supervisor: dict {nome, papel, label} ou None.
        self._sessao = None
        # Inicializa lista fixa de supervisores no JSON (idempotente).
        try:
            auth_mod.inicializar_supervisores_padrao(BASE_DIR)
        except Exception as e:
            logging.error(f"[BOOT] Falha ao inicializar supervisores: {e}")
        # Aplica credenciais salvas em config.json ao singleton db logo no boot.
        try:
            cfg = self._ler_config()
            db.configurar(
                dsn=cfg.get("db_dsn"),
                uid=cfg.get("db_uid"),
                pwd=cfg.get("db_pwd") or db.pwd,
                timeout=cfg.get("db_timeout", 5),
            )
        except Exception as e:
            logging.error(f"[BOOT] Falha ao aplicar credenciais do config: {e}")
    # ── Autenticação ────────────────────────────────────────────

    def _autorizado(self, modulo):
        """Verifica se a sessão atual pode executar o módulo informado."""
        if not self._sessao:
            return False
        return auth_mod.pode_executar(self._sessao.get("papel"), modulo)

    def verificar_estado_login(self):
        return {"sessao": self._sessao}

    def abrir_guia_uso(self):
        """Abre o guia de uso no navegador padrão (fora da janela do app)."""
        import webbrowser
        url = ("https://any.coop/AAKr8LpD6zoPuFhRs2EXVjQQEVdcaqM2JW8XPhRpS9JjL4FV/"
               "guia-de-uso-controle-de-horas-dmf")
        try:
            webbrowser.open(url, new=2)
            return {"ok": True}
        except Exception as e:
            logging.error(f"[GUIA] Falha ao abrir navegador: {e}")
            return {"ok": False, "erro": str(e)}

    def login(self, nome, senha):
        r = auth_mod.autenticar(BASE_DIR, nome, senha)
        if r.get("ok") and not r.get("precisa_trocar_senha"):
            self._sessao = {
                "nome": r["nome"], "label": r.get("label"),
                "papel": r.get("papel"), "maquina": r.get("maquina"),
            }
            logging.info(f"[AUTH] Sessão iniciada: {r['nome']} ({r.get('papel')})")
        return r

    def trocar_senha_primeiro_acesso(self, nome, senha_atual, senha_nova):
        r = auth_mod.trocar_senha_primeiro_acesso(BASE_DIR, nome, senha_atual, senha_nova)
        if r.get("ok"):
            self._sessao = {
                "nome": r["nome"], "label": r.get("label"),
                "papel": r.get("papel"), "maquina": r.get("maquina"),
            }
        return r

    def logout(self):
        anterior = self._sessao.get("nome") if self._sessao else None
        self._sessao = None
        logging.info(f"[AUTH] Sessão encerrada: {anterior}")
        return {"ok": True}

    def trocar_senha(self, senha_atual, senha_nova):
        if not self._sessao:
            return {"ok": False, "erro": "Sessão não autenticada."}
        return auth_mod.trocar_senha(BASE_DIR, self._sessao["nome"], senha_atual, senha_nova)

    # — Admin (apenas papel='admin')
    def admin_listar_usuarios(self):
        if not self._sessao or self._sessao.get("papel") != "admin":
            return {"ok": False, "erro": "Apenas o admin pode listar usuários."}
        return {"ok": True, "usuarios": auth_mod.listar_supervisores(BASE_DIR)}

    def admin_liberar_maquina(self, nome):
        if not self._sessao or self._sessao.get("papel") != "admin":
            return {"ok": False, "erro": "Apenas o admin pode liberar máquinas."}
        return auth_mod.liberar_maquina(BASE_DIR, nome)

    def admin_resetar_senha(self, nome):
        if not self._sessao or self._sessao.get("papel") != "admin":
            return {"ok": False, "erro": "Apenas o admin pode resetar senhas."}
        return auth_mod.resetar_senha(BASE_DIR, nome)

    # ── Parâmetros ──────────────────────────────────────────────

    def salvar_parametros(self, params: dict):
        cfg = self._ler_config()
        cfg.update(params)
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2, ensure_ascii=False)
        logging.info(f"[CONFIG] Parâmetros salvos: {list(params.keys())}")
        return {"ok": True}

    def carregar_parametros(self):
        return self._ler_config()

    def _ler_config(self):
        defaults = {
            "fiscal_adicional_pct": 80,
            "dp_fator_carga": 0.33,
            "dp_overhead_fixo": 1.5,
            "dp_apenas_socios": 1.0,
            "dp_tempo_minimo": 5.0,
            "dp_consultoria_horas": 1.5,
            "contabil_origem_path": r"C:\Users\DMF-AUTOMACAO\OneDrive - DMF\DMF - Documentos\Administrativo\HORAS CONTABEIS.xlsx",
            "contabil_aplicar_excecoes": True,
            "db_dsn": "Contabil",
            "db_uid": "EXTERNO",
            "db_timeout": 5,
            "governanca_match_minimo": 2,
            "governanca_bloquear_cnpj_dup": True,
            "governanca_gravar_zero": True,
        }
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, encoding="utf-8") as f:
                    saved = json.load(f)
                defaults.update(saved)
            except Exception:
                pass
        return defaults

    # ── Conexão BD ──────────────────────────────────────────────

    def testar_conexao(self):
        try:
            cfg = self._ler_config()
            db.configurar(
                dsn=cfg.get("db_dsn"),
                uid=cfg.get("db_uid"),
                pwd=cfg.get("db_pwd") or db.pwd,
                timeout=cfg.get("db_timeout", 5),
            )
            ok = db.connect()
            if ok:
                msg = f"Conectado ao Domínio (DSN={db.dsn}, UID={db.uid})"
            else:
                msg = f"Falha ODBC: {db.ultimo_erro or 'erro desconhecido'}"
            self.salvar_parametros({
                "db_ultimo_teste_ok": bool(ok),
                "db_ultimo_teste_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
                "db_ultimo_erro": db.ultimo_erro if not ok else None,
            })
            return {"ok": ok, "msg": msg, "dsn": db.dsn, "uid": db.uid, "erro": db.ultimo_erro}
        except Exception as e:
            self.salvar_parametros({
                "db_ultimo_teste_ok": False,
                "db_ultimo_teste_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
                "db_ultimo_erro": str(e),
            })
            return {"ok": False, "msg": str(e), "erro": str(e)}

    def diagnosticar_odbc(self):
        """Retorna informações do ambiente ODBC para diagnosticar problemas
        de driver, arquitetura (32/64-bit) e DSNs disponíveis."""
        import platform
        info = {
            "python_bits": platform.architecture()[0],
            "python_version": platform.python_version(),
            "drivers": [],
            "dsns_user": {},
            "dsns_system": {},
            "ultimo_erro": None,
        }
        try:
            import pyodbc
            info["drivers"] = list(pyodbc.drivers())
            try: info["dsns_user"] = dict(pyodbc.dataSources())
            except Exception: pass
            cfg = self._ler_config()
            info["dsn_configurado"] = cfg.get("db_dsn")
            info["dsn_existe"] = cfg.get("db_dsn") in info["dsns_user"]
            info["ultimo_erro"] = cfg.get("db_ultimo_erro")
        except ImportError:
            info["erro"] = "pyodbc não instalado neste Python."
        except Exception as e:
            info["erro"] = str(e)
        return info

    def obter_status_geral(self):
        """Resumo leve para o rodapé: master configurada e último teste de BD."""
        cfg = self._ler_config()
        master_path = cfg.get("master_path") or os.path.join(PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm")
        return {
            "master_path": master_path if os.path.exists(master_path) else None,
            "master_nome": os.path.basename(master_path) if os.path.exists(master_path) else None,
            "db_dsn": cfg.get("db_dsn"),
            "db_ultimo_teste_ok": cfg.get("db_ultimo_teste_ok"),
            "db_ultimo_teste_em": cfg.get("db_ultimo_teste_em"),
        }

    # ── Fluxo Contábil em 2 etapas (Spec_Contabil.md) ───────────

    def selecionar_horas_contabeis(self):
        """File dialog específico para a planilha HORAS CONTABEIS."""
        if not self._autorizado("contabil"):
            return {"ok": False, "erro": "Apenas James (contábil) ou Carol (admin) podem operar este módulo."}
        tipos = ["Planilha Contábil (*.xlsx)", "Todos os arquivos (*.*)"]
        result = window.create_file_dialog(
            webview.OPEN_DIALOG, allow_multiple=False, file_types=tipos
        )
        if result and result[0]:
            caminho = result[0]
            self.salvar_parametros({"contabil_origem_path": caminho})
            return {"ok": True, "caminho": caminho}
        return {"ok": False}

    def processar_horas_contabeis(self, opcoes=None):
        """Fase 2: escreve F/I/O na HORAS CONTABEIS.xlsx (com validação tripla)."""
        if not self._autorizado("contabil"):
            return {"ok": False, "erro": "Apenas James (contábil) ou Carol (admin) podem executar este módulo."}
        cfg = self._ler_config()
        opcoes = opcoes or {}
        caminho = opcoes.get("caminho") or cfg.get("contabil_origem_path")
        if not caminho or not os.path.exists(caminho):
            return {"ok": False, "erro": "Selecione antes a planilha HORAS CONTABEIS.xlsx."}

        data_inicio = opcoes.get("data_inicio")
        data_fim = opcoes.get("data_fim")
        if not data_inicio or not data_fim:
            return {"ok": False, "erro": "Competência ausente."}

        def _run():
            def progresso(pct, label):
                if window:
                    window.evaluate_js(f"window.atualizarProgresso({pct}, '{label.replace(chr(39), chr(92)+chr(39))}')")
            try:
                progresso(10, "Conectando ao Domínio...")
                if not db.connect():
                    if window:
                        window.evaluate_js(
                            "window.contabilFase2Concluido("
                            + _json_safe({"ok": False, "erro": "Falha na conexão ODBC com o Domínio."})
                            + ")"
                        )
                    return
                progresso(30, "Consultando lançamentos / faturamento / folha...")
                resultado = preencher_horas_contabeis(caminho, data_inicio, data_fim, raiz_relatorios=PROJECT_ROOT)
                progresso(95, "Finalizando...")
                if resultado.get("ok"):
                    competencia = data_inicio[:7]
                    cfg_atual = self._ler_config()
                    master_path_estado = cfg_atual.get("master_path") or os.path.join(
                        PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm"
                    )
                    usuario_atual, host_atual = self._id_label_host()
                    estado_sh.marcar(master_path_estado, "contabil", competencia,
                                     "processado", por=usuario_atual, host=host_atual)
                    estado_sh.remover(master_path_estado, "contabil", competencia, evento="lancado")
                    # Mantém também no config local por compatibilidade com instalações em transição
                    self.salvar_parametros({
                        f"contabil_processado_{competencia}": True,
                        f"contabil_processado_em_{competencia}": datetime.now().strftime("%d/%m/%Y %H:%M"),
                        f"contabil_lancado_{competencia}": False,
                    })
                progresso(100, "Concluído.")
                if window:
                    window.evaluate_js(
                        "window.contabilFase2Concluido(" + _json_safe(resultado) + ")"
                    )
            except Exception as e:
                logging.error(f"[CONTÁBIL-FASE2] Erro: {traceback.format_exc()}")
                if window:
                    window.evaluate_js(
                        "window.contabilFase2Concluido("
                        + _json_safe({"ok": False, "erro": str(e)})
                        + ")"
                    )

        threading.Thread(target=_run, daemon=True).start()
        return {"ok": True, "msg": "Processamento iniciado em background."}

    def injetar_horas_contabeis_master(self, opcoes=None):
        """Fase 5: lê coluna R (horas validadas) da HORAS CONTABEIS e lança na master."""
        if not self._autorizado("contabil"):
            return {"ok": False, "erro": "Apenas James (contábil) ou Carol (admin) podem executar este módulo."}
        cfg = self._ler_config()
        opcoes = opcoes or {}
        caminho = opcoes.get("caminho") or cfg.get("contabil_origem_path")
        master_path = opcoes.get("master_path") or cfg.get("master_path") or os.path.join(
            PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm"
        )
        data_inicio = opcoes.get("data_inicio")
        data_fim = opcoes.get("data_fim")
        if not all([caminho, data_inicio, data_fim]):
            return {"ok": False, "erro": "Parâmetros incompletos (caminho/competência)."}
        if not os.path.exists(caminho):
            return {"ok": False, "erro": "Planilha HORAS CONTABEIS não encontrada."}
        if not os.path.exists(master_path):
            return {"ok": False, "erro": "Planilha master não encontrada."}

        usuario, host = self._id_usuario_host()
        ok_lock, info_lock = adquirir_lock(master_path, usuario, host, "Contábil (lançamento)")
        if not ok_lock:
            return self._erro_lock_alheio(info_lock)

        def _run():
            def progresso(pct, label):
                if window:
                    window.evaluate_js(f"window.atualizarProgresso({pct}, '{label.replace(chr(39), chr(92)+chr(39))}')")
            try:
                lockfile = os.path.join(
                    os.path.dirname(master_path),
                    "~$" + os.path.basename(master_path),
                )
                if os.path.exists(lockfile):
                    if window:
                        window.evaluate_js(
                            "window.contabilFase5Concluido("
                            + _json_safe({"ok": False, "erro": "Master está aberta no Excel. Feche o arquivo.", "tipo": "locked"})
                            + ")"
                        )
                    return

                progresso(20, "Abrindo planilha master...")
                writer = MasterWriter(master_path)
                if not writer.carregar():
                    if window:
                        window.evaluate_js(
                            "window.contabilFase5Concluido("
                            + _json_safe({"ok": False, "erro": "Falha ao abrir master."})
                            + ")"
                        )
                    return

                progresso(50, "Lendo horas validadas e lançando na master...")
                resultado = injetar_contabil_na_master(writer, caminho, data_inicio, data_fim)

                progresso(80, "Recalculando totais...")
                writer.recalcular_totais()

                progresso(95, "Salvando master...")
                salvo = writer.salvar()
                if not salvo and writer.ultimo_erro:
                    if window:
                        window.evaluate_js(
                            "window.contabilFase5Concluido("
                            + _json_safe({"ok": False, **writer.ultimo_erro})
                            + ")"
                        )
                    return

                if resultado.get("ok"):
                    competencia = data_inicio[:7]
                    usuario_atual, host_atual = self._id_label_host()
                    estado_sh.marcar(master_path, "contabil", competencia,
                                     "lancado", por=usuario_atual, host=host_atual)
                    self.salvar_parametros({
                        f"contabil_lancado_{competencia}": True,
                        f"contabil_lancado_em_{competencia}": datetime.now().strftime("%d/%m/%Y %H:%M"),
                    })
                progresso(100, "Concluído.")
                if window:
                    window.evaluate_js(
                        "window.contabilFase5Concluido(" + _json_safe(resultado) + ")"
                    )
            except Exception as e:
                logging.error(f"[CONTÁBIL-FASE5] Erro: {traceback.format_exc()}")
                if window:
                    window.evaluate_js(
                        "window.contabilFase5Concluido("
                        + _json_safe({"ok": False, "erro": str(e)})
                        + ")"
                    )
            finally:
                liberar_lock(master_path, usuario, host)

        threading.Thread(target=_run, daemon=True).start()
        return {"ok": True, "msg": "Lançamento iniciado em background."}

    def obter_estado_contabil(self, data_inicio):
        """Retorna estado persistente da competência (processado / lançado).
        Lê o estado compartilhado no OneDrive (visível para todos) e cai no
        config local como fallback de compatibilidade."""
        cfg = self._ler_config()
        comp = (data_inicio or "")[:7]
        master_path = cfg.get("master_path") or os.path.join(PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm")
        sh = estado_sh.obter(master_path, "contabil", comp)
        return {
            "competencia": comp,
            "processado": bool(sh.get("processado")) or bool(cfg.get(f"contabil_processado_{comp}")),
            "processado_em": (sh.get("processado") or {}).get("em") or cfg.get(f"contabil_processado_em_{comp}"),
            "processado_por": (sh.get("processado") or {}).get("por"),
            "lancado": bool(sh.get("lancado")) or bool(cfg.get(f"contabil_lancado_{comp}")),
            "lancado_em": (sh.get("lancado") or {}).get("em") or cfg.get(f"contabil_lancado_em_{comp}"),
            "lancado_por": (sh.get("lancado") or {}).get("por"),
            "caminho": cfg.get("contabil_origem_path"),
        }

    # ── Variações mês a mês ────────────────────────────────────

    def obter_variacoes_mes_a_mes(self, competencia=None, top_n=8):
        """Compara coluna R (total) entre a aba da competência e a aba do mês anterior.
        Retorna top N clientes com maior aumento e maior queda em horas."""
        cfg = self._ler_config()
        master_path = (
            cfg.get("master_path")
            or os.path.join(PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm")
        )
        vazio = {"ok": False, "top_altas": [], "top_quedas": [],
                 "aba_atual": None, "aba_anterior": None}
        if not master_path or not os.path.exists(master_path):
            vazio["msg"] = "Master não encontrada"
            return vazio
        if not competencia or len(competencia) < 7:
            vazio["msg"] = "Competência inválida"
            return vazio

        ano_a, mes_a = int(competencia[:4]), int(competencia[5:7])
        mes_p = mes_a - 1
        ano_p = ano_a
        if mes_p == 0:
            mes_p, ano_p = 12, ano_a - 1
        aba_atual = f"{mes_a:02d}.{ano_a}"
        aba_anterior = f"{mes_p:02d}.{ano_p}"

        try:
            import openpyxl
            wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
            if aba_atual not in wb.sheetnames or aba_anterior not in wb.sheetnames:
                wb.close()
                vazio["msg"] = f"Aba ausente ({aba_atual} ou {aba_anterior})"
                vazio["aba_atual"] = aba_atual if aba_atual in wb.sheetnames else None
                vazio["aba_anterior"] = aba_anterior if aba_anterior in wb.sheetnames else None
                return vazio

            def _ler_aba(sheet):
                """código → (nome, horas_total_dia)."""
                out = {}
                for row in sheet.iter_rows(min_row=10, values_only=True):
                    if not row or len(row) < 18:
                        continue
                    cod = row[7]
                    if cod is None:
                        continue
                    try:
                        cod_str = str(int(float(str(cod).strip())))
                    except (ValueError, TypeError):
                        continue
                    nome = row[10] if row[10] not in (None, "") else row[8]
                    total = row[17]
                    if isinstance(total, timedelta):
                        total_dia = total.total_seconds() / 86400.0
                    elif isinstance(total, (int, float)) and not isinstance(total, bool):
                        total_dia = float(total)
                    else:
                        # soma O+P+Q como fallback
                        partes = []
                        for idx in (14, 15, 16):
                            v = row[idx]
                            if isinstance(v, timedelta):
                                partes.append(v.total_seconds() / 86400.0)
                            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                                partes.append(float(v))
                        total_dia = sum(partes) if partes else 0.0
                    out[cod_str] = (str(nome or ""), total_dia)
                return out

            atual = _ler_aba(wb[aba_atual])
            anterior = _ler_aba(wb[aba_anterior])
            wb.close()

            comparacoes = []
            for cod, (nome, t_atual) in atual.items():
                t_ant = anterior.get(cod, (nome, 0.0))[1]
                delta_dia = t_atual - t_ant
                comparacoes.append({
                    "codigo": cod,
                    "nome": nome[:40],
                    "horas_atual": round(t_atual * 24.0, 2),
                    "horas_anterior": round(t_ant * 24.0, 2),
                    "delta_horas": round(delta_dia * 24.0, 2),
                    "delta_pct": round((delta_dia / t_ant * 100.0) if t_ant > 0.001 else 0.0, 1),
                })

            altas = sorted([c for c in comparacoes if c["delta_horas"] > 0.1],
                           key=lambda c: c["delta_horas"], reverse=True)[:top_n]
            quedas = sorted([c for c in comparacoes if c["delta_horas"] < -0.1],
                            key=lambda c: c["delta_horas"])[:top_n]

            return {
                "ok": True,
                "aba_atual": aba_atual,
                "aba_anterior": aba_anterior,
                "top_altas": altas,
                "top_quedas": quedas,
            }
        except Exception as e:
            logging.error(f"[VARIACOES] {e}")
            vazio["msg"] = str(e)
            return vazio

    # ── Execução individual Fiscal ──────────────────────────────

    def _id_usuario_host(self):
        """Retorna (usuario, host) da sessão atual para uso no lock."""
        if not self._sessao:
            return None, None
        return self._sessao.get("nome"), self._sessao.get("maquina") or auth_mod.identificar_maquina()

    def _id_label_host(self):
        """B07: snapshot atômico de (label, host) — uso em estado_sh.marcar dentro
        de threads daemon. Evita TOCTOU se o usuário fizer logout durante a operação
        (que zera self._sessao mid-flight, levando self._sessao.get(...) a AttributeError
        ou label/host divergentes entre duas linhas consecutivas)."""
        s = dict(self._sessao) if self._sessao else {}
        label = s.get("label") or s.get("nome") or "?"
        host = s.get("maquina") or auth_mod.identificar_maquina()
        return label, host

    def _erro_lock_alheio(self, info_lock):
        """Monta mensagem padronizada quando o lock pertence a outro usuário."""
        return {
            "ok": False, "tipo": "locked_concorrente",
            "erro": (f"A planilha master está sendo usada por {info_lock.get('usuario', '?')} "
                     f"({info_lock.get('modulo', '?')}) desde {info_lock.get('iniciado_em', '?')}. "
                     f"Aguarde ele(a) terminar (máx. 5 min)."),
            "lock": info_lock,
        }

    def obter_lock_master(self):
        """Para a UI consultar quem (se alguém) está com a master agora."""
        cfg = self._ler_config()
        master_path = cfg.get("master_path") or os.path.join(
            PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm"
        )
        if not os.path.exists(master_path):
            return {"lock": None}
        return {"lock": verificar_lock(master_path)}

    def executar_fiscal_individual(self, opcoes=None):
        """Roda apenas o módulo Fiscal na janela informada (ou mês_atual - 2)."""
        if not self._autorizado("fiscal"):
            return {"ok": False, "erro": "Apenas Nayane (fiscal) ou Carol (admin) podem executar este módulo."}
        opcoes = opcoes or {}
        cfg = self._ler_config()
        master_path = opcoes.get("master_path") or cfg.get("master_path") or os.path.join(
            PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm"
        )
        data_inicio = opcoes.get("data_inicio")
        data_fim = opcoes.get("data_fim")
        if not data_inicio or not data_fim:
            return {"ok": False, "erro": "Competência ausente."}
        if not os.path.exists(master_path):
            return {"ok": False, "erro": "Planilha master não encontrada."}

        usuario, host = self._id_usuario_host()
        ok_lock, info_lock = adquirir_lock(master_path, usuario, host, "Fiscal")
        if not ok_lock:
            return self._erro_lock_alheio(info_lock)

        def _run():
            def progresso(pct, label):
                if window:
                    window.evaluate_js(f"window.atualizarProgresso({pct}, '{label.replace(chr(39), chr(92)+chr(39))}')")
            try:
                lockfile = os.path.join(os.path.dirname(master_path),
                                        "~$" + os.path.basename(master_path))
                if os.path.exists(lockfile):
                    if window: window.evaluate_js(
                        "window.fiscalIndividualConcluido(" + _json_safe(
                            {"ok": False, "tipo": "locked",
                             "erro": "Master aberta no Excel. Feche e tente de novo."}) + ")")
                    return

                progresso(15, "Conectando ao Domínio...")
                if not db.connect():
                    if window: window.evaluate_js(
                        "window.fiscalIndividualConcluido(" + _json_safe(
                            {"ok": False, "erro": "Falha na conexão ODBC."}) + ")")
                    return

                progresso(30, "Abrindo master...")
                writer = MasterWriter(master_path)
                if not writer.carregar():
                    if window: window.evaluate_js(
                        "window.fiscalIndividualConcluido(" + _json_safe(
                            {"ok": False, "erro": "Falha ao abrir master."}) + ")")
                    return

                progresso(55, f"Executando Fiscal ({data_inicio[:7]})...")
                ok = extrair_e_preencher_fiscal(
                    writer, data_inicio, data_fim,
                    adicional_pct=cfg.get("fiscal_adicional_pct", 80),
                )

                if cfg.get("governanca_gravar_zero", True):
                    progresso(80, "Limpando resíduos do mês anterior na col. O...")
                    writer.limpar_nao_tocadas([15])

                progresso(88, "Recalculando totais...")
                writer.recalcular_totais()

                progresso(95, "Salvando master...")
                salvo = writer.salvar()
                if not salvo and writer.ultimo_erro:
                    if window: window.evaluate_js(
                        "window.fiscalIndividualConcluido(" + _json_safe(
                            {"ok": False, **writer.ultimo_erro}) + ")")
                    return

                if ok:
                    usuario_atual, host_atual = self._id_label_host()
                    estado_sh.marcar(master_path, "fiscal", data_inicio[:7], "lancado",
                                     por=usuario_atual, host=host_atual)
                progresso(100, "Concluído.")
                if window: window.evaluate_js(
                    "window.fiscalIndividualConcluido(" + _json_safe(
                        {"ok": bool(ok), "competencia": data_inicio[:7]}) + ")")
            except Exception as e:
                logging.error(f"[FISCAL-IND] {traceback.format_exc()}")
                if window: window.evaluate_js(
                    "window.fiscalIndividualConcluido(" + _json_safe(
                        {"ok": False, "erro": str(e)}) + ")")
            finally:
                liberar_lock(master_path, usuario, host)

        threading.Thread(target=_run, daemon=True).start()
        return {"ok": True, "msg": "Execução iniciada em background."}

    # ── Fluxo DP em 2 etapas ────────────────────────────────────

    def importar_planilha_carol(self, opcoes=None):
        """Fase 1: usuário seleciona a Carol; lê e mapeia empresas."""
        if not self._autorizado("dp"):
            return {"ok": False, "erro": "Apenas Jailton (DP) ou Carol (admin) podem operar este módulo."}
        opcoes = opcoes or {}
        data_inicio = opcoes.get("data_inicio")
        if not data_inicio:
            return {"ok": False, "erro": "Competência ausente."}

        tipos = ["Planilha Carol (*.xls;*.xlsx;*.xlsm)", "Todos os arquivos (*.*)"]
        result = window.create_file_dialog(
            webview.OPEN_DIALOG, allow_multiple=False, file_types=tipos
        )
        if not result or not result[0]:
            return {"ok": False, "cancelado": True}
        caminho = result[0]

        # Validação flexível do nome — gera aviso se a competência não aparecer.
        nome = os.path.basename(caminho)
        mes = data_inicio[5:7]
        ano = data_inicio[:4]
        alvos = [f"{mes}{ano}", f"{mes}.{ano}", f"{mes}/{ano}", f"{mes}-{ano}", f"{mes}_{ano}"]
        aviso = None
        if not any(a in nome for a in alvos):
            aviso = (f"O nome do arquivo ('{nome}') não contém a competência "
                     f"{mes}/{ano}. Confirme se é a Carol correta.")

        # Lê e mapeia
        from engine.excel_parser import ExcelParser
        dados = ExcelParser.ler_planilha_carol(caminho)
        if dados is None:
            return {"ok": False, "erro": f"Falha ao ler a planilha em {caminho}."}

        total_empresas = len(dados)
        com_ativos = sum(1 for v in dados.values() if v.get("total_ativos", 0) > 0)
        sem_ativos = total_empresas - com_ativos

        competencia = data_inicio[:7]
        cfg_now = self._ler_config()
        master_path_estado = cfg_now.get("master_path") or os.path.join(
            PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm"
        )
        usuario_atual, host_atual = self._id_label_host()
        estado_sh.marcar(master_path_estado, "dp", competencia, "carol_importada",
                         por=usuario_atual, host=host_atual,
                         total=total_empresas, com_ativos=com_ativos)
        estado_sh.remover(master_path_estado, "dp", competencia, evento="lancado")
        self.salvar_parametros({
            f"dp_carol_path_{competencia}": caminho,
            f"dp_carol_importado_{competencia}": True,
            f"dp_carol_importado_em_{competencia}": datetime.now().strftime("%d/%m/%Y %H:%M"),
            f"dp_carol_total_{competencia}": total_empresas,
            f"dp_carol_com_ativos_{competencia}": com_ativos,
            f"dp_lancado_{competencia}": False,
        })

        return {
            "ok": True,
            "caminho": caminho,
            "nome": nome,
            "total": total_empresas,
            "com_ativos": com_ativos,
            "sem_ativos": sem_ativos,
            "aviso": aviso,
        }

    def injetar_dp_master(self, opcoes=None):
        """Fase 2: lê Carol importada + Domínio, calcula e grava Coluna Q da master."""
        if not self._autorizado("dp"):
            return {"ok": False, "erro": "Apenas Jailton (DP) ou Carol (admin) podem executar este módulo."}
        opcoes = opcoes or {}
        cfg = self._ler_config()
        data_inicio = opcoes.get("data_inicio")
        data_fim = opcoes.get("data_fim")
        if not data_inicio or not data_fim:
            return {"ok": False, "erro": "Competência ausente."}

        competencia = data_inicio[:7]
        caminho_carol = cfg.get(f"dp_carol_path_{competencia}")
        if not caminho_carol or not os.path.exists(caminho_carol):
            return {"ok": False, "erro": "Importe a planilha Carol antes (Passo 1)."}

        master_path = opcoes.get("master_path") or cfg.get("master_path") or os.path.join(
            PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm"
        )
        if not os.path.exists(master_path):
            return {"ok": False, "erro": "Planilha master não encontrada."}

        usuario, host = self._id_usuario_host()
        ok_lock, info_lock = adquirir_lock(master_path, usuario, host, "DP")
        if not ok_lock:
            return self._erro_lock_alheio(info_lock)

        def _run():
            def progresso(pct, label):
                if window:
                    window.evaluate_js(f"window.atualizarProgresso({pct}, '{label.replace(chr(39), chr(92)+chr(39))}')")
            try:
                lockfile = os.path.join(os.path.dirname(master_path),
                                        "~$" + os.path.basename(master_path))
                if os.path.exists(lockfile):
                    if window: window.evaluate_js(
                        "window.dpFase2Concluido(" + _json_safe(
                            {"ok": False, "tipo": "locked",
                             "erro": "Master aberta no Excel. Feche e tente de novo."}) + ")")
                    return

                progresso(15, "Conectando ao Domínio (se disponível)...")
                db.connect()  # se falhar, o DP usa a Carol como única fonte

                progresso(35, "Abrindo master...")
                writer = MasterWriter(master_path)
                if not writer.carregar():
                    if window: window.evaluate_js(
                        "window.dpFase2Concluido(" + _json_safe(
                            {"ok": False, "erro": "Falha ao abrir master."}) + ")")
                    return

                progresso(60, f"Calculando e gravando DP ({competencia})...")
                ok = extrair_e_preencher_dp(
                    writer, data_inicio, data_fim,
                    fator_carga=cfg.get("dp_fator_carga", 0.33),
                    overhead_fixo=cfg.get("dp_overhead_fixo", 1.5),
                    tempo_minimo_minutos=cfg.get("dp_tempo_minimo", 5.0),
                    consultoria_horas=cfg.get("dp_consultoria_horas", 1.5),
                    tempo_fixo_apenas_socios=cfg.get("dp_apenas_socios", 1.0),
                    caminho_carol=caminho_carol,
                )

                if cfg.get("governanca_gravar_zero", True):
                    progresso(82, "Limpando resíduos do mês anterior na col. Q...")
                    writer.limpar_nao_tocadas([17])

                progresso(90, "Recalculando totais...")
                writer.recalcular_totais()

                progresso(95, "Salvando master...")
                salvo = writer.salvar()
                if not salvo and writer.ultimo_erro:
                    if window: window.evaluate_js(
                        "window.dpFase2Concluido(" + _json_safe(
                            {"ok": False, **writer.ultimo_erro}) + ")")
                    return

                if ok:
                    usuario_atual, host_atual = self._id_label_host()
                    estado_sh.marcar(master_path, "dp", competencia, "lancado",
                                     por=usuario_atual, host=host_atual)
                    self.salvar_parametros({
                        f"dp_lancado_{competencia}": True,
                        f"dp_lancado_em_{competencia}": datetime.now().strftime("%d/%m/%Y %H:%M"),
                    })

                progresso(100, "Concluído.")
                if window: window.evaluate_js(
                    "window.dpFase2Concluido(" + _json_safe(
                        {"ok": bool(ok), "competencia": competencia}) + ")")
            except Exception as e:
                logging.error(f"[DP-IND] {traceback.format_exc()}")
                if window: window.evaluate_js(
                    "window.dpFase2Concluido(" + _json_safe(
                        {"ok": False, "erro": str(e)}) + ")")
            finally:
                liberar_lock(master_path, usuario, host)

        threading.Thread(target=_run, daemon=True).start()
        return {"ok": True, "msg": "Execução iniciada em background."}

    def obter_estado_dp(self, data_inicio):
        """Estado persistente do fluxo DP para a competência.
        Combina estado compartilhado (visível para todos) + caminho local da Carol."""
        cfg = self._ler_config()
        comp = (data_inicio or "")[:7]
        master_path = cfg.get("master_path") or os.path.join(PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm")
        sh = estado_sh.obter(master_path, "dp", comp)
        carol = sh.get("carol_importada") or {}
        lancado = sh.get("lancado") or {}
        return {
            "competencia": comp,
            "importado": bool(carol) or bool(cfg.get(f"dp_carol_importado_{comp}")),
            "importado_em": carol.get("em") or cfg.get(f"dp_carol_importado_em_{comp}"),
            "importado_por": carol.get("por"),
            # `caminho` continua local — cada máquina pode ter a Carol em local diferente
            "caminho": cfg.get(f"dp_carol_path_{comp}"),
            "total": carol.get("total") or cfg.get(f"dp_carol_total_{comp}"),
            "com_ativos": carol.get("com_ativos") or cfg.get(f"dp_carol_com_ativos_{comp}"),
            "lancado": bool(lancado) or bool(cfg.get(f"dp_lancado_{comp}")),
            "lancado_em": lancado.get("em") or cfg.get(f"dp_lancado_em_{comp}"),
            "lancado_por": lancado.get("por"),
        }

    def obter_estado_fiscal(self, data_inicio):
        """Estado de lançamento do Fiscal por competência."""
        cfg = self._ler_config()
        master_path = cfg.get("master_path") or os.path.join(PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm")
        comp = (data_inicio or "")[:7]
        sh = estado_sh.obter(master_path, "fiscal", comp)
        lancado = sh.get("lancado") or {}
        return {
            "competencia": comp,
            "lancado": bool(lancado),
            "lancado_em": lancado.get("em"),
            "lancado_por": lancado.get("por"),
        }

    # ── Relatório consolidado ───────────────────────────────────

    def obter_resumo_competencia(self, data_inicio):
        """Resumo consolidado da competência: estado de cada módulo
        (visível para todos os supervisores via estado compartilhado)."""
        comp = (data_inicio or "")[:7]
        # Fiscal usa janela -1 da competência alvo
        try:
            ano_f, mes_f = int(comp[:4]), int(comp[5:7]) - 1
            if mes_f == 0:
                mes_f, ano_f = 12, ano_f - 1
            comp_fiscal = f"{ano_f}-{mes_f:02d}"
        except Exception:
            comp_fiscal = comp
        return {
            "competencia": comp,
            "competencia_fiscal": comp_fiscal,
            "fiscal": self.obter_estado_fiscal(f"{comp_fiscal}-01"),
            "contabil": self.obter_estado_contabil(data_inicio),
            "dp": self.obter_estado_dp(data_inicio),
        }

    def listar_relatorios_md(self):
        """Lista os arquivos .md gerados pelo motor (relatórios contábeis e equivalentes)."""
        padroes = ("relatorio_contabil_", "log_execucao", "relatorio_dp_", "relatorio_fiscal_")
        encontrados = []
        try:
            for nome in os.listdir(PROJECT_ROOT):
                if not nome.lower().endswith(".md"):
                    continue
                if not any(nome.startswith(p) for p in padroes):
                    continue
                caminho = os.path.join(PROJECT_ROOT, nome)
                try:
                    st = os.stat(caminho)
                    encontrados.append({
                        "nome": nome,
                        "caminho": caminho,
                        "tamanho_kb": round(st.st_size / 1024, 1),
                        "modificado_em": datetime.fromtimestamp(st.st_mtime).strftime("%d/%m/%Y %H:%M"),
                    })
                except Exception:
                    pass
        except Exception as e:
            logging.error(f"[RELATORIOS] {e}")
        encontrados.sort(key=lambda d: d["modificado_em"], reverse=True)
        return {"relatorios": encontrados}

    def ler_relatorio_md(self, nome):
        """Lê o conteúdo de um relatório .md específico."""
        if not nome or ".." in nome or "/" in nome or "\\" in nome or not nome.lower().endswith(".md"):
            return {"ok": False, "erro": "Nome inválido."}
        caminho = os.path.join(PROJECT_ROOT, nome)
        if not os.path.exists(caminho):
            return {"ok": False, "erro": "Arquivo não encontrado."}
        try:
            with open(caminho, encoding="utf-8") as f:
                return {"ok": True, "conteudo": f.read(), "nome": nome}
        except Exception as e:
            return {"ok": False, "erro": str(e)}

    # ── Exceções por setor (DP / Contábil / Fiscal) ─────────────

    def listar_excecoes(self, setor=None):
        """Retorna metadados + lista de códigos para um setor (ou todos)."""
        if setor in ("dp", "contabil", "fiscal"):
            info = metadados_setor(setor)
            info["codigos"] = sorted(ler_codigos_setor(setor), key=lambda c: int(c))
            return info
        return {s: {**metadados_setor(s), "codigos": sorted(ler_codigos_setor(s), key=lambda c: int(c))}
                for s in ("dp", "contabil", "fiscal")}

    def importar_excecoes(self, setor):
        """Abre file dialog, copia o TXT selecionado para o caminho padrão do setor."""
        if setor not in ("dp", "contabil", "fiscal"):
            return {"ok": False, "erro": f"Setor inválido: {setor}"}
        tipos = ["Arquivo texto (*.txt)", "Todos os arquivos (*.*)"]
        result = window.create_file_dialog(
            webview.OPEN_DIALOG, allow_multiple=False, file_types=tipos
        )
        if not result or not result[0]:
            return {"ok": False, "cancelado": True}
        origem = result[0]
        destino_dir = os.path.join(PROJECT_ROOT, PASTA_EXCECOES)
        os.makedirs(destino_dir, exist_ok=True)
        destino = os.path.join(PROJECT_ROOT, caminho_excecao(setor))
        try:
            import shutil
            shutil.copyfile(origem, destino)
        except Exception as e:
            logging.error(f"[EXCECOES] Falha ao copiar {origem} → {destino}: {e}")
            return {"ok": False, "erro": str(e)}
        info = metadados_setor(setor)
        info["codigos"] = sorted(ler_codigos_setor(setor), key=lambda c: int(c))
        info["ok"] = True
        return info

    def abrir_pasta_excecoes(self):
        """Abre o Explorer na pasta de exceções."""
        pasta = os.path.join(PROJECT_ROOT, PASTA_EXCECOES)
        try:
            os.startfile(pasta)
            return {"ok": True, "pasta": pasta}
        except Exception as e:
            return {"ok": False, "erro": str(e)}

    # ── Diagnóstico Contábil ────────────────────────────────────

    def obter_status_contabil(self):
        """Reporta o estado da planilha-fonte do módulo contábil
        (caminho configurado em `contabil_origem_path`, com fallback local)."""
        cfg = self._ler_config()
        candidatos = [
            cfg.get("contabil_origem_path"),
            os.path.join(PROJECT_ROOT, "ENTRADAS_MANUAIS", "HORAS CONTABEIS.xlsx"),
        ]
        caminho = next((p for p in candidatos if p and os.path.exists(p)), candidatos[0])
        info = {
            "caminho": caminho,
            "configurado": cfg.get("contabil_origem_path"),
            "existe": bool(caminho and os.path.exists(caminho)),
            "tamanho_kb": None,
            "modificado_em": None,
            "linhas": None,
            "abas": [],
            "online_only": False,
        }
        if info["existe"]:
            try:
                if esta_online_only(caminho):
                    info["online_only"] = True
                    forcar_download(caminho, timeout=20)
                st = os.stat(caminho)
                info["tamanho_kb"] = round(st.st_size / 1024, 1)
                info["modificado_em"] = datetime.fromtimestamp(st.st_mtime).strftime("%d/%m/%Y %H:%M")
                import openpyxl
                wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
                info["abas"] = wb.sheetnames
                info["linhas"] = wb.active.max_row
                wb.close()
            except Exception as e:
                logging.error(f"[CONTABIL_STATUS] {e}")
                info["erro"] = str(e)
        return info

    # ── Seleção de Arquivos ─────────────────────────────────────

    def selecionar_arquivo(self, tipos=None):
        if tipos is None:
            tipos = ["Planilha Excel (*.xlsx;*.xlsm)", "Todos os arquivos (*.*)"]
        result = window.create_file_dialog(
            webview.OPEN_DIALOG, allow_multiple=False, file_types=tipos
        )
        if result and result[0]:
            caminho = result[0]
            self.salvar_parametros({"master_path": caminho})
            return caminho
        return None

    def validar_arquivos(self, caminhos: dict):
        return {k: os.path.exists(v) if v else False for k, v in caminhos.items()}

    # ── Execução do Ciclo ───────────────────────────────────────

    def executar_ciclo(self, opcoes: dict):
        def _run():
            def progresso(pct, label):
                if window:
                    window.evaluate_js(f"window.atualizarProgresso({pct}, '{label}')")

            try:
                cfg = self._ler_config()
                # Tenta pegar da requisicao -> config salva -> pasta raiz automatica
                master_path = (
                    opcoes.get("master_path")
                    or cfg.get("master_path")
                    or os.path.join(PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm")
                )

                if not master_path or not os.path.exists(master_path):
                    root_fallback = os.path.join(PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm")
                    if os.path.exists(root_fallback):
                        master_path = root_fallback
                    else:
                        progresso(0, "Erro: planilha master não encontrada")
                        return

                progresso(5, "Carregando planilha master...")
                lockfile = os.path.join(
                    os.path.dirname(master_path),
                    "~$" + os.path.basename(master_path),
                )
                if os.path.exists(lockfile):
                    progresso(0, "Master aberta no Excel — feche o arquivo para continuar.")
                    if window:
                        window.evaluate_js(
                            "window.execucaoFalhou("
                            + _json_safe({"tipo": "locked", "msg": "Planilha master está aberta no Excel. Feche o arquivo antes de executar o ciclo."})
                            + ")"
                        )
                    return

                writer = MasterWriter(master_path)
                if not writer.carregar():
                    progresso(0, "Erro: falha ao abrir planilha master")
                    if window:
                        window.evaluate_js(
                            "window.execucaoFalhou("
                            + _json_safe({"tipo": "erro", "msg": "Não foi possível abrir a planilha master. Verifique o caminho e se ela não está corrompida."})
                            + ")"
                        )
                    return

                # Janela "alvo" do ciclo (Contábil/DP). Fallback defensivo: mês atual - 1.
                hoje = datetime.now()
                mes_alvo = hoje.month - 1
                ano_alvo = hoje.year
                if mes_alvo == 0:
                    mes_alvo, ano_alvo = 12, ano_alvo - 1
                from calendar import monthrange
                ultimo_dia = monthrange(ano_alvo, mes_alvo)[1]
                data_inicio = opcoes.get("data_inicio") or f"{ano_alvo}-{mes_alvo:02d}-01"
                data_fim = opcoes.get("data_fim") or f"{ano_alvo}-{mes_alvo:02d}-{ultimo_dia:02d}"

                # Janela específica do Fiscal — mês alvo - 1 (regra -2 vs mês corrente).
                # Se o front não passar, derivamos a partir da janela alvo.
                fiscal_data_inicio = opcoes.get("fiscal_data_inicio")
                fiscal_data_fim = opcoes.get("fiscal_data_fim")
                if not fiscal_data_inicio or not fiscal_data_fim:
                    ano_f = int(data_inicio[:4])
                    mes_f = int(data_inicio[5:7]) - 1
                    if mes_f == 0:
                        mes_f, ano_f = 12, ano_f - 1
                    ult_f = monthrange(ano_f, mes_f)[1]
                    fiscal_data_inicio = f"{ano_f}-{mes_f:02d}-01"
                    fiscal_data_fim = f"{ano_f}-{mes_f:02d}-{ult_f:02d}"

                status = {"fiscal": None, "dp": None, "contabil": None}

                if opcoes.get("fiscal"):
                    progresso(15, f"Executando Fiscal ({fiscal_data_inicio[:7]})...")
                    try:
                        status["fiscal"] = extrair_e_preencher_fiscal(
                            writer,
                            fiscal_data_inicio,
                            fiscal_data_fim,
                            adicional_pct=cfg.get("fiscal_adicional_pct", 80),
                        )
                    except Exception as e:
                        status["fiscal"] = False
                        logging.error(f"[FISCAL] {e}")

                if opcoes.get("dp"):
                    progresso(40, "Executando DP / Folha...")
                    try:
                        status["dp"] = extrair_e_preencher_dp(
                            writer,
                            data_inicio,
                            data_fim,
                            fator_carga=cfg.get("dp_fator_carga", 0.33),
                            overhead_fixo=cfg.get("dp_overhead_fixo", 1.5),
                            tempo_minimo_minutos=cfg.get("dp_tempo_minimo", 5.0),
                            consultoria_horas=cfg.get("dp_consultoria_horas", 1.5),
                            tempo_fixo_apenas_socios=cfg.get("dp_apenas_socios", 1.0),
                        )
                    except Exception as e:
                        status["dp"] = False
                        logging.error(f"[DP] {e}")

                # Contábil tem fluxo próprio em 2 etapas (preencher + injetar) acionado
                # pela aba dedicada da UI — não entra no ciclo global. Aqui apenas marcamos
                # se o lançamento na master já foi feito para a competência atual.
                if cfg.get(f"contabil_lancado_{data_inicio[:7]}"):
                    status["contabil"] = True

                if cfg.get("governanca_gravar_zero", True):
                    colunas_para_limpar = []
                    if status.get("fiscal") is not None:
                        colunas_para_limpar.append(15)
                    if status.get("contabil") is not None:
                        colunas_para_limpar.append(16)
                    if status.get("dp") is not None:
                        colunas_para_limpar.append(17)
                    if colunas_para_limpar:
                        progresso(80, "Limpando resíduos do mês anterior...")
                        writer.limpar_nao_tocadas(colunas_para_limpar)

                if opcoes.get("totais"):
                    progresso(85, "Recalculando totais (col. R)...")
                    writer.recalcular_totais()

                progresso(95, "Salvando planilha master...")
                salvo = writer.salvar()
                if not salvo and writer.ultimo_erro and window:
                    window.evaluate_js(
                        "window.execucaoFalhou(" + _json_safe(writer.ultimo_erro) + ")"
                    )

                # Geração de log consolidado em markdown (Fase 4 concluída)
                log_md_path = os.path.join(PROJECT_ROOT, "log_execucao.md")
                try:
                    with open(log_md_path, "w", encoding="utf-8") as f_md:
                        f_md.write("# Relatório de Execução Consolidado — DMF Engine\n\n")
                        f_md.write(f"**Arquivo Master:** `{master_path}`\n")
                        f_md.write(f"**Período Processado:** `{data_inicio}` até `{data_fim}`\n\n")
                        f_md.write("## Status dos Submódulos\n")
                        for mod, st in status.items():
                            st_ico = "✅ Sucesso" if st else ("⚠️ Sem atualizações / Aviso" if st is False else "⏸️ Ignorado")
                            f_md.write(f"- **{mod.upper()}**: {st_ico}\n")
                        f_md.write("\n## Finalização\nTotais gravados e fórmulas protegidas com resiliência.\n")
                except Exception as e_md:
                    logging.error(f"Falha ao escrever log markdown: {e_md}")

                estado_path = os.path.join(BASE_DIR, "ultimo_ciclo.json")
                try:
                    with open(estado_path, "w", encoding="utf-8") as f_est:
                        json.dump(
                            {
                                "modulos": status,
                                "totais": bool(opcoes.get("totais")),
                                "salvo": salvo,
                                "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M"),
                                "competencia": data_inicio[:7],         # alvo Contábil/DP
                                "competencia_fiscal": fiscal_data_inicio[:7],
                                "master_path": master_path,
                            },
                            f_est,
                            indent=2,
                            ensure_ascii=False,
                        )
                except Exception as e_est:
                    logging.error(f"[CICLO] Falha ao persistir estado: {e_est}")

                progresso(100, "Concluído!")
                resultado = _json_safe({"status": status, "salvo": salvo})
                if window:
                    window.evaluate_js(f"window.execucaoConcluida({resultado})")

            except Exception as e:
                logging.error(f"[CICLO] Erro fatal: {traceback.format_exc()}")
                progresso(0, f"Erro fatal: {e}")

        threading.Thread(target=_run, daemon=True).start()
        return {"ok": True, "msg": "Execução iniciada em background"}

    # ── Dashboard (dados reais) ─────────────────────────────────

    def obter_metricas_dashboard(self, competencia=None):
        """Wrapper retrocompatível — apenas métricas (sem dados de gráfico)."""
        d = self.obter_dashboard_completo(competencia)
        return {
            "ok": d.get("ok", False),
            "clientes": d.get("clientes", 0),
            "matches": d.get("matches", 0),
            "pendencias": d.get("pendencias", 0),
            "cnpj_dup": d.get("cnpj_dup", 0),
            "master_path": d.get("master_path"),
            "msg": d.get("msg"),
        }

    def obter_dashboard_completo(self, competencia=None):
        """Faz UMA leitura da master e devolve métricas + dados para todos os gráficos.

        `competencia`: string 'YYYY-MM' ou None. Tenta ler a aba 'MM.YYYY';
        cai em wb.active se ausente.
        """
        cfg = self._ler_config()
        master_path = (
            cfg.get("master_path")
            or os.path.join(PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm")
        )
        vazio = {
            "ok": False, "clientes": 0, "matches": 0, "pendencias": 0, "cnpj_dup": 0,
            "horas_por_setor": {"fiscal": 0, "contabil": 0, "dp": 0},
            "top_clientes": [],
            "distribuicao": [],
            "preenchimento_setor": {"fiscal": 0, "contabil": 0, "dp": 0},
            "aba_usada": None,
            "master_path": master_path,
        }
        if not master_path or not os.path.exists(master_path):
            vazio["msg"] = "Planilha master não encontrada"
            return vazio

        # Pré-checagem OneDrive: força download se Online-only.
        if esta_online_only(master_path):
            forcar_download(master_path, timeout=20)

        # Aba alvo: 'YYYY-MM' → 'MM.YYYY'
        aba_alvo = None
        if competencia and len(competencia) >= 7:
            aba_alvo = f"{competencia[5:7]}.{competencia[:4]}"

        try:
            import openpyxl
            wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
            if aba_alvo and aba_alvo in wb.sheetnames:
                ws = wb[aba_alvo]
                aba_usada = aba_alvo
            else:
                ws = wb.active
                aba_usada = ws.title

            clientes = 0
            matches = 0
            pendencias = 0
            cnpj_count = {}
            codigo_count = {}
            soma_fiscal = 0.0
            soma_contabil = 0.0
            soma_dp = 0.0
            pres_fiscal = 0
            pres_contabil = 0
            pres_dp = 0
            buckets = [0, 0, 0, 0, 0, 0]  # 0h, 0-2h, 2-5h, 5-10h, 10-20h, 20+
            clientes_buf = []  # (total_dia, nome, fiscal, contabil, dp, codigo, linha, cnpj)
            pendentes_buf = []  # listagem para drill-down: linhas com 0/1 campo

            def _num(v):
                """Converte célula numérica para fração de dia (formato Excel).
                Aceita float/int direto e datetime.timedelta (formato [h]:mm:ss),
                que é como o openpyxl entrega células formatadas como duração."""
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    return float(v)
                if isinstance(v, timedelta):
                    return v.total_seconds() / 86400.0
                return None

            for linha_idx, row in enumerate(ws.iter_rows(min_row=10, values_only=True), start=10):
                if not row or len(row) < 18:
                    continue
                cod = row[7]
                # Razão social está em K (índice 10). I (índice 8) é o grupo
                # empresarial (opcional), não o nome do cliente.
                nome = row[10] if row[10] not in (None, "") else row[8]
                cnpj = row[9]
                if cod is None and cnpj is None and nome is None:
                    continue
                if cod is None:
                    continue
                clientes += 1
                campos = sum(1 for v in (cod, nome, cnpj) if v not in (None, ""))
                if campos >= 2:
                    matches += 1
                else:
                    pendencias += 1
                    pendentes_buf.append({
                        "linha": linha_idx,
                        "codigo": str(cod) if cod is not None else "",
                        "nome": str(nome or ""),
                        "cnpj": str(cnpj or ""),
                        "campos_preenchidos": campos,
                    })
                if cnpj:
                    cnpj_limpo = re.sub(r"\D", "", str(cnpj))
                    if cnpj_limpo:
                        cnpj_count.setdefault(cnpj_limpo, []).append(
                            {"linha": linha_idx, "codigo": str(cod), "nome": str(nome or ""), "cnpj": cnpj_limpo}
                        )
                try:
                    codigo_norm = str(int(float(str(cod).strip())))
                    codigo_count.setdefault(codigo_norm, []).append(
                        {"linha": linha_idx, "codigo": codigo_norm, "nome": str(nome or ""), "cnpj": str(cnpj or "")}
                    )
                except (ValueError, TypeError):
                    pass

                v_fiscal = _num(row[14]) or 0.0
                v_contabil = _num(row[15]) or 0.0
                v_dp = _num(row[16]) or 0.0
                soma_fiscal += v_fiscal
                soma_contabil += v_contabil
                soma_dp += v_dp
                if v_fiscal > 0: pres_fiscal += 1
                if v_contabil > 0: pres_contabil += 1
                if v_dp > 0: pres_dp += 1

                # Coluna R pode ser fórmula → preferimos o cacheado se houver,
                # senão somamos O+P+Q manualmente.
                total_r = _num(row[17])
                if total_r is None:
                    total_r = v_fiscal + v_contabil + v_dp
                total_horas = total_r * 24.0   # fração de dia → horas

                if total_horas <= 0:
                    buckets[0] += 1
                elif total_horas <= 2:
                    buckets[1] += 1
                elif total_horas <= 5:
                    buckets[2] += 1
                elif total_horas <= 10:
                    buckets[3] += 1
                elif total_horas <= 20:
                    buckets[4] += 1
                else:
                    buckets[5] += 1

                try:
                    codigo_str = str(int(float(str(cod).strip())))
                except (ValueError, TypeError):
                    # Linha sem código numérico (cabeçalhos extras, totalizadores
                    # ou anotações textuais). Conta nas somas/distribuição mas
                    # não entra no Top N de clientes.
                    continue
                clientes_buf.append((
                    total_r, str(nome or "")[:42], v_fiscal, v_contabil, v_dp, codigo_str
                ))

            wb.close()

            # Top 10 ordenado por total decrescente
            clientes_buf.sort(key=lambda t: t[0], reverse=True)
            top = clientes_buf[:10]
            top_clientes = [
                {
                    "codigo": c[5],
                    "nome": c[1] or c[5],
                    "horas_total": round(c[0] * 24.0, 2),
                    "horas_fiscal": round(c[2] * 24.0, 2),
                    "horas_contabil": round(c[3] * 24.0, 2),
                    "horas_dp": round(c[4] * 24.0, 2),
                }
                for c in top
            ]
            # Listas detalhadas (drill-down) — clientes_buf serializável
            cnpj_dup_lista = [
                {"cnpj": k, "ocorrencias": ocs}
                for k, ocs in cnpj_count.items() if len(ocs) > 1
            ]
            codigo_dup_lista = [
                {"codigo": k, "ocorrencias": ocs}
                for k, ocs in codigo_count.items() if len(ocs) > 1
            ]
            cnpj_dup = len(cnpj_dup_lista)
            codigo_dup = len(codigo_dup_lista)

            return {
                "ok": True,
                "aba_usada": aba_usada,
                "competencia_solicitada": competencia,
                "clientes": clientes,
                "matches": matches,
                "pendencias": pendencias,
                "cnpj_dup": cnpj_dup,
                "codigo_dup": codigo_dup,
                "horas_por_setor": {
                    "fiscal":   round(soma_fiscal * 24.0, 2),
                    "contabil": round(soma_contabil * 24.0, 2),
                    "dp":       round(soma_dp * 24.0, 2),
                },
                "preenchimento_setor": {
                    "fiscal": pres_fiscal,
                    "contabil": pres_contabil,
                    "dp": pres_dp,
                },
                "top_clientes": top_clientes,
                "distribuicao": [
                    {"faixa": "0h",     "qtd": buckets[0]},
                    {"faixa": "0–2h",   "qtd": buckets[1]},
                    {"faixa": "2–5h",   "qtd": buckets[2]},
                    {"faixa": "5–10h",  "qtd": buckets[3]},
                    {"faixa": "10–20h", "qtd": buckets[4]},
                    {"faixa": "20h+",   "qtd": buckets[5]},
                ],
                "drill": {
                    "pendencias": pendentes_buf[:200],
                    "cnpj_dup": cnpj_dup_lista[:100],
                    "codigo_dup": codigo_dup_lista[:100],
                },
                "master_path": master_path,
            }
        except Exception as e:
            logging.error(f"[DASHBOARD] {e}")
            vazio["msg"] = str(e)
            return vazio

    def obter_estado_modulos(self):
        """Lê o resultado persistido do último ciclo executado."""
        estado_path = os.path.join(BASE_DIR, "ultimo_ciclo.json")
        if not os.path.exists(estado_path):
            return {"modulos": {}, "timestamp": None, "competencia": None}
        try:
            with open(estado_path, encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logging.error(f"[ESTADO] {e}")
            return {"modulos": {}, "timestamp": None, "competencia": None}

    def listar_competencias_master(self):
        """Retorna abas MM.YYYY da planilha Master como lista de competências, mais recente primeiro."""
        cfg = self._ler_config()
        master_path = cfg.get('master_path') or os.path.join(PROJECT_ROOT, "CONTROLE DE HORAS DMF.xlsm")
        if not master_path or not os.path.exists(master_path):
            return {"ok": False, "msg": "Master não encontrada", "competencias": []}
        if esta_online_only(master_path):
            forcar_download(master_path, timeout=20)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
            padrao = re.compile(r'^(\d{2})\.(\d{4})$')
            result = []
            for aba in wb.sheetnames:
                m = padrao.match(aba)
                if m:
                    mes, ano = m.group(1), m.group(2)
                    result.append({
                        "aba": aba,
                        "value": f"{ano}-{mes}",
                        "label": f"{mes}/{ano}"
                    })
            wb.close()
            result.sort(key=lambda x: x["value"], reverse=True)
            return {"ok": True, "competencias": result}
        except Exception as e:
            logging.error(f"[COMPETENCIAS] {e}")
            return {"ok": False, "msg": str(e), "competencias": []}

    def obter_ultima_execucao(self):
        """Extrai os últimos eventos por módulo do dmf_engine.log."""
        log_path = os.path.join(BASE_DIR, "dmf_engine.log")
        if not os.path.exists(log_path):
            return {"eventos": []}
        try:
            with open(log_path, encoding="utf-8") as f:
                linhas = f.readlines()[-400:]
            padrao = re.compile(
                r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}).+?\[(FISCAL|DP|CONTABIL|CONTÁBIL|MASTER|CICLO|METRICAS)\]\s*(.+)"
            )
            eventos = []
            for ln in linhas:
                m = padrao.search(ln)
                if not m:
                    continue
                ts, modulo, msg = m.groups()
                eventos.append(
                    {
                        "hora": ts[11:16],
                        "data": ts[:10],
                        "modulo": modulo.replace("Á", "A").title(),
                        "msg": msg.strip(),
                    }
                )
            return {"eventos": eventos[-6:]}
        except Exception as e:
            logging.error(f"[ULTIMA_EXEC] {e}")
            return {"eventos": [], "erro": str(e)}

    # ── Log ─────────────────────────────────────────────────────

    def ler_log(self):
        log_path = os.path.join(BASE_DIR, "dmf_engine.log")
        if os.path.exists(log_path):
            with open(log_path, encoding="utf-8") as f:
                lines = f.readlines()
            return {"conteudo": "".join(lines[-100:])}  # últimas 100 linhas
        return {"conteudo": "Nenhuma execução registrada."}

    def ler_log_erros(self, limite=40):
        """Últimas N entradas do log de erros (WARNING+)."""
        log_path = os.path.join(BASE_DIR, "dmf_engine_errors.log")
        if not os.path.exists(log_path):
            return {"eventos": [], "conteudo": ""}
        try:
            with open(log_path, encoding="utf-8") as f:
                linhas = f.readlines()
        except Exception as e:
            return {"eventos": [], "erro": str(e)}

        # Agrupa por entrada (mensagens multi-linha de traceback).
        padrao = re.compile(r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ \[(WARNING|ERROR|CRITICAL)\]")
        eventos = []
        atual = None
        for linha in linhas:
            m = padrao.match(linha)
            if m:
                if atual is not None:
                    eventos.append(atual)
                atual = {"timestamp": m.group(1), "nivel": m.group(2), "linha": linha.rstrip("\n")}
            elif atual is not None:
                atual["linha"] += "\n" + linha.rstrip("\n")
        if atual is not None:
            eventos.append(atual)
        return {"eventos": eventos[-limite:], "total_arquivo": len(eventos)}

    def limpar_log_erros(self):
        """Trunca o arquivo de erros (após o supervisor confirmar revisão)."""
        log_path = os.path.join(BASE_DIR, "dmf_engine_errors.log")
        try:
            open(log_path, "w", encoding="utf-8").close()
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "erro": str(e)}


# ── Inicialização ────────────────────────────────────────────────

api = Api()

window = webview.create_window(
    title="DMF Engine — Automação de Horas",
    url=os.path.join(RESOURCES_DIR, "ui", "index.html"),
    js_api=api,
    width=1100,
    height=720,
    min_size=(900, 600),
    resizable=True,
)

if __name__ == "__main__":
    webview.start(debug=False, icon=os.path.join(RESOURCES_DIR, "ui", "logo.ico"))
