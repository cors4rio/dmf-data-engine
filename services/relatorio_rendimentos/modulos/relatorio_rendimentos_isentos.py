"""
modulos/relatorio_rendimentos_isentos.py
Extrai dados de Rendimentos Isentos (EFD-Reinf) do banco Domínio via ODBC
e gera relatório Excel formatado.

Adaptado do projeto externo efd_contabil para integração com a Central DMF:
- Credenciais recebidas como parâmetros (dsn, user, pwd) em vez de .env
- Função main() CLI mantida como fallback
"""

import os
import sys
import logging
from pathlib import Path
from datetime import datetime, date

import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger("RendimentosIsentos")

# ---------------------------------------------------------------------------
# Constantes de estilo Excel
# ---------------------------------------------------------------------------
COR_CABECALHO_BG = "1F3864"
COR_CABECALHO_FG = "FFFFFF"
COR_LINHA_PAR    = "E8EFF7"
COR_LINHA_IMPAR  = "FFFFFF"
COR_BORDA        = "B8CCE4"

COLUNAS = [
    ("Código",              12),
    ("Razão Social",        45),
    ("Data de Emissão",     16),
    ("Beneficiário (CPF)",  18),
    ("Nome Beneficiário",   45),
    ("Nat. Rend.",           55),
    ("Valor",               15),
]

QUERY_RENDIMENTOS_ISENTOS = """
SELECT
    e.codi_emp                      AS Codigo,
    e.nome_emp                      AS Razao_Social,
    r.DATA_RENDIMENTO               AS Data_Emissao,
    r.BENEFICIARIO                  AS Beneficiario,
    COALESCE(RTRIM(s.nome), '')     AS Nome_Beneficiario,
    COALESCE(n.DESCRICAO, CAST(r.CODI_NAT_RENDIMENTO AS VARCHAR(20)))
                                    AS Nat_Rend,
    r.VALOR                         AS Valor
FROM bethadba.EFOUTROSDADOS_REINF_RENDIMENTOS_ISENTOS r
JOIN bethadba.geempre e
    ON e.codi_emp = r.CODI_EMP
LEFT JOIN bethadba.EFNATUREZA_RENDIMENTOS_ISENTOS n
    ON n.CODI_NAT_RENDIMENTO = r.CODI_NAT_RENDIMENTO
LEFT JOIN bethadba.gesocios s
    ON s.inscricao = r.BENEFICIARIO
WHERE r.COMPETENCIA = ?
  {filtro_cliente}
ORDER BY e.codi_emp ASC, r.DATA_RENDIMENTO ASC, r.I_RENDIMENTO_ISENTO ASC
"""


# ---------------------------------------------------------------------------
# Conexão ODBC
# ---------------------------------------------------------------------------
def conectar_dominio(dsn: str, user: str, pwd: str) -> pyodbc.Connection:
    """Conecta ao banco Domínio via ODBC usando credenciais explícitas."""
    if not dsn or not user or not pwd:
        raise ValueError("DSN, usuário e senha são obrigatórios para conectar ao Domínio.")

    conn_str = f"DSN={dsn};UID={user};PWD={pwd}"
    log.info("Conectando ao banco Domínio (DSN=%s, usuário=%s)...", dsn, user)

    conn = pyodbc.connect(conn_str, timeout=30)
    conn.setdecoding(pyodbc.SQL_CHAR,  encoding="latin-1")
    conn.setdecoding(pyodbc.SQL_WCHAR, encoding="utf-8")

    log.info("Conexão estabelecida com sucesso.")
    return conn


# ---------------------------------------------------------------------------
# Extração de dados
# ---------------------------------------------------------------------------
def extrair_dados(conn: pyodbc.Connection, competencia: str, codi_emp=None) -> list:
    """
    Executa a query e retorna lista de dicionários com os dados.

    Args:
        conn:        Conexão ODBC ativa
        competencia: Data no formato 'YYYY-MM-DD' (ex: '2026-03-01')
        codi_emp:    Opcional — código da empresa para filtrar
    """
    filtro_cliente = ""
    params = [competencia]

    if codi_emp:
        log.info("Filtrando extração para o cliente: %s", codi_emp)
        filtro_cliente = "AND e.codi_emp = ?"
        try:
            params.append(int(codi_emp))
        except (ValueError, TypeError):
            params.append(codi_emp)

    query = QUERY_RENDIMENTOS_ISENTOS.format(filtro_cliente=filtro_cliente)
    log.info("Extraindo dados para competência %s...", competencia)

    cursor = conn.cursor()
    cursor.execute(query, params)
    colunas = [desc[0] for desc in cursor.description]
    linhas = []

    for row in cursor.fetchall():
        registro = {}
        for col, val in zip(colunas, row):
            if val is None:
                val = 0.0 if col == "Valor" else ""
            if isinstance(val, str):
                val = val.strip()
            registro[col] = val
        linhas.append(registro)

    log.info("Total de registros encontrados: %d", len(linhas))
    return linhas


# ---------------------------------------------------------------------------
# Geração do Excel
# ---------------------------------------------------------------------------
def _borda_fina(cor: str = COR_BORDA) -> Border:
    lado = Side(style="thin", color=cor)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def gerar_excel(dados: list, competencia: str, caminho_saida: Path) -> Path:
    """
    Cria o arquivo Excel formatado com os dados extraídos.

    Returns:
        Path completo do arquivo gerado
    """
    dt_comp  = date.fromisoformat(competencia)
    mes_ano  = dt_comp.strftime("%m/%Y")
    sufixo   = dt_comp.strftime("%m%Y")
    nome_arq = f"rendimentos_isentos_{sufixo}.xlsx"
    arquivo  = Path(caminho_saida) / nome_arq

    wb = Workbook()
    ws = wb.active
    ws.title = f"Rend Isentos {dt_comp.strftime('%m-%Y')}"

    # Linha 1: Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUNAS))
    c = ws.cell(row=1, column=1)
    c.value     = f"EFD-Reinf — Rendimentos Isentos — Competência {mes_ano}"
    c.font      = Font(name="Calibri", bold=True, size=13, color=COR_CABECALHO_FG)
    c.fill      = PatternFill("solid", fgColor=COR_CABECALHO_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Linha 2: Subtítulo
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUNAS))
    c = ws.cell(row=2, column=1)
    c.value     = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   Registros: {len(dados)}"
    c.font      = Font(name="Calibri", italic=True, size=10, color="555555")
    c.fill      = PatternFill("solid", fgColor="D6E4F0")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Linha 3: Cabeçalhos
    for col_idx, (titulo, largura) in enumerate(COLUNAS, start=1):
        cel = ws.cell(row=3, column=col_idx)
        cel.value     = titulo
        cel.font      = Font(name="Calibri", bold=True, size=11, color=COR_CABECALHO_FG)
        cel.fill      = PatternFill("solid", fgColor=COR_CABECALHO_BG)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border    = _borda_fina()
        ws.column_dimensions[get_column_letter(col_idx)].width = largura
    ws.row_dimensions[3].height = 22

    # Linhas de dados
    fmt_data  = "DD/MM/YYYY"
    fmt_moeda = "#,##0.00"

    for i, reg in enumerate(dados):
        linha  = 4 + i
        cor_bg = COR_LINHA_PAR if i % 2 == 0 else COR_LINHA_IMPAR
        fill   = PatternFill("solid", fgColor=cor_bg)
        fonte  = Font(name="Calibri", size=10)
        borda  = _borda_fina()

        valores = [
            reg.get("Codigo",            ""),
            reg.get("Razao_Social",      ""),
            reg.get("Data_Emissao",      ""),
            reg.get("Beneficiario",      ""),
            reg.get("Nome_Beneficiario", ""),
            reg.get("Nat_Rend",          ""),
            reg.get("Valor",             0.0),
        ]
        alinhamentos = [
            Alignment(horizontal="center"),
            Alignment(horizontal="left"),
            Alignment(horizontal="center"),
            Alignment(horizontal="center"),
            Alignment(horizontal="left"),
            Alignment(horizontal="left", wrap_text=True),
            Alignment(horizontal="right"),
        ]

        for col_idx, (valor, align) in enumerate(zip(valores, alinhamentos), start=1):
            cel = ws.cell(row=linha, column=col_idx, value=valor)
            cel.font      = fonte
            cel.fill      = fill
            cel.border    = borda
            cel.alignment = align
            if col_idx == 3 and isinstance(valor, (date, datetime)):
                cel.number_format = fmt_data
            if col_idx == 7:
                cel.number_format = fmt_moeda

    ws.freeze_panes = ws.cell(row=4, column=1)
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUNAS))}{3 + len(dados)}"

    Path(caminho_saida).mkdir(parents=True, exist_ok=True)
    wb.save(arquivo)
    log.info("Arquivo Excel salvo em: %s", arquivo)
    return arquivo


# ---------------------------------------------------------------------------
# CLI (fallback — não usado pela plataforma DMF)
# ---------------------------------------------------------------------------
def main() -> None:
    import argparse
    from datetime import date as _date

    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    parser = argparse.ArgumentParser(description="Extrai Rendimentos Isentos via ODBC.")
    parser.add_argument("--dsn",          default="Contabil")
    parser.add_argument("--user",         required=True)
    parser.add_argument("--pwd",          required=True)
    parser.add_argument("--competencia",  default=_date.today().replace(day=1).isoformat())
    parser.add_argument("--cliente",      type=int, default=None)
    parser.add_argument("--output",       default="./output")
    args = parser.parse_args()

    conn  = conectar_dominio(args.dsn, args.user, args.pwd)
    dados = extrair_dados(conn, args.competencia, args.cliente)
    conn.close()

    if not dados:
        print("Nenhum dado encontrado.")
        return

    arquivo = gerar_excel(dados, args.competencia, Path(args.output))
    print(f"Relatório gerado: {arquivo}")


if __name__ == "__main__":
    main()
