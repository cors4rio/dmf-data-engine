"""
sm_engine/sm_planilha.py — Leitura do lote de empresas (CNPJ + senha).

Aceita .xlsx (openpyxl) ou .txt. Colunas: CNPJ | Senha.
TXT usa separador ';' ou TAB, uma empresa por linha. Normaliza o CNPJ para só dígitos.

O nome da empresa é extraído do portal após login (extrair_nome_contribuinte).
A senha é retornada em claro (necessária para o login). A máscara para exibição é
responsabilidade da fronteira com o JS (api.py) — aqui nunca se mascara nem se loga.
"""
import os
import re
import logging

log = logging.getLogger("SemMovimento.Planilha")


def _so_digitos(valor) -> str:
    return re.sub(r"\D", "", str(valor or ""))


def _cnpj_valido(cnpj: str) -> bool:
    return len(cnpj) == 14


def carregar(caminho: str) -> dict:
    """
    Retorna {"empresas": [{"cnpj","senha","nome","linha"}], "invalidas": [...]}.
    Lança ValueError se o arquivo não existir ou a extensão não for suportada.
    """
    if not caminho or not os.path.isfile(caminho):
        raise ValueError(f"Arquivo não encontrado: {caminho}")

    ext = os.path.splitext(caminho)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        linhas = _ler_xlsx(caminho)
    elif ext == ".txt":
        linhas = _ler_txt(caminho)
    else:
        raise ValueError(f"Extensão não suportada: {ext} (use .txt, .xlsx)")

    empresas, invalidas = [], []
    for num, (raw_cnpj, raw_senha) in linhas:
        cnpj  = _so_digitos(raw_cnpj)
        senha = str(raw_senha or "").strip()

        if not cnpj and not senha:
            continue
        if not _cnpj_valido(cnpj):
            invalidas.append({"linha": num, "motivo": "CNPJ inválido (esperado 14 dígitos)",
                              "conteudo": str(raw_cnpj)})
            continue
        if not senha:
            invalidas.append({"linha": num, "motivo": "Senha ausente", "conteudo": cnpj})
            continue
        empresas.append({"cnpj": cnpj, "senha": senha, "linha": num})

    log.info(f"Planilha lida: {len(empresas)} empresa(s) válida(s), {len(invalidas)} inválida(s).")
    return {"empresas": empresas, "invalidas": invalidas}


def _ler_xlsx(caminho: str):
    """Gera (numero_linha, (cnpj, senha)). Pula header se 1ª linha não tiver CNPJ numérico."""
    from openpyxl import load_workbook
    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        col_a = row[0] if len(row) > 0 else None
        col_b = row[1] if len(row) > 1 else None
        if i == 1 and not _so_digitos(col_a):
            continue
        out.append((i, (col_a, col_b)))
    wb.close()
    return out


def _ler_txt(caminho: str):
    """Gera (numero_linha, (cnpj, senha)). Separador ';' ou TAB."""
    out = []
    with open(caminho, "r", encoding="utf-8-sig", errors="replace") as f:
        for i, linha in enumerate(f, start=1):
            linha = linha.rstrip("\n").rstrip("\r")
            if not linha.strip() or linha.strip().startswith("#"):
                continue
            partes = linha.split(";") if ";" in linha else linha.split("\t")
            cnpj  = partes[0] if len(partes) > 0 else ""
            senha = partes[1] if len(partes) > 1 else ""
            if i == 1 and not _so_digitos(cnpj):
                continue
            out.append((i, (cnpj, senha)))
    return out
