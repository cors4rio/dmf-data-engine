"""
sm_engine/sm_resumo.py — Gera resumo consolidado do lote em Excel.

Colunas: CNPJ | Status | Qtd Emitidas | Qtd Recebidas | Arquivo Emitidas | Arquivo Recebidas | Detalhe
"""
import os
import logging
from datetime import datetime

log = logging.getLogger("SemMovimento.Resumo")


def gerar(resultados: list, mes: int, ano: int, pasta_destino: str) -> str:
    """
    resultados: lista de dicts retornados por empresa_cb:
        {cnpj, status, emitidas:{arquivo,qtd}, recebidas:{arquivo,qtd}, detalhe}

    Retorna o caminho do .xlsx gerado.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    # Cabeçalho
    cabecalho = ["CNPJ", "Status", "Qtd Emitidas", "Qtd Recebidas",
                 "Arquivo Emitidas", "Arquivo Recebidas", "Detalhe"]
    ws.append(cabecalho)

    # Estilo cabeçalho
    hdr_fill = PatternFill("solid", fgColor="2E4A7A")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    # Cores de status
    cores = {
        "ok":             "D4EDDA",
        "erro":           "F8D7DA",
        "captcha_falhou": "FFF3CD",
        "pendente":       "E2E3E5",
    }

    for r in resultados:
        cnpj       = r.get("cnpj", "")
        status     = r.get("status", "")
        emit       = r.get("emitidas",  {})
        receb      = r.get("recebidas", {})
        detalhe    = r.get("detalhe", "")
        arq_emit   = emit.get("arquivo") or ""
        arq_receb  = receb.get("arquivo") or ""
        qtd_emit   = emit.get("qtd", 0)
        qtd_receb  = receb.get("qtd", 0)

        # CNPJ formatado
        cnpj_fmt = _fmt_cnpj(cnpj)

        linha = [cnpj_fmt, status, qtd_emit, qtd_receb,
                 os.path.basename(arq_emit),
                 os.path.basename(arq_receb),
                 detalhe]
        ws.append(linha)

        # Cor por status
        cor = cores.get(status, "FFFFFF")
        fill = PatternFill("solid", fgColor=cor)
        for cell in ws[ws.max_row]:
            cell.fill = fill

    # Larguras de coluna
    larguras = [20, 16, 14, 14, 40, 40, 50]
    for i, larg in enumerate(larguras, start=1):
        ws.column_dimensions[_col_letter(i)].width = larg

    nome = f"resumo_sem_movimento_{mes:02d}{ano}.xlsx"
    caminho = os.path.join(pasta_destino, nome)
    wb.save(caminho)
    log.info(f"Resumo salvo: {caminho}")
    return caminho


def _fmt_cnpj(cnpj: str) -> str:
    d = cnpj.replace(".", "").replace("/", "").replace("-", "")
    if len(d) == 14:
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
    return cnpj


def _col_letter(n: int) -> str:
    """Converte número de coluna (1-based) para letra (A, B, ..., Z, AA, ...)."""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s
