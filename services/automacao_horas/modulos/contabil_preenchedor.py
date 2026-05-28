"""
Preenchedor da planilha HORAS CONTABEIS.xlsx (Fase 2 do fluxo contábil).

Conforme Spec_Contabil.md:
- Escreve EXCLUSIVAMENTE nas colunas F (QTD LANÇAMENTOS), I (TEM FOLHA?) e O (TOTAL FATURAMENTO MÊS).
- Aplica validação tripla (codi_emp + CNPJ + nome) com mínimo 2/3 antes de escrever.
- Gera relatório .md de auditoria com anomalias.
- NÃO altera nenhuma outra coluna sob NENHUMA hipótese.
"""

import os
import re
import logging
from datetime import datetime
import openpyxl

from engine.database import db
from engine.onedrive_helper import preparar_para_uso, chamar_com_retry

# Colunas-alvo na planilha HORAS CONTABEIS (1-indexadas para openpyxl)
COL_CODIGO = 1   # A
COL_CNPJ = 3     # C
COL_NOME = 4     # D
COL_LANCAMENTOS = 6  # F  ← escrita
COL_TEM_FOLHA = 9    # I  ← escrita
COL_FATURAMENTO = 15 # O  ← escrita


def _normalizar_cnpj(valor):
    if valor is None:
        return ""
    return re.sub(r"\D", "", str(valor))


def _normalizar_nome(valor):
    if valor is None:
        return ""
    nome = str(valor).strip().upper()
    nome = re.sub(r"\(.*?\)", "", nome)
    for sufixo in [" LTDA", " S.A.", " S/A", " ME", " EPP", " SCP", "."]:
        nome = nome.replace(sufixo, "")
    return nome.strip()


def _nome_aba_competencia(data_inicio):
    """'2026-05-01' → '05.2026'."""
    return f"{data_inicio[5:7]}.{data_inicio[:4]}"


def _consultar_cadastros():
    """Query 4.1 da Spec — base para validação tripla."""
    rows = db.fetch_all("SELECT codi_emp, cgce_emp, nome_emp FROM bethadba.geempre")
    cad = {}
    for r in rows:
        cod = r.get("codi_emp")
        if cod is None:
            continue
        try:
            chave = str(int(float(str(cod))))
        except (ValueError, TypeError):
            continue
        cad[chave] = {
            "cnpj": _normalizar_cnpj(r.get("cgce_emp")),
            "nome": _normalizar_nome(r.get("nome_emp")),
        }
    return cad


def _consultar_lancamentos(data_inicio, data_fim):
    """Query 4.2 — F: QTD LANÇAMENTOS CONTÁBEIS."""
    sql = """
    SELECT codi_emp, COUNT(*) AS qtd
    FROM bethadba.ctlancto
    WHERE data_lan >= ? AND data_lan <= ?
      AND orig_lan IN (1, 39)
    GROUP BY codi_emp
    """
    out = {}
    for r in db.fetch_all(sql, (data_inicio, data_fim)):
        cod = r.get("codi_emp")
        if cod is None:
            continue
        try:
            chave = str(int(float(str(cod))))
            out[chave] = int(r.get("qtd") or 0)
        except (ValueError, TypeError):
            continue
    return out


def _consultar_faturamento(data_inicio, data_fim):
    """Query 4.3 — O: TOTAL FATURAMENTO MÊS (efsaidas + efservicos)."""
    sql = """
    SELECT codi_emp, SUM(total_contabil) AS faturamento
    FROM (
        SELECT codi_emp, SUM(vcon_sai) AS total_contabil
        FROM bethadba.efsaidas
        WHERE dsai_sai >= ? AND dsai_sai <= ?
        GROUP BY codi_emp
        UNION ALL
        SELECT codi_emp, SUM(vcon_ser) AS total_contabil
        FROM bethadba.efservicos
        WHERE dser_ser >= ? AND dser_ser <= ?
        GROUP BY codi_emp
    ) base
    GROUP BY codi_emp
    """
    out = {}
    for r in db.fetch_all(sql, (data_inicio, data_fim, data_inicio, data_fim)):
        cod = r.get("codi_emp")
        if cod is None:
            continue
        try:
            chave = str(int(float(str(cod))))
            out[chave] = float(r.get("faturamento") or 0.0)
        except (ValueError, TypeError):
            continue
    return out


def _consultar_tem_folha(data_inicio, data_fim):
    """Query 4.4 — I: TEM FOLHA? (foempregados cruzado com forescisoes)."""
    sql = """
    SELECT e.codi_emp,
           SUM(CASE WHEN e.vinculo IN (1, 6, 11) THEN 1 ELSE 0 END) AS total
    FROM bethadba.foempregados e
    LEFT JOIN bethadba.forescisoes r
        ON r.codi_emp = e.codi_emp
       AND r.i_empregados = e.i_empregados
       AND r.demissao < ?
    WHERE e.admissao <= ?
      AND r.i_empregados IS NULL
    GROUP BY e.codi_emp
    """
    out = {}
    for r in db.fetch_all(sql, (data_inicio, data_fim)):
        cod = r.get("codi_emp")
        if cod is None:
            continue
        try:
            chave = str(int(float(str(cod))))
            out[chave] = int(r.get("total") or 0) > 0
        except (ValueError, TypeError):
            continue
    return out


def _validacao_tripla(codigo_planilha, cnpj_planilha, nome_planilha, cadastro):
    """Retorna (score, divergencias) onde score = 0..3 campos batidos."""
    ref = cadastro.get(codigo_planilha)
    if not ref:
        return 0, ["codigo_inexistente_no_dominio"]
    divergencias = []
    score = 1  # código já bate (pelo menos a chave existe no cadastro)
    cnpj_p = _normalizar_cnpj(cnpj_planilha)
    if cnpj_p and ref["cnpj"] and cnpj_p == ref["cnpj"]:
        score += 1
    else:
        divergencias.append(f"cnpj({cnpj_p!r} vs {ref['cnpj']!r})")
    nome_p = _normalizar_nome(nome_planilha)
    if nome_p and ref["nome"]:
        if nome_p == ref["nome"] or nome_p in ref["nome"] or ref["nome"] in nome_p:
            score += 1
        else:
            divergencias.append(f"nome({nome_p!r} vs {ref['nome']!r})")
    else:
        divergencias.append("nome_vazio")
    return score, divergencias


def _gerar_relatorio(caminho_md, data_inicio, data_fim, aba, stats, anomalias):
    with open(caminho_md, "w", encoding="utf-8") as f:
        f.write(f"# Relatório Contábil — Aba {aba}\n\n")
        f.write(f"**Período:** {data_inicio} a {data_fim}\n")
        f.write(f"**Execução:** {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n\n")
        f.write("## Resumo\n")
        f.write(f"- Linhas processadas: **{stats['total']}**\n")
        f.write(f"- Aprovadas (3/3): **{stats['ok']}**\n")
        f.write(f"- Match parcial (2/3): **{stats['parcial']}**\n")
        f.write(f"- Rejeitadas (0-1/3): **{stats['rejeitadas']}**\n")
        f.write(f"- F/I/O preenchidas: **{stats['escritas']}**\n\n")
        if anomalias:
            f.write("## Anomalias\n\n")
            f.write("| Linha | Código | CNPJ | Nome | Score | Divergências |\n")
            f.write("|-------|--------|------|------|-------|--------------|\n")
            for a in anomalias:
                f.write(
                    f"| {a['linha']} | {a['codigo']} | {a['cnpj']} | "
                    f"{a['nome'][:40]} | {a['score']}/3 | {'; '.join(a['divergencias'])} |\n"
                )


def preencher_horas_contabeis(caminho_planilha, data_inicio, data_fim, raiz_relatorios="."):
    """
    Fase 2 do fluxo contábil. Recebe o caminho da `HORAS CONTABEIS.xlsx`,
    a competência, e popula F/I/O na aba correspondente.

    Retorna dict:
      {ok, escritas, total, ok_count, parcial, rejeitadas, relatorio, aba}
    """
    if not os.path.exists(caminho_planilha):
        return {"ok": False, "erro": f"Arquivo não encontrado: {caminho_planilha}"}

    aba_alvo = _nome_aba_competencia(data_inicio)
    logging.info(f"[CONTÁBIL-PRE] Iniciando preenchimento. Arquivo={caminho_planilha} Aba={aba_alvo}")

    ok_arq, motivo = preparar_para_uso(caminho_planilha)
    if not ok_arq:
        if motivo == "arquivo_aberto_no_excel":
            return {"ok": False, "tipo": "locked",
                    "erro": "Planilha está aberta no Excel. Feche e tente novamente."}
        if motivo == "online_only_sem_download":
            return {"ok": False, "tipo": "online_only",
                    "erro": "Arquivo está como 'somente online' no OneDrive e não pôde ser baixado. "
                            "Clique com botão direito → 'Sempre manter neste dispositivo' e tente de novo."}
        return {"ok": False, "erro": f"Falha de pré-checagem: {motivo}"}

    try:
        wb = chamar_com_retry(
            lambda: openpyxl.load_workbook(caminho_planilha),
            tentativas=4, intervalo_inicial=0.6,
            descricao="load_workbook(contabil)",
        )
    except Exception as e:
        return {"ok": False, "erro": f"Falha ao abrir planilha: {e}"}

    if aba_alvo not in wb.sheetnames:
        wb.close()
        return {"ok": False, "erro": f"Aba '{aba_alvo}' não existe. Abas disponíveis: {wb.sheetnames}"}

    ws = wb[aba_alvo]

    # Consultas no Domínio
    cadastros = _consultar_cadastros()
    if not cadastros:
        wb.close()
        return {"ok": False, "erro": "Não foi possível obter cadastros do Domínio (geempre vazio)."}

    lancamentos = _consultar_lancamentos(data_inicio, data_fim)
    faturamento = _consultar_faturamento(data_inicio, data_fim)
    tem_folha = _consultar_tem_folha(data_inicio, data_fim)

    stats = {"total": 0, "ok": 0, "parcial": 0, "rejeitadas": 0, "escritas": 0}
    anomalias = []

    for r in range(2, ws.max_row + 1):
        cod = ws.cell(row=r, column=COL_CODIGO).value
        if cod is None:
            continue
        try:
            codigo = str(int(float(str(cod).strip())))
        except (ValueError, TypeError):
            continue
        stats["total"] += 1
        cnpj = ws.cell(row=r, column=COL_CNPJ).value
        nome = ws.cell(row=r, column=COL_NOME).value

        score, div = _validacao_tripla(codigo, cnpj, nome, cadastros)
        if score >= 3:
            stats["ok"] += 1
        elif score == 2:
            stats["parcial"] += 1
            anomalias.append({"linha": r, "codigo": codigo, "cnpj": cnpj or "",
                              "nome": str(nome or ""), "score": score, "divergencias": div})
        else:
            stats["rejeitadas"] += 1
            anomalias.append({"linha": r, "codigo": codigo, "cnpj": cnpj or "",
                              "nome": str(nome or ""), "score": score, "divergencias": div})
            continue  # 0/3 ou 1/3 → não escreve

        # Escrita SOMENTE em F, I, O
        ws.cell(row=r, column=COL_LANCAMENTOS).value = int(lancamentos.get(codigo, 0))
        ws.cell(row=r, column=COL_TEM_FOLHA).value = "SIM" if tem_folha.get(codigo, False) else "NAO"
        ws.cell(row=r, column=COL_FATURAMENTO).value = float(faturamento.get(codigo, 0.0))
        stats["escritas"] += 1

    try:
        chamar_com_retry(
            lambda: wb.save(caminho_planilha),
            tentativas=4, intervalo_inicial=0.8,
            descricao="wb.save(contabil)",
        )
    except PermissionError:
        wb.close()
        return {"ok": False, "tipo": "locked",
                "erro": "Permissão negada ao salvar. Planilha em uso ou cliente OneDrive bloqueando."}
    except Exception as e:
        wb.close()
        return {"ok": False, "erro": f"Falha ao salvar: {e}"}
    finally:
        try: wb.close()
        except Exception: pass

    relatorio = os.path.join(raiz_relatorios, f"relatorio_contabil_{aba_alvo}.md")
    try:
        _gerar_relatorio(relatorio, data_inicio, data_fim, aba_alvo, stats, anomalias)
    except Exception as e:
        logging.error(f"[CONTÁBIL-PRE] Falha ao gerar relatório: {e}")
        relatorio = None

    logging.info(
        f"[CONTÁBIL-PRE] Concluído. Aba={aba_alvo} · "
        f"{stats['escritas']}/{stats['total']} escritas · "
        f"{stats['parcial']} parciais · {stats['rejeitadas']} rejeitadas."
    )
    return {
        "ok": True,
        "aba": aba_alvo,
        "total": stats["total"],
        "escritas": stats["escritas"],
        "ok_count": stats["ok"],
        "parcial": stats["parcial"],
        "rejeitadas": stats["rejeitadas"],
        "relatorio": relatorio,
    }
