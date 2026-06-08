import os
import logging
import openpyxl
import xlrd

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class ExcelParser:
    """
    Motor resiliente para leitura de planilhas locais.
    Absorve falhas como arquivo inexistente, travado, e converte formatos antigos (.xls).
    """
    
    @staticmethod
    def ler_planilha_carol(caminho):
        """
        Lê a planilha de Controle de Empregados em qualquer formato Excel suportado:
        .xls (xlrd), .xlsx (openpyxl), .xlsm (openpyxl).
        Retorna um dicionário indexado pelo Código Domínio (codi_emp).
        """
        if not os.path.exists(caminho):
            logging.warning(f"[DP] Planilha da Carol não encontrada em: {caminho}")
            return None

        ext = os.path.splitext(caminho)[1].lower()
        logging.info(f"[DP] Lendo Planilha da Carol ({ext}): {caminho}")

        try:
            if ext == ".xls":
                return ExcelParser._ler_carol_xls(caminho)
            elif ext in (".xlsx", ".xlsm"):
                return ExcelParser._ler_carol_xlsx(caminho)
            else:
                logging.error(f"[DP] Formato de planilha Carol não suportado: '{ext}'. "
                              f"Use .xls, .xlsx ou .xlsm.")
                return None
        except Exception as e:
            logging.error(f"[DP] Erro ao processar planilha da Carol ({ext}): {e}")
            return None

    @staticmethod
    def _ler_carol_xls(caminho):
        """Backend xlrd para .xls (Excel 97-2003)."""
        wb = xlrd.open_workbook(caminho)
        sheet = wb.sheet_by_index(0)

        dados = {}
        # Ignora as 2 primeiras linhas (cabeçalhos)
        for rowx in range(2, sheet.nrows):
            row = sheet.row_values(rowx)
            # Na planilha, a coluna 0 está vazia, o código está na coluna 1
            if not row or len(row) < 2 or not row[1]:
                continue
            try:
                cod_str = str(int(float(str(row[1]).strip())))
                funcionarios = int(float(row[7] or 0)) if len(row) > 7 and str(row[7]).strip() else 0
                estagiarios = int(float(row[9] or 0)) if len(row) > 9 and str(row[9]).strip() else 0
                contribuintes = int(float(row[11] or 0)) if len(row) > 11 and str(row[11]).strip() else 0
                total_ativos = funcionarios + estagiarios + contribuintes

                dados[cod_str] = {
                    'func': funcionarios,
                    'estag': estagiarios,
                    'contrib': contribuintes,
                    'total_ativos': total_ativos
                }
            except Exception:
                continue  # Pula cabeçalhos repetidos ou rodapés
        return dados

    @staticmethod
    def _ler_carol_xlsx(caminho):
        """Backend openpyxl para .xlsx e .xlsm (formato moderno).
        Mesmas posições de coluna do .xls: idx 1=código, 7=func, 9=estag, 11=contrib.
        """
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        try:
            sheet = wb.active
            dados = {}

            def _num(row, idx):
                """Converte célula em int >= 0. Tolera None, string vazia, decimais."""
                if idx >= len(row) or row[idx] is None:
                    return 0
                s = str(row[idx]).strip()
                if not s:
                    return 0
                try:
                    return int(float(s))
                except (ValueError, TypeError):
                    return 0

            # iter_rows com min_row=3 = ignora linhas 1-2 (cabeçalhos), igual ao .xls
            for row in sheet.iter_rows(min_row=3, values_only=True):
                if not row or len(row) < 2 or row[1] in (None, ""):
                    continue
                try:
                    cod_str = str(int(float(str(row[1]).strip())))
                    funcionarios = _num(row, 7)
                    estagiarios = _num(row, 9)
                    contribuintes = _num(row, 11)
                    total_ativos = funcionarios + estagiarios + contribuintes

                    dados[cod_str] = {
                        'func': funcionarios,
                        'estag': estagiarios,
                        'contrib': contribuintes,
                        'total_ativos': total_ativos
                    }
                except Exception:
                    continue
            return dados
        finally:
            wb.close()

    @staticmethod
    def ler_pre_planilha_contabil(caminho):
        """
        Lê a pré-planilha Contábil (.xlsx).
        Retorna dicionário { 'cod_emp': horas_float } extraído da coluna R.
        """
        if not os.path.exists(caminho):
            logging.warning(f"[CONTÁBIL] Pré-planilha Contábil não encontrada em: {caminho}")
            return None
            
        logging.info(f"[CONTÁBIL] Lendo Pré-planilha Contábil: {caminho}")
        try:
            # data_only=True garante a leitura do valor calculado
            wb = openpyxl.load_workbook(caminho, data_only=True)
            sheet = wb.active
            
            dados = {}
            for row in range(2, sheet.max_row + 1):
                cod = sheet.cell(row=row, column=1).value # Coluna A
                total_horas = sheet.cell(row=row, column=18).value # Coluna R
                
                if cod is not None:
                    try:
                        cod_str = str(int(float(str(cod).strip())))
                        if total_horas is not None:
                            if isinstance(total_horas, str) and "REF!" in total_horas:
                                logging.warning(f"[CONTÁBIL] Valor quebrado (#REF!) na linha {row} para cliente {cod_str}.")
                                continue
                            try:
                                dados[cod_str] = float(total_horas)
                            except ValueError:
                                pass # ignora texto
                    except Exception:
                        pass
            return dados
        except Exception as e:
            logging.error(f"[CONTÁBIL] Erro ao ler pré-planilha contábil: {e}")
            return None
