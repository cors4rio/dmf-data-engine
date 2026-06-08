import openpyxl
import logging
import os
import re
from ah_engine.onedrive_helper import preparar_para_uso, chamar_com_retry, esta_online_only, forcar_download

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class MasterWriter:
    """
    Componente central responsável por ler e escrever de forma segura na Planilha Master.
    Aplica validação dupla (Codi_emp e CNPJ) e formatação correta de horas.
    """
    def __init__(self, caminho_master):
        self.caminho = caminho_master
        self.wb = None
        self.ws = None
        self.row_map = {}
        # Rastreio de linhas efetivamente escritas por coluna (para limpeza de resíduos).
        # Chaves: 15=O (Fiscal), 16=P (Contábil), 17=Q (DP).
        self.linhas_escritas = {15: set(), 16: set(), 17: set()}
        # Última falha de salvamento (consultada pelo orquestrador para UI).
        self.ultimo_erro = None
        
    def carregar(self):
        if not os.path.exists(self.caminho):
            logging.error(f"[MASTER] Planilha Master não encontrada: {self.caminho}")
            self.ultimo_erro = {"tipo": "nao_encontrado",
                                "msg": f"Planilha master não encontrada: {self.caminho}"}
            return False

        # Pipeline OneDrive: bloqueia se estiver aberta no Excel e força download
        # se for placeholder Online-only.
        ok, motivo = preparar_para_uso(self.caminho)
        if not ok:
            if motivo == "arquivo_aberto_no_excel":
                msg = "Planilha está aberta no Excel. Feche o arquivo e tente novamente."
                tipo = "locked"
            elif motivo == "online_only_sem_download":
                msg = ("Arquivo está como 'somente online' no OneDrive e não foi possível "
                       "baixá-lo. Clique com botão direito no arquivo no Explorer → "
                       "'Sempre manter neste dispositivo'.")
                tipo = "online_only"
            else:
                msg = f"Falha de pré-checagem: {motivo}"
                tipo = "erro"
            logging.error(f"[MASTER] {msg}")
            self.ultimo_erro = {"tipo": tipo, "msg": msg}
            return False

        logging.info(f"[MASTER] Carregando Planilha Master na memória: {self.caminho}")
        try:
            self.wb = chamar_com_retry(
                lambda: openpyxl.load_workbook(self.caminho, keep_vba=True),
                tentativas=4, intervalo_inicial=0.6,
                descricao="load_workbook(master)",
            )
            self.ws = self.wb.active
            self._mapear_linhas()
            self.ultimo_erro = None
            return True
        except PermissionError as e:
            logging.error(f"[MASTER] PermissionError ao abrir após retry: {e}")
            self.ultimo_erro = {"tipo": "locked",
                                "msg": "Arquivo bloqueado por outro processo. Feche o Excel e o cliente OneDrive e tente novamente."}
            return False
        except Exception as e:
            logging.error(f"[MASTER] Erro crítico ao abrir Planilha Master: {e}")
            self.ultimo_erro = {"tipo": "erro", "msg": str(e)}
            return False
            
    def _mapear_linhas(self):
        """Mapeia os códigos e CNPJs para as linhas correspondentes."""
        self.row_map = {}
        # Na Master, os dados reais começam na linha 10
        for row in range(10, self.ws.max_row + 1):
            cod = self.ws.cell(row=row, column=8).value  # Coluna H
            cnpj = self.ws.cell(row=row, column=10).value # Coluna J
            
            # Mapeamento Primário (Código)
            if cod is not None:
                try:
                    cod_str = str(int(float(str(cod).strip())))
                    if cod_str not in self.row_map:
                        self.row_map[cod_str] = []
                    self.row_map[cod_str].append(row)
                except ValueError:
                    pass
                    
            # Mapeamento Secundário (CNPJ)
            if cnpj is not None:
                cnpj_limpo = re.sub(r'\D', '', str(cnpj))
                if cnpj_limpo:
                    if cnpj_limpo not in self.row_map:
                        self.row_map[cnpj_limpo] = []
                    # Evitar adicionar a mesma linha 2 vezes na mesma chave
                    if row not in self.row_map[cnpj_limpo]:
                        self.row_map[cnpj_limpo].append(row)
                        
        logging.info(f"[MASTER] Mapeamento concluído. Indexadas linhas seguras.")

    def obter_linhas(self, cod_str=None, cnpj_str=None):
        linhas = []
        if cod_str and cod_str in self.row_map:
            linhas.extend(self.row_map[cod_str])
        elif cnpj_str: # Fallback seguro
            cnpj_limpo = re.sub(r'\D', '', str(cnpj_str))
            if cnpj_limpo in self.row_map:
                linhas.extend(self.row_map[cnpj_limpo])
        return list(set(linhas))

    def preencher_fiscal(self, cod_str, valor_horas, cnpj_str=None):
        """Escreve na Coluna O"""
        self._escrever(cod_str, cnpj_str, 15, valor_horas)

    def preencher_contabil(self, cod_str, valor_horas, cnpj_str=None):
        """Escreve na Coluna P"""
        self._escrever(cod_str, cnpj_str, 16, valor_horas)
        
    def preencher_dp(self, cod_str, valor_horas, cnpj_str=None):
        """Escreve na Coluna Q"""
        self._escrever(cod_str, cnpj_str, 17, valor_horas)

    def _escrever(self, cod_str, cnpj_str, col_idx, valor):
        linhas = self.obter_linhas(cod_str, cnpj_str)
        for r in linhas:
            c = self.ws.cell(row=r, column=col_idx)
            c.value = valor
            if isinstance(valor, (float, int)):
                c.number_format = '[h]:mm:ss'
            if col_idx in self.linhas_escritas:
                self.linhas_escritas[col_idx].add(r)

    def limpar_nao_tocadas(self, colunas):
        """Zera as células das colunas indicadas em linhas com codi_emp que NÃO foram
        escritas neste ciclo. Limpa resíduos de competências anteriores."""
        if not colunas:
            return 0
        total = 0
        for r in range(10, self.ws.max_row + 1):
            cod = self.ws.cell(row=r, column=8).value
            if cod is None:
                continue
            for col_idx in colunas:
                if r in self.linhas_escritas.get(col_idx, set()):
                    continue
                c = self.ws.cell(row=r, column=col_idx)
                c.value = 0
                c.number_format = '[h]:mm:ss'
                total += 1
        logging.info(f"[MASTER] {total} célula(s) zeradas em {colunas} (limpeza de resíduo).")
        return total
                
    @staticmethod
    def _celula_e_numerica(val):
        import datetime
        if val is None:
            return False
        if isinstance(val, (float, int, datetime.time, datetime.timedelta, datetime.datetime)):
            return True
        if isinstance(val, str):
            s = val.strip()
            if not s:
                return False
            if re.match(r'^\d+:\d+:\d+$', s):
                return True
            try:
                float(s)
                return True
            except ValueError:
                return False
        return False

    def recalcular_totais(self):
        """Coluna R = soma das colunas O/P/Q que tiverem valor numérico (evita #VALUE!
        quando alguma coluna contém string como 'NAO FAZ CONTABIL')."""
        max_r = self.ws.max_row
        for r in range(10, max_r + 1):
            cod = self.ws.cell(row=r, column=8).value
            if cod is None:
                continue
            partes = []
            if self._celula_e_numerica(self.ws.cell(row=r, column=15).value):
                partes.append(f"O{r}")
            if self._celula_e_numerica(self.ws.cell(row=r, column=16).value):
                partes.append(f"P{r}")
            if self._celula_e_numerica(self.ws.cell(row=r, column=17).value):
                partes.append(f"Q{r}")
            c = self.ws.cell(row=r, column=18)
            c.value = ("=" + "+".join(partes)) if partes else ""
            c.number_format = '[h]:mm:ss'

        for col_idx, letra in [(15, 'O'), (16, 'P'), (17, 'Q'), (18, 'R')]:
            c = self.ws.cell(row=7, column=col_idx)
            c.value = f"=SUBTOTAL(9,{letra}10:{letra}{max_r})"
            c.number_format = '[h]:mm:ss'

        logging.info("[MASTER] Fórmulas de total recalculadas (blindagem por tipo).")

    def salvar(self):
        """Salva com retry tolerante a locks transitórios do cliente OneDrive."""
        logging.info("[MASTER] Salvando Planilha Master...")
        self.ultimo_erro = None
        try:
            chamar_com_retry(
                lambda: self.wb.save(self.caminho),
                tentativas=4, intervalo_inicial=0.8,
                descricao="wb.save(master)",
            )
            logging.info(f"[MASTER] Salvo com sucesso em: {self.caminho}")
            return True
        except PermissionError:
            logging.error("[MASTER] ERRO CRÍTICO: Planilha aberta no Excel ou bloqueada após retry.")
            self.ultimo_erro = {
                "tipo": "locked",
                "msg": ("Planilha master está aberta no Excel ou bloqueada por outro "
                        "processo (cliente OneDrive pode estar sincronizando). Feche o "
                        "arquivo, aguarde 10s e tente de novo."),
            }
            return False
        except Exception as e:
            logging.error(f"[MASTER] Erro ao salvar: {e}")
            self.ultimo_erro = {"tipo": "erro", "msg": str(e)}
            return False
