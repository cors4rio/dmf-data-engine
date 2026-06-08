"""
Integrador da Contábil para a Master (Fase 5 do fluxo contábil).

Lê a coluna R (Horas Validadas) da `HORAS CONTABEIS.xlsx` para a aba da competência,
e lança na Coluna P da Master via double match (código + CNPJ), conforme Spec_Contabil 8.1.

Pré-requisito: a Fase 2 (contabil_preenchedor) deve ter sido executada — caso contrário a
coluna R provavelmente estará vazia (ou refletindo o mês anterior).
"""

import os
import re
import datetime
import logging
import openpyxl

from ah_modulos.excecoes import ler_codigos_setor
from ah_engine.onedrive_helper import esta_online_only, forcar_download


# Layout da HORAS CONTABEIS (Spec 1.2)
COL_CODIGO_ORIGEM = 1   # A
COL_CNPJ_ORIGEM = 3     # C
COL_HORAS_VALIDADAS = 18  # R (preenchida manualmente pelo supervisor após Fase 2)


def _nome_aba_competencia(data_inicio):
    return f"{data_inicio[5:7]}.{data_inicio[:4]}"


def _norm_cnpj(v):
    return re.sub(r"\D", "", str(v)) if v else ""


def _ler_horas_validadas(caminho, aba):
    """Lê (codigo, cnpj) → horas_float a partir da coluna R."""
    if esta_online_only(caminho):
        forcar_download(caminho, timeout=20)
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    if aba not in wb.sheetnames:
        wb.close()
        return None, f"Aba '{aba}' não encontrada. Disponíveis: {wb.sheetnames}"
    sh = wb[aba]
    dados = {}
    for row in sh.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < COL_HORAS_VALIDADAS:
            continue
        cod = row[COL_CODIGO_ORIGEM - 1]
        cnpj = row[COL_CNPJ_ORIGEM - 1] if len(row) >= COL_CNPJ_ORIGEM else None
        horas = row[COL_HORAS_VALIDADAS - 1]
        if cod is None:
            continue
        try:
            cod_str = str(int(float(str(cod).strip())))
        except (ValueError, TypeError):
            continue
        horas_float = None
        if isinstance(horas, datetime.timedelta):
            horas_float = horas.total_seconds() / 86400.0
        elif isinstance(horas, (float, int)):
            horas_float = float(horas)
        if horas_float is None or horas_float == 0:
            continue
        dados[cod_str] = {"horas": horas_float, "cnpj": _norm_cnpj(cnpj)}
    wb.close()
    return dados, None


def injetar_contabil_na_master(writer, caminho_origem, data_inicio, data_fim):
    """
    Fase 5. Aplica horas validadas na Coluna P da master via double match.
    """
    if not os.path.exists(caminho_origem):
        return {"ok": False, "erro": f"Planilha origem não encontrada: {caminho_origem}"}

    aba = _nome_aba_competencia(data_inicio)
    logging.info(f"[CONTÁBIL-INT] Lançando na Master. Origem={caminho_origem} Aba={aba}")

    dados, erro = _ler_horas_validadas(caminho_origem, aba)
    if dados is None:
        return {"ok": False, "erro": erro}
    if not dados:
        return {"ok": False, "erro": "Nenhuma hora validada encontrada na coluna R."}

    excecoes = ler_codigos_setor("contabil")
    if excecoes:
        logging.info(f"[CONTÁBIL-INT] {len(excecoes)} código(s) na lista CONTABIL_NAO serão ignorados.")

    gravadas = 0
    sem_match = 0
    excluidas_excecao = 0
    for cod_str, info in dados.items():
        if cod_str in excecoes:
            excluidas_excecao += 1
            continue
        cnpj = info.get("cnpj") or None
        linhas = writer.obter_linhas(cod_str, cnpj)
        if not linhas:
            sem_match += 1
            continue
        # Double match: confere se a linha encontrada bate código E CNPJ
        for r in linhas:
            cod_master = writer.ws.cell(row=r, column=8).value
            cnpj_master = _norm_cnpj(writer.ws.cell(row=r, column=10).value)
            try:
                cod_master_str = str(int(float(str(cod_master).strip())))
            except (ValueError, TypeError):
                cod_master_str = None
            confere_codigo = cod_master_str == cod_str
            confere_cnpj = cnpj and cnpj_master and cnpj == cnpj_master
            if confere_codigo and (not cnpj or confere_cnpj):
                writer.preencher_contabil(cod_str, info["horas"])
                gravadas += 1

    logging.info(
        f"[CONTÁBIL-INT] Concluído. {gravadas} linha(s) gravadas em P · "
        f"{sem_match} sem match · {excluidas_excecao} excluídas por CONTABIL_NAO · "
        f"fonte={caminho_origem}."
    )
    return {"ok": True, "gravadas": gravadas, "sem_match": sem_match,
            "excluidas_excecao": excluidas_excecao, "aba": aba}


# Alias para compatibilidade enquanto o orquestrador ainda referencia o nome antigo.
def extrair_e_preencher_contabil(writer, *args, **kwargs):
    logging.warning(
        "[CONTÁBIL-INT] extrair_e_preencher_contabil() foi descontinuado. "
        "Use o fluxo em 2 etapas: contabil_preenchedor.preencher_horas_contabeis() "
        "seguido de contabil_integrador.injetar_contabil_na_master()."
    )
    return False
