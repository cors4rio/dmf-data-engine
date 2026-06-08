"""
tf_engine/tf_planilha.py — Leitura do lote de clientes TFF.

Colunas esperadas (ordem fixa, sem cabeçalho obrigatório):
  A: Razão Social
  B: CNPJ
  C: Município
  D: CGA

Aceita .xlsx/.xlsm ou .txt (separador ';' ou TAB).
O CGA é normalizado para apenas dígitos (remove pontos, traços, barras, etc.),
conforme instrução da prefeitura ("não utilize separador entre o número e o dígito").
"""
import os
import re
import logging

log = logging.getLogger("TffSalvador.Planilha")


def _so_digitos(valor) -> str:
    return re.sub(r"\D", "", str(valor or ""))


def _cga_valido(cga: str) -> bool:
    return len(cga) >= 1


def carregar(caminho: str) -> dict:
    """
    Retorna {"clientes": [{"razao_social","cnpj","municipio","cga","linha"}],
             "invalidos": [{"linha","motivo","conteudo"}]}.
    Lança ValueError se o arquivo não existir ou a extensão não for suportada.
    """
    if not caminho or not os.path.isfile(caminho):
        raise ValueError(f"Arquivo não encontrado: {caminho}")

    ext = os.path.splitext(caminho)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        linhas = _ler_xlsx(caminho)
    elif ext == ".txt":
        linhas = _ler_txt(caminho)
    else:
        raise ValueError(f"Extensão não suportada: {ext} (use .txt, .xlsx)")

    clientes, invalidos = [], []
    for num, (raw_razao, raw_cnpj, raw_municipio, raw_cga) in linhas:
        razao     = str(raw_razao or "").strip()
        cnpj      = str(raw_cnpj or "").strip()
        municipio = str(raw_municipio or "").strip()
        cga       = _so_digitos(raw_cga)  # remove qualquer caractere especial

        if not cga and not razao:
            continue

        if not _cga_valido(cga):
            invalidos.append({
                "linha":    num,
                "motivo":   "CGA inválido (deve conter ao menos 1 dígito)",
                "conteudo": str(raw_cga),
            })
            continue

        clientes.append({
            "razao_social": razao,
            "cnpj":         cnpj,
            "municipio":    municipio,
            "cga":          cga,
            "linha":        num,
        })

    log.info(f"Planilha lida: {len(clientes)} cliente(s) válido(s), {len(invalidos)} inválido(s).")
    return {"clientes": clientes, "invalidos": invalidos}


def _ler_xlsx(caminho: str):
    """Gera (numero_linha, (razao, cnpj, municipio, cga)). Pula header se col A for texto puro."""
    from openpyxl import load_workbook
    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        col_a = row[0] if len(row) > 0 else None
        col_b = row[1] if len(row) > 1 else None
        col_c = row[2] if len(row) > 2 else None
        col_d = row[3] if len(row) > 3 else None
        # Pula linha de cabeçalho: col D (CGA) sem nenhum dígito
        if i == 1 and not _so_digitos(col_d):
            continue
        out.append((i, (col_a, col_b, col_c, col_d)))
    wb.close()
    return out


def _ler_txt(caminho: str):
    """Gera (numero_linha, (razao, cnpj, municipio, cga)). Separador ';' ou TAB."""
    out = []
    with open(caminho, "r", encoding="utf-8-sig", errors="replace") as f:
        for i, linha in enumerate(f, start=1):
            linha = linha.rstrip("\n").rstrip("\r")
            if not linha.strip() or linha.strip().startswith("#"):
                continue
            partes    = linha.split(";") if ";" in linha else linha.split("\t")
            razao     = partes[0] if len(partes) > 0 else ""
            cnpj      = partes[1] if len(partes) > 1 else ""
            municipio = partes[2] if len(partes) > 2 else ""
            cga       = partes[3] if len(partes) > 3 else ""
            # Pula header: col D sem dígito
            if i == 1 and not _so_digitos(cga):
                continue
            out.append((i, (razao, cnpj, municipio, cga)))
    return out
