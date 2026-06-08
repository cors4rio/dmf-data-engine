"""
tf_engine/tf_resumo.py — Gera resumo consolidado do lote TFF em Excel.

Colunas: Razão Social | CNPJ | Município | CGA | Status | Cotas OK | Arquivos | Detalhe
"""
import os
import logging

log = logging.getLogger("TffSalvador.Resumo")


def gerar(resultados: list, ano: int, pasta_destino: str) -> str:
    """
    resultados: lista de dicts retornados por cliente_cb:
        {cga, razao_social, cnpj, municipio, status, guias:[{nome,arquivo}], detalhe}

    Retorna o caminho do .xlsx gerado.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo TFF"

    cabecalho = ["Razão Social", "CNPJ", "Município", "CGA",
                 "Status", "Cotas OK", "Arquivos", "Detalhe"]
    ws.append(cabecalho)

    hdr_fill = PatternFill("solid", fgColor="4A3080")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    cores = {
        "ok":          "D4EDDA",
        "parcial":     "FFF3CD",
        "pago":        "D4EDDA",
        "sem_divida":  "E2E3E5",
        "nao_existe":  "F8D7DA",
        "captcha_erro":"FFE8B2",
        "erro":        "F8D7DA",
    }

    for r in resultados:
        razao     = r.get("razao_social", "")
        cnpj      = r.get("cnpj", "")
        municipio = r.get("municipio", "")
        cga       = r.get("cga", "")
        status    = r.get("status", "")
        guias     = r.get("guias", [])
        detalhe   = r.get("detalhe", "")

        cotas_ok  = sum(1 for g in guias if g.get("arquivo"))
        arquivos  = "; ".join(
            os.path.basename(g["arquivo"]) for g in guias if g.get("arquivo")
        )

        linha = [razao, cnpj, municipio, cga, status, cotas_ok, arquivos, detalhe]
        ws.append(linha)

        cor  = cores.get(status, "FFFFFF")
        fill = PatternFill("solid", fgColor=cor)
        for cell in ws[ws.max_row]:
            cell.fill = fill

    larguras = [40, 20, 20, 14, 14, 10, 70, 50]
    for i, larg in enumerate(larguras, start=1):
        ws.column_dimensions[_col_letter(i)].width = larg

    nome    = f"resumo_tff_{ano}.xlsx"
    caminho = os.path.join(pasta_destino, nome)
    wb.save(caminho)
    log.info(f"Resumo salvo: {caminho}")
    return caminho


def _col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s
